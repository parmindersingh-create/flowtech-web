from fastapi import FastAPI, HTTPException, Request, Response, Depends, UploadFile, File, Body, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from motor.motor_asyncio import AsyncIOMotorClient
from zoneinfo import ZoneInfo
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import os
import sys
import uuid
import httpx
import asyncio
import logging
import io
import re
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from ai_audit import AIAuditEngine

# Import modular routers
from routes.moulds import router as moulds_router
from routes.cnc import router as cnc_router
from routes.salary import router as salary_router
from routes.leave import router as leave_router
from routes.gauges import router as gauges_router
from routes.attendance import router as attendance_router
from utils.database import set_db as set_shared_db

# Configure logging for production
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    stream=sys.stdout
)
logger = logging.getLogger(__name__)

load_dotenv()

# IST Timezone
IST = ZoneInfo("Asia/Kolkata")

def get_ist_now():
    """Get current datetime in IST timezone"""
    return datetime.now(IST)

def to_ist(dt):
    """Convert a datetime to IST timezone"""
    if dt is None:
        return None
    if isinstance(dt, str):
        dt = datetime.fromisoformat(dt.replace('Z', '+00:00'))
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(IST)

def format_ist_datetime(dt):
    """Format datetime as IST string"""
    ist_dt = to_ist(dt)
    if ist_dt is None:
        return None
    return ist_dt.strftime('%d/%m/%Y %I:%M %p')

# Shift Management Functions
def get_current_shift():
    """Get current shift based on IST time"""
    now = get_ist_now()
    hour = now.hour
    
    # Morning shift: 8 AM to 8 PM (8:00 - 19:59)
    # Night shift: 8 PM to 6 AM (20:00 - 05:59)
    if 8 <= hour < 20:
        return "morning"
    else:
        return "night"

def get_shift_end_time():
    """Get the end time of current shift in IST"""
    now = get_ist_now()
    hour = now.hour
    
    if 8 <= hour < 20:
        # Morning shift ends at 8 PM today
        shift_end = now.replace(hour=20, minute=0, second=0, microsecond=0)
    else:
        # Night shift ends at 6 AM next day
        if hour >= 20:
            # After 8 PM, shift ends at 6 AM next day
            shift_end = (now + timedelta(days=1)).replace(hour=6, minute=0, second=0, microsecond=0)
        else:
            # Before 6 AM, shift ends at 6 AM today
            shift_end = now.replace(hour=6, minute=0, second=0, microsecond=0)
    
    return shift_end

# Early stub for get_shift_for_time - full implementation in attendance section
def get_shift_for_time(dt):
    """Get shift config dict for a given datetime - early stub"""
    if dt is None:
        return {"name": "Unknown", "start": "08:00", "end": "20:00", "late_grace_minutes": 15}
    try:
        ist_dt = to_ist(dt)
        hour = ist_dt.hour
    except:
        hour = 12
    if 5 <= hour < 17:
        return {"name": "Day Shift", "start": "08:00", "end": "20:00", "late_grace_minutes": 15}
    return {"name": "Night Shift", "start": "20:00", "end": "06:00", "late_grace_minutes": 15}

# Log startup
logger.info("Starting Shop Floor Track Backend...")
logger.info(f"Python version: {sys.version}")

# Initialize APScheduler for background CRON tasks
scheduler = AsyncIOScheduler(timezone="Asia/Kolkata")

app = FastAPI()

# Include modular routers
app.include_router(moulds_router)
app.include_router(cnc_router)
app.include_router(salary_router)
app.include_router(leave_router)
app.include_router(gauges_router)
app.include_router(attendance_router)

# New username/password auth (v2) — overrides Google OAuth on the web
from auth_v2 import router as auth_v2_router
app.include_router(auth_v2_router)

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== APP VERSION TRACKING ====================
# This middleware tracks app versions and can optionally block old versions
# Version blocking is DISABLED by default - must be enabled in settings

def compare_versions(version1: str, version2: str) -> int:
    """Compare two version strings. Returns -1 if v1 < v2, 0 if equal, 1 if v1 > v2"""
    try:
        v1_parts = [int(x) for x in version1.split('.')]
        v2_parts = [int(x) for x in version2.split('.')]
        # Pad shorter version with zeros
        while len(v1_parts) < len(v2_parts):
            v1_parts.append(0)
        while len(v2_parts) < len(v1_parts):
            v2_parts.append(0)
        for i in range(len(v1_parts)):
            if v1_parts[i] < v2_parts[i]:
                return -1
            elif v1_parts[i] > v2_parts[i]:
                return 1
        return 0
    except:
        return 0

@app.middleware("http")
async def version_tracking_middleware(request: Request, call_next):
    """Track app version from requests and optionally block old versions"""
    # Skip version check for non-API routes, health checks, and static files
    path = request.url.path
    if not path.startswith("/api/") or path == "/api/health" or path.startswith("/api/public"):
        return await call_next(request)
    
    # Extract version info from headers
    app_version = request.headers.get("X-App-Version", "unknown")
    app_platform = request.headers.get("X-App-Platform", "unknown")
    
    # Check if version blocking is enabled (check settings collection)
    # Only do this check if db is available
    try:
        if db is not None:
            version_settings = await db.settings.find_one({"key": "app_version_control"})
            if version_settings and version_settings.get("blocking_enabled", False):
                min_version = version_settings.get("minimum_version", "0.0.0")
                if app_version != "unknown" and compare_versions(app_version, min_version) < 0:
                    # Version is too old - block request
                    return Response(
                        content=json.dumps({
                            "detail": f"App version {app_version} is outdated. Please update to version {min_version} or later.",
                            "error_code": "APP_VERSION_OUTDATED",
                            "minimum_version": min_version,
                            "current_version": app_version
                        }),
                        status_code=426,  # Upgrade Required
                        media_type="application/json"
                    )
    except Exception:
        # Don't block requests if settings check fails - just log and continue
        pass
    
    # Continue with request
    response = await call_next(request)
    return response

# ==================== END VERSION TRACKING ====================


# Health check endpoint - must respond quickly for deployment checks
@app.get("/api/health")
@app.head("/api/health")
async def health_check():
    """Simple health check endpoint for deployment verification"""
    return {"status": "healthy", "service": "shop-floor-track-backend", "timestamp": datetime.now(timezone.utc).isoformat()}

# Root endpoint - redirect to web app or handle OAuth callback
@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    """
    Handle root endpoint:
    1. If session_id in hash: Process OAuth callback
    2. Otherwise: Serve Expo web app (redirect or proxy)
    """
    # Check if this is an OAuth callback (has session_id in URL)
    url = str(request.url)
    if '#session_id=' in url or '?session_id=' in url:
        # This is OAuth callback - serve the callback handler
        html_content = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>VMC Job Shop - Logging In</title>
            <style>
                body {
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    background: #0F172A;
                    color: white;
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    min-height: 100vh;
                    margin: 0;
                }
                .container { text-align: center; }
                .logo { font-size: 64px; margin-bottom: 20px; }
                .spinner {
                    border: 4px solid #334155;
                    border-top: 4px solid #3B82F6;
                    border-radius: 50%;
                    width: 40px;
                    height: 40px;
                    animation: spin 1s linear infinite;
                    margin: 20px auto;
                }
                @keyframes spin { 100% { transform: rotate(360deg); } }
                .success { color: #22C55E; }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="logo">🏭</div>
                <h1>VMC Job Shop</h1>
                <p id="message">Logging you in...</p>
                <div class="spinner" id="spinner"></div>
            </div>
            <script>
                (function() {
                    var hash = window.location.hash;
                    var sessionId = hash ? hash.split('session_id=')[1]?.split('&')[0] : null;
                    
                    if (sessionId) {
                        var isMobile = /iPhone|iPad|iPod|Android/i.test(navigator.userAgent);
                        
                        if (isMobile) {
                            window.location.href = 'vmcjobshop://auth?session_id=' + sessionId;
                        } else {
                            fetch('/api/auth/session', {
                                method: 'POST',
                                headers: { 'Content-Type': 'application/json' },
                                body: JSON.stringify({ session_id: sessionId }),
                                credentials: 'include'
                            })
                            .then(function(r) {
                                if (r.ok) {
                                    return r.json();
                                } else {
                                    throw new Error('Login failed');
                                }
                            })
                            .then(function(userData) {
                                document.getElementById('message').textContent = 'Login successful! Welcome, ' + (userData.name || 'User') + '!';
                                document.getElementById('message').className = 'success';
                                document.getElementById('spinner').style.display = 'none';
                                
                                // Store auth data in localStorage for the Expo app
                                localStorage.setItem('auth_token', userData.session_token || '');
                                localStorage.setItem('user_data', JSON.stringify(userData));
                                
                                // Clear hash and redirect to dashboard
                                window.location.hash = '';
                                setTimeout(function() {
                                    window.location.replace('/dashboard');
                                }, 500);
                            })
                            .catch(function(err) {
                                document.getElementById('message').textContent = 'Login failed. Redirecting...';
                                setTimeout(function() { window.location.href = '/'; }, 2000);
                            });
                        }
                    } else {
                        window.location.reload();
                    }
                })();
            </script>
        </body>
        </html>
        """
        return HTMLResponse(content=html_content)
    
    # No session_id - this is a regular request
    # For web, the Expo app handles the login UI via index.tsx
    # Return a simple redirect page that checks auth status
    html_content = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>VMC Job Shop</title>
        <style>
            body {
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                background: #0F172A;
                color: white;
                display: flex;
                justify-content: center;
                align-items: center;
                min-height: 100vh;
                margin: 0;
            }
            .container { text-align: center; max-width: 400px; padding: 20px; }
            .logo { font-size: 64px; margin-bottom: 20px; }
            h1 { font-size: 24px; margin-bottom: 10px; }
            p { color: #94A3B8; margin-bottom: 20px; }
            .error { color: #EF4444; }
            .success { color: #22C55E; }
            .btn {
                display: inline-block;
                background: #3B82F6;
                color: white;
                padding: 16px 32px;
                border-radius: 12px;
                text-decoration: none;
                font-weight: 600;
                font-size: 16px;
                cursor: pointer;
                border: none;
            }
            .btn:hover { background: #2563EB; }
            .spinner {
                border: 4px solid #334155;
                border-top: 4px solid #3B82F6;
                border-radius: 50%;
                width: 40px;
                height: 40px;
                animation: spin 1s linear infinite;
                margin: 20px auto;
                display: none;
            }
            @keyframes spin { 100% { transform: rotate(360deg); } }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="logo">🏭</div>
            <h1>VMC Job Shop</h1>
            <p id="message">Checking login status...</p>
            <div class="spinner" id="spinner" style="display: block;"></div>
            <div id="buttons" style="display: none;">
                <button class="btn" onclick="login()">Login with Google</button>
            </div>
        </div>
        <script>
            function login() {
                var redirectUrl = window.location.origin;
                window.location.href = 'https://auth.emergentagent.com/?redirect=' + encodeURIComponent(redirectUrl);
            }
            
            function showError(msg) {
                document.getElementById('message').textContent = msg;
                document.getElementById('message').className = 'error';
                document.getElementById('spinner').style.display = 'none';
                document.getElementById('buttons').style.display = 'block';
            }
            
            function showSuccess(msg) {
                document.getElementById('message').textContent = msg;
                document.getElementById('message').className = 'success';
                document.getElementById('spinner').style.display = 'none';
            }
            
            // FIRST: Check if session_id is in URL hash (OAuth callback)
            var hash = window.location.hash;
            var sessionId = null;
            
            if (hash && hash.includes('session_id=')) {
                var params = hash.substring(1).split('&');
                for (var i = 0; i < params.length; i++) {
                    if (params[i].startsWith('session_id=')) {
                        sessionId = params[i].split('=')[1];
                        break;
                    }
                }
            }
            
            if (sessionId) {
                // Process OAuth callback
                document.getElementById('message').textContent = 'Logging you in...';
                
                var isMobile = /iPhone|iPad|iPod|Android/i.test(navigator.userAgent);
                
                if (isMobile) {
                    // Redirect to app scheme for mobile
                    window.location.href = 'vmcjobshop://auth?session_id=' + sessionId;
                } else {
                    // Web: call backend API
                    fetch('/api/auth/session', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: JSON.stringify({ session_id: sessionId }),
                        credentials: 'include'
                    })
                    .then(function(r) {
                        if (r.ok) {
                            return r.json();
                        } else {
                            return r.text().then(function(text) {
                                throw new Error(text || 'Login failed');
                            });
                        }
                    })
                    .then(function(userData) {
                        showSuccess('Login successful! Welcome, ' + (userData.name || 'User') + '!');
                        
                        // Store auth data in localStorage
                        localStorage.setItem('auth_token', userData.session_token || '');
                        localStorage.setItem('user_data', JSON.stringify(userData));
                        
                        // Clear hash and redirect to dashboard
                        window.history.replaceState(null, '', window.location.pathname);
                        setTimeout(function() {
                            window.location.replace('/dashboard');
                        }, 500);
                    })
                    .catch(function(err) {
                        console.error('Login error:', err);
                        // Clear the hash to allow retry
                        window.history.replaceState(null, '', window.location.pathname);
                        showError('Login failed. Please try again.');
                    });
                }
            } else {
                // No session_id - check if user is already logged in
                fetch('/api/auth/me', { credentials: 'include' })
                .then(function(r) {
                    if (r.ok) {
                        return r.json();
                    } else {
                        throw new Error('Not logged in');
                    }
                })
                .then(function(userData) {
                    // User is logged in - store token and redirect
                    showSuccess('Welcome back, ' + (userData.name || 'User') + '!');
                    document.getElementById('spinner').style.display = 'block';
                    
                    // Store auth data in localStorage
                    localStorage.setItem('auth_token', userData.session_token || '');
                    localStorage.setItem('user_data', JSON.stringify(userData));
                    
                    setTimeout(function() {
                        window.location.replace('/dashboard');
                    }, 500);
                })
                .catch(function(err) {
                    // Not logged in - show login button
                    document.getElementById('message').textContent = 'Welcome! Please login to continue.';
                    document.getElementById('spinner').style.display = 'none';
                    document.getElementById('buttons').style.display = 'block';
                });
            }
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html_content)

from fastapi.responses import RedirectResponse  # add to imports at top (line 3)

@app.get("/api/auth/mobile-callback")
async def mobile_auth_callback(session_id: str):
    """
    Redirect to the mobile app custom scheme with the session_id.
    This uses an HTTP 302 redirect which Chrome Custom Tabs CAN intercept,
    unlike JavaScript window.location.href redirects.
    """
    app_url = f"vmcjobshop://auth?session_id={session_id}"
    return RedirectResponse(url=app_url, status_code=302)

# MongoDB setup with proper error handling for production
MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.getenv("DB_NAME", "test_database")

logger.info(f"Connecting to MongoDB: {MONGO_URL[:30]}...")
logger.info(f"Database name: {DB_NAME}")

try:
    # Set connection timeout for faster failure detection
    client = AsyncIOMotorClient(
        MONGO_URL,
        serverSelectionTimeoutMS=5000,  # 5 second timeout
        connectTimeoutMS=5000,
        socketTimeoutMS=10000
    )
    db = client[DB_NAME]
    # Share DB with modular routes
    set_shared_db(db)
    logger.info("MongoDB client initialized successfully")
except Exception as e:
    logger.error(f"Failed to initialize MongoDB client: {e}")
    # Create a dummy client that will fail gracefully
    client = None
    db = None

# Pydantic Models
class User(BaseModel):
    user_id: str
    email: str
    name: str
    role: str  # Admin, TL, Operator
    picture: Optional[str] = None
    created_at: datetime

class SessionData(BaseModel):
    session_id: str

class Job(BaseModel):
    job_id: str
    category: str  # cnc_lathe, vmc, tool_room, assembly
    machine_name: str
    operator_name: Optional[str] = None
    operator_id: Optional[str] = None
    job_details: str
    cycle_time: Optional[float] = None
    setting_time: Optional[float] = None
    production_lot: Optional[str] = None
    remarks: Optional[str] = None
    status: str  # pending, in_progress, completed
    assigned_to: str  # TL user_id
    assigned_by: str  # Admin user_id
    created_date: datetime
    updated_date: datetime
    # New fields for Assembly & Tool Room tracking
    work_in_action: Optional[str] = None
    pending_work: Optional[str] = None
    estimated_completion_date: Optional[datetime] = None
    status_image: Optional[str] = None  # base64 image
    # New fields for progress tracking
    target_qty: Optional[int] = None
    produced_qty: Optional[int] = 0
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None
    expected_hours: Optional[float] = None  # Expected time to complete in hours

class JobCreate(BaseModel):
    category: str  # cnc_lathe, vmc, tool_room, assembly, programmer, designer
    sub_category: Optional[str] = None  # For programmer/designer: fixture, mould, production
    machine_name: str
    job_details: str
    assigned_to: str  # TL user_id
    target_qty: Optional[int] = None
    expected_hours: Optional[float] = None

class JobUpdate(BaseModel):
    operator_name: Optional[str] = None
    operator_id: Optional[str] = None
    cycle_time: Optional[float] = None
    setting_time: Optional[float] = None
    production_lot: Optional[str] = None
    remarks: Optional[str] = None
    status: Optional[str] = None
    # New fields for Assembly & Tool Room tracking
    work_in_action: Optional[str] = None
    pending_work: Optional[str] = None
    estimated_completion_date: Optional[datetime] = None
    status_image: Optional[str] = None
    # Progress tracking fields
    target_qty: Optional[int] = None
    produced_qty: Optional[int] = None
    expected_hours: Optional[float] = None

class ProductionEntry(BaseModel):
    entry_id: str
    date: datetime
    category: str  # cnc_lathe or vmc
    machine_name: str
    operator_name: str
    operator_id: str
    job_details: str
    production_qty: int
    cycle_time: float
    setting_time: float
    production_lot: str
    shift: str  # day or night
    remarks: Optional[str] = None
    image: Optional[str] = None  # base64 encoded image
    created_at: datetime

class ProductionEntryCreate(BaseModel):
    date: datetime
    category: str  # cnc_lathe or vmc
    machine_name: str
    operator_name: str  # Selected from dropdown
    job_details: str
    production_qty: int
    cycle_time: float
    setting_time: float
    shift: str  # day or night
    remarks: Optional[str] = None
    image: Optional[str] = None  # base64 encoded image

# Machine Breakdown Models
class MachineBreakdown(BaseModel):
    breakdown_id: str
    machine_name: str
    date: datetime
    alarm_photo: Optional[str] = None  # base64 encoded
    machine_photo: Optional[str] = None  # base64 encoded
    action_taken: str
    action_taken_by: str
    action_taken_by_id: str
    reported_by: str
    reported_by_id: str
    status: str  # pending, resolved
    created_at: datetime

class MachineBreakdownCreate(BaseModel):
    machine_name: str
    date: datetime
    alarm_photo: Optional[str] = None
    machine_photo: Optional[str] = None
    action_taken: str

class Machine(BaseModel):
    machine_id: str
    name: str
    category: str  # cnc_lathe, vmc, or moulding
    status: str  # running, idle, breakdown
    current_job_id: Optional[str] = None  # Currently running job
    current_operator_id: Optional[str] = None  # Current operator
    last_status_update: Optional[datetime] = None
    created_at: datetime

class MachineCreate(BaseModel):
    name: str
    category: str  # cnc_lathe, vmc, or moulding

class MachineStatusUpdate(BaseModel):
    status: str  # running, idle, breakdown

# Daily Machine Status Model - for tracking daily records
class DailyMachineStatus(BaseModel):
    record_id: str
    date: str  # YYYY-MM-DD format
    machine_id: str
    machine_name: str
    operator_id: str
    operator_name: str
    job_id: str
    job_details: str
    status: str  # running, idle, breakdown
    start_time: datetime
    end_time: Optional[datetime] = None
    created_at: datetime

# Job Start/Complete Models
class JobStartRequest(BaseModel):
    job_id: Optional[str] = None  # Optional - can create new job with manual details
    machine_id: Optional[str] = None  # Optional - some jobs (programmer/designer) don't need machine
    job_details: Optional[str] = None  # Manual job details
    target_qty: Optional[int] = None
    expected_hours: Optional[float] = None
    estimated_completion_minutes: Optional[int] = None  # Estimated time to complete in minutes

class JobCompleteRequest(BaseModel):
    job_id: str
    produced_qty: Optional[int] = None
    remarks: Optional[str] = None  # Remarks on completion

# ============================================
# TOOLS & INSERTS INVENTORY MODELS
# ============================================

# VMC Tools Inventory
class ToolCreate(BaseModel):
    tool_type: str  # End Mill, Drill, Tap, etc.
    subtype: Optional[str] = None  # Flat, Ball Nose, Corner Radius
    diameter: float  # in mm
    material: str  # Carbide, HSS
    work_material: str  # SS, Aluminium, MS, etc.
    quantity: int
    min_quantity: int
    location: Optional[str] = None
    remarks: Optional[str] = None

class ToolUpdate(BaseModel):
    quantity: Optional[int] = None
    min_quantity: Optional[int] = None
    location: Optional[str] = None
    remarks: Optional[str] = None

# CNC Lathe Inserts Inventory
class InsertCreate(BaseModel):
    insert_type: str  # CNMG, VNMG, WNMG, TNMG, etc.
    grade: str  # Insert grade
    work_material: str  # SS, Aluminium, MS, etc.
    quantity: int
    min_quantity: int
    location: Optional[str] = None
    remarks: Optional[str] = None

class InsertUpdate(BaseModel):
    quantity: Optional[int] = None
    min_quantity: Optional[int] = None
    location: Optional[str] = None
    remarks: Optional[str] = None

# Tool/Insert Issue Model
class ToolIssueRequest(BaseModel):
    item_id: str  # tool_id or insert_id
    item_type: str  # 'tool' or 'insert'
    quantity: int
    issued_to: str  # user_id
    machine_id: Optional[str] = None
    purpose: Optional[str] = None
    signature: Optional[str] = None  # base64 signature image

# Job Comment Model
class JobComment(BaseModel):
    comment: str
    image: Optional[str] = None  # Base64 encoded image



# Material Request Models
class MaterialRequestCreate(BaseModel):
    posted_by: str
    material_required: str
    material_details: str
    required_for: str

class MaterialRequestStatusUpdate(BaseModel):
    status: str  # pending, approved, rejected

class UserBlockUpdate(BaseModel):
    blocked: bool

class PushTokenUpdate(BaseModel):
    push_token: str


class ProfileImageUpdate(BaseModel):
    image: str  # base64 encoded image

class InjectionMouldingEntry(BaseModel):
    machine: str  # Pratishna or Hatian
    mould_no: str
    material: str
    entry_type: str  # Trial or Production
    operator_name: str
    production_qty: int
    remarks: Optional[str] = None

class StorageEntry(BaseModel):
    assembly_name: str
    part_no: str
    product_details: str
    quantity: int = 1
    storage_place: str
    crate_no: str
    stored_by: str
    image: Optional[str] = None

# Plastic Raw Material Storage Models
class PlasticRawMaterialEntry(BaseModel):
    material: str  # PPS, Nylon, IXEF, PP, Delrin, PC, Santoprene
    sub_type: Optional[str] = None  # For Nylon: 6, 66, MOS2
    glass_percentage: Optional[str] = None  # For 6/66: Plain, 15%, 33%
    make: str
    grade: str
    colour: str
    quantity_kg: float  # Quantity in KGs
    storage_place: str
    image: Optional[str] = None

class PlasticRawMaterialDeductRequest(BaseModel):
    quantity_kg: float
    given_to: str  # User ID
    given_to_name: str  # User name
    for_moulding: str  # Which mould/job
    remarks: Optional[str] = None
    signature: Optional[str] = None  # base64 signature image

class PlasticRawMaterialAddRequest(BaseModel):
    quantity_kg: float
    remarks: Optional[str] = None

# Gauge Storage Models
class GaugeEntry(BaseModel):
    gauge_type: str  # Thread, Plug
    sub_type: str  # ID, OD
    gauge_details: str
    make: str
    stored_at: str
    image: Optional[str] = None

class GaugeIssueRequest(BaseModel):
    issued_to_id: str
    issued_to_name: str
    remarks: Optional[str] = None
    signature: Optional[str] = None  # base64 signature image

class GaugeReturnRequest(BaseModel):
    remarks: Optional[str] = None
    signature: Optional[str] = None  # base64 signature image

class UserRoleUpdate(BaseModel):
    email: str
    role: str

class RoleSelection(BaseModel):
    role: str

# Authentication helper
async def get_current_user(request: Request) -> User:
    """Get current authenticated user from session token"""
    # Check cookie first
    session_token = request.cookies.get("session_token")
    
    # Fallback to Authorization header
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header.replace("Bearer ", "")
    
    if not session_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Find session in database
    session_doc = await db.user_sessions.find_one(
        {"session_token": session_token},
        {"_id": 0}
    )
    
    if not session_doc:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    # Check expiry
    expires_at = session_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    
    # Get user data
    user_doc = await db.users.find_one(
        {"user_id": session_doc["user_id"]},
        {"_id": 0}
    )
    
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Track app version (non-blocking, fire and forget)
    try:
        app_version = request.headers.get("X-App-Version", "unknown")
        app_platform = request.headers.get("X-App-Platform", "unknown")
        if app_version != "unknown":
            await db.users.update_one(
                {"user_id": session_doc["user_id"]},
                {"$set": {
                    "last_app_version": app_version,
                    "last_app_platform": app_platform,
                    "last_seen_at": get_ist_now()
                }}
            )
    except Exception as e:
        # Don't fail request if tracking fails
        logger.warning(f"Failed to track app version: {e}")
    
    return User(**user_doc)


async def get_optional_user(request: Request) -> Optional[User]:
    """Get current user if authenticated, otherwise return None (for public endpoints)"""
    try:
        return await get_current_user(request)
    except HTTPException:
        return None


# Auth Endpoints
@app.post("/api/auth/session")
async def create_session(session_data: SessionData, response: Response):
    """[DEPRECATED] Google OAuth removed. Use POST /api/auth/login (username+password)."""
    raise HTTPException(
        status_code=410,
        detail="Google login has been removed. Please use username/password login.",
    )

async def _legacy_create_session_DEAD(session_data: SessionData, response: Response):
    """Exchange session_id for user data and create session"""
    try:
        print(f"Received session_id: {session_data.session_id}")
        
        # Call Emergent Auth API
        async with httpx.AsyncClient() as client:
            auth_response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": session_data.session_id}
            )
            
            print(f"Auth response status: {auth_response.status_code}")
            print(f"Auth response body: {auth_response.text}")
            
            if auth_response.status_code != 200:
                raise HTTPException(status_code=401, detail=f"Invalid session_id: {auth_response.text}")
            
            auth_data = auth_response.json()
        
        # Check if user exists
        existing_user = await db.users.find_one(
            {"email": auth_data["email"]},
            {"_id": 0}
        )
        
        if existing_user:
            # Check if user is blocked
            if existing_user.get("blocked"):
                raise HTTPException(status_code=403, detail="Your account has been blocked. Please contact admin.")
            
            user_id = existing_user["user_id"]
            # Update user data
            await db.users.update_one(
                {"user_id": user_id},
                {"$set": {
                    "name": auth_data["name"],
                    "picture": auth_data.get("picture")
                }}
            )
        else:
            # Create new user with role based on email
            user_id = f"user_{uuid.uuid4().hex[:12]}"
            
            # Auto-assign Admin to specific emails (case-insensitive)
            admin_emails = ["parmindersingh@flowtechindia.net", "mitesh.gandhi@flowtechindia.net"]
            user_email_lower = auth_data["email"].lower()
            if user_email_lower in [e.lower() for e in admin_emails]:
                default_role = "Admin"
            else:
                default_role = "pending"  # User needs to select role
            
            await db.users.insert_one({
                "user_id": user_id,
                "email": auth_data["email"],
                "name": auth_data["name"],
                "role": default_role,
                "picture": auth_data.get("picture"),
                "created_at": datetime.now(timezone.utc)
            })
        
        # Create session
        session_token = auth_data["session_token"]
        await db.user_sessions.insert_one({
            "user_id": user_id,
            "session_token": session_token,
            "expires_at": datetime.now(timezone.utc) + timedelta(days=7),
            "created_at": datetime.now(timezone.utc)
        })
        
        # Set cookie
        response.set_cookie(
            key="session_token",
            value=session_token,
            httponly=True,
            secure=True,
            samesite="none",
            path="/",
            max_age=7 * 24 * 60 * 60
        )
        
        # Get updated user
        user_doc = await db.users.find_one(
            {"user_id": user_id},
            {"_id": 0}
        )
        
        # Include session_token in response for localStorage fallback
        user_doc["session_token"] = session_token
        
        print(f"Returning user: {user_doc}")
        return user_doc
        
    except HTTPException:
        # Re-raise HTTP exceptions as-is
        raise
    except Exception as e:
        import traceback
        print(f"Auth session error: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/auth/me")
async def get_me(current_user: User = Depends(get_current_user)):
    """Get current user data"""
    return current_user

@app.post("/api/auth/logout")
async def logout(request: Request, response: Response):
    """Logout user and reset their role to pending (except Admin)"""
    session_token = request.cookies.get("session_token")
    
    # Also check Authorization header for mobile
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        session_token = auth_header.replace("Bearer ", "")
    
    if session_token:
        # Find the user session to get user_id
        session = await db.user_sessions.find_one({"session_token": session_token})
        if session:
            user_id = session.get("user_id")
            # Get current user to check role
            user = await db.users.find_one({"user_id": user_id})
            # Reset user's role to 'pending' only if not Admin
            if user and user.get("role") != "Admin":
                await db.users.update_one(
                    {"user_id": user_id},
                    {"$set": {"role": "pending"}}
                )
        # Delete the session
        await db.user_sessions.delete_one({"session_token": session_token})
    
    response.delete_cookie(key="session_token", path="/")
    return {"message": "Logged out successfully"}

@app.put("/api/auth/update-name")
async def update_display_name(
    name_update: dict,
    current_user: User = Depends(get_current_user)
):
    """Allow users to update their display name"""
    new_name = name_update.get("name", "").strip()
    
    if not new_name:
        raise HTTPException(status_code=400, detail="Name cannot be empty")
    
    result = await db.users.update_one(
        {"user_id": current_user.user_id},
        {"$set": {"name": new_name}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Name updated successfully", "name": new_name}

# User Management Endpoints
@app.get("/api/users")
async def get_users(current_user: User = Depends(get_current_user)):
    """Get all users (Admin only) with login status"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    users = await db.users.find({}, {"_id": 0}).limit(1000).to_list(length=1000)
    
    # Get active sessions (logged in within last 24 hours)
    now = datetime.now(timezone.utc)
    yesterday = now - timedelta(hours=24)
    
    active_sessions = await db.user_sessions.find(
        {"last_active": {"$gte": yesterday}},
        {"user_id": 1, "last_active": 1}
    ).to_list(length=1000)
    
    active_user_ids = {s["user_id"] for s in active_sessions}
    
    # Enrich users with login status
    for user in users:
        user["is_active"] = user["user_id"] in active_user_ids
        # Don't include sensitive fields
        user.pop("session_token", None)
    
    # Sort: active users first, then by created_at
    users.sort(key=lambda x: (not x.get("is_active", False), x.get("created_at", datetime.min)), reverse=False)
    
    return users



# ==================== APP VERSION CONTROL ENDPOINTS ====================
@app.get("/api/admin/version-settings")
async def get_version_settings(current_user: User = Depends(get_current_user)):
    """Get current app version control settings (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    settings = await db.settings.find_one({"key": "app_version_control"}, {"_id": 0})
    if not settings:
        return {
            "blocking_enabled": False,
            "minimum_version": "1.0.0",
            "current_app_version": "1.0.0",
            "message": "Version blocking is disabled"
        }
    return settings


@app.post("/api/admin/version-settings")
async def update_version_settings(request: Request, current_user: User = Depends(get_current_user)):
    """Update app version control settings (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    data = await request.json()
    blocking_enabled = data.get("blocking_enabled", False)
    minimum_version = data.get("minimum_version", "1.0.0")
    
    await db.settings.update_one(
        {"key": "app_version_control"},
        {"$set": {
            "key": "app_version_control",
            "blocking_enabled": blocking_enabled,
            "minimum_version": minimum_version,
            "updated_at": get_ist_now(),
            "updated_by": current_user.user_id
        }},
        upsert=True
    )
    
    return {
        "success": True,
        "blocking_enabled": blocking_enabled,
        "minimum_version": minimum_version
    }


@app.get("/api/admin/user-versions")
async def get_user_versions(current_user: User = Depends(get_current_user)):
    """Get all users with their app version info (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    users = await db.users.find(
        {},
        {"_id": 0, "user_id": 1, "name": 1, "role": 1, "last_app_version": 1, "last_app_platform": 1, "last_seen_at": 1}
    ).to_list(length=500)
    
    # Get current minimum version
    settings = await db.settings.find_one({"key": "app_version_control"}, {"_id": 0})
    min_version = settings.get("minimum_version", "1.0.0") if settings else "1.0.0"
    
    # Mark users with outdated versions
    for user in users:
        user_version = user.get("last_app_version", "unknown")
        if user_version != "unknown":
            user["is_outdated"] = compare_versions(user_version, min_version) < 0
        else:
            user["is_outdated"] = False
    
    return {
        "users": users,
        "minimum_version": min_version,
        "blocking_enabled": settings.get("blocking_enabled", False) if settings else False
    }
# ==================== END VERSION CONTROL ====================



@app.get("/api/users/stats")
async def get_user_stats(current_user: User = Depends(get_current_user)):
    """Get user statistics (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    total_users = await db.users.count_documents({})
    pending_users = await db.users.count_documents({"role": "pending"})
    blocked_users = await db.users.count_documents({"is_blocked": True})
    
    # Active sessions in last 24 hours
    now = datetime.now(timezone.utc)
    yesterday = now - timedelta(hours=24)
    active_sessions = await db.user_sessions.count_documents({"last_active": {"$gte": yesterday}})
    
    # Count by role
    role_counts = {}
    pipeline = [
        {"$group": {"_id": "$role", "count": {"$sum": 1}}}
    ]
    async for doc in db.users.aggregate(pipeline):
        role_counts[doc["_id"]] = doc["count"]
    
    return {
        "total_users": total_users,
        "active_now": active_sessions,
        "pending_users": pending_users,
        "blocked_users": blocked_users,
        "by_role": role_counts
    }

@app.put("/api/users/role")
async def update_user_role(
    role_update: UserRoleUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update user role (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    valid_roles = [
        "Admin", "Designer", "Die Maker", "Fitter",
        "Programmer", "Programmer (VMC)", "Programmer (CNC)",
        "Operator", "Operator (VMC)", "Operator (CNC)"
    ]
    if role_update.role not in valid_roles:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    result = await db.users.update_one(
        {"email": role_update.email},
        {"$set": {"role": role_update.role}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Role updated successfully"}

@app.put("/api/users/{user_id}/role")
async def update_user_role_by_id(
    user_id: str,
    role_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Update user role by user_id (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    role = role_data.get("role")
    valid_roles = [
        "Admin", "Supervisor", "Team Lead", "Designer", "Die Maker", "Fitter",
        "Programmer", "Programmer (VMC)", "Programmer (CNC)",
        "Operator", "Operator (VMC)", "Operator (CNC)", "Operator (Moulding)",
        "Store Manager", "Quality Inspector", "Viewer", "pending"
    ]
    if role not in valid_roles:
        raise HTTPException(status_code=400, detail=f"Invalid role: {role}")
    
    result = await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"role": role}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": f"Role updated to {role} successfully"}

@app.post("/api/users/select-role")
async def select_role(
    role_selection: RoleSelection,
    current_user: User = Depends(get_current_user)
):
    """Allow users with 'pending' role to select their own role"""
    if current_user.role != "pending":
        raise HTTPException(status_code=400, detail="Role already selected")
    
    valid_roles = [
        "Designer", "Die Maker", "Fitter",
        "Programmer (VMC)", "Programmer (CNC)",
        "Operator (VMC)", "Operator (CNC)"
    ]
    if role_selection.role not in valid_roles:
        raise HTTPException(status_code=400, detail=f"Invalid role. Must be one of: {', '.join(valid_roles)}")
    
    # Update the user's role in the database
    result = await db.users.update_one(
        {"user_id": current_user.user_id},
        {"$set": {"role": role_selection.role}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Failed to update role")
    
    return {"message": "Role selected successfully", "role": role_selection.role}

@app.delete("/api/users/{user_id}")
async def delete_user(
    user_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete user (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Prevent self-deletion
    if user_id == current_user.user_id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    # Check if user exists
    user = await db.users.find_one({"user_id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Delete user sessions
    await db.user_sessions.delete_many({"user_id": user_id})
    
    # Delete user
    result = await db.users.delete_one({"user_id": user_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

@app.post("/api/users/{user_id}/reset-role")
async def reset_user_role(
    user_id: str,
    current_user: User = Depends(get_current_user)
):
    """Reset user's role to pending (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Prevent resetting own role
    if user_id == current_user.user_id:
        raise HTTPException(status_code=400, detail="Cannot reset your own role")
    
    # Check if user exists
    user = await db.users.find_one({"user_id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Reset role to pending
    result = await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"role": "pending"}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=400, detail="Role is already pending")
    
    return {"message": "User role reset to pending successfully"}

@app.post("/api/users/{user_id}/block")
async def block_user(
    user_id: str,
    block_update: UserBlockUpdate,
    current_user: User = Depends(get_current_user)
):
    """Block or unblock a user (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Prevent blocking own account
    if user_id == current_user.user_id:
        raise HTTPException(status_code=400, detail="Cannot block your own account")
    
    # Check if user exists
    user = await db.users.find_one({"user_id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update blocked status
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"blocked": block_update.blocked}}
    )
    
    # If blocking, delete their sessions to force logout
    if block_update.blocked:
        await db.user_sessions.delete_many({"user_id": user_id})
    
    status = "blocked" if block_update.blocked else "unblocked"
    return {"message": f"User {status} successfully"}


@app.get("/api/users/assignable")
async def get_assignable_users(current_user: User = Depends(get_current_user)):
    """Get all users that can be assigned jobs (excluding pending users)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get all users except those with 'pending' role
    users = await db.users.find(
        {"role": {"$ne": "pending"}},
        {"_id": 0}
    ).to_list(length=1000)
    return users

# Machine Endpoints
@app.get("/api/machines")
async def get_machines(current_user: User = Depends(get_current_user)):
    """Get all machines"""
    machines = await db.machines.find({}, {"_id": 0}).to_list(length=500)
    return machines

@app.post("/api/machines")
async def create_machine(
    machine: MachineCreate,
    current_user: User = Depends(get_current_user)
):
    """Create new machine (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    machine_id = f"machine_{uuid.uuid4().hex[:12]}"
    machine_doc = {
        "machine_id": machine_id,
        "name": machine.name,
        "category": machine.category,  # cnc_lathe or vmc
        "status": "available",
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.machines.insert_one(machine_doc)
    
    # Return without _id field
    return {
        "machine_id": machine_id,
        "name": machine.name,
        "category": machine.category,
        "status": "available",
        "created_at": datetime.now(timezone.utc).isoformat()
    }

# Get machines by category
@app.get("/api/machines/category/{category}")
async def get_machines_by_category(
    category: str,
    current_user: User = Depends(get_current_user)
):
    """Get machines filtered by category (cnc_lathe or vmc)"""
    machines = await db.machines.find(
        {"category": category}, 
        {"_id": 0}
    ).to_list(length=500)
    return machines

# Get operators/staff for production entry dropdown
@app.get("/api/operators")
async def get_operators(current_user: User = Depends(get_current_user)):
    """Get all active users (including Admin) for assignment"""
    # Return all users who are not pending
    operators = await db.users.find(
        {
            "role": {"$nin": ["pending"]},
        },
        {"_id": 0, "user_id": 1, "name": 1, "email": 1, "role": 1}
    ).sort("name", 1).to_list(length=500)
    return operators

@app.delete("/api/machines/{machine_id}")
async def delete_machine(
    machine_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete machine (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await db.machines.delete_one({"machine_id": machine_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Machine not found")
    
    return {"message": "Machine deleted successfully"}


@app.post("/api/machines/{machine_id}/reset-status")
async def reset_machine_status(
    machine_id: str,
    current_user: User = Depends(get_current_user)
):
    """Reset machine status to idle (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    machine = await db.machines.find_one({"machine_id": machine_id})
    if not machine:
        raise HTTPException(status_code=404, detail="Machine not found")
    
    now = datetime.now(timezone.utc)
    
    # Reset machine to idle
    await db.machines.update_one(
        {"machine_id": machine_id},
        {"$set": {
            "status": "idle",
            "current_job_id": None,
            "current_operator_id": None,
            "last_status_update": now
        }}
    )
    
    return {"message": f"Machine {machine['name']} status reset to idle"}


@app.post("/api/machines/reset-all-status")
async def reset_all_machine_status(
    current_user: User = Depends(get_current_user)
):
    """Reset all machines to idle (Admin only) - for shift change"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    now = datetime.now(timezone.utc)
    
    # Reset all machines to idle
    result = await db.machines.update_many(
        {},
        {"$set": {
            "status": "idle",
            "current_job_id": None,
            "current_operator_id": None,
            "last_status_update": now
        }}
    )
    
    return {"message": f"Reset {result.modified_count} machines to idle"}


# Job Endpoints
@app.get("/api/jobs")
async def get_jobs(
    status: Optional[str] = None,
    category: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    current_user: User = Depends(get_current_user)
):
    """Get all jobs with pagination"""
    query = {}
    
    # Everyone sees all jobs now
    if status:
        query["status"] = status
    if category:
        query["category"] = category
    
    # Calculate skip for pagination
    skip = (page - 1) * limit
    
    # Get total count
    total_count = await db.jobs.count_documents(query)
    
    jobs = await db.jobs.find(query, {"_id": 0}).sort("created_date", -1).skip(skip).limit(limit).to_list(length=limit)
    
    # Batch fetch machine_status and comments to avoid N+1 queries
    job_ids = [job.get("job_id") for job in jobs]
    
    # Batch fetch all machine statuses for these jobs
    machine_statuses = await db.machine_status.find(
        {"job_id": {"$in": job_ids}},
        {"_id": 0}
    ).to_list(length=500)
    
    # Create lookup dict for machine status
    status_by_job = {ms.get("job_id"): ms for ms in machine_statuses}
    
    # Batch fetch all comments for these jobs (limited to 5 per job)
    all_comments = await db.job_comments.find(
        {"job_id": {"$in": job_ids}},
        {"_id": 0}
    ).sort("created_at", -1).to_list(length=2500)  # 5 comments * 500 jobs max
    
    # Group comments by job_id
    comments_by_job = {}
    for comment in all_comments:
        jid = comment.get("job_id")
        if jid not in comments_by_job:
            comments_by_job[jid] = []
        if len(comments_by_job[jid]) < 5:  # Limit to 5 per job
            comments_by_job[jid].append(comment)
    
    # Enrich jobs with fetched data
    enriched_jobs = []
    for job in jobs:
        job_id = job.get("job_id")
        
        # Add machine status info
        machine_status = status_by_job.get(job_id)
        if machine_status:
            job["started_by_name"] = machine_status.get("operator_name")
            job["started_at"] = machine_status.get("start_time")
        
        # Add comments
        job["comments"] = comments_by_job.get(job_id, [])
        
        enriched_jobs.append(job)
    
    # Return array directly for backward compatibility
    return enriched_jobs

@app.get("/api/jobs/pipeline")
async def get_pipeline_jobs(current_user: User = Depends(get_current_user)):
    """Get pending jobs (pipeline/to-do list) - everyone sees all"""
    query = {"status": "pending"}
    
    # Everyone sees all pending jobs now - limited to 500 for performance
    
    jobs = await db.jobs.find(query, {"_id": 0}).sort("created_date", -1).limit(500).to_list(length=500)
    return jobs


@app.get("/api/jobs/available")
async def get_available_jobs(
    current_user: User = Depends(get_current_user)
):
    """Get jobs available for the operator to start"""
    query = {
        "status": {"$in": ["pending", "in_progress"]},
        "$or": [
            {"assigned_to": current_user.user_id},
            {"operator_id": current_user.user_id}
        ]
    }
    
    # Admins can see all pending jobs
    if current_user.role == "Admin":
        query = {"status": {"$in": ["pending", "in_progress"]}}
    
    jobs = await db.jobs.find(query).sort("created_date", -1).to_list(length=100)
    
    result = []
    for job in jobs:
        progress = 0
        if job.get("target_qty") and job.get("target_qty") > 0:
            progress = min(100, int((job.get("produced_qty", 0) / job["target_qty"]) * 100))
        
        result.append({
            "job_id": job["job_id"],
            "category": job.get("category", ""),
            "machine_name": job.get("machine_name", ""),
            "job_details": job.get("job_details", ""),
            "status": job.get("status", "pending"),
            "target_qty": job.get("target_qty"),
            "produced_qty": job.get("produced_qty", 0),
            "progress": progress,
            "start_time": job.get("start_time").isoformat() if job.get("start_time") else None,
            "operator_name": job.get("operator_name"),
            "assigned_to": job.get("assigned_to")
        })
    
    return result


@app.get("/api/jobs/my-assigned")
async def get_my_assigned_jobs(
    current_user: User = Depends(get_current_user)
):
    """Get all pending jobs for tool room/assembly/programming/designing categories.
    Any user can start any available job."""
    
    # Categories that don't require machine selection
    non_machine_categories = ['tool_room', 'assembly', 'programmer', 'designer', 'programming', 'designing']
    
    # Show ALL pending jobs in non-machine categories for any user
    query = {
        "$or": [
            {"category": {"$in": non_machine_categories}},
            {"status": {"$in": ["pending", "assigned"]}}
        ],
        "status": {"$in": ["pending", "assigned"]}
    }
    
    jobs = await db.jobs.find(query).sort("created_date", -1).to_list(length=100)
    
    result = []
    for job in jobs:
        progress = 0
        if job.get("target_qty") and job.get("target_qty") > 0:
            progress = min(100, int((job.get("produced_qty", 0) / job["target_qty"]) * 100))
        
        result.append({
            "job_id": job["job_id"],
            "category": job.get("category", ""),
            "sub_category": job.get("sub_category", ""),
            "machine_name": job.get("machine_name", ""),
            "job_details": job.get("job_details", ""),
            "status": job.get("status", "pending"),
            "target_qty": job.get("target_qty"),
            "produced_qty": job.get("produced_qty", 0),
            "progress": progress,
            "start_time": job.get("start_time").isoformat() if job.get("start_time") else None,
            "operator_name": job.get("operator_name"),
            "assigned_to": job.get("assigned_to"),
            "assigned_to_name": job.get("assigned_to_name", "")
        })
    
    return result


@app.get("/api/jobs/active")
async def get_active_jobs(
    current_user: User = Depends(get_current_user)
):
    """Get active work entries from active_work collection - ONLY for machines currently running"""
    query = {"status": "in_progress"}
    
    # Non-admins only see their own work
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        query["operator_id"] = current_user.user_id
    
    entries = await db.active_work.find(query, {"_id": 0}).sort("start_time", -1).to_list(length=100)
    
    # Get all running machines (case-insensitive)
    running_machines = await db.machines.find(
        {"status": {"$regex": "^running$", "$options": "i"}}, 
        {"machine_id": 1, "name": 1}
    ).to_list(length=100)
    running_machine_names = {m.get("name", "").lower() for m in running_machines}
    running_machine_ids = {m.get("machine_id", "").lower() for m in running_machines}
    
    # Filter entries to only include those on running machines
    filtered_entries = []
    for entry in entries:
        machine_name = (entry.get("machine_name") or "").lower()
        machine_id = (entry.get("machine_id") or "").lower()
        # Check if machine is running (by name or ID)
        if machine_name in running_machine_names or machine_id in running_machine_ids:
            filtered_entries.append(entry)
    
    # Serialize datetime fields - CONVERT UTC TO IST
    for entry in filtered_entries:
        if entry.get("start_time"):
            start_dt = entry["start_time"]
            if hasattr(start_dt, 'isoformat'):
                ist_start = to_ist(start_dt)
                entry["start_time"] = ist_start.isoformat()
                entry["start_time_ist"] = format_ist_datetime(ist_start)
        if entry.get("created_at"):
            if hasattr(entry["created_at"], 'isoformat'):
                entry["created_at"] = to_ist(entry["created_at"]).isoformat()
    
    return filtered_entries


@app.post("/api/jobs")
async def create_job(
    job: JobCreate,
    current_user: User = Depends(get_current_user)
):
    """Create new job (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Validate category
    valid_categories = ["cnc_lathe", "vmc", "tool_room", "assembly", "programmer", "designer"]
    if job.category not in valid_categories:
        raise HTTPException(status_code=400, detail=f"Invalid category. Must be one of: {', '.join(valid_categories)}")
    
    # Verify assigned_to user exists and is TL
    assigned_user = await db.users.find_one(
        {"user_id": job.assigned_to},
        {"_id": 0}
    )
    
    if not assigned_user:
        raise HTTPException(status_code=404, detail="Assigned user not found")
    
    allowed_assign_roles = [
        "Admin",  # Admin can assign to other Admins
        "Designer", "Die Maker", "Fitter",
        "Programmer", "Programmer (VMC)", "Programmer (CNC)",
        "Operator", "Operator (VMC)", "Operator (CNC)"
    ]
    if assigned_user["role"] not in allowed_assign_roles:
        raise HTTPException(status_code=400, detail="Can only assign to valid roles")
    
    job_id = f"job_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    job_doc = {
        "job_id": job_id,
        "category": job.category,
        "sub_category": job.sub_category,  # For programmer/designer: fixture, mould, production
        "machine_name": job.machine_name,
        "operator_name": None,
        "operator_id": None,
        "job_details": job.job_details,
        "cycle_time": None,
        "setting_time": None,
        "production_lot": None,
        "remarks": None,
        "status": "pending",
        "assigned_to": job.assigned_to,
        "assigned_to_name": assigned_user.get("name", ""),
        "assigned_by": current_user.user_id,
        "created_date": now,
        "updated_date": now,
        "work_in_action": None,
        "pending_work": None,
        "estimated_completion_date": None,
        "status_image": None,
        "last_updated_by": None,
        "last_updated_by_name": None
    }
    
    await db.jobs.insert_one(job_doc)
    
    # Send push notification to assigned user about new job
    await send_push_notification(
        job.assigned_to,
        "📋 New Job Assigned / नया काम सौंपा गया",
        f"Machine: {job.machine_name} - {job.job_details[:50]}..."
    )
    
    # Return without _id field and with serialized dates
    return {
        "job_id": job_id,
        "category": job.category,
        "machine_name": job.machine_name,
        "operator_name": None,
        "operator_id": None,
        "job_details": job.job_details,
        "cycle_time": None,
        "setting_time": None,
        "production_lot": None,
        "remarks": None,
        "status": "pending",
        "assigned_to": job.assigned_to,
        "assigned_by": current_user.user_id,
        "created_date": now.isoformat(),
        "updated_date": now.isoformat()
    }

@app.put("/api/jobs/{job_id}")
async def update_job(
    job_id: str,
    job_update: JobUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update job (Everyone can update any job now)"""
    # Find job
    job = await db.jobs.find_one({"job_id": job_id}, {"_id": 0})
    
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Everyone can now update any job - no permission check
    
    # Build update dict
    update_data = {
        "updated_date": datetime.now(timezone.utc),
        "last_updated_by": current_user.user_id,
        "last_updated_by_name": current_user.name
    }
    
    if job_update.operator_name is not None:
        update_data["operator_name"] = job_update.operator_name
        update_data["operator_id"] = current_user.user_id
    if job_update.cycle_time is not None:
        update_data["cycle_time"] = job_update.cycle_time
    if job_update.setting_time is not None:
        update_data["setting_time"] = job_update.setting_time
    if job_update.production_lot is not None:
        update_data["production_lot"] = job_update.production_lot
    if job_update.remarks is not None:
        update_data["remarks"] = job_update.remarks
    if job_update.status is not None:
        if job_update.status not in ["pending", "in_progress", "completed"]:
            raise HTTPException(status_code=400, detail="Invalid status")
        update_data["status"] = job_update.status
    # New fields for Assembly & Tool Room
    if job_update.work_in_action is not None:
        update_data["work_in_action"] = job_update.work_in_action
    if job_update.pending_work is not None:
        update_data["pending_work"] = job_update.pending_work
    if job_update.estimated_completion_date is not None:
        update_data["estimated_completion_date"] = job_update.estimated_completion_date
    if job_update.status_image is not None:
        update_data["status_image"] = job_update.status_image
    
    await db.jobs.update_one(
        {"job_id": job_id},
        {"$set": update_data}
    )
    
    # Return updated job
    updated_job = await db.jobs.find_one({"job_id": job_id}, {"_id": 0})
    return updated_job

@app.delete("/api/jobs/{job_id}")
async def delete_job(
    job_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete job (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await db.jobs.delete_one({"job_id": job_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Job not found")
    
    return {"message": "Job deleted successfully"}

# Get single job with comments
@app.get("/api/jobs/{job_id}")
async def get_job(
    job_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get a single job with its comments"""
    job = await db.jobs.find_one({"job_id": job_id}, {"_id": 0})
    
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Get comments for this job
    comments = await db.job_comments.find(
        {"job_id": job_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(length=100)
    
    # Serialize dates
    if job.get("created_date"):
        job["created_date"] = job["created_date"].isoformat() if hasattr(job["created_date"], 'isoformat') else job["created_date"]
    if job.get("updated_date"):
        job["updated_date"] = job["updated_date"].isoformat() if hasattr(job["updated_date"], 'isoformat') else job["updated_date"]
    
    for comment in comments:
        if comment.get("created_at"):
            comment["created_at"] = comment["created_at"].isoformat() if hasattr(comment["created_at"], 'isoformat') else comment["created_at"]
    
    job["comments"] = comments
    return job

# Add comment to job
@app.post("/api/jobs/{job_id}/comments")
async def add_job_comment(
    job_id: str,
    comment_data: JobComment,
    current_user: User = Depends(get_current_user)
):
    """Add a comment/remark to a job (Everyone can comment)"""
    # Check if job exists
    job = await db.jobs.find_one({"job_id": job_id})
    
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    comment_id = f"comment_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    comment_doc = {
        "comment_id": comment_id,
        "job_id": job_id,
        "comment": comment_data.comment,
        "image": comment_data.image,  # Store image if provided
        "created_by_id": current_user.user_id,
        "created_by_name": current_user.name,
        "created_by_role": current_user.role,
        "created_at": now
    }
    
    await db.job_comments.insert_one(comment_doc)
    
    # Also update the job's updated_date
    await db.jobs.update_one(
        {"job_id": job_id},
        {"$set": {
            "updated_date": now,
            "last_updated_by": current_user.user_id,
            "last_updated_by_name": current_user.name
        }}
    )
    
    return {
        "comment_id": comment_id,
        "job_id": job_id,
        "comment": comment_data.comment,
        "image": comment_data.image,
        "created_by_id": current_user.user_id,
        "created_by_name": current_user.name,
        "created_by_role": current_user.role,
        "created_at": now.isoformat()
    }

# Get comments for a job
@app.get("/api/jobs/{job_id}/comments")
async def get_job_comments(
    job_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get all comments for a job"""
    comments = await db.job_comments.find(
        {"job_id": job_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(length=100)
    
    for comment in comments:
        if comment.get("created_at"):
            comment["created_at"] = comment["created_at"].isoformat() if hasattr(comment["created_at"], 'isoformat') else comment["created_at"]
    
    return comments

# Dashboard Stats
@app.get("/api/stats")
async def get_stats(current_user: User = Depends(get_current_user)):
    """Get dashboard statistics with optimized queries"""
    if current_user.role == "Admin":
        # Use aggregation for better performance
        pipeline = [
            {
                "$facet": {
                    "total": [{"$count": "count"}],
                    "by_status": [{"$group": {"_id": "$status", "count": {"$sum": 1}}}],
                    "by_category": [{"$group": {"_id": "$category", "count": {"$sum": 1}}}]
                }
            }
        ]
        
        result = await db.jobs.aggregate(pipeline).to_list(length=1)
        
        # Parse aggregation results
        total_jobs = result[0]["total"][0]["count"] if result[0]["total"] else 0
        
        status_counts = {item["_id"]: item["count"] for item in result[0]["by_status"]}
        pending_jobs = status_counts.get("pending", 0)
        in_progress_jobs = status_counts.get("in_progress", 0)
        completed_jobs = status_counts.get("completed", 0)
        
        category_counts = {item["_id"]: item["count"] for item in result[0]["by_category"]}
        cnc_lathe_jobs = category_counts.get("cnc_lathe", 0)
        vmc_jobs = category_counts.get("vmc", 0)
        tool_room_jobs = category_counts.get("tool_room", 0)
        assembly_jobs = category_counts.get("assembly", 0)
        
        total_machines = await db.machines.count_documents({})
        total_users = await db.users.count_documents({})
    else:
        # For non-admins, use optimized queries with their user_id filter
        pipeline = [
            {"$match": {"assigned_to": current_user.user_id}},
            {
                "$facet": {
                    "total": [{"$count": "count"}],
                    "by_status": [{"$group": {"_id": "$status", "count": {"$sum": 1}}}],
                    "by_category": [{"$group": {"_id": "$category", "count": {"$sum": 1}}}]
                }
            }
        ]
        
        result = await db.jobs.aggregate(pipeline).to_list(length=1)
        
        total_jobs = result[0]["total"][0]["count"] if result[0]["total"] else 0
        
        status_counts = {item["_id"]: item["count"] for item in result[0]["by_status"]}
        pending_jobs = status_counts.get("pending", 0)
        in_progress_jobs = status_counts.get("in_progress", 0)
        completed_jobs = status_counts.get("completed", 0)
        
        category_counts = {item["_id"]: item["count"] for item in result[0]["by_category"]}
        cnc_lathe_jobs = category_counts.get("cnc_lathe", 0)
        vmc_jobs = category_counts.get("vmc", 0)
        tool_room_jobs = category_counts.get("tool_room", 0)
        assembly_jobs = category_counts.get("assembly", 0)
        
        total_machines = await db.machines.count_documents({})
        total_users = None
    
    return {
        "total_jobs": total_jobs,
        "pending_jobs": pending_jobs,
        "in_progress_jobs": in_progress_jobs,
        "completed_jobs": completed_jobs,
        "total_machines": total_machines,
        "total_users": total_users,
        "categories": {
            "cnc_lathe": cnc_lathe_jobs,
            "vmc": vmc_jobs,
            "tool_room": tool_room_jobs,
            "assembly": assembly_jobs
        }
    }

@app.get("/api/health")
async def health():
    """Health check endpoint"""
    return {"status": "healthy"}

# Reports Endpoints
@app.get("/api/reports/summary")
async def get_reports_summary(
    period: str,  # 'weekly' or 'monthly'
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get summary reports - machine-wise and operator-wise"""
    # Only Admin and those with 'TL' role (Team Leaders with Designer designation) can access
    allowed_roles = ["Admin", "Designer", "Programmer"]
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="Access denied. Admin or TL only.")
    
    # Calculate date range based on period
    now = datetime.now(timezone.utc)
    if period == "weekly":
        if not date_from:
            date_from = (now - timedelta(days=7)).isoformat()
        if not date_to:
            date_to = now.isoformat()
    elif period == "monthly":
        if not date_from:
            date_from = (now - timedelta(days=30)).isoformat()
        if not date_to:
            date_to = now.isoformat()
    else:
        raise HTTPException(status_code=400, detail="Invalid period. Use 'weekly' or 'monthly'")
    
    date_filter = {
        "updated_date": {
            "$gte": datetime.fromisoformat(date_from),
            "$lte": datetime.fromisoformat(date_to)
        }
    }
    
    # Machine-wise report (for CNC and VMC jobs)
    machine_pipeline = [
        {"$match": {**date_filter, "category": {"$in": ["cnc_lathe", "vmc"]}}},
        {
            "$group": {
                "_id": "$machine_name",
                "total_jobs": {"$sum": 1},
                "completed": {"$sum": {"$cond": [{"$eq": ["$status", "completed"]}, 1, 0]}},
                "in_progress": {"$sum": {"$cond": [{"$eq": ["$status", "in_progress"]}, 1, 0]}},
                "pending": {"$sum": {"$cond": [{"$eq": ["$status", "pending"]}, 1, 0]}},
                "avg_cycle_time": {"$avg": "$cycle_time"},
                "total_cycle_time": {"$sum": "$cycle_time"}
            }
        },
        {"$sort": {"total_jobs": -1}}
    ]
    
    machine_stats = await db.jobs.aggregate(machine_pipeline).to_list(length=100)
    
    # Operator-wise report (from jobs)
    operator_pipeline = [
        {"$match": {**date_filter, "operator_name": {"$ne": None}}},
        {
            "$group": {
                "_id": "$operator_name",
                "operator_id": {"$first": "$operator_id"},
                "total_jobs": {"$sum": 1},
                "completed": {"$sum": {"$cond": [{"$eq": ["$status", "completed"]}, 1, 0]}},
                "in_progress": {"$sum": {"$cond": [{"$eq": ["$status", "in_progress"]}, 1, 0]}},
                "pending": {"$sum": {"$cond": [{"$eq": ["$status", "pending"]}, 1, 0]}},
                "avg_cycle_time": {"$avg": "$cycle_time"}
            }
        },
        {"$sort": {"total_jobs": -1}}
    ]
    
    operator_stats = await db.jobs.aggregate(operator_pipeline).to_list(length=100)
    
    # Production entries summary (daily production entries)
    production_pipeline = [
        {
            "$match": {
                "date": {
                    "$gte": datetime.fromisoformat(date_from),
                    "$lte": datetime.fromisoformat(date_to)
                }
            }
        },
        {
            "$group": {
                "_id": {
                    "machine": "$machine_name",
                    "operator": "$operator_name"
                },
                "entries_count": {"$sum": 1},
                "total_cycle_time": {"$sum": "$cycle_time"},
                "total_setting_time": {"$sum": "$setting_time"},
                "avg_cycle_time": {"$avg": "$cycle_time"}
            }
        },
        {"$sort": {"entries_count": -1}}
    ]
    
    production_stats = await db.production_entries.aggregate(production_pipeline).to_list(length=100)
    
    return {
        "period": period,
        "date_from": date_from,
        "date_to": date_to,
        "machine_wise": machine_stats,
        "operator_wise": operator_stats,
        "production_entries": production_stats
    }

# Production Entry Endpoints
@app.post("/api/production-entries")
async def create_production_entry(
    entry: ProductionEntryCreate,
    current_user: User = Depends(get_current_user)
):
    """Create daily production entry (Anyone can create)"""
    # Validate category - now includes all work types
    valid_categories = ["cnc_lathe", "cnc", "vmc", "moulding", "tool_room", "assembly"]
    if entry.category not in valid_categories:
        raise HTTPException(status_code=400, detail=f"Invalid category. Must be one of: {', '.join(valid_categories)}")
    
    entry_id = f"prod_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    entry_doc = {
        "entry_id": entry_id,
        "date": entry.date,
        "category": entry.category,
        "machine_name": entry.machine_name,
        "operator_name": entry.operator_name,  # From dropdown selection
        "job_details": entry.job_details,
        "production_qty": entry.production_qty,
        "cycle_time": entry.cycle_time,
        "setting_time": entry.setting_time,
        "shift": entry.shift,  # day or night
        "remarks": entry.remarks,
        "image": entry.image,
        "created_by_id": current_user.user_id,
        "created_by_name": current_user.name,
        "created_at": now,
        "status": "in_progress"  # Can be: in_progress, completed
    }
    
    await db.production_entries.insert_one(entry_doc)
    
    return {
        "entry_id": entry_id,
        "date": entry.date.isoformat(),
        "category": entry.category,
        "machine_name": entry.machine_name,
        "operator_name": entry.operator_name,
        "job_details": entry.job_details,
        "production_qty": entry.production_qty,
        "cycle_time": entry.cycle_time,
        "setting_time": entry.setting_time,
        "shift": entry.shift,
        "remarks": entry.remarks,
        "created_at": now.isoformat()
    }


# Start work endpoint - for the new Start Work page
class ProductionStartRequest(BaseModel):
    category: str
    sub_category: Optional[str] = None  # For programming/designing: fixture, mould, production
    machine_id: Optional[str] = None
    machine_name: Optional[str] = None
    job_id: Optional[str] = None
    job_details: str
    quantity: Optional[int] = None
    setting_time: Optional[float] = None
    cycle_time: Optional[float] = None
    remarks: Optional[str] = None
    image: Optional[str] = None
    start_time: str
    operator_id: Optional[str] = None
    operator_name: Optional[str] = None
    estimated_time: Optional[int] = None  # Estimated completion time in minutes
    spindle_rpm: Optional[int] = None  # CNC spindle RPM
    feed_rate: Optional[float] = None  # CNC feed rate in mm/min


@app.post("/api/production/start")
async def start_production_work(
    request: ProductionStartRequest,
    current_user: User = Depends(get_current_user)
):
    """Start a new production work entry"""
    valid_categories = ["cnc", "vmc", "moulding", "tool_room", "assembly", "programming", "designing"]
    if request.category not in valid_categories:
        raise HTTPException(status_code=400, detail=f"Invalid category. Must be one of: {', '.join(valid_categories)}")
    
    entry_id = f"work_{uuid.uuid4().hex[:12]}"
    now = get_ist_now()  # Use IST timezone
    
    # Determine shift based on IST time
    hour = now.hour
    shift = "day" if 6 <= hour < 18 else "night"
    
    # Fetch assigned_to_name from the job if job_id is provided
    assigned_to_name = ""
    if request.job_id:
        job = await db.jobs.find_one({"job_id": request.job_id})
        if job:
            assigned_to_name = job.get("assigned_to_name", "")
    
    work_entry = {
        "entry_id": entry_id,
        "date": now.date().isoformat(),
        "category": request.category,
        "sub_category": request.sub_category,  # For programming/designing: fixture, mould, production
        "machine_id": request.machine_id,
        "machine_name": request.machine_name,
        "job_id": request.job_id,
        "job_details": request.job_details,
        "quantity": request.quantity,
        "setting_time": request.setting_time,
        "cycle_time": request.cycle_time,
        "remarks": request.remarks,
        "image": request.image,
        "shift": shift,
        "start_time": now,
        "start_time_ist": format_ist_datetime(now),  # Human readable IST format
        "end_time": None,
        "estimated_time": request.estimated_time,  # Estimated completion time in minutes
        "spindle_rpm": request.spindle_rpm,  # CNC spindle RPM
        "feed_rate": request.feed_rate,  # CNC feed rate in mm/min
        "status": "in_progress",
        "operator_id": request.operator_id or current_user.user_id,
        "operator_name": request.operator_name or current_user.name,
        "assigned_to_name": assigned_to_name,  # Who the job was originally assigned to
        "done_by_name": current_user.name,  # Who actually started/did the work
        "created_by_id": current_user.user_id,
        "created_by_name": current_user.name,
        "created_at": now
    }
    
    await db.active_work.insert_one(work_entry)
    
    # Update job status to in_progress if job_id is provided
    if request.job_id:
        await db.jobs.update_one(
            {"job_id": request.job_id},
            {"$set": {
                "status": "in_progress",
                "start_time": now,
                "started_by_id": current_user.user_id,
                "started_by_name": current_user.name,
                "updated_date": now
            }}
        )
    
    # Update machine status if applicable
    if request.machine_id:
        await db.machines.update_one(
            {"machine_id": request.machine_id},
            {"$set": {
                "status": "running",
                "current_job_id": entry_id,
                "current_operator_id": current_user.user_id,
                "last_status_update": now
            }}
        )
    
    return {
        "message": "Work started successfully",
        "entry_id": entry_id,
        "category": request.category,
        "start_time": now.isoformat()
    }


@app.get("/api/production/active")
async def get_active_work(
    current_user: User = Depends(get_current_user)
):
    """Get all active/in-progress work entries for the current user"""
    query = {"status": "in_progress"}
    
    # Non-admins only see their own work
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        query["operator_id"] = current_user.user_id
    
    entries = await db.active_work.find(query, {"_id": 0}).sort("start_time", -1).to_list(length=100)
    
    # Serialize datetime fields - CONVERT UTC TO IST
    for entry in entries:
        if entry.get("start_time"):
            start_dt = entry["start_time"]
            if hasattr(start_dt, 'isoformat'):
                # MongoDB stores as UTC, convert to IST for display
                ist_start = to_ist(start_dt)
                entry["start_time"] = ist_start.isoformat()  # IST ISO string
                entry["start_time_ist"] = format_ist_datetime(ist_start)  # Human readable
        if entry.get("end_time") and hasattr(entry["end_time"], 'isoformat'):
            ist_end = to_ist(entry["end_time"])
            entry["end_time"] = ist_end.isoformat()
            entry["end_time_ist"] = format_ist_datetime(ist_end)
        if entry.get("created_at") and hasattr(entry["created_at"], 'isoformat'):
            entry["created_at"] = to_ist(entry["created_at"]).isoformat()
    
    return entries

@app.delete("/api/admin/active-work/{entry_id}")
async def delete_active_work(entry_id: str, current_user: User = Depends(get_current_user)):
    """Delete an orphaned/stuck active_work entry. Admin only."""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Find and delete
    result = await db.active_work.delete_one({"entry_id": entry_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    # Also reset machine status if applicable
    await db.machines.update_many(
        {"current_job_id": entry_id},
        {"$set": {"status": "idle", "current_job_id": None, "current_operator_id": None, "current_operator": None, "current_job": None}}
    )
    
    logger.info(f"✅ Admin {current_user.name} deleted orphaned active_work entry {entry_id}")
    
    return {"success": True, "message": f"Entry {entry_id} deleted", "entry_id": entry_id}


# Public endpoint for TV Mode - No authentication required
@app.get("/api/production/active-public")
async def get_active_work_public():
    """Get all active/in-progress work entries - PUBLIC for TV display"""
    query = {"status": "in_progress"}
    
    entries = await db.active_work.find(query, {"_id": 0}).sort("start_time", -1).to_list(length=100)
    
    # Serialize datetime fields - CONVERT UTC TO IST
    for entry in entries:
        if entry.get("start_time"):
            start_dt = entry["start_time"]
            if hasattr(start_dt, 'isoformat'):
                ist_start = to_ist(start_dt)
                entry["start_time"] = ist_start.isoformat()
                entry["start_time_ist"] = format_ist_datetime(ist_start)
        if entry.get("created_at") and hasattr(entry["created_at"], 'isoformat'):
            entry["created_at"] = to_ist(entry["created_at"]).isoformat()
    
    return entries


@app.get("/api/production/active-by-machine/{machine_id}")
async def get_active_work_by_machine(
    machine_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get active/in-progress work entry for a specific machine"""
    query = {"status": "in_progress", "machine_id": machine_id}
    
    entry = await db.active_work.find_one(query, {"_id": 0})
    return entry


@app.post("/api/production/{entry_id}/post-qty")
async def post_production_qty(
    entry_id: str,
    post_data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    Post intermediate production quantity during work.
    Records the qty produced since last post (or start time if first post).
    """
    work = await db.active_work.find_one({"entry_id": entry_id})
    if not work:
        raise HTTPException(status_code=404, detail="Active work entry not found")
    
    qty = post_data.get("qty")
    if not qty or qty <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be greater than 0")
    
    now = get_ist_now()
    
    # Get start time for this period (last post time or work start time)
    last_post_time = work.get("last_post_time")
    if last_post_time:
        if isinstance(last_post_time, str):
            period_start = datetime.fromisoformat(last_post_time.replace('Z', '+00:00'))
        else:
            period_start = last_post_time
        if period_start.tzinfo is None:
            period_start = period_start.replace(tzinfo=timezone.utc)
    else:
        period_start = work.get("start_time")
        if isinstance(period_start, str):
            period_start = datetime.fromisoformat(period_start.replace('Z', '+00:00'))
        if period_start.tzinfo is None:
            period_start = period_start.replace(tzinfo=timezone.utc)
    
    # Calculate duration
    duration_seconds = (now - period_start).total_seconds()
    duration_minutes = int(duration_seconds / 60)
    
    # Calculate average cycle time for this period (in seconds per piece)
    avg_cycle_time_seconds = duration_seconds / qty if qty > 0 else 0
    
    # Create post entry
    post_entry = {
        "post_id": f"post_{uuid.uuid4().hex[:8]}",
        "start_time": period_start.isoformat(),
        "start_time_ist": format_ist_datetime(to_ist(period_start)),
        "end_time": now.isoformat(),
        "end_time_ist": format_ist_datetime(now),
        "duration_minutes": duration_minutes,
        "duration_display": f"{duration_minutes // 60}h {duration_minutes % 60}m",
        "qty": qty,
        "avg_cycle_time_seconds": round(avg_cycle_time_seconds, 2),
        "avg_cycle_time_display": f"{int(avg_cycle_time_seconds // 60)}m {int(avg_cycle_time_seconds % 60)}s",
        "posted_by_id": current_user.user_id,
        "posted_by_name": current_user.name,
        "posted_at": now.isoformat()
    }
    
    # Get existing posts or initialize
    production_posts = work.get("production_posts", [])
    production_posts.append(post_entry)
    
    # Calculate total posted qty
    total_posted_qty = sum(p.get("qty", 0) for p in production_posts)
    
    # Update active_work
    await db.active_work.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "production_posts": production_posts,
            "last_post_time": now,
            "total_posted_qty": total_posted_qty
        }}
    )
    
    logger.info(f"Production post recorded: {qty} pcs by {current_user.name} for entry {entry_id}")
    
    return {
        "success": True,
        "message": f"Production qty posted: {qty} pcs in {post_entry['duration_display']}",
        "post": post_entry,
        "total_posted_qty": total_posted_qty,
        "total_posts": len(production_posts)
    }


@app.get("/api/production/{entry_id}/posts")
async def get_production_posts(
    entry_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get all production posts for a work entry (active or completed)"""
    # Check active work first
    work = await db.active_work.find_one({"entry_id": entry_id}, {"_id": 0})
    if work:
        return {
            "entry_id": entry_id,
            "status": "active",
            "production_posts": work.get("production_posts", []),
            "total_posted_qty": work.get("total_posted_qty", 0),
            "last_post_time": work.get("last_post_time")
        }
    
    # Check production entries
    entry = await db.production_entries.find_one({"entry_id": entry_id}, {"_id": 0})
    if entry:
        return {
            "entry_id": entry_id,
            "status": "completed",
            "production_posts": entry.get("production_posts", []),
            "total_posted_qty": entry.get("total_posted_qty", 0),
            "final_qty": entry.get("production_qty"),
            "production_loss": entry.get("production_loss"),
            "production_loss_percent": entry.get("production_loss_percent")
        }
    
    raise HTTPException(status_code=404, detail="Entry not found")


@app.post("/api/production/end/{entry_id}")
async def end_production_work(
    entry_id: str,
    end_data: dict,
    current_user: User = Depends(get_current_user)
):
    """End a work entry and move it to production sheet"""
    work = await db.active_work.find_one({"entry_id": entry_id})
    
    if not work:
        raise HTTPException(status_code=404, detail="Work entry not found")
    
    # Check permission
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"} and work.get("operator_id") != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not authorized to end this work")
    
    now = get_ist_now()  # Use IST timezone
    
    # Update with end data
    final_qty = end_data.get("quantity", work.get("quantity"))
    final_remarks = end_data.get("remarks", work.get("remarks"))
    
    # Handle production posts - add final period if there were previous posts
    production_posts = work.get("production_posts", [])
    total_posted_qty = work.get("total_posted_qty", 0)
    qty_since_last_post = end_data.get("qty_since_last_post", 0)
    
    # If there were previous posts, add the final period post
    if production_posts and qty_since_last_post > 0:
        last_post_time = work.get("last_post_time")
        if last_post_time:
            if isinstance(last_post_time, str):
                period_start = datetime.fromisoformat(last_post_time.replace('Z', '+00:00'))
            else:
                period_start = last_post_time
            if period_start.tzinfo is None:
                period_start = period_start.replace(tzinfo=timezone.utc)
            
            # Calculate duration for final period
            duration_seconds = (now - period_start).total_seconds()
            duration_minutes = int(duration_seconds / 60)
            avg_cycle_time_seconds = duration_seconds / qty_since_last_post if qty_since_last_post > 0 else 0
            
            # Add final post entry
            final_post = {
                "post_id": f"post_final_{uuid.uuid4().hex[:8]}",
                "start_time": period_start.isoformat(),
                "start_time_ist": format_ist_datetime(to_ist(period_start)),
                "end_time": now.isoformat(),
                "end_time_ist": format_ist_datetime(now),
                "duration_minutes": duration_minutes,
                "duration_display": f"{duration_minutes // 60}h {duration_minutes % 60}m",
                "qty": qty_since_last_post,
                "avg_cycle_time_seconds": round(avg_cycle_time_seconds, 2),
                "avg_cycle_time_display": f"{int(avg_cycle_time_seconds // 60)}m {int(avg_cycle_time_seconds % 60)}s",
                "posted_by_id": current_user.user_id,
                "posted_by_name": current_user.name,
                "posted_at": now.isoformat(),
                "is_final": True
            }
            production_posts.append(final_post)
            total_posted_qty += qty_since_last_post
    
    # Calculate production loss if cycle_time is available
    production_loss = None
    production_loss_percent = None
    expected_production = None
    calculated_avg_cycle_time = None
    
    start_time = work.get("start_time")
    if start_time:
        if isinstance(start_time, str):
            start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
        if start_time.tzinfo is None:
            start_time = start_time.replace(tzinfo=timezone.utc)
        
        total_duration_seconds = (now - start_time).total_seconds()
        
        # Calculate overall average cycle time
        if final_qty and final_qty > 0:
            calculated_avg_cycle_time = round(total_duration_seconds / final_qty, 2)
        
        # Calculate expected production based on standard cycle time
        standard_cycle_time = work.get("cycle_time")  # in seconds
        if standard_cycle_time and standard_cycle_time > 0 and final_qty:
            expected_production = int(total_duration_seconds / standard_cycle_time)
            production_loss = expected_production - final_qty
            if expected_production > 0:
                production_loss_percent = round((production_loss / expected_production) * 100, 1)
    
    # Create production entry from work
    production_entry = {
        "entry_id": f"prod_{uuid.uuid4().hex[:12]}",
        "date": work.get("date"),
        "category": work.get("category"),
        "sub_category": work.get("sub_category"),
        "machine_id": work.get("machine_id"),
        "machine_name": work.get("machine_name"),
        "job_id": work.get("job_id"),
        "job_details": work.get("job_details"),
        "production_qty": final_qty,
        "cycle_time": work.get("cycle_time"),
        "setting_time": work.get("setting_time"),
        "spindle_rpm": work.get("spindle_rpm"),
        "feed_rate": work.get("feed_rate"),
        "shift": work.get("shift"),
        "remarks": final_remarks,
        "end_remarks": end_data.get("end_remarks", ""),
        "image": work.get("image"),
        "start_time": work.get("start_time"),
        "start_time_ist": work.get("start_time_ist"),
        "end_time": now,
        "end_time_ist": format_ist_datetime(now),
        "estimated_time": work.get("estimated_time"),
        "operator_id": work.get("operator_id"),
        "operator_name": work.get("operator_name"),
        "assigned_to_name": work.get("assigned_to_name", ""),
        "done_by_name": work.get("done_by_name", work.get("created_by_name", "")),
        "created_by_id": work.get("created_by_id"),
        "created_by_name": work.get("created_by_name"),
        "created_at": work.get("created_at"),
        "completed_at": now,
        "ended_by": current_user.name,
        "system_ended": False,
        "status": "completed",
        # Production posts tracking
        "production_posts": production_posts,
        "total_posted_qty": total_posted_qty,
        "qty_since_last_post": qty_since_last_post,
        "calculated_avg_cycle_time": calculated_avg_cycle_time,
        "expected_production": expected_production,
        "production_loss": production_loss,
        "production_loss_percent": production_loss_percent
    }
    
    await db.production_entries.insert_one(production_entry)
    
    # Update work entry status
    await db.active_work.update_one(
        {"entry_id": entry_id},
        {"$set": {"status": "completed", "end_time": now}}
    )
    
    # Reset machine status if applicable
    if work.get("machine_id"):
        await db.machines.update_one(
            {"machine_id": work.get("machine_id")},
            {"$set": {
                "status": "idle",
                "current_job_id": None,
                "current_operator_id": None,
                "last_status_update": now
            }}
        )
    
    return {
        "message": "Work completed and saved to production sheet",
        "production_entry_id": production_entry["entry_id"]
    }


@app.put("/api/production/edit/{entry_id}")
async def edit_production_work(
    entry_id: str,
    update_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Edit an active work entry"""
    work = await db.active_work.find_one({"entry_id": entry_id})
    
    if not work:
        raise HTTPException(status_code=404, detail="Work entry not found")
    
    # Check permission - operators can edit their own, admins can edit all
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"} and work.get("operator_id") != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not authorized to edit this work")
    
    allowed_fields = ["job_details", "quantity", "setting_time", "cycle_time", "remarks", "image"]
    update_dict = {k: v for k, v in update_data.items() if k in allowed_fields}
    update_dict["updated_at"] = datetime.now(timezone.utc)
    
    await db.active_work.update_one({"entry_id": entry_id}, {"$set": update_dict})
    
    return {"message": "Work updated successfully"}


@app.delete("/api/production/delete/{entry_id}")
async def delete_work_entry(
    entry_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete active work entry - Admin only"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    work = await db.active_work.find_one({"entry_id": entry_id})
    
    if not work:
        raise HTTPException(status_code=404, detail="Work entry not found")
    
    # Reset machine status
    if work.get("machine_id"):
        await db.machines.update_one(
            {"machine_id": work.get("machine_id")},
            {"$set": {"status": "idle", "current_operator": None, "current_job": None}}
        )
    
    # Delete the active work entry
    await db.active_work.delete_one({"entry_id": entry_id})
    
    logger.info(f"Admin {current_user.name} deleted work entry {entry_id}")
    
    return {"message": "Work entry deleted successfully"}


# ============ SHIFT HANDOVER SYSTEM ============

@app.get("/api/shift/current")
async def get_current_shift_info():
    """Get current shift information"""
    now = get_ist_now()
    shift = get_current_shift()
    shift_end = get_shift_end_time()
    time_remaining = shift_end - now
    
    return {
        "current_shift": shift,
        "shift_end_time": shift_end.strftime('%I:%M %p'),
        "shift_end_date": shift_end.strftime('%d/%m/%Y'),
        "time_remaining_minutes": int(time_remaining.total_seconds() / 60),
        "time_remaining_display": f"{int(time_remaining.total_seconds() // 3600)}h {int((time_remaining.total_seconds() % 3600) // 60)}m"
    }

@app.post("/api/production/handover/{entry_id}")
async def handover_work(
    entry_id: str,
    handover_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Handover active work to another operator"""
    work = await db.active_work.find_one({"entry_id": entry_id})
    
    if not work:
        raise HTTPException(status_code=404, detail="Work entry not found")
    
    # Check permission - only current operator or admin can handover
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"} and work.get("operator_id") != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not authorized to handover this work")
    
    new_operator_id = handover_data.get("new_operator_id")
    new_operator_name = handover_data.get("new_operator_name")
    handover_remarks = handover_data.get("remarks", "")
    
    if not new_operator_name:
        raise HTTPException(status_code=400, detail="New operator name is required")
    
    now = datetime.now(timezone.utc)
    
    # Create handover record
    handover_record = {
        "handover_id": str(uuid.uuid4()),
        "entry_id": entry_id,
        "from_operator_id": work.get("operator_id"),
        "from_operator_name": work.get("operator_name"),
        "to_operator_id": new_operator_id,
        "to_operator_name": new_operator_name,
        "handover_time": now,
        "handover_time_ist": format_ist_datetime(now),
        "shift": get_current_shift(),
        "remarks": handover_remarks,
        "status": "pending",  # pending, accepted, rejected
        "job_details": work.get("job_details"),
        "machine_name": work.get("machine_name"),
        "category": work.get("category")
    }
    
    await db.handovers.insert_one(handover_record)
    
    # Update active work with pending handover
    await db.active_work.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "handover_pending": True,
            "handover_to": new_operator_name,
            "handover_time": now
        }}
    )
    
    # Create notification for new operator
    notification = {
        "notification_id": str(uuid.uuid4()),
        "user_id": new_operator_id,
        "type": "handover_request",
        "title": "Handover Request",
        "message": f"{work.get('operator_name')} wants to handover work on {work.get('machine_name', 'job')} to you",
        "data": {"entry_id": entry_id, "handover_id": handover_record["handover_id"]},
        "is_read": False,
        "created_at": now
    }
    await db.notifications.insert_one(notification)
    
    # Update operator stats
    await db.operator_stats.update_one(
        {"user_id": current_user.user_id},
        {"$inc": {"handovers_given": 1}, "$set": {"last_activity": now}},
        upsert=True
    )
    
    return {"message": "Handover request sent", "handover_id": handover_record["handover_id"]}

@app.post("/api/production/takeover/{entry_id}")
async def takeover_work(
    entry_id: str,
    takeover_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Take over active work from another operator (operator-initiated)"""
    work = await db.active_work.find_one({"entry_id": entry_id})
    
    if not work:
        raise HTTPException(status_code=404, detail="Work entry not found")
    
    # Can't take over your own work
    if work.get("operator_id") == current_user.user_id:
        raise HTTPException(status_code=400, detail="Cannot take over your own work")
    
    now = datetime.now(timezone.utc)
    previous_operator = work.get("operator_name")
    
    # Create handover record (auto-accepted takeover)
    handover_record = {
        "handover_id": str(uuid.uuid4()),
        "entry_id": entry_id,
        "from_operator_id": work.get("operator_id"),
        "from_operator_name": previous_operator,
        "to_operator_id": current_user.user_id,
        "to_operator_name": current_user.name,
        "handover_time": now,
        "handover_time_ist": format_ist_datetime(now),
        "shift": get_current_shift(),
        "remarks": takeover_data.get("remarks", "Operator-initiated takeover"),
        "status": "accepted",
        "type": "takeover",
        "job_details": work.get("job_details"),
        "machine_name": work.get("machine_name"),
        "category": work.get("category")
    }
    
    await db.handovers.insert_one(handover_record)
    
    # Update active work with new operator
    await db.active_work.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "operator_id": current_user.user_id,
            "operator_name": current_user.name,
            "previous_operator": previous_operator,
            "takeover_time": now,
            "handover_pending": False,
            "handover_to": None
        },
        "$push": {
            "operator_history": {
                "operator_name": previous_operator,
                "end_time": now,
                "reason": "takeover"
            }
        }}
    )
    
    # Notify previous operator
    notification = {
        "notification_id": str(uuid.uuid4()),
        "user_id": work.get("operator_id"),
        "type": "work_taken_over",
        "title": "Work Taken Over",
        "message": f"{current_user.name} has taken over your work on {work.get('machine_name', 'job')}",
        "data": {"entry_id": entry_id},
        "is_read": False,
        "created_at": now
    }
    await db.notifications.insert_one(notification)
    
    # Update operator stats
    await db.operator_stats.update_one(
        {"user_id": current_user.user_id},
        {"$inc": {"handovers_received": 1}, "$set": {"last_activity": now}},
        upsert=True
    )
    
    return {"message": f"Work taken over from {previous_operator}"}

@app.post("/api/production/handover/accept/{handover_id}")
async def accept_handover(
    handover_id: str,
    current_user: User = Depends(get_current_user)
):
    """Accept a pending handover request"""
    handover = await db.handovers.find_one({"handover_id": handover_id})
    
    if not handover:
        raise HTTPException(status_code=404, detail="Handover not found")
    
    if handover.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Handover already processed")
    
    # Verify current user is the intended recipient
    if handover.get("to_operator_name") != current_user.name and (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Not authorized to accept this handover")
    
    now = datetime.now(timezone.utc)
    entry_id = handover.get("entry_id")
    
    # Update handover status
    await db.handovers.update_one(
        {"handover_id": handover_id},
        {"$set": {"status": "accepted", "accepted_at": now}}
    )
    
    # Update active work with new operator
    await db.active_work.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "operator_id": current_user.user_id,
            "operator_name": current_user.name,
            "previous_operator": handover.get("from_operator_name"),
            "handover_accepted_at": now,
            "handover_pending": False,
            "handover_to": None
        },
        "$push": {
            "operator_history": {
                "operator_name": handover.get("from_operator_name"),
                "end_time": now,
                "reason": "handover_accepted"
            }
        }}
    )
    
    # Update operator stats for receiver
    await db.operator_stats.update_one(
        {"user_id": current_user.user_id},
        {"$inc": {"handovers_received": 1}, "$set": {"last_activity": now}},
        upsert=True
    )
    
    return {"message": "Handover accepted"}

@app.get("/api/production/pending-handovers")
async def get_pending_handovers(current_user: User = Depends(get_current_user)):
    """Get pending handover requests for current user"""
    handovers = await db.handovers.find({
        "to_operator_name": current_user.name,
        "status": "pending"
    }, {"_id": 0}).to_list(length=50)
    
    return handovers

@app.get("/api/production/available-for-takeover")
async def get_available_for_takeover(current_user: User = Depends(get_current_user)):
    """Get active work entries that can be taken over (VMC, CNC, Moulding only)"""
    # Get all active VMC, CNC, Moulding work
    entries = await db.active_work.find({
        "status": "active",
        "category": {"$in": ["vmc", "cnc", "moulding", "VMC", "CNC", "Moulding"]}
    }, {"_id": 0}).to_list(length=50)
    
    # Mark which ones belong to current user (can't take over own work)
    for entry in entries:
        entry["is_own_work"] = entry.get("operator_id") == current_user.user_id
    
    return entries

@app.get("/api/production/all-active-machine-work")
async def get_all_active_machine_work():
    """Get all active VMC, CNC, Moulding work (public for takeover view)"""
    entries = await db.active_work.find({
        "status": "active",
        "category": {"$in": ["vmc", "cnc", "moulding", "VMC", "CNC", "Moulding"]}
    }, {"_id": 0}).to_list(length=50)
    
    return entries

# ============ AUTO-END SHIFT WORK ============

async def scheduled_auto_end_shift():
    """
    Scheduled function called by APScheduler at shift boundaries (8PM and 6AM IST).
    This function auto-ends all active work and updates operator stats with system_ended count.
    """
    logger.info("⏰ CRON JOB: Auto-end shift triggered by scheduler")
    now = datetime.now(timezone.utc)
    
    try:
        # Find all active work entries
        active_entries = await db.active_work.find({
            "$or": [{"status": "active"}, {"status": "in_progress"}]
        }).to_list(length=1000)
        
        if not active_entries:
            logger.info("⏰ CRON JOB: No active work entries to auto-end")
            return
        
        ended_count = 0
        operators_affected = set()
        
        for work in active_entries:
            # Generate new entry_id for production sheet - CRITICAL FIX: always generate new UUID
            production_entry_id = f"auto_{uuid.uuid4().hex[:12]}"
            start_time = work.get("start_time")
            operator_id = work.get("operator_id")
            
            # Calculate duration
            if start_time:
                if isinstance(start_time, str):
                    start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                # Ensure both datetimes are timezone-aware for comparison
                if start_time.tzinfo is None:
                    start_time = start_time.replace(tzinfo=timezone.utc)
                duration_seconds = (now - start_time).total_seconds()
                duration_minutes = int(duration_seconds / 60)
                hours = duration_minutes // 60
                mins = duration_minutes % 60
                duration_display = f"{hours}h {mins}m"
            else:
                duration_display = "Unknown"
            
            # Create production entry with system_ended flag
            production_entry = {
                "entry_id": production_entry_id,
                "category": work.get("category"),
                "sub_category": work.get("sub_category"),
                "machine_id": work.get("machine_id"),
                "machine_name": work.get("machine_name"),
                "job_id": work.get("job_id"),
                "job_details": work.get("job_details"),
                "operator_id": operator_id,
                "operator_name": work.get("operator_name"),
                "operator_history": work.get("operator_history", []),
                "quantity": work.get("quantity"),
                "produced_qty": 0,
                "rejection_qty": 0,
                "setting_time": work.get("setting_time"),
                "cycle_time": work.get("cycle_time"),
                "spindle_rpm": work.get("spindle_rpm"),  # CNC spindle RPM
                "feed_rate": work.get("feed_rate"),  # CNC feed rate
                "start_time": work.get("start_time"),
                "start_time_ist": work.get("start_time_ist"),
                "end_time": now,
                "end_time_ist": format_ist_datetime(now),
                "duration": duration_display,
                "shift": get_shift_for_time(start_time),
                "remarks": work.get("remarks"),
                "end_remarks": "⏰ Auto-Ended by System at Shift Boundary",
                "ended_by": "System",
                "system_ended": True,  # Flag for tracking auto-ended entries
                "image": work.get("image"),
                "date": to_ist(now).strftime('%Y-%m-%d'),
                "created_at": now,
                "created_by_id": operator_id,
                "created_by_name": work.get("operator_name"),
                "status": "completed"  # Required for Daily Summary display
            }
            
            await db.production_entries.insert_one(production_entry)
            logger.info(f"⏰ CRON JOB: Created production entry {production_entry_id} for operator {work.get('operator_name')}")
            
            # Track affected operators for stats update
            if operator_id:
                operators_affected.add(operator_id)
            
            # Reset machine status
            if work.get("machine_id"):
                await db.machines.update_one(
                    {"machine_id": work.get("machine_id")},
                    {"$set": {"status": "idle", "current_operator": None, "current_job": None}}
                )
            
            # Remove from active work - use _id as fallback if entry_id is missing
            work_entry_id = work.get("entry_id")
            if work_entry_id:
                await db.active_work.delete_one({"entry_id": work_entry_id})
            else:
                await db.active_work.delete_one({"_id": work.get("_id")})
            ended_count += 1
        
        # Update system_ended count for affected operators
        for op_id in operators_affected:
            await db.users.update_one(
                {"user_id": op_id},
                {"$inc": {"system_ended_count": 1}}
            )
        
        logger.info(f"⏰ CRON JOB: Auto-ended {ended_count} work entries for {len(operators_affected)} operators")
        
    except Exception as e:
        logger.error(f"⏰ CRON JOB ERROR: Failed to auto-end shift work: {e}")

@app.get("/api/scheduler/status")
async def get_scheduler_status(current_user: User = Depends(get_current_user)):
    """Get APScheduler status (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    jobs = []
    if scheduler.running:
        for job in scheduler.get_jobs():
            jobs.append({
                "id": job.id,
                "next_run": str(job.next_run_time) if job.next_run_time else None,
                "trigger": str(job.trigger)
            })
    
    return {
        "running": scheduler.running,
        "timezone": str(scheduler.timezone),
        "jobs": jobs
    }

@app.post("/api/scheduler/trigger-auto-end")
async def manually_trigger_auto_end(current_user: User = Depends(get_current_user)):
    """Manually trigger auto-end for testing (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    await scheduled_auto_end_shift()
    return {"message": "Auto-end triggered manually", "timestamp": get_ist_now().strftime('%d/%m/%Y %I:%M %p')}


@app.get("/api/admin/migrate-factory-numbers", response_class=HTMLResponse)
async def migrate_factory_numbers_page(current_user: User = Depends(get_current_user)):
    """Show migration page with a button to run the migration"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    return HTMLResponse(content="""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Migrate Factory Numbers</title>
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <style>
            body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #0F172A; color: white; padding: 20px; margin: 0; }
            .container { max-width: 600px; margin: 0 auto; padding-top: 30px; }
            h1 { color: #3B82F6; }
            .info { background: #1E293B; padding: 20px; border-radius: 12px; margin-bottom: 20px; }
            .warning { background: #7F1D1D; padding: 15px; border-radius: 8px; margin-bottom: 20px; border: 1px solid #EF4444; }
            button { background: #3B82F6; color: white; border: none; padding: 15px 30px; border-radius: 8px; font-size: 16px; cursor: pointer; width: 100%; }
            button:hover { background: #2563EB; }
            button:disabled { background: #475569; cursor: not-allowed; }
            #result { margin-top: 20px; padding: 15px; border-radius: 8px; display: none; }
            .success { background: #166534; }
            .error { background: #7F1D1D; }
            pre { white-space: pre-wrap; word-wrap: break-word; font-size: 12px; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🔧 Migrate Factory Numbers</h1>
            
            <div class="info">
                <h3>What this does:</h3>
                <ul>
                    <li>Finds all parts with duplicate F-numbers across assemblies</li>
                    <li>Keeps the FIRST occurrence of each F-number</li>
                    <li>Renumbers duplicate parts to have unique F-numbers globally</li>
                    <li>Updates all references (assembly_parts, storage, production)</li>
                </ul>
            </div>
            
            <div class="warning">
                ⚠️ <strong>Run this ONCE only!</strong> This is a one-time migration to fix the duplicate F-numbers issue.
            </div>
            
            <button id="migrateBtn" onclick="runMigration()">
                🚀 Run Migration
            </button>
            
            <div id="result"></div>
        </div>
        
        <script>
            async function runMigration() {
                const btn = document.getElementById('migrateBtn');
                const result = document.getElementById('result');
                
                btn.disabled = true;
                btn.textContent = '⏳ Running migration...';
                result.style.display = 'none';
                
                try {
                    const response = await fetch('/api/admin/migrate-factory-numbers', {
                        method: 'POST',
                        credentials: 'include',
                        headers: { 'Content-Type': 'application/json' }
                    });
                    
                    const data = await response.json();
                    
                    result.style.display = 'block';
                    if (response.ok) {
                        result.className = 'success';
                        result.innerHTML = '<h3>✅ Migration Complete!</h3><pre>' + JSON.stringify(data, null, 2) + '</pre>';
                    } else {
                        result.className = 'error';
                        result.innerHTML = '<h3>❌ Error</h3><pre>' + JSON.stringify(data, null, 2) + '</pre>';
                    }
                } catch (err) {
                    result.style.display = 'block';
                    result.className = 'error';
                    result.innerHTML = '<h3>❌ Network Error</h3><p>' + err.message + '</p>';
                }
                
                btn.disabled = false;
                btn.textContent = '🚀 Run Migration Again';
            }
        </script>
    </body>
    </html>
    """)

@app.post("/api/admin/migrate-factory-numbers")
async def migrate_factory_numbers(current_user: User = Depends(get_current_user)):
    """
    Migrate existing parts to have GLOBALLY UNIQUE factory numbers (F0001, F0002...).
    This fixes the bug where each assembly's parts started from F0001.
    
    Admin only endpoint. Run this ONCE after the fix is deployed.
    """
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        # Find all parts with F-number pattern, sorted by created_at (oldest first)
        parts = await db.parts_library.find(
            {"part_id": {"$regex": r"-F\d+"}},
            {"part_id": 1, "name": 1, "created_at": 1}
        ).sort("created_at", 1).to_list(length=10000)
        
        if not parts:
            return {
                "success": True,
                "message": "No parts with F-numbers found",
                "migrated_count": 0
            }
        
        # Track used F-numbers to detect duplicates
        f_number_usage = {}  # f_number -> list of part_ids
        for p in parts:
            part_id = p.get("part_id", "")
            match = re.search(r'-F(\d+)', part_id)
            if match:
                f_num = int(match.group(1))
                if f_num not in f_number_usage:
                    f_number_usage[f_num] = []
                f_number_usage[f_num].append(part_id)
        
        # Find duplicates (F-numbers used by more than one part)
        duplicates = {k: v for k, v in f_number_usage.items() if len(v) > 1}
        
        if not duplicates:
            return {
                "success": True,
                "message": "No duplicate F-numbers found - all numbers are unique",
                "total_parts": len(parts),
                "migrated_count": 0
            }
        
        # Get the highest existing F-number
        max_f_number = max(f_number_usage.keys())
        next_f_number = max_f_number + 1
        
        migrated_parts = []
        
        # Process duplicates - keep the FIRST occurrence, renumber others
        for f_num, part_ids in sorted(duplicates.items()):
            # Skip the first part (keep its F-number)
            for part_id in part_ids[1:]:
                # Generate new part_id with unique F-number
                old_part_id = part_id
                new_part_id = re.sub(r'-F\d+', f'-F{next_f_number:04d}', old_part_id)
                
                # Update part_id in parts_library
                await db.parts_library.update_one(
                    {"part_id": old_part_id},
                    {"$set": {
                        "part_id": new_part_id,
                        "factory_no": next_f_number,
                        "migrated_from": old_part_id,
                        "migrated_at": get_ist_now()
                    }}
                )
                
                # Update part_id in assembly_parts (where-used tracking)
                await db.assembly_parts.update_many(
                    {"part_id": old_part_id},
                    {"$set": {"part_id": new_part_id}}
                )
                
                # Update any storage locations referencing this part
                await db.storage_locations.update_many(
                    {"parts.part_id": old_part_id},
                    {"$set": {"parts.$.part_id": new_part_id}}
                )
                
                # Update any production entries referencing this part
                await db.production_entries.update_many(
                    {"part_id": old_part_id},
                    {"$set": {"part_id": new_part_id}}
                )
                
                migrated_parts.append({
                    "old_part_id": old_part_id,
                    "new_part_id": new_part_id,
                    "old_factory_no": f_num,
                    "new_factory_no": next_f_number
                })
                
                next_f_number += 1
        
        return {
            "success": True,
            "message": f"Successfully migrated {len(migrated_parts)} parts to unique F-numbers",
            "total_parts_with_f_numbers": len(parts),
            "duplicate_f_numbers_found": len(duplicates),
            "migrated_count": len(migrated_parts),
            "next_available_f_number": next_f_number,
            "migrated_parts": migrated_parts[:50]  # Return first 50 for reference
        }
        
    except Exception as e:
        logger.error(f"Error migrating factory numbers: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Migration failed: {str(e)}")


@app.post("/api/admin/undo-auto-end")
async def undo_auto_end(current_user: User = Depends(get_current_user)):
    """
    Undo auto-ended entries and restore them to active work.
    Admin only endpoint.
    """
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        # Find all system-ended entries
        auto_ended = await db.production_entries.find({
            "$or": [
                {"system_ended": True},
                {"ended_by": "System"}
            ]
        }).to_list(length=500)
        
        if not auto_ended:
            return {
                "success": True,
                "message": "No auto-ended entries found to restore",
                "restored_count": 0
            }
        
        restored_count = 0
        restored_entries = []
        now = get_ist_now()
        
        for entry in auto_ended:
            # Create active_work entry with ALL required fields
            active_work = {
                "entry_id": entry.get("entry_id"),
                "date": entry.get("date") or now.date().isoformat(),
                "category": entry.get("category"),
                "sub_category": entry.get("sub_category"),
                "machine_id": entry.get("machine_id"),
                "machine_name": entry.get("machine_name"),
                "job_id": entry.get("job_id"),
                "job_details": entry.get("job_details"),
                "operator_id": entry.get("operator_id"),
                "operator_name": entry.get("operator_name"),
                "quantity": entry.get("quantity"),
                "setting_time": entry.get("setting_time"),
                "cycle_time": entry.get("cycle_time"),
                "start_time": entry.get("start_time"),
                "start_time_ist": entry.get("start_time_ist"),
                "end_time": None,
                "estimated_time": entry.get("estimated_time"),
                "shift": entry.get("shift"),
                "remarks": entry.get("remarks"),
                "image": entry.get("image"),
                "created_at": entry.get("created_at"),
                "created_by_id": entry.get("created_by_id"),
                "created_by_name": entry.get("created_by_name"),
                "assigned_to_name": entry.get("assigned_to_name", ""),
                "done_by_name": entry.get("done_by_name", ""),
                "status": "in_progress"
            }
            
            # Insert back to active_work
            await db.active_work.insert_one(active_work)
            
            # Delete from production_entries
            await db.production_entries.delete_one({"entry_id": entry.get("entry_id")})
            
            # Reset machine status with ALL required fields
            if entry.get("machine_id"):
                await db.machines.update_one(
                    {"machine_id": entry.get("machine_id")},
                    {"$set": {
                        "status": "running",
                        "current_job_id": entry.get("entry_id"),
                        "current_operator_id": entry.get("operator_id"),
                        "current_operator": entry.get("operator_name"),
                        "current_job": entry.get("job_details"),
                        "last_status_update": now
                    }}
                )
            
            restored_entries.append({
                "entry_id": entry.get("entry_id"),
                "machine_name": entry.get("machine_name"),
                "operator_name": entry.get("operator_name"),
                "job_details": entry.get("job_details")
            })
            restored_count += 1
        
        # Decrement system_ended_count for affected operators
        operator_counts = {}
        for entry in auto_ended:
            op_id = entry.get("operator_id")
            if op_id:
                operator_counts[op_id] = operator_counts.get(op_id, 0) + 1
        
        for op_id, count in operator_counts.items():
            await db.users.update_one(
                {"user_id": op_id},
                {"$inc": {"system_ended_count": -count}}
            )
        
        logger.info(f"✅ Admin {current_user.name} restored {restored_count} auto-ended entries")
        
        return {
            "success": True,
            "message": f"Successfully restored {restored_count} entries to active work",
            "restored_count": restored_count,
            "restored_entries": restored_entries,
            "timestamp": get_ist_now().strftime('%d/%m/%Y %I:%M %p')
        }
        
    except Exception as e:
        logger.error(f"Error undoing auto-end: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to undo auto-end: {str(e)}")

@app.post("/api/admin/undo-auto-end/{entry_id}")
async def undo_single_auto_end(entry_id: str, current_user: User = Depends(get_current_user)):
    """
    Undo a SINGLE auto-ended entry and restore it to active work.
    Admin only endpoint.
    """
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        # Find the specific entry
        entry = await db.production_entries.find_one({
            "entry_id": entry_id,
            "$or": [
                {"system_ended": True},
                {"ended_by": "System"}
            ]
        })
        
        if not entry:
            raise HTTPException(status_code=404, detail="Entry not found or not a system-ended entry")
        
        now = get_ist_now()
        
        # Create active_work entry with ALL required fields
        active_work = {
            "entry_id": entry.get("entry_id"),
            "date": entry.get("date") or now.date().isoformat(),
            "category": entry.get("category"),
            "sub_category": entry.get("sub_category"),
            "machine_id": entry.get("machine_id"),
            "machine_name": entry.get("machine_name"),
            "job_id": entry.get("job_id"),
            "job_details": entry.get("job_details"),
            "operator_id": entry.get("operator_id"),
            "operator_name": entry.get("operator_name"),
            "quantity": entry.get("quantity"),
            "setting_time": entry.get("setting_time"),
            "cycle_time": entry.get("cycle_time"),
            "start_time": entry.get("start_time"),
            "start_time_ist": entry.get("start_time_ist"),
            "end_time": None,
            "estimated_time": entry.get("estimated_time"),
            "shift": entry.get("shift"),
            "remarks": entry.get("remarks"),
            "image": entry.get("image"),
            "created_at": entry.get("created_at"),
            "created_by_id": entry.get("created_by_id"),
            "created_by_name": entry.get("created_by_name"),
            "assigned_to_name": entry.get("assigned_to_name", ""),
            "done_by_name": entry.get("done_by_name", ""),
            "status": "in_progress"
        }
        
        # Insert back to active_work
        await db.active_work.insert_one(active_work)
        
        # Delete from production_entries
        await db.production_entries.delete_one({"entry_id": entry_id})
        
        # Reset machine status with ALL required fields
        if entry.get("machine_id"):
            await db.machines.update_one(
                {"machine_id": entry.get("machine_id")},
                {"$set": {
                    "status": "running",
                    "current_job_id": entry.get("entry_id"),
                    "current_operator_id": entry.get("operator_id"),
                    "current_operator": entry.get("operator_name"),
                    "current_job": entry.get("job_details"),
                    "last_status_update": now
                }}
            )
        
        # Decrement system_ended_count for operator
        if entry.get("operator_id"):
            await db.users.update_one(
                {"user_id": entry.get("operator_id")},
                {"$inc": {"system_ended_count": -1}}
            )
        
        logger.info(f"✅ Admin {current_user.name} restored entry {entry_id} to active work")
        
        return {
            "success": True,
            "message": "Entry restored to active work",
            "entry_id": entry_id,
            "machine_name": entry.get("machine_name"),
            "operator_name": entry.get("operator_name"),
            "timestamp": get_ist_now().strftime('%d/%m/%Y %I:%M %p')
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error undoing single auto-end: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to restore entry: {str(e)}")

@app.post("/api/shift/auto-end")
async def auto_end_shift_work():
    """Auto-end all active work at shift end (called by cron/scheduler)"""
    now = datetime.now(timezone.utc)
    
    # Find all active work entries (check both status values)
    active_entries = await db.active_work.find({
        "$or": [{"status": "active"}, {"status": "in_progress"}]
    }).to_list(length=1000)
    
    ended_count = 0
    for work in active_entries:
        entry_id = work.get("entry_id")
        start_time = work.get("start_time")
        
        # Calculate duration
        if start_time:
            if isinstance(start_time, str):
                start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
            duration_seconds = (now - start_time).total_seconds()
            duration_minutes = int(duration_seconds / 60)
            hours = duration_minutes // 60
            mins = duration_minutes % 60
            duration_display = f"{hours}h {mins}m"
        else:
            duration_display = "Unknown"
        
        # Create production entry
        production_entry = {
            "entry_id": entry_id,
            "category": work.get("category"),
            "sub_category": work.get("sub_category"),
            "machine_id": work.get("machine_id"),
            "machine_name": work.get("machine_name"),
            "job_id": work.get("job_id"),
            "job_details": work.get("job_details"),
            "operator_id": work.get("operator_id"),
            "operator_name": work.get("operator_name"),
            "operator_history": work.get("operator_history", []),
            "quantity": work.get("quantity"),
            "produced_qty": 0,
            "rejection_qty": 0,
            "setting_time": work.get("setting_time"),
            "cycle_time": work.get("cycle_time"),
            "start_time": work.get("start_time"),
            "start_time_ist": work.get("start_time_ist"),
            "end_time": now,
            "end_time_ist": format_ist_datetime(now),
            "duration": duration_display,
            "shift": get_shift_for_time(start_time),
            "remarks": work.get("remarks"),
            "end_remarks": "Shift Auto-End - Work not ended manually",
            "ended_by": "System",
            "image": work.get("image"),
            "date": to_ist(now).strftime('%Y-%m-%d'),
            "created_at": now,
            "created_by_id": work.get("operator_id"),
            "created_by_name": work.get("operator_name")
        }
        
        await db.production_entries.insert_one(production_entry)
        
        # Reset machine status
        if work.get("machine_id"):
            await db.machines.update_one(
                {"machine_id": work.get("machine_id")},
                {"$set": {"status": "idle", "current_operator": None, "current_job": None}}
            )
        
        # Remove from active work
        await db.active_work.delete_one({"entry_id": entry_id})
        ended_count += 1
    
    logger.info(f"Auto-ended {ended_count} work entries at shift change")
    return {"message": f"Auto-ended {ended_count} work entries", "count": ended_count}


@app.post("/api/shift/check-and-auto-end")
async def check_and_auto_end_expired_work():
    """Check for work from previous shift and auto-end them. Called on dashboard load."""
    now = get_ist_now()
    current_shift = get_current_shift()
    
    # Find all active work entries (check both status values)
    active_entries = await db.active_work.find({
        "$or": [{"status": "active"}, {"status": "in_progress"}]
    }).to_list(length=1000)
    
    ended_count = 0
    for work in active_entries:
        start_time = work.get("start_time")
        
        if not start_time:
            continue
            
        # Parse start time
        if isinstance(start_time, str):
            start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
        
        # Get the shift when work started
        work_shift = get_shift_for_time(start_time)
        start_time_ist = to_ist(start_time)
        
        # Check if work should be auto-ended
        should_auto_end = False
        
        # If work started in morning shift (8AM-8PM) and current time is past 8PM
        if work_shift == "morning":
            # Morning shift ends at 8PM
            shift_end = start_time_ist.replace(hour=20, minute=0, second=0, microsecond=0)
            if now > shift_end:
                should_auto_end = True
        
        # If work started in night shift (8PM-6AM) 
        elif work_shift == "night":
            # Night shift ends at 6AM next day
            if start_time_ist.hour >= 20:
                # Started after 8PM, ends at 6AM next day
                shift_end = (start_time_ist + timedelta(days=1)).replace(hour=6, minute=0, second=0, microsecond=0)
            else:
                # Started before 6AM (like 2AM), ends at 6AM same day
                shift_end = start_time_ist.replace(hour=6, minute=0, second=0, microsecond=0)
            
            if now > shift_end:
                should_auto_end = True
        
        if should_auto_end:
            entry_id = work.get("entry_id")
            end_time = datetime.now(timezone.utc)
            
            # Ensure start_time is timezone-aware for calculation
            start_time_calc = start_time
            if start_time_calc.tzinfo is None:
                start_time_calc = start_time_calc.replace(tzinfo=timezone.utc)
            
            # Calculate duration
            duration_seconds = (end_time - start_time_calc).total_seconds()
            duration_minutes = int(duration_seconds / 60)
            hours = duration_minutes // 60
            mins = duration_minutes % 60
            duration_display = f"{hours}h {mins}m"
            
            # Create production entry
            production_entry = {
                "entry_id": entry_id,
                "category": work.get("category"),
                "sub_category": work.get("sub_category"),
                "machine_id": work.get("machine_id"),
                "machine_name": work.get("machine_name"),
                "job_id": work.get("job_id"),
                "job_details": work.get("job_details"),
                "operator_id": work.get("operator_id"),
                "operator_name": work.get("operator_name"),
                "operator_history": work.get("operator_history", []),
                "quantity": work.get("quantity"),
                "produced_qty": 0,
                "rejection_qty": 0,
                "setting_time": work.get("setting_time"),
                "cycle_time": work.get("cycle_time"),
                "start_time": work.get("start_time"),
                "start_time_ist": work.get("start_time_ist"),
                "end_time": end_time,
                "end_time_ist": format_ist_datetime(end_time),
                "duration": duration_display,
                "shift": work_shift,
                "remarks": work.get("remarks"),
                "end_remarks": f"Auto-ended: Shift ({work_shift}) ended, work not completed",
                "ended_by": "System",
                "image": work.get("image"),
                "date": to_ist(end_time).strftime('%Y-%m-%d'),
                "created_at": end_time,
                "created_by_id": work.get("operator_id"),
                "created_by_name": work.get("operator_name")
            }
            
            await db.production_entries.insert_one(production_entry)
            
            # Reset machine status
            if work.get("machine_id"):
                await db.machines.update_one(
                    {"machine_id": work.get("machine_id")},
                    {"$set": {"status": "idle", "current_operator": None, "current_job": None}}
                )
            
            # Remove from active work
            await db.active_work.delete_one({"entry_id": entry_id})
            ended_count += 1
            
            logger.info(f"Auto-ended work {entry_id} for operator {work.get('operator_name')} - shift expired")
    
    return {
        "message": f"Checked and auto-ended {ended_count} expired work entries",
        "count": ended_count,
        "current_shift": current_shift
    }

# ============ OPERATOR STATS ============

@app.get("/api/operator-stats")
async def get_operator_stats(current_user: User = Depends(get_current_user)):
    """Get operator activity stats - MONTHLY stats with reset on 1st (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    now = get_ist_now()
    current_month = now.strftime('%Y-%m')
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Get all operators
    operators = await db.users.find(
        {"role": {"$regex": "Operator|Team Lead|Supervisor", "$options": "i"}},
        {"_id": 0, "user_id": 1, "name": 1, "email": 1, "role": 1}
    ).to_list(length=200)
    
    stats_list = []
    for op in operators:
        user_id = op.get("user_id")
        name = op.get("name")
        
        # Build query for current month entries only
        month_query = {
            "$and": [
                {"$or": [{"operator_name": name}, {"operator_id": user_id}]},
                {"date": {"$gte": month_start.strftime('%Y-%m-%d')}}
            ]
        }
        
        # Get stats from operator_stats collection (for handovers)
        stats = await db.operator_stats.find_one({"user_id": user_id}) or {}
        
        # Get production entry counts for current month
        total_entries = await db.production_entries.count_documents(month_query)
        
        proper_query = {
            "$and": [
                {"$or": [{"operator_name": name}, {"operator_id": user_id}]},
                {"date": {"$gte": month_start.strftime('%Y-%m-%d')}},
                {"$or": [
                    {"ended_by": {"$exists": False}},
                    {"ended_by": None},
                    {"ended_by": {"$ne": "System"}}
                ]}
            ]
        }
        proper_entries = await db.production_entries.count_documents(proper_query)
        
        system_query = {
            "$and": [
                {"$or": [{"operator_name": name}, {"operator_id": user_id}]},
                {"date": {"$gte": month_start.strftime('%Y-%m-%d')}},
                {"ended_by": "System"}
            ]
        }
        system_ended = await db.production_entries.count_documents(system_query)
        
        # Calculate accuracy
        accuracy = round((proper_entries / total_entries * 100), 1) if total_entries > 0 else 0
        
        # Calculate active score (higher is better)
        # Proper entries add points, system-ended subtract points
        active_score = (proper_entries * 10) - (system_ended * 5)
        active_score += stats.get("handovers_received", 0) * 2
        active_score -= stats.get("handovers_given", 0) * 1
        
        stats_list.append({
            "user_id": user_id,
            "name": op.get("name"),
            "email": op.get("email"),
            "role": op.get("role"),
            "total_entries": total_entries,
            "proper_entries": proper_entries,
            "system_ended": system_ended,
            "accuracy": accuracy,
            "handovers_given": stats.get("handovers_given", 0),
            "handovers_received": stats.get("handovers_received", 0),
            "active_score": active_score,
            "last_activity": stats.get("last_activity"),
            "month": current_month
        })
    
    # Sort by active score
    stats_list.sort(key=lambda x: x.get("active_score", 0), reverse=True)
    
    return {
        "current_month": current_month,
        "month_start": month_start.strftime('%Y-%m-%d'),
        "stats": stats_list
    }

@app.get("/api/operator-stats/top")
async def get_top_operators(month: Optional[str] = None):
    """Get star and lazy operator for dashboard - MONTHLY stats with reset on 1st
    
    Args:
        month: Optional month in YYYY-MM format. If not provided, uses current month.
    """
    now = get_ist_now()
    
    # Use provided month or current month
    if month:
        try:
            target_month = datetime.strptime(month + "-01", '%Y-%m-%d')
            current_month = month
        except ValueError:
            current_month = now.strftime('%Y-%m')
            target_month = now.replace(day=1)
    else:
        current_month = now.strftime('%Y-%m')
        target_month = now.replace(day=1)
    
    month_start = target_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Calculate month end (first day of next month)
    if month_start.month == 12:
        month_end = month_start.replace(year=month_start.year + 1, month=1)
    else:
        month_end = month_start.replace(month=month_start.month + 1)
    
    # Get all operators
    operators = await db.users.find(
        {"role": {"$regex": "Operator|Team Lead|Supervisor", "$options": "i"}},
        {"_id": 0, "user_id": 1, "name": 1, "role": 1}
    ).to_list(length=200)
    
    stats_list = []
    for op in operators:
        name = op.get("name")
        user_id = op.get("user_id")
        
        # Build query for the specified month entries only
        month_query_base = {
            "$or": [
                {"operator_name": name},
                {"operator_id": user_id}
            ],
            "date": {"$gte": month_start.strftime('%Y-%m-%d'), "$lt": month_end.strftime('%Y-%m-%d')}
        }
        
        # Count entries for the month
        total_entries = await db.production_entries.count_documents(month_query_base)
        
        # Count proper entries (not ended by System) for the month
        proper_query = {
            "$and": [
                {"$or": [{"operator_name": name}, {"operator_id": user_id}]},
                {"date": {"$gte": month_start.strftime('%Y-%m-%d'), "$lt": month_end.strftime('%Y-%m-%d')}},
                {"$or": [
                    {"ended_by": {"$exists": False}},
                    {"ended_by": None},
                    {"ended_by": {"$ne": "System"}}
                ]}
            ]
        }
        proper_entries = await db.production_entries.count_documents(proper_query)
        
        # Count system ended for the month
        system_query = {
            "$and": [
                {"$or": [{"operator_name": name}, {"operator_id": user_id}]},
                {"date": {"$gte": month_start.strftime('%Y-%m-%d'), "$lt": month_end.strftime('%Y-%m-%d')}},
                {"ended_by": "System"}
            ]
        }
        system_ended = await db.production_entries.count_documents(system_query)
        
        # Calculate accuracy (proper entries / total entries * 100)
        accuracy = round((proper_entries / total_entries * 100), 1) if total_entries > 0 else 0
        
        # Calculate score: based on entry count and accuracy
        # Score = (total_entries * 2) + (accuracy * 0.5) - (system_ended * 5)
        active_score = (total_entries * 2) + (accuracy * 0.5) - (system_ended * 5)
        
        stats_list.append({
            "user_id": user_id,
            "name": name,
            "role": op.get("role"),
            "total_entries": total_entries,
            "proper_entries": proper_entries,
            "system_ended": system_ended,
            "accuracy": accuracy,
            "active_score": round(active_score, 1),
            "month": current_month
        })
    
    # Sort by score (highest first)
    stats_list.sort(key=lambda x: x.get("active_score", 0), reverse=True)
    
    # Filter only operators with entries
    operators_with_entries = [op for op in stats_list if op.get("total_entries", 0) > 0]
    
    star_operator = operators_with_entries[0] if operators_with_entries else None
    lazy_operator = operators_with_entries[-1] if len(operators_with_entries) > 1 else None
    
    return {
        "current_month": current_month,
        "star_operator": {
            "name": star_operator.get("name") if star_operator else None,
            "score": star_operator.get("active_score") if star_operator else 0,
            "entries": star_operator.get("total_entries") if star_operator else 0,
            "system_ended": star_operator.get("system_ended") if star_operator else 0,
            "accuracy": star_operator.get("accuracy") if star_operator else 0
        } if star_operator else None,
        "lazy_operator": {
            "name": lazy_operator.get("name") if lazy_operator else None,
            "score": lazy_operator.get("active_score") if lazy_operator else 0,
            "entries": lazy_operator.get("total_entries") if lazy_operator else 0,
            "system_ended": lazy_operator.get("system_ended") if lazy_operator else 0,
            "accuracy": lazy_operator.get("accuracy") if lazy_operator else 0
        } if lazy_operator else None,
        "all_stats": stats_list
    }


@app.get("/api/operator-stats/weekly")
async def get_weekly_operator_stats():
    """Get operator stats for the current week (Monday to Sunday)"""
    now = get_ist_now()
    
    # Calculate week start (Monday) and end (Sunday)
    days_since_monday = now.weekday()
    week_start = (now - timedelta(days=days_since_monday)).replace(hour=0, minute=0, second=0, microsecond=0)
    week_end = week_start + timedelta(days=7)
    
    # Get all operators
    operators = await db.users.find(
        {"role": {"$regex": "Operator|Team Lead|Supervisor", "$options": "i"}},
        {"_id": 0, "user_id": 1, "name": 1, "role": 1}
    ).to_list(length=200)
    
    stats_list = []
    for op in operators:
        name = op.get("name")
        user_id = op.get("user_id")
        
        # Count entries for this week
        week_query = {
            "$or": [
                {"operator_name": name},
                {"operator_id": user_id}
            ],
            "date": {"$gte": week_start.strftime('%Y-%m-%d'), "$lt": week_end.strftime('%Y-%m-%d')}
        }
        total_entries = await db.production_entries.count_documents(week_query)
        
        # Count proper entries
        proper_query = {
            "$and": [
                {"$or": [{"operator_name": name}, {"operator_id": user_id}]},
                {"date": {"$gte": week_start.strftime('%Y-%m-%d'), "$lt": week_end.strftime('%Y-%m-%d')}},
                {"$or": [
                    {"ended_by": {"$exists": False}},
                    {"ended_by": None},
                    {"ended_by": {"$ne": "System"}}
                ]}
            ]
        }
        proper_entries = await db.production_entries.count_documents(proper_query)
        
        # Count system ended
        system_query = {
            "$and": [
                {"$or": [{"operator_name": name}, {"operator_id": user_id}]},
                {"date": {"$gte": week_start.strftime('%Y-%m-%d'), "$lt": week_end.strftime('%Y-%m-%d')}},
                {"ended_by": "System"}
            ]
        }
        system_ended = await db.production_entries.count_documents(system_query)
        
        accuracy = round((proper_entries / total_entries * 100), 1) if total_entries > 0 else 0
        active_score = (total_entries * 2) + (accuracy * 0.5) - (system_ended * 5)
        
        if total_entries > 0:
            stats_list.append({
                "user_id": user_id,
                "name": name,
                "role": op.get("role"),
                "total_entries": total_entries,
                "proper_entries": proper_entries,
                "system_ended": system_ended,
                "accuracy": accuracy,
                "active_score": round(active_score, 1)
            })
    
    # Sort by score
    stats_list.sort(key=lambda x: x.get("active_score", 0), reverse=True)
    
    star_operator = stats_list[0] if stats_list else None
    
    return {
        "week_start": week_start.strftime('%Y-%m-%d'),
        "week_end": (week_end - timedelta(days=1)).strftime('%Y-%m-%d'),
        "star_operator": {
            "name": star_operator.get("name") if star_operator else None,
            "score": star_operator.get("active_score") if star_operator else 0,
            "entries": star_operator.get("total_entries") if star_operator else 0,
            "system_ended": star_operator.get("system_ended") if star_operator else 0,
            "accuracy": star_operator.get("accuracy") if star_operator else 0
        } if star_operator else None,
        "all_stats": stats_list
    }


@app.post("/api/operator-stats/archive-month")
async def archive_month_stats(current_user: User = Depends(get_current_user)):
    """Archive current month stats before reset (Admin only). Call on 1st of month."""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    now = get_ist_now()
    # Archive previous month stats
    prev_month = (now.replace(day=1) - timedelta(days=1))
    archive_month = prev_month.strftime('%Y-%m')
    month_start = prev_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_end = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(seconds=1)
    
    # Get all operators
    operators = await db.users.find(
        {"role": {"$regex": "Operator|Team Lead|Supervisor", "$options": "i"}},
        {"_id": 0, "user_id": 1, "name": 1, "role": 1}
    ).to_list(length=200)
    
    archived_count = 0
    for op in operators:
        name = op.get("name")
        user_id = op.get("user_id")
        
        # Count entries for the previous month
        month_query = {
            "$and": [
                {"$or": [{"operator_name": name}, {"operator_id": user_id}]},
                {"date": {"$gte": month_start.strftime('%Y-%m-%d'), "$lte": month_end.strftime('%Y-%m-%d')}}
            ]
        }
        total_entries = await db.production_entries.count_documents(month_query)
        
        if total_entries == 0:
            continue
        
        # Count proper entries
        proper_query = {
            "$and": [
                {"$or": [{"operator_name": name}, {"operator_id": user_id}]},
                {"date": {"$gte": month_start.strftime('%Y-%m-%d'), "$lte": month_end.strftime('%Y-%m-%d')}},
                {"$or": [
                    {"ended_by": {"$exists": False}},
                    {"ended_by": None},
                    {"ended_by": {"$ne": "System"}}
                ]}
            ]
        }
        proper_entries = await db.production_entries.count_documents(proper_query)
        
        # Count system ended
        system_query = {
            "$and": [
                {"$or": [{"operator_name": name}, {"operator_id": user_id}]},
                {"date": {"$gte": month_start.strftime('%Y-%m-%d'), "$lte": month_end.strftime('%Y-%m-%d')}},
                {"ended_by": "System"}
            ]
        }
        system_ended = await db.production_entries.count_documents(system_query)
        
        accuracy = round((proper_entries / total_entries * 100), 1) if total_entries > 0 else 0
        active_score = (total_entries * 2) + (accuracy * 0.5) - (system_ended * 5)
        
        # Archive to operator_stats_history collection
        archive_entry = {
            "user_id": user_id,
            "name": name,
            "month": archive_month,
            "total_entries": total_entries,
            "proper_entries": proper_entries,
            "system_ended": system_ended,
            "accuracy": accuracy,
            "active_score": round(active_score, 1),
            "archived_at": now
        }
        
        # Upsert to avoid duplicates
        await db.operator_stats_history.update_one(
            {"user_id": user_id, "month": archive_month},
            {"$set": archive_entry},
            upsert=True
        )
        archived_count += 1
    
    return {
        "message": f"Archived {archived_count} operator stats for {archive_month}",
        "month": archive_month,
        "count": archived_count
    }


@app.get("/api/operator-stats/history")
async def get_operator_stats_history(
    user_id: Optional[str] = None,
    months: int = 6,
    current_user: User = Depends(get_current_user)
):
    """Get operator stats history (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if user_id:
        query["user_id"] = user_id
    
    history = await db.operator_stats_history.find(
        query,
        {"_id": 0}
    ).sort("month", -1).limit(months * 100).to_list(length=1000)
    
    return history


@app.get("/api/production/daily-summary")
async def get_daily_production_summary(
    date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get daily production summary grouped by category (CNC, VMC, Moulding) with running/ended status"""
    now = get_ist_now()
    target_date = date or now.strftime('%Y-%m-%d')
    
    # Parse target date to create IST day boundaries
    try:
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
        # Create IST day start and end
        ist_day_start = target_dt.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=IST)
        ist_day_end = target_dt.replace(hour=23, minute=59, second=59, microsecond=999999, tzinfo=IST)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    # Query completed entries - match by date string OR by completed_at/end_time within IST day
    # FIX: Include ALL entries for the date (don't filter by status - old entries may not have it)
    completed_entries = await db.production_entries.find({
        "$or": [
            {"date": target_date},
            {"date": {"$regex": f"^{target_date}"}},
            {"completed_at": {"$gte": ist_day_start, "$lte": ist_day_end}},
            {"end_time": {"$gte": ist_day_start, "$lte": ist_day_end}}
        ]
    }, {"_id": 0}).sort("end_time", -1).to_list(length=500)
    
    # Get active/running entries for the day
    active_entries = await db.active_work.find(
        {"status": {"$in": ["active", "in_progress"]}},
        {"_id": 0}
    ).to_list(length=100)
    
    # Filter active entries for today
    active_today = []
    for entry in active_entries:
        start_time = entry.get("start_time")
        entry_date = entry.get("date")
        
        # Check by date field first
        if entry_date == target_date:
            entry["status"] = "running"
            active_today.append(entry)
            continue
        
        # Fall back to start_time check
        if start_time:
            if isinstance(start_time, str):
                start_date = start_time[:10]
            else:
                start_date = to_ist(start_time).strftime('%Y-%m-%d')
            if start_date == target_date:
                entry["status"] = "running"
                active_today.append(entry)
    
    # Mark completed entries
    for entry in completed_entries:
        entry["status"] = "ended"
    
    # Combine all entries
    all_entries = active_today + completed_entries
    
    # Group by category
    cnc_entries = [e for e in all_entries if e.get("category") in ["cnc_lathe", "cnc", "CNC"]]
    vmc_entries = [e for e in all_entries if e.get("category") in ["vmc", "VMC"]]
    moulding_entries = [e for e in all_entries if e.get("category") in ["moulding", "Moulding", "injection"]]
    other_entries = [e for e in all_entries if e.get("category") not in ["cnc_lathe", "cnc", "CNC", "vmc", "VMC", "moulding", "Moulding", "injection"]]
    
    # Calculate totals
    def calc_totals(entries):
        running = len([e for e in entries if e.get("status") == "running"])
        ended = len([e for e in entries if e.get("status") == "ended"])
        total_qty = sum(e.get("quantity", 0) or e.get("production_qty", 0) or 0 for e in entries)
        return {"count": len(entries), "running": running, "ended": ended, "total_qty": total_qty}
    
    return {
        "date": target_date,
        "total_entries": len(all_entries),
        "total_produced": sum(e.get("quantity", 0) or e.get("production_qty", 0) or 0 for e in all_entries),
        "summary": {
            "cnc": calc_totals(cnc_entries),
            "vmc": calc_totals(vmc_entries),
            "moulding": calc_totals(moulding_entries),
            "other": calc_totals(other_entries),
            "total": calc_totals(all_entries)
        },
        "entries": {
            "cnc": cnc_entries,
            "vmc": vmc_entries,
            "moulding": moulding_entries,
            "other": other_entries
        }
    }


@app.get("/api/production-entries")
async def get_production_entries(
    category: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    machine_name: Optional[str] = None,
    operator_name: Optional[str] = None,
    job_details: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get production entries with filtering and pagination"""
    query = {}
    
    # Non-admins only see their own entries
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        query["$or"] = [
            {"created_by_id": current_user.user_id},
            {"operator_name": current_user.name}
        ]
    
    if category:
        query["category"] = category
    
    if machine_name:
        query["machine_name"] = {"$regex": machine_name, "$options": "i"}
    
    if operator_name:
        query["operator_name"] = {"$regex": operator_name, "$options": "i"}
    
    if job_details:
        query["job_details"] = {"$regex": job_details, "$options": "i"}
    
    # Default to last 30 days if no date range specified
    if not date_from and not date_to:
        default_from = (datetime.now(timezone.utc) - timedelta(days=30)).date().isoformat()
        query["date"] = {"$gte": default_from}
    elif date_from or date_to:
        query["date"] = {}
        if date_from:
            query["date"]["$gte"] = date_from
        if date_to:
            query["date"]["$lte"] = date_to
    
    # Calculate skip for pagination
    skip = (page - 1) * limit
    
    entries = await db.production_entries.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(limit).to_list(length=limit)
    
    # Return array directly for backward compatibility
    return entries
    
    # Convert UTC datetimes to IST for display
    for entry in entries:
        if entry.get("start_time") and hasattr(entry["start_time"], 'isoformat'):
            ist_start = to_ist(entry["start_time"])
            entry["start_time"] = ist_start.isoformat()
            entry["start_time_ist"] = format_ist_datetime(ist_start)
        if entry.get("end_time") and hasattr(entry["end_time"], 'isoformat'):
            ist_end = to_ist(entry["end_time"])
            entry["end_time"] = ist_end.isoformat()
            entry["end_time_ist"] = format_ist_datetime(ist_end)
        if entry.get("created_at") and hasattr(entry["created_at"], 'isoformat'):
            entry["created_at"] = to_ist(entry["created_at"]).isoformat()
        if entry.get("completed_at") and hasattr(entry["completed_at"], 'isoformat'):
            entry["completed_at"] = to_ist(entry["completed_at"]).isoformat()
    
    # Calculate total count for pagination info
    total_count = await db.production_entries.count_documents(query)
    total_pages = (total_count + limit - 1) // limit
    
    return {
        "entries": entries,
        "pagination": {
            "page": page,
            "limit": limit,
            "total_entries": total_count,
            "total_pages": total_pages,
            "has_more": page < total_pages
        }
    }

@app.delete("/api/production-entries/{entry_id}")
async def delete_production_entry(
    entry_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete production entry (Admin or entry creator only)"""
    entry = await db.production_entries.find_one({"entry_id": entry_id}, {"_id": 0})
    
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    # Only admin or the creator can delete
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"} and entry["operator_id"] != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this entry")
    
    await db.production_entries.delete_one({"entry_id": entry_id})
    return {"message": "Production entry deleted successfully"}


@app.put("/api/production-entries/{entry_id}")
async def update_production_entry(
    entry_id: str,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Update production entry (Admin or entry creator only)"""
    data = await request.json()
    
    entry = await db.production_entries.find_one({"entry_id": entry_id}, {"_id": 0})
    
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    # Only admin or the creator can update
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"} and entry.get("operator_id") != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not authorized to update this entry")
    
    update_data = {}
    if "production_qty" in data:
        update_data["production_qty"] = data["production_qty"]
    if "remarks" in data:
        update_data["remarks"] = data["remarks"]
    if "cycle_time" in data:
        update_data["cycle_time"] = data["cycle_time"]
    if "setting_time" in data:
        update_data["setting_time"] = data["setting_time"]
    if "spindle_rpm" in data:
        update_data["spindle_rpm"] = data["spindle_rpm"]
    if "feed_rate" in data:
        update_data["feed_rate"] = data["feed_rate"]
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.production_entries.update_one(
        {"entry_id": entry_id},
        {"$set": update_data}
    )
    
    return {"message": "Production entry updated successfully"}


# Machine Breakdown Endpoints
@app.post("/api/breakdowns")
async def create_breakdown(
    breakdown: MachineBreakdownCreate,
    current_user: User = Depends(get_current_user)
):
    """Report a machine breakdown (Any user can report)"""
    breakdown_id = f"brk_{uuid.uuid4().hex[:12]}"
    now = get_ist_now()  # Use IST timezone
    
    breakdown_doc = {
        "breakdown_id": breakdown_id,
        "machine_name": breakdown.machine_name,
        "date": breakdown.date,
        "date_ist": format_ist_datetime(now),  # Human readable IST format
        "alarm_photo": breakdown.alarm_photo,
        "machine_photo": breakdown.machine_photo,
        "action_taken": breakdown.action_taken,
        "action_taken_by": current_user.name,
        "action_taken_by_id": current_user.user_id,
        "reported_by": current_user.name,
        "reported_by_id": current_user.user_id,
        "status": "pending",
        "created_at": now
    }
    
    await db.breakdowns.insert_one(breakdown_doc)
    
    # Mark the machine as under maintenance/breakdown
    await db.machines.update_one(
        {"name": breakdown.machine_name},
        {"$set": {
            "status": "breakdown",
            "breakdown_id": breakdown_id,
            "breakdown_reported_at": now,
            "last_status_update": now
        }}
    )
    
    # Send push notification to all admins about the breakdown (using breakdowns channel for high priority alert)
    admins = await db.users.find({"role": "Admin"}).to_list(length=100)
    for admin in admins:
        if admin.get("push_token"):
            await send_push_notification(
                admin["user_id"], 
                "🚨 Machine Breakdown Alert",
                f"{breakdown.machine_name} breakdown reported by {current_user.name}",
                "breakdowns"  # Use breakdowns channel for ringtone/vibration
            )
    
    return {
        "breakdown_id": breakdown_id,
        "machine_name": breakdown.machine_name,
        "date": breakdown.date.isoformat(),
        "action_taken": breakdown.action_taken,
        "action_taken_by": current_user.name,
        "reported_by": current_user.name,
        "status": "pending",
        "created_at": now.isoformat()
    }

@app.get("/api/breakdowns")
async def get_breakdowns(
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all breakdowns (Admin sees all, others see their own)"""
    query = {}
    
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        query["reported_by_id"] = current_user.user_id
    
    if status:
        query["status"] = status
    
    breakdowns = await db.breakdowns.find(query, {"_id": 0}).sort("created_at", -1).limit(200).to_list(length=200)
    
    # Enrich breakdowns with comments
    enriched = []
    for breakdown in breakdowns:
        breakdown_id = breakdown.get("breakdown_id")
        comments = await db.breakdown_comments.find(
            {"breakdown_id": breakdown_id},
            {"_id": 0}
        ).sort("created_at", -1).limit(10).to_list(length=10)
        breakdown["comments"] = comments
        enriched.append(breakdown)
    
    return enriched

@app.put("/api/breakdowns/{breakdown_id}")
async def update_breakdown(
    breakdown_id: str,
    update_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Update breakdown status (Admin only can mark as resolved)"""
    breakdown = await db.breakdowns.find_one({"breakdown_id": breakdown_id})
    
    if not breakdown:
        raise HTTPException(status_code=404, detail="Breakdown not found")
    
    # Only admin can change status to resolved
    if update_data.get("status") == "resolved" and (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Only Admin can mark breakdown as resolved")
    
    allowed_fields = ["status", "action_taken"]
    update_dict = {k: v for k, v in update_data.items() if k in allowed_fields}
    update_dict["updated_at"] = datetime.now(timezone.utc)
    
    # If resolving breakdown, update machine status back to available
    if update_data.get("status") == "resolved":
        update_dict["resolved_at"] = datetime.now(timezone.utc)
        update_dict["resolved_by"] = current_user.user_id
        update_dict["resolved_by_name"] = current_user.name
        
        # Update machine status back to available
        machine_name = breakdown.get("machine_name")
        if machine_name:
            await db.machines.update_one(
                {"name": machine_name},
                {"$set": {"status": "available", "updated_at": datetime.now(timezone.utc)}}
            )
            logger.info(f"Machine '{machine_name}' status updated to 'available' after breakdown resolved")
    
    await db.breakdowns.update_one(
        {"breakdown_id": breakdown_id},
        {"$set": update_dict}
    )
    
    updated = await db.breakdowns.find_one({"breakdown_id": breakdown_id}, {"_id": 0})
    return updated

@app.delete("/api/breakdowns/{breakdown_id}")
async def delete_breakdown(
    breakdown_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete breakdown (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get breakdown first to update machine status
    breakdown = await db.breakdowns.find_one({"breakdown_id": breakdown_id})
    if breakdown:
        machine_name = breakdown.get("machine_name")
        if machine_name:
            # Check if there are any other pending breakdowns for this machine
            other_breakdowns = await db.breakdowns.find_one({
                "machine_name": machine_name,
                "status": "pending",
                "breakdown_id": {"$ne": breakdown_id}
            })
            if not other_breakdowns:
                # No other pending breakdowns, set machine to available
                await db.machines.update_one(
                    {"name": machine_name},
                    {"$set": {"status": "available", "updated_at": datetime.now(timezone.utc)}}
                )
    
    result = await db.breakdowns.delete_one({"breakdown_id": breakdown_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Breakdown not found")
    
    return {"message": "Breakdown deleted successfully"}


@app.post("/api/machines/sync-breakdown-status")
async def sync_machine_breakdown_status(
    current_user: User = Depends(get_current_user)
):
    """Sync all machine statuses based on breakdown records (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    updated_count = 0
    
    # Get all machines
    machines = await db.machines.find({}).to_list(length=500)
    
    for machine in machines:
        machine_name = machine.get("name")
        
        # Check if there's a pending breakdown for this machine
        pending_breakdown = await db.breakdowns.find_one({
            "machine_name": machine_name,
            "status": "pending"
        })
        
        if pending_breakdown:
            # Machine should be in breakdown status
            if machine.get("status") != "breakdown":
                await db.machines.update_one(
                    {"machine_id": machine["machine_id"]},
                    {"$set": {"status": "breakdown", "updated_at": datetime.now(timezone.utc)}}
                )
                updated_count += 1
                logger.info(f"Machine '{machine_name}' set to 'breakdown' (has pending breakdown)")
        else:
            # Machine should NOT be in breakdown status
            if machine.get("status") == "breakdown":
                await db.machines.update_one(
                    {"machine_id": machine["machine_id"]},
                    {"$set": {"status": "available", "updated_at": datetime.now(timezone.utc)}}
                )
                updated_count += 1
                logger.info(f"Machine '{machine_name}' set to 'available' (no pending breakdown)")
    
    return {
        "message": "Machine statuses synced successfully",
        "updated_count": updated_count
    }


class BreakdownCommentCreate(BaseModel):
    comment: str

@app.post("/api/breakdowns/{breakdown_id}/comments")
async def add_breakdown_comment(
    breakdown_id: str,
    comment_data: BreakdownCommentCreate,
    current_user: User = Depends(get_current_user)
):
    """Add comment to a breakdown"""
    breakdown = await db.breakdowns.find_one({"breakdown_id": breakdown_id})
    if not breakdown:
        raise HTTPException(status_code=404, detail="Breakdown not found")
    
    comment_id = f"bdcmt_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    comment_doc = {
        "comment_id": comment_id,
        "breakdown_id": breakdown_id,
        "comment": comment_data.comment,
        "created_by_id": current_user.user_id,
        "created_by_name": current_user.name,
        "created_at": now
    }
    
    await db.breakdown_comments.insert_one(comment_doc)
    
    return {
        "message": "Comment added successfully",
        "comment_id": comment_id,
        "comment": comment_data.comment,
        "created_by_name": current_user.name,
        "created_at": now.isoformat()
    }


@app.get("/api/breakdowns/{breakdown_id}/comments")
async def get_breakdown_comments(
    breakdown_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get comments for a breakdown"""
    comments = await db.breakdown_comments.find(
        {"breakdown_id": breakdown_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(length=100)
    
    return comments


# Notifications Endpoints
@app.get("/api/notifications")
async def get_notifications(current_user: User = Depends(get_current_user)):
    """Get unread notifications for admin (new jobs, breakdowns)"""
    notifications = []
    
    if current_user.role == "Admin":
        # Get recent pending breakdowns (last 24 hours)
        from datetime import timedelta
        yesterday = datetime.now(timezone.utc) - timedelta(hours=24)
        
        pending_breakdowns = await db.breakdowns.find(
            {"status": "pending", "created_at": {"$gte": yesterday}},
            {"_id": 0}
        ).sort("created_at", -1).limit(10).to_list(length=10)
        
        for bd in pending_breakdowns:
            notifications.append({
                "id": bd.get("breakdown_id"),
                "type": "breakdown",
                "message": f"Machine breakdown reported: {bd.get('machine_name')}",
                "created_at": bd.get("created_at").isoformat() if bd.get("created_at") else None
            })
        
        # Get recently assigned jobs (last 24 hours)
        recent_jobs = await db.jobs.find(
            {"created_at": {"$gte": yesterday}},
            {"_id": 0}
        ).sort("created_at", -1).limit(10).to_list(length=10)
        
        for job in recent_jobs:
            notifications.append({
                "id": job.get("job_id"),
                "type": "job",
                "message": f"New job created: {job.get('machine_name')} - {job.get('category', '').replace('_', ' ').title()}",
                "created_at": job.get("created_at").isoformat() if job.get("created_at") else None
            })
    else:
        # For non-admins, show jobs assigned to them
        user_jobs = await db.jobs.find(
            {"assigned_to": current_user.name, "status": "pending"},
            {"_id": 0}
        ).sort("created_at", -1).limit(5).to_list(length=5)
        
        for job in user_jobs:
            notifications.append({
                "id": job.get("job_id"),
                "type": "job_assigned",
                "message": f"Job assigned to you: {job.get('machine_name')}",
                "created_at": job.get("created_at").isoformat() if job.get("created_at") else None
            })
    
    # Sort by created_at
    notifications.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    
    return {
        "count": len(notifications),
        "notifications": notifications[:10]
    }


# =====================
# MATERIAL REQUESTS
# =====================

@app.post("/api/material-requests")
async def create_material_request(
    request: MaterialRequestCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new material request"""
    request_id = f"mat_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    request_doc = {
        "request_id": request_id,
        "posted_by": request.posted_by,
        "material_required": request.material_required,
        "material_details": request.material_details,
        "required_for": request.required_for,
        "status": "pending",
        "created_by_id": current_user.user_id,
        "created_by_name": current_user.name,
        "created_at": now
    }
    
    await db.material_requests.insert_one(request_doc)
    
    # Create notification for admins
    admins = await db.users.find({"role": "Admin"}).to_list(length=100)
    for admin in admins:
        await db.notifications.insert_one({
            "notification_id": f"notif_{uuid.uuid4().hex[:12]}",
            "user_id": admin["user_id"],
            "type": "material_request",
            "message": f"New material request from {request.posted_by}: {request.material_required}",
            "read": False,
            "created_at": now
        })
    
    # Send push notification to all admins about the material request
    await notify_admins(
        title="📦 New Material Request",
        body=f"{request.posted_by} requested: {request.material_required}"
    )
    
    return {"message": "Request submitted successfully", "request_id": request_id}

@app.get("/api/material-requests")
async def get_material_requests(
    current_user: User = Depends(get_current_user)
):
    """Get all material requests (Admin only for all, others see their own)"""
    if current_user.role == "Admin":
        requests = await db.material_requests.find({}, {"_id": 0}).sort("created_at", -1).to_list(length=500)
    else:
        requests = await db.material_requests.find(
            {"created_by_id": current_user.user_id},
            {"_id": 0}
        ).sort("created_at", -1).to_list(length=500)
    
    return requests

@app.put("/api/material-requests/{request_id}/status")
async def update_material_request_status(
    request_id: str,
    status_update: MaterialRequestStatusUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update material request status (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if status_update.status not in ["pending", "approved", "rejected", "fulfilled"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    update_data = {
        "status": status_update.status, 
        "updated_at": datetime.now(timezone.utc)
    }
    
    # If marking as fulfilled, add fulfilled timestamp
    if status_update.status == "fulfilled":
        update_data["fulfilled_at"] = datetime.now(timezone.utc)
        update_data["fulfilled_by"] = current_user.user_id
        update_data["fulfilled_by_name"] = current_user.name
    
    result = await db.material_requests.update_one(
        {"request_id": request_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Request not found")
    
    return {"message": f"Request {status_update.status}"}


# =====================
# PUSH NOTIFICATIONS
# =====================

@app.post("/api/users/push-token")
async def save_push_token(
    token_update: PushTokenUpdate,
    current_user: User = Depends(get_current_user)
):
    """Save user's push notification token"""
    await db.users.update_one(
        {"user_id": current_user.user_id},
        {"$set": {"push_token": token_update.push_token}}
    )
    return {"message": "Push token saved successfully"}

async def send_push_notification(user_id: str, title: str, body: str, channel_id: str = "default"):
    """Send push notification to a specific user via Expo Push API"""
    user = await db.users.find_one({"user_id": user_id})
    if not user or not user.get("push_token"):
        return
    
    push_token = user["push_token"]
    
    try:
        import httpx
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://exp.host/--/api/v2/push/send",
                json={
                    "to": push_token,
                    "title": title,
                    "body": body,
                    "sound": "default",
                    "priority": "high",
                    "channelId": channel_id,
                    "_displayInForeground": True,
                    "badge": 1
                },
                headers={"Content-Type": "application/json"}
            )
            print(f"Push notification sent: {response.status_code}")
    except Exception as e:
        print(f"Failed to send push notification: {e}")

async def notify_admins(title: str, body: str):
    """Send push notifications to all admins"""
    admins = await db.users.find({"role": "Admin"}).to_list(length=100)
    for admin in admins:
        if admin.get("push_token"):
            await send_push_notification(admin["user_id"], title, body)


# Background task for job reminders (12 hour interval)
async def send_incomplete_job_reminders():
    """Send reminders for incomplete jobs every 12 hours"""
    while True:
        try:
            # Wait 12 hours (43200 seconds)
            await asyncio.sleep(43200)
            
            # Find all incomplete jobs (pending or in_progress)
            incomplete_jobs = await db.jobs.find({
                "status": {"$in": ["pending", "in_progress"]}
            }).to_list(length=1000)
            
            for job in incomplete_jobs:
                assigned_to = job.get("assigned_to")
                if assigned_to:
                    job_details = job.get("job_details", "No details")[:30]
                    machine = job.get("machine_name", "Unknown")
                    status = job.get("status", "pending")
                    
                    await send_push_notification(
                        assigned_to,
                        "⏰ Job Reminder / काम याद दिलाना",
                        f"{machine}: {job_details}... | Status: {status}"
                    )
            
            print(f"Sent reminders for {len(incomplete_jobs)} incomplete jobs")
        except Exception as e:
            print(f"Error in job reminder task: {e}")
            await asyncio.sleep(3600)  # Wait 1 hour on error before retrying


# Background task for machine status reset at shift changes (8am and 8pm IST)
async def auto_reset_machine_status():
    """Automatically reset all machine status to idle at 8am and 8pm IST"""
    ist = ZoneInfo('Asia/Kolkata')
    
    while True:
        try:
            now_ist = datetime.now(ist)
            current_hour = now_ist.hour
            
            # Check if it's 8am (08:00) or 8pm (20:00) IST
            if current_hour in [8, 20]:
                # Reset all machines to idle
                result = await db.machines.update_many(
                    {},
                    {"$set": {
                        "status": "idle",
                        "current_job_id": None,
                        "current_operator_id": None,
                        "last_status_update": datetime.now(timezone.utc)
                    }}
                )
                
                print(f"[Shift Change] Auto-reset {result.modified_count} machines to idle at {now_ist.strftime('%Y-%m-%d %H:%M:%S')} IST")
                
                # Wait 1 hour to avoid multiple resets
                await asyncio.sleep(3600)
            else:
                # Check every 5 minutes
                await asyncio.sleep(300)
        except Exception as e:
            print(f"Error in machine reset task: {e}")
            await asyncio.sleep(300)


# Start background task on app startup
@app.on_event("startup")
async def startup_event():
    """Verify server is ready for production"""
    logger.info("Running startup checks...")

    # Bootstrap super_admin (idempotent — production first-boot safety net)
    try:
        from auth_v2 import ensure_super_admin_seeded, deactivate_users_without_username
        await ensure_super_admin_seeded()
        await deactivate_users_without_username()
    except Exception as e:
        logger.error(f"super_admin seed failed: {e}")

    # Verify MongoDB connection
    try:
        if db is not None:
            # Ping the database to verify connection
            await client.admin.command('ping')
            logger.info("MongoDB connection verified successfully")
            
            # Create indexes for faster queries (Speed Optimization)
            logger.info("Creating database indexes for performance...")
            try:
                # Jobs indexes
                await db.jobs.create_index("job_id")
                await db.jobs.create_index("status")
                await db.jobs.create_index("category")
                await db.jobs.create_index("created_date")
                await db.jobs.create_index([("status", 1), ("category", 1)])
                
                # Machines indexes
                await db.machines.create_index("machine_id")
                await db.machines.create_index("category")
                await db.machines.create_index("status")
                
                # Production entries indexes
                await db.production_entries.create_index("entry_id")
                await db.production_entries.create_index("date")
                await db.production_entries.create_index("machine_name")
                await db.production_entries.create_index("operator_name")
                await db.production_entries.create_index([("date", -1)])
                
                # Storage indexes
                await db.storage.create_index("item_id")
                await db.storage.create_index("category")
                await db.storage.create_index("created_at")
                await db.storage.create_index([("category", 1), ("created_at", -1)])
                
                # Users indexes
                await db.users.create_index("user_id")
                await db.users.create_index("email")
                await db.users.create_index("role")
                
                # Active work indexes
                await db.active_work.create_index("entry_id")
                await db.active_work.create_index("machine_name")
                await db.active_work.create_index("status")
                
                # Plastic raw materials indexes
                await db.plastic_raw_materials.create_index("material_id")
                await db.plastic_raw_materials.create_index("created_at")
                
                # Tools/Inserts indexes
                await db.tools_inserts.create_index("tool_id")
                await db.tools_inserts.create_index("created_at")
                
                # Gauges indexes
                await db.gauges.create_index("gauge_id")
                await db.gauges.create_index("created_at")
                
                # Attendance records indexes (CRITICAL for 79K+ records)
                await db.attendance_records.create_index("record_id")  # Not unique - some records may have null
                await db.attendance_records.create_index([("synced_at", -1)])
                await db.attendance_records.create_index("biometric_id")
                await db.attendance_records.create_index("date")
                await db.attendance_records.create_index([("biometric_id", 1), ("date", 1)])
                
                # Biometric employees indexes
                await db.biometric_employees.create_index("user_id")  # Not unique - might have nulls
                await db.biometric_employees.create_index("linked_user_id")
                
                logger.info("Database indexes created successfully")
            except Exception as e:
                logger.warning(f"Index creation warning (may already exist): {e}")
        else:
            logger.warning("MongoDB client not initialized")
    except Exception as e:
        logger.error(f"MongoDB connection check failed: {e}")
    
    # Background tasks disabled to reduce memory usage
    # Use manual API endpoints instead:
    # - POST /api/machines/reset-all (manual reset)
    # - The 8am/8pm auto-reset is disabled
    logger.info("Startup complete - Background tasks disabled for memory optimization")
    logger.info("Use POST /api/machines/reset-all for manual machine reset")
    
    # Start APScheduler for CRON jobs
    try:
        # Add job to auto-end shift at 8:00 PM IST (Morning shift end)
        # FIX: Use timezone parameter to ensure IST timing
        scheduler.add_job(
            scheduled_auto_end_shift,
            CronTrigger(hour=20, minute=0, timezone='Asia/Kolkata'),  # 8:00 PM IST
            id="auto_end_morning_shift",
            replace_existing=True
        )
        
        # Add job to auto-end shift at 6:00 AM IST (Night shift end)
        scheduler.add_job(
            scheduled_auto_end_shift,
            CronTrigger(hour=6, minute=0, timezone='Asia/Kolkata'),  # 6:00 AM IST
            id="auto_end_night_shift",
            replace_existing=True
        )
        
        scheduler.start()
        logger.info("APScheduler started - Auto-end jobs scheduled at 8:00 PM & 6:00 AM IST (Asia/Kolkata timezone)")
    except Exception as e:
        logger.error(f"Failed to start APScheduler: {e}")
    
    logger.info("Server ready to accept requests")


@app.on_event("shutdown")
async def shutdown_event():
    """Cleanup on server shutdown"""
    logger.info("Shutting down server...")
    
    # Shutdown APScheduler gracefully
    try:
        if scheduler.running:
            scheduler.shutdown(wait=False)
            logger.info("APScheduler shutdown complete")
    except Exception as e:
        logger.error(f"Error shutting down scheduler: {e}")


# Test push notification endpoint
@app.post("/api/test-push-notification")
async def test_push_notification(
    current_user: User = Depends(get_current_user)
):
    """Send a test push notification to the current user"""
    user = await db.users.find_one({"user_id": current_user.user_id})
    
    if not user or not user.get("push_token"):
        return {"success": False, "message": "No push token registered for this user. Please allow notifications in the app."}
    
    try:
        import httpx
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://exp.host/--/api/v2/push/send",
                json={
                    "to": user["push_token"],
                    "title": "🎉 Test Notification",
                    "body": "Push notifications are working!",
                    "sound": "default",
                    "priority": "high"
                },
                headers={"Content-Type": "application/json"}
            )
            result = response.json()
            print(f"Test push notification result: {result}")
            return {"success": True, "message": "Test notification sent!", "response": result}
    except Exception as e:
        print(f"Failed to send test push notification: {e}")
        return {"success": False, "message": f"Failed to send: {str(e)}"}


# =====================
# PROFILE IMAGE
# =====================

@app.post("/api/users/profile-image")
async def update_profile_image(
    image_update: ProfileImageUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update user's profile image"""
    await db.users.update_one(
        {"user_id": current_user.user_id},
        {"$set": {"profile_image": image_update.image}}
    )
    return {"message": "Profile image updated successfully"}

# =====================
# INJECTION MOULDING
# =====================

@app.post("/api/injection-moulding")
async def create_injection_moulding_entry(
    entry: InjectionMouldingEntry,
    current_user: User = Depends(get_current_user)
):
    """Create injection moulding production entry"""
    entry_id = f"inj_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    entry_doc = {
        "entry_id": entry_id,
        "date": now,
        "machine": entry.machine,
        "mould_no": entry.mould_no,
        "material": entry.material,
        "entry_type": entry.entry_type,
        "operator_name": entry.operator_name,
        "production_qty": entry.production_qty,
        "remarks": entry.remarks,
        "created_by_id": current_user.user_id,
        "created_by_name": current_user.name,
        "created_at": now
    }
    
    await db.injection_moulding.insert_one(entry_doc)
    return {"message": "Entry created successfully", "entry_id": entry_id}

@app.get("/api/injection-moulding")
async def get_injection_moulding_entries(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    machine: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get injection moulding entries"""
    query = {}
    
    # Default to last 7 days
    if not date_from and not date_to:
        default_from = datetime.now(timezone.utc) - timedelta(days=7)
        query["date"] = {"$gte": default_from}
    elif date_from or date_to:
        query["date"] = {}
        if date_from:
            query["date"]["$gte"] = datetime.fromisoformat(date_from)
        if date_to:
            query["date"]["$lte"] = datetime.fromisoformat(date_to)
    
    if machine:
        query["machine"] = machine
    
    entries = await db.injection_moulding.find(query, {"_id": 0}).sort("date", -1).to_list(length=500)
    return entries

# =====================
# MATERIAL STORAGE
# =====================

@app.post("/api/storage")
async def create_storage_entry(
    entry: StorageEntry,
    current_user: User = Depends(get_current_user)
):
    """Create storage inventory entry"""
    entry_id = f"stor_{uuid.uuid4().hex[:12]}"
    now = get_ist_now()  # Use IST timezone
    
    entry_doc = {
        "entry_id": entry_id,
        "date": now,
        "date_ist": format_ist_datetime(now),  # Human readable IST format
        "assembly_name": entry.assembly_name,
        "part_no": entry.part_no,
        "product_details": entry.product_details,
        "quantity": entry.quantity,
        "storage_place": entry.storage_place,
        "crate_no": entry.crate_no,
        "stored_by": entry.stored_by,
        "image": entry.image,
        "created_by_id": current_user.user_id,
        "created_by_name": current_user.name,
        "created_at": now
    }
    
    await db.storage.insert_one(entry_doc)
    return {"message": "Storage entry created successfully", "entry_id": entry_id}

@app.get("/api/storage")
async def get_storage_entries(
    part_no: Optional[str] = None,
    assembly_name: Optional[str] = None,
    crate_no: Optional[str] = None,
    keyword: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get storage entries with search filters and pagination"""
    query = {}
    
    if part_no:
        query["part_no"] = {"$regex": part_no, "$options": "i"}
    
    if assembly_name:
        query["assembly_name"] = {"$regex": assembly_name, "$options": "i"}
    
    if crate_no:
        query["crate_no"] = {"$regex": crate_no, "$options": "i"}
    
    if keyword:
        query["$or"] = [
            {"part_no": {"$regex": keyword, "$options": "i"}},
            {"assembly_name": {"$regex": keyword, "$options": "i"}},
            {"product_details": {"$regex": keyword, "$options": "i"}},
            {"storage_place": {"$regex": keyword, "$options": "i"}},
            {"crate_no": {"$regex": keyword, "$options": "i"}}
        ]
    
    # Calculate skip for pagination
    skip = (page - 1) * limit
    
    entries = await db.storage.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(limit).to_list(length=limit)
    
    # Return array directly for backward compatibility
    return entries


# Storage Deduction Model
class StorageDeductRequest(BaseModel):
    quantity: int
    taken_by: str  # user_id of the person taking the item
    signature: Optional[str] = None  # base64 signature image


# Storage Add Quantity Model
class StorageAddQtyRequest(BaseModel):
    quantity: int


@app.post("/api/storage/{entry_id}/deduct")
async def deduct_storage_item(
    entry_id: str,
    deduct_request: StorageDeductRequest,
    current_user: User = Depends(get_current_user)
):
    """Deduct quantity from storage item and log who took it, when"""
    # Find the storage entry
    entry = await db.storage.find_one({"entry_id": entry_id})
    
    if not entry:
        raise HTTPException(status_code=404, detail="Storage entry not found")
    
    current_qty = entry.get("quantity", 1)
    
    if deduct_request.quantity > current_qty:
        raise HTTPException(status_code=400, detail=f"Cannot take {deduct_request.quantity}. Only {current_qty} available.")
    
    if deduct_request.quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be greater than 0")
    
    # Get taker user info
    taker_user = await db.users.find_one({"user_id": deduct_request.taken_by})
    taker_name = taker_user.get("name", "Unknown") if taker_user else "Unknown"
    
    new_qty = current_qty - deduct_request.quantity
    now = datetime.now(timezone.utc)
    
    # Update the storage quantity
    await db.storage.update_one(
        {"entry_id": entry_id},
        {"$set": {"quantity": new_qty, "updated_at": now}}
    )
    
    # Log the deduction in a separate collection for tracking
    deduction_log = {
        "log_id": f"dedlog_{uuid.uuid4().hex[:12]}",
        "entry_id": entry_id,
        "part_no": entry.get("part_no"),
        "assembly_name": entry.get("assembly_name"),
        "quantity_taken": deduct_request.quantity,
        "remaining_qty": new_qty,
        "taken_by_id": deduct_request.taken_by,
        "taken_by_name": taker_name,
        "recorded_by_id": current_user.user_id,
        "recorded_by_name": current_user.name,
        "taken_at": now,
        "signature": deduct_request.signature
    }
    
    await db.storage_deductions.insert_one(deduction_log)
    
    return {
        "message": "Item deducted successfully",
        "remaining_qty": new_qty,
        "taken_by": taker_name,
        "taken_at": now.isoformat()
    }


@app.post("/api/storage/{entry_id}/add-qty")
async def add_storage_quantity(
    entry_id: str,
    add_request: StorageAddQtyRequest,
    current_user: User = Depends(get_current_user)
):
    """Add more quantity to an existing storage item"""
    # Find the storage entry
    entry = await db.storage.find_one({"entry_id": entry_id})
    
    if not entry:
        raise HTTPException(status_code=404, detail="Storage entry not found")
    
    if add_request.quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be greater than 0")
    
    current_qty = entry.get("quantity", 1)
    new_qty = current_qty + add_request.quantity
    now = datetime.now(timezone.utc)
    
    # Update the storage quantity
    await db.storage.update_one(
        {"entry_id": entry_id},
        {"$set": {"quantity": new_qty, "updated_at": now}}
    )
    
    return {
        "message": "Quantity added successfully",
        "previous_qty": current_qty,
        "added_qty": add_request.quantity,
        "new_qty": new_qty
    }


@app.get("/api/storage/{entry_id}/deduction-logs")
async def get_storage_deduction_logs(
    entry_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get deduction history for a storage item with full details"""
    logs = await db.storage_deductions.find(
        {"entry_id": entry_id},
        {"_id": 0}
    ).sort("taken_at", -1).to_list(length=100)
    
    # Enrich each log with formatted date and user details
    enriched_logs = []
    for log in logs:
        # Format date in IST
        taken_at = log.get("taken_at")
        if taken_at:
            if isinstance(taken_at, datetime):
                ist_time = to_ist(taken_at)
                log["taken_at_formatted"] = format_ist_datetime(ist_time)
                log["taken_date"] = ist_time.strftime('%d/%m/%Y')
                log["taken_time"] = ist_time.strftime('%I:%M %p')
        
        # Get signature if exists
        log["signature"] = log.get("signature") or log.get("sign") or None
        
        # Add transaction summary
        log["transaction_summary"] = f"{log.get('quantity_taken', 0)} items from {log.get('rack', 'Unknown rack')}"
        log["from_user"] = log.get("recorded_by_name") or log.get("taken_by_name") or "Unknown"
        log["to_user"] = log.get("taken_by_name") or "Unknown"
        
        enriched_logs.append(log)
    
    return enriched_logs


@app.get("/api/storage/{entry_id}/full-history")
async def get_storage_full_history(
    entry_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get complete transaction history for a storage item"""
    # Get the storage entry
    entry = await db.storage.find_one({"entry_id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Storage entry not found")
    
    # Get deduction logs
    deductions = await db.storage_deductions.find(
        {"entry_id": entry_id},
        {"_id": 0}
    ).sort("taken_at", -1).to_list(length=100)
    
    # Get addition logs (if any)
    additions = await db.storage_additions.find(
        {"entry_id": entry_id},
        {"_id": 0}
    ).sort("added_at", -1).to_list(length=100) if "storage_additions" in await db.list_collection_names() else []
    
    # Combine and format all transactions
    all_transactions = []
    
    # Add deductions
    for log in deductions:
        taken_at = log.get("taken_at")
        if isinstance(taken_at, datetime):
            ist_time = to_ist(taken_at)
            formatted_date = format_ist_datetime(ist_time)
        else:
            formatted_date = str(taken_at) if taken_at else "Unknown"
        
        all_transactions.append({
            "type": "deduction",
            "log_id": log.get("log_id"),
            "quantity": log.get("quantity_taken", 0),
            "date": formatted_date,
            "date_short": ist_time.strftime('%d/%m/%Y') if isinstance(taken_at, datetime) else "",
            "time": ist_time.strftime('%I:%M %p') if isinstance(taken_at, datetime) else "",
            "from_user": log.get("recorded_by_name") or log.get("issued_by_name") or "Admin",
            "to_user": log.get("taken_by_name") or "Unknown",
            "rack": log.get("rack"),
            "reason": log.get("reason", ""),
            "signature": log.get("signature") or log.get("sign"),
            "can_undo": log.get("can_undo", False),
            "timestamp": taken_at
        })
    
    # Add additions
    for log in additions:
        added_at = log.get("added_at")
        if isinstance(added_at, datetime):
            ist_time = to_ist(added_at)
            formatted_date = format_ist_datetime(ist_time)
        else:
            formatted_date = str(added_at) if added_at else "Unknown"
        
        all_transactions.append({
            "type": "addition",
            "log_id": log.get("log_id"),
            "quantity": log.get("quantity", 0),
            "date": formatted_date,
            "date_short": ist_time.strftime('%d/%m/%Y') if isinstance(added_at, datetime) else "",
            "time": ist_time.strftime('%I:%M %p') if isinstance(added_at, datetime) else "",
            "from_user": "Stock Addition",
            "to_user": log.get("added_by_name") or "Admin",
            "rack": log.get("rack"),
            "reason": "Stock Added",
            "signature": None,
            "timestamp": added_at
        })
    
    # Sort by timestamp descending
    all_transactions.sort(key=lambda x: x.get("timestamp") or datetime.min.replace(tzinfo=timezone.utc), reverse=True)
    
    return {
        "entry": entry,
        "transactions": all_transactions,
        "total_deductions": len(deductions),
        "total_additions": len(additions),
        "current_quantity": entry.get("quantity", 0)
    }


@app.put("/api/storage/{entry_id}")
async def update_storage_entry(
    entry_id: str,
    update_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Update a storage entry"""
    entry = await db.storage.find_one({"entry_id": entry_id})
    
    if not entry:
        raise HTTPException(status_code=404, detail="Storage entry not found")
    
    # Only admins or the creator can edit
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"} and entry.get("created_by_id") != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not authorized to edit this entry")
    
    allowed_fields = ["part_no", "assembly_name", "crate_no", "storage_place", 
                      "quantity", "remarks", "image_base64"]
    
    update_dict = {k: v for k, v in update_data.items() if k in allowed_fields}
    update_dict["updated_at"] = datetime.now(timezone.utc)
    
    await db.storage.update_one({"entry_id": entry_id}, {"$set": update_dict})
    
    updated_entry = await db.storage.find_one({"entry_id": entry_id}, {"_id": 0})
    return updated_entry


@app.delete("/api/storage/{entry_id}")
async def delete_storage_entry(
    entry_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a storage entry (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    entry = await db.storage.find_one({"entry_id": entry_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Storage entry not found")
    
    await db.storage.delete_one({"entry_id": entry_id})
    return {"message": "Storage entry deleted successfully"}


# =============================================
# PART LIBRARY <-> STORAGE SYNC HELPERS
# =============================================

async def sync_part_stock_from_storage(part_id: str):
    """Sync part library stock with storage entries linked to this part"""
    # Find all storage entries linked to this part
    storage_entries = await db.storage.find({"part_id": part_id}).to_list(length=100)
    
    # Calculate total stock and locations
    total_stock = 0
    locations = []
    
    for entry in storage_entries:
        qty = entry.get("quantity", 0)
        rack = entry.get("storage_place") or entry.get("crate_no") or "Unknown"
        total_stock += qty
        
        # Aggregate by rack
        found = False
        for loc in locations:
            if loc["rack"].lower() == rack.lower():
                loc["qty"] += qty
                found = True
                break
        if not found and qty > 0:
            locations.append({"rack": rack, "qty": qty, "entry_id": entry.get("entry_id")})
    
    # Update part library
    await db.parts_library.update_one(
        {"part_id": part_id},
        {
            "$set": {
                "stock": total_stock,
                "locations": locations,
                "updated_at": get_ist_now()
            }
        }
    )
    
    return {"stock": total_stock, "locations": locations}


@app.post("/api/parts-library/{part_id}/sync-stock")
async def sync_part_stock(part_id: str, current_user: User = Depends(get_current_user)):
    """Manually sync part stock from storage entries"""
    part = await db.parts_library.find_one({"part_id": part_id})
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    
    result = await sync_part_stock_from_storage(part_id)
    return {"message": "Stock synced successfully", **result}


@app.get("/api/parts-library/{part_id}/storage-entries")
async def get_part_storage_entries(part_id: str, current_user: User = Depends(get_current_user)):
    """Get all storage entries linked to a part"""
    entries = await db.storage.find(
        {"part_id": part_id},
        {"_id": 0}
    ).sort("date", -1).to_list(length=100)
    
    return entries


@app.post("/api/parts-library/{part_id}/add-to-storage")
async def add_part_to_storage(
    part_id: str,
    rack: str = "",
    qty: int = 0,
    current_user: User = Depends(get_current_user)
):
    """Add stock for a part - creates/updates storage entry and syncs to part library"""
    if qty <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be positive")
    if not rack:
        raise HTTPException(status_code=400, detail="Rack/location is required")
    
    part = await db.parts_library.find_one({"part_id": part_id})
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    
    # Check if storage entry exists for this part + rack
    existing = await db.storage.find_one({
        "part_id": part_id,
        "$or": [
            {"storage_place": {"$regex": f"^{rack}$", "$options": "i"}},
            {"crate_no": {"$regex": f"^{rack}$", "$options": "i"}}
        ]
    })
    
    now = get_ist_now()
    
    if existing:
        # Update existing storage entry
        new_qty = existing.get("quantity", 0) + qty
        await db.storage.update_one(
            {"entry_id": existing["entry_id"]},
            {"$set": {"quantity": new_qty, "updated_at": now}}
        )
        entry_id = existing["entry_id"]
    else:
        # Create new storage entry
        entry_id = f"stor_{uuid.uuid4().hex[:12]}"
        entry_doc = {
            "entry_id": entry_id,
            "date": now,
            "date_ist": format_ist_datetime(now),
            "part_id": part_id,
            "part_no": part_id,
            "assembly_name": part.get("source_assembly", ""),
            "product_details": part.get("name", ""),
            "quantity": qty,
            "storage_place": rack,
            "crate_no": "",
            "stored_by": current_user.name,
            "created_by_id": current_user.user_id,
            "created_by_name": current_user.name,
            "created_at": now
        }
        await db.storage.insert_one(entry_doc)
    
    # Sync part library stock
    result = await sync_part_stock_from_storage(part_id)
    
    return {
        "message": f"Added {qty} to {rack}",
        "entry_id": entry_id,
        "new_stock": result["stock"],
        "locations": result["locations"]
    }


@app.post("/api/parts-library/{part_id}/deduct-from-storage")
async def deduct_part_from_storage(
    part_id: str,
    rack: str = "",
    qty: int = 0,
    reason: str = "",
    current_user: User = Depends(get_current_user)
):
    """Deduct stock for a part from specific rack - updates storage entry and syncs"""
    if qty <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be positive")
    if not rack:
        raise HTTPException(status_code=400, detail="Rack/location is required")
    
    part = await db.parts_library.find_one({"part_id": part_id})
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    
    # Find storage entry for this part + rack
    existing = await db.storage.find_one({
        "part_id": part_id,
        "$or": [
            {"storage_place": {"$regex": f"^{rack}$", "$options": "i"}},
            {"crate_no": {"$regex": f"^{rack}$", "$options": "i"}}
        ]
    })
    
    if not existing:
        raise HTTPException(status_code=404, detail=f"No storage entry found for {part_id} at {rack}")
    
    current_qty = existing.get("quantity", 0)
    if current_qty < qty:
        raise HTTPException(status_code=400, detail=f"Insufficient stock at {rack}. Available: {current_qty}")
    
    now = get_ist_now()
    new_qty = current_qty - qty
    
    # Update storage entry
    await db.storage.update_one(
        {"entry_id": existing["entry_id"]},
        {"$set": {"quantity": new_qty, "updated_at": now}}
    )
    
    # Log the deduction
    deduction_log = {
        "log_id": f"dedlog_{uuid.uuid4().hex[:12]}",
        "entry_id": existing["entry_id"],
        "part_id": part_id,
        "part_no": part_id,
        "assembly_name": existing.get("assembly_name"),
        "quantity_taken": qty,
        "remaining_qty": new_qty,
        "rack": rack,
        "reason": reason,
        "taken_by_id": current_user.user_id,
        "taken_by_name": current_user.name,
        "recorded_by_id": current_user.user_id,
        "recorded_by_name": current_user.name,
        "taken_at": now,
        "can_undo": True,
        "undo_expires_at": now + timedelta(hours=24)
    }
    await db.storage_deductions.insert_one(deduction_log)
    
    # Sync part library stock
    result = await sync_part_stock_from_storage(part_id)
    
    return {
        "message": f"Deducted {qty} from {rack}",
        "log_id": deduction_log["log_id"],
        "new_stock": result["stock"],
        "locations": result["locations"]
    }


# =============================================
# ASSEMBLY PRODUCTION PLANNING WITH STOCK CHECK
# =============================================

class AssemblyProductionRequest(BaseModel):
    assembly_id: str
    quantity: int  # Number of assemblies to produce
    reason: Optional[str] = "Production"

@app.post("/api/assemblies/{assembly_id}/check-stock")
async def check_assembly_stock(
    assembly_id: str,
    quantity: int = 1,
    current_user: User = Depends(get_current_user)
):
    """Check if sufficient stock exists for assembly production"""
    # Get assembly
    assembly = await db.assemblies.find_one({"assembly_id": assembly_id})
    if not assembly:
        raise HTTPException(status_code=404, detail="Assembly not found")
    
    # Get BOM parts
    assembly_parts = await db.assembly_parts.find({"assembly_id": assembly_id}).to_list(length=100)
    
    shortages = []
    parts_detail = []
    
    for ap in assembly_parts:
        part = await db.parts_library.find_one({"part_id": ap["part_id"]})
        if not part:
            continue
        
        required_qty = ap["quantity"] * quantity
        available_stock = part.get("stock", 0)
        locations = part.get("locations", [])
        
        part_info = {
            "part_id": ap["part_id"],
            "part_name": part.get("name", ""),
            "required_qty": required_qty,
            "available_stock": available_stock,
            "locations": locations,
            "sufficient": available_stock >= required_qty,
            "shortage": max(0, required_qty - available_stock)
        }
        parts_detail.append(part_info)
        
        if available_stock < required_qty:
            shortages.append({
                "part_id": ap["part_id"],
                "part_name": part.get("name", ""),
                "required": required_qty,
                "available": available_stock,
                "shortage": required_qty - available_stock
            })
    
    return {
        "assembly_id": assembly_id,
        "assembly_name": assembly.get("name", ""),
        "production_qty": quantity,
        "can_proceed": len(shortages) == 0,
        "shortages": shortages,
        "parts": parts_detail
    }


@app.post("/api/assemblies/{assembly_id}/deduct-for-production")
async def deduct_assembly_for_production(
    assembly_id: str,
    request: AssemblyProductionRequest,
    current_user: User = Depends(get_current_user)
):
    """Deduct all parts from storage for assembly production"""
    quantity = request.quantity
    
    # First check stock
    stock_check = await check_assembly_stock(assembly_id, quantity, current_user)
    
    if not stock_check["can_proceed"]:
        raise HTTPException(
            status_code=400, 
            detail={
                "message": "Insufficient stock for production",
                "shortages": stock_check["shortages"]
            }
        )
    
    # Get BOM parts
    assembly_parts = await db.assembly_parts.find({"assembly_id": assembly_id}).to_list(length=100)
    
    now = get_ist_now()
    batch_id = f"prod_{uuid.uuid4().hex[:12]}"
    deductions = []
    
    for ap in assembly_parts:
        part = await db.parts_library.find_one({"part_id": ap["part_id"]})
        if not part:
            continue
        
        required_qty = ap["quantity"] * quantity
        locations = part.get("locations", [])
        remaining_to_deduct = required_qty
        
        # Deduct from each location until fulfilled
        for loc in locations:
            if remaining_to_deduct <= 0:
                break
            
            rack = loc.get("rack", "")
            available = loc.get("qty", 0)
            deduct_from_this = min(available, remaining_to_deduct)
            
            if deduct_from_this > 0:
                # Find the storage entry for this rack
                storage_entry = await db.storage.find_one({
                    "part_id": ap["part_id"],
                    "$or": [
                        {"storage_place": {"$regex": f"^{rack}$", "$options": "i"}},
                        {"crate_no": {"$regex": f"^{rack}$", "$options": "i"}}
                    ]
                })
                
                if storage_entry:
                    new_qty = storage_entry.get("quantity", 0) - deduct_from_this
                    await db.storage.update_one(
                        {"entry_id": storage_entry["entry_id"]},
                        {"$set": {"quantity": max(0, new_qty), "updated_at": now}}
                    )
                    
                    # Log deduction
                    log = {
                        "log_id": f"dedlog_{uuid.uuid4().hex[:12]}",
                        "batch_id": batch_id,
                        "entry_id": storage_entry["entry_id"],
                        "part_id": ap["part_id"],
                        "part_name": part.get("name", ""),
                        "assembly_id": assembly_id,
                        "production_qty": quantity,
                        "quantity_taken": deduct_from_this,
                        "rack": rack,
                        "reason": request.reason or "Production",
                        "taken_by_id": current_user.user_id,
                        "taken_by_name": current_user.name,
                        "taken_at": now,
                        "can_undo": True,
                        "undo_expires_at": now + timedelta(hours=24)
                    }
                    await db.storage_deductions.insert_one(log)
                    deductions.append(log)
                
                remaining_to_deduct -= deduct_from_this
        
        # Sync part library stock
        await sync_part_stock_from_storage(ap["part_id"])
    
    # Save production record
    production_record = {
        "batch_id": batch_id,
        "assembly_id": assembly_id,
        "assembly_name": stock_check["assembly_name"],
        "production_qty": quantity,
        "reason": request.reason,
        "deductions_count": len(deductions),
        "created_by_id": current_user.user_id,
        "created_by_name": current_user.name,
        "created_at": now,
        "can_undo": True,
        "undo_expires_at": now + timedelta(hours=24)
    }
    await db.production_deductions.insert_one(production_record)
    
    return {
        "success": True,
        "message": f"Deducted parts for {quantity} x {stock_check['assembly_name']}",
        "batch_id": batch_id,
        "deductions": len(deductions),
        "can_undo_until": (now + timedelta(hours=24)).isoformat()
    }


@app.post("/api/production/{batch_id}/undo")
async def undo_production_deduction(
    batch_id: str,
    current_user: User = Depends(get_current_user)
):
    """Undo a production deduction (within 24 hours)"""
    # Find production record
    production = await db.production_deductions.find_one({"batch_id": batch_id})
    if not production:
        raise HTTPException(status_code=404, detail="Production batch not found")
    
    # Check if can undo
    if not production.get("can_undo", False):
        raise HTTPException(status_code=400, detail="This production cannot be undone")
    
    now = get_ist_now()
    undo_expires = production.get("undo_expires_at")
    if undo_expires:
        # Ensure undo_expires is timezone-aware for comparison
        if isinstance(undo_expires, str):
            undo_expires = datetime.fromisoformat(undo_expires)
        if undo_expires.tzinfo is None:
            undo_expires = undo_expires.replace(tzinfo=timezone.utc)
        
        if now > undo_expires:
            raise HTTPException(status_code=400, detail="Undo period has expired (24 hours)")
    
    # Find all deduction logs for this batch
    deduction_logs = await db.storage_deductions.find({"batch_id": batch_id}).to_list(length=500)
    
    parts_restored = set()
    
    for log in deduction_logs:
        # Restore quantity to storage
        entry_id = log.get("entry_id")
        qty_taken = log.get("quantity_taken", 0)
        
        if entry_id and qty_taken > 0:
            storage_entry = await db.storage.find_one({"entry_id": entry_id})
            if storage_entry:
                new_qty = storage_entry.get("quantity", 0) + qty_taken
                await db.storage.update_one(
                    {"entry_id": entry_id},
                    {"$set": {"quantity": new_qty, "updated_at": now}}
                )
        
        parts_restored.add(log.get("part_id"))
        
        # Mark log as undone
        await db.storage_deductions.update_one(
            {"log_id": log["log_id"]},
            {"$set": {"undone": True, "undone_at": now, "undone_by": current_user.name}}
        )
    
    # Sync all affected parts
    for part_id in parts_restored:
        if part_id:
            await sync_part_stock_from_storage(part_id)
    
    # Mark production as undone
    await db.production_deductions.update_one(
        {"batch_id": batch_id},
        {"$set": {"can_undo": False, "undone": True, "undone_at": now, "undone_by": current_user.name}}
    )
    
    return {
        "success": True,
        "message": f"Production deduction undone. Restored {len(deduction_logs)} entries for {len(parts_restored)} parts.",
        "parts_restored": list(parts_restored)
    }


@app.get("/api/production/recent")
async def get_recent_productions(
    current_user: User = Depends(get_current_user)
):
    """Get recent production deductions with undo status"""
    productions = await db.production_deductions.find(
        {},
        {"_id": 0}
    ).sort("created_at", -1).limit(50).to_list(length=50)
    
    now = get_ist_now()
    for prod in productions:
        # Check if still can undo
        undo_expires = prod.get("undo_expires_at")
        if prod.get("can_undo") and undo_expires:
            prod["can_undo"] = now < undo_expires
    
    return productions


# =============================================
# PLASTIC RAW MATERIAL STORAGE ENDPOINTS
# =============================================

@app.post("/api/plastic-raw-material")
async def create_plastic_raw_material(
    entry: PlasticRawMaterialEntry,
    current_user: User = Depends(get_current_user)
):
    """Create a new plastic raw material entry"""
    now = get_ist_now()
    entry_id = f"prm_{uuid.uuid4().hex[:12]}"
    
    entry_doc = {
        "entry_id": entry_id,
        "date": now,
        "date_ist": format_ist_datetime(now),
        "material": entry.material,
        "sub_type": entry.sub_type,
        "glass_percentage": entry.glass_percentage,
        "make": entry.make,
        "grade": entry.grade,
        "colour": entry.colour,
        "quantity_kg": entry.quantity_kg,
        "storage_place": entry.storage_place,
        "stored_by_id": current_user.user_id,
        "stored_by_name": current_user.name,
        "image": entry.image,
        "created_at": now,
        "updated_at": now
    }
    
    await db.plastic_raw_materials.insert_one(entry_doc)
    return {"message": "Plastic raw material added", "entry_id": entry_id}


@app.get("/api/plastic-raw-material")
async def get_plastic_raw_materials(
    keyword: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all plastic raw material entries"""
    query = {}
    
    if keyword:
        query["$or"] = [
            {"material": {"$regex": keyword, "$options": "i"}},
            {"make": {"$regex": keyword, "$options": "i"}},
            {"grade": {"$regex": keyword, "$options": "i"}},
            {"colour": {"$regex": keyword, "$options": "i"}},
            {"storage_place": {"$regex": keyword, "$options": "i"}},
        ]
    
    entries = await db.plastic_raw_materials.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=500)
    
    # Convert datetimes to IST
    for entry in entries:
        if entry.get("date") and hasattr(entry["date"], 'isoformat'):
            entry["date"] = to_ist(entry["date"]).isoformat()
        if entry.get("created_at") and hasattr(entry["created_at"], 'isoformat'):
            entry["created_at"] = to_ist(entry["created_at"]).isoformat()
    
    return entries


@app.post("/api/plastic-raw-material/{entry_id}/deduct")
async def deduct_plastic_raw_material(
    entry_id: str,
    deduct_request: PlasticRawMaterialDeductRequest,
    current_user: User = Depends(get_current_user)
):
    """Deduct quantity from plastic raw material and log moulding details"""
    entry = await db.plastic_raw_materials.find_one({"entry_id": entry_id})
    
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    current_qty = entry.get("quantity_kg", 0)
    if deduct_request.quantity_kg > current_qty:
        raise HTTPException(status_code=400, detail=f"Cannot deduct {deduct_request.quantity_kg}kg. Only {current_qty}kg available")
    
    new_qty = current_qty - deduct_request.quantity_kg
    now = get_ist_now()
    
    # Update the quantity
    await db.plastic_raw_materials.update_one(
        {"entry_id": entry_id},
        {"$set": {"quantity_kg": new_qty, "updated_at": now}}
    )
    
    # Log the deduction with moulding details
    deduction_log = {
        "log_id": f"prm_deduct_{uuid.uuid4().hex[:8]}",
        "entry_id": entry_id,
        "material": entry.get("material"),
        "quantity_kg": deduct_request.quantity_kg,
        "given_to_id": deduct_request.given_to,
        "given_to_name": deduct_request.given_to_name,
        "for_moulding": deduct_request.for_moulding,
        "remarks": deduct_request.remarks,
        "deducted_by_id": current_user.user_id,
        "deducted_by_name": current_user.name,
        "deducted_at": now,
        "deducted_at_ist": format_ist_datetime(now),
        "previous_qty": current_qty,
        "new_qty": new_qty,
        "signature": deduct_request.signature
    }
    
    await db.plastic_raw_material_transactions.insert_one(deduction_log)
    
    return {
        "message": "Stock deducted successfully",
        "previous_qty": current_qty,
        "deducted_qty": deduct_request.quantity_kg,
        "new_qty": new_qty,
        "for_moulding": deduct_request.for_moulding
    }


@app.post("/api/plastic-raw-material/{entry_id}/add")
async def add_plastic_raw_material_qty(
    entry_id: str,
    add_request: PlasticRawMaterialAddRequest,
    current_user: User = Depends(get_current_user)
):
    """Add quantity to plastic raw material"""
    entry = await db.plastic_raw_materials.find_one({"entry_id": entry_id})
    
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    current_qty = entry.get("quantity_kg", 0)
    new_qty = current_qty + add_request.quantity_kg
    now = get_ist_now()
    
    # Update the quantity
    await db.plastic_raw_materials.update_one(
        {"entry_id": entry_id},
        {"$set": {"quantity_kg": new_qty, "updated_at": now}}
    )
    
    # Log the addition
    add_log = {
        "log_id": f"prm_add_{uuid.uuid4().hex[:8]}",
        "entry_id": entry_id,
        "material": entry.get("material"),
        "quantity_kg": add_request.quantity_kg,
        "remarks": add_request.remarks,
        "added_by_id": current_user.user_id,
        "added_by_name": current_user.name,
        "added_at": now,
        "added_at_ist": format_ist_datetime(now),
        "previous_qty": current_qty,
        "new_qty": new_qty,
        "type": "add"
    }
    
    await db.plastic_raw_material_transactions.insert_one(add_log)
    
    return {
        "message": "Stock added successfully",
        "previous_qty": current_qty,
        "added_qty": add_request.quantity_kg,
        "new_qty": new_qty
    }


@app.get("/api/plastic-raw-material/{entry_id}/transactions")
async def get_plastic_raw_material_transactions(
    entry_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get transaction history for a plastic raw material entry"""
    transactions = await db.plastic_raw_material_transactions.find(
        {"entry_id": entry_id},
        {"_id": 0}
    ).sort("deducted_at", -1).to_list(length=100)
    
    return transactions


@app.put("/api/plastic-raw-material/{entry_id}")
async def update_plastic_raw_material(
    entry_id: str,
    update_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Update a plastic raw material entry"""
    entry = await db.plastic_raw_materials.find_one({"entry_id": entry_id})
    
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"} and entry.get("stored_by_id") != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    allowed_fields = ["material", "sub_type", "glass_percentage", "make", "grade", 
                      "colour", "storage_place", "image"]
    
    update_dict = {k: v for k, v in update_data.items() if k in allowed_fields}
    update_dict["updated_at"] = get_ist_now()
    
    await db.plastic_raw_materials.update_one({"entry_id": entry_id}, {"$set": update_dict})
    
    updated_entry = await db.plastic_raw_materials.find_one({"entry_id": entry_id}, {"_id": 0})
    return updated_entry
# =============================================
# GAUGES STORAGE ENDPOINTS - MOVED TO: routes/gauges.py
# =============================================

    return {"message": "Gauge deleted successfully"}


@app.delete("/api/plastic-raw-materials/{entry_id}")
async def delete_plastic_raw_material(
    entry_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a plastic raw material entry (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    entry = await db.plastic_raw_materials.find_one({"entry_id": entry_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    await db.plastic_raw_materials.delete_one({"entry_id": entry_id})
    return {"message": "Plastic raw material entry deleted successfully"}


# =====================
# EXPORT ENDPOINTS
# =====================

@app.get("/api/export/production-entries")
async def export_production_entries(
    format: str = "json",  # json, csv, pdf
    category: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    machine: Optional[str] = None,
    operator: Optional[str] = None,
    job_details: Optional[str] = None
):
    """Export production entries - Public access for downloading reports"""
    query = {}
    if category:
        query["category"] = category
    if date_from:
        query["date"] = {"$gte": date_from}
    if date_to:
        if "date" not in query:
            query["date"] = {}
        query["date"]["$lte"] = date_to
    if machine:
        query["machine_name"] = machine
    if operator:
        query["operator_name"] = {"$regex": operator, "$options": "i"}
    if job_details:
        query["job_details"] = {"$regex": job_details, "$options": "i"}
    
    entries = await db.production_entries.find(query, {"_id": 0}).sort("date", -1).to_list(length=5000)
    
    if format == "csv":
        import io
        output = io.StringIO()
        if entries:
            headers = ["Date", "Machine", "Shift", "Operator", "Job Details", "Production Qty", "Cycle Time (s)", "Setting Time", "Spindle RPM", "Feed Rate", "Remarks"]
            output.write(",".join(headers) + "\n")
            for entry in entries:
                row = [
                    str(entry.get("date", ""))[:10],
                    entry.get("machine_name", ""),
                    entry.get("shift", ""),
                    entry.get("operator_name", ""),
                    entry.get("job_details", "").replace(",", ";"),
                    str(entry.get("production_qty", 0)),
                    str(entry.get("cycle_time", 0)),
                    str(entry.get("setting_time", 0)),
                    str(entry.get("spindle_rpm", "") or ""),
                    str(entry.get("feed_rate", "") or ""),
                    (entry.get("remarks") or "").replace(",", ";")
                ]
                output.write(",".join(row) + "\n")
        
        from fastapi.responses import StreamingResponse
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=production_entries.csv"}
        )
    
    if format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        import io
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.darkblue, spaceAfter=20)
        elements.append(Paragraph(f"Production Report - {category.upper() if category else 'All'}", title_style))
        elements.append(Spacer(1, 10))
        
        # Table data
        table_data = [["Date", "Machine", "Shift", "Operator", "Job Details", "Qty", "Cycle", "Setting", "RPM", "Feed"]]
        for entry in entries[:200]:  # Limit to 200 for PDF
            cycle_sec = entry.get("cycle_time", 0) or 0
            cycle_display = f"{cycle_sec//60}m{cycle_sec%60}s" if cycle_sec > 0 else "-"
            table_data.append([
                str(entry.get("date", ""))[:10],
                (entry.get("machine_name", "") or "-")[:12],
                entry.get("shift", "day")[:3],
                (entry.get("operator_name", "") or "-")[:10],
                (entry.get("job_details", "") or "-")[:20],
                str(entry.get("production_qty", 0)),
                cycle_display,
                str(entry.get("setting_time", 0) or "-"),
                str(entry.get("spindle_rpm", "") or "-"),
                str(entry.get("feed_rate", "") or "-")
            ])
        
        # Create table
        col_widths = [55, 60, 30, 55, 100, 30, 40, 40, 40, 40]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F1F5F9')]),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=production_report_{category or 'all'}.pdf"}
        )
    
    return entries

@app.get("/api/export/storage")
async def export_storage_entries(
    format: str = "json"
):
    """Export storage entries - Public access for downloading reports"""
    entries = await db.storage.find({}, {"_id": 0, "image": 0}).sort("date", -1).to_list(length=5000)
    
    if format == "csv":
        import io
        output = io.StringIO()
        if entries:
            headers = ["Date", "Assembly Name", "Part No", "Product Details", "Storage Place", "Crate No", "Quantity", "Stored By"]
            output.write(",".join(headers) + "\n")
            for entry in entries:
                row = [
                    str(entry.get("date", ""))[:19],
                    entry.get("assembly_name", "").replace(",", ";"),
                    entry.get("part_no", ""),
                    entry.get("product_details", "").replace(",", ";"),
                    entry.get("storage_place", ""),
                    entry.get("crate_no", ""),
                    str(entry.get("quantity", 1)),
                    entry.get("stored_by", "")
                ]
                output.write(",".join(row) + "\n")
        
        from fastapi.responses import StreamingResponse
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=storage_inventory.csv"}
        )
    
    return entries


@app.get("/api/export/tools-inserts")
async def export_tools_inserts(
    format: str = "json",
    item_type: Optional[str] = None
):
    """Export tools and inserts - Public access for downloading reports"""
    query = {}
    if item_type:
        query["type"] = item_type
    
    items = await db.tools_inserts.find(query, {"_id": 0, "image": 0}).sort("created_at", -1).to_list(length=5000)
    
    if format == "csv":
        import io
        output = io.StringIO()
        if items:
            headers = ["Type", "Material", "Diameter", "Length", "Grade", "Endmill Type", "Stock Qty", "Scrap Qty", "Remarks"]
            output.write(",".join(headers) + "\n")
            for item in items:
                row = [
                    item.get("type", ""),
                    item.get("material", ""),
                    str(item.get("diameter", "")),
                    str(item.get("length", "")),
                    item.get("grade", ""),
                    item.get("endmill_type", ""),
                    str(item.get("stock_qty", 0)),
                    str(item.get("scrap_qty", 0)),
                    (item.get("remarks") or "").replace(",", ";")
                ]
                output.write(",".join(row) + "\n")
        
        from fastapi.responses import StreamingResponse
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=tools_inserts.csv"}
        )
    
    return items


# ============================================
# MACHINE SPARES MANAGEMENT SYSTEM
# ============================================

class MachineSpareCreate(BaseModel):
    category: str
    name: str
    specs: Optional[str] = None
    make: Optional[str] = None
    quantity: int = 0
    location: Optional[str] = None
    min_quantity: Optional[int] = 5
    image: Optional[str] = None  # base64 encoded image

class SpareStockUpdate(BaseModel):
    quantity: int
    remarks: Optional[str] = None

class SpareTakeRequest(BaseModel):
    quantity: int
    remarks: Optional[str] = None
    signature: Optional[str] = None


@app.get("/api/machines/{machine_id}/spares")
async def get_machine_spares(
    machine_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get all spares for a specific machine"""
    spares = await db.machine_spares.find({"machine_id": machine_id}).to_list(length=500)
    
    for spare in spares:
        spare["_id"] = str(spare["_id"])
        spare["is_low_stock"] = spare.get("quantity", 0) <= spare.get("min_quantity", 5)
    
    return spares


@app.post("/api/machines/{machine_id}/spares")
async def add_machine_spare(
    machine_id: str,
    spare: MachineSpareCreate,
    current_user: User = Depends(get_current_user)
):
    """Add a new spare to a machine"""
    # Get machine details
    machine = await db.machines.find_one({"machine_id": machine_id})
    if not machine:
        raise HTTPException(status_code=404, detail="Machine not found")
    
    spare_id = f"SP-{machine_id[:8]}-{str(uuid.uuid4())[:8].upper()}"
    
    spare_doc = {
        "spare_id": spare_id,
        "machine_id": machine_id,
        "machine_name": machine.get("name", "Unknown"),
        "category": spare.category,
        "name": spare.name,
        "specs": spare.specs,
        "make": spare.make,
        "quantity": spare.quantity,
        "location": spare.location,
        "min_quantity": spare.min_quantity,
        "image": spare.image,  # base64 image
        "stored_by_id": current_user.user_id,
        "stored_by_name": current_user.name,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.machine_spares.insert_one(spare_doc)
    
    # Log the addition
    if spare.quantity > 0:
        log_entry = {
            "log_id": str(uuid.uuid4()),
            "spare_id": spare_id,
            "machine_id": machine_id,
            "action": "added",
            "quantity": spare.quantity,
            "added_by_id": current_user.user_id,
            "added_by_name": current_user.name,
            "timestamp": datetime.now(timezone.utc),
            "timestamp_ist": format_ist_datetime(datetime.now(timezone.utc)),
            "remarks": "Initial stock"
        }
        await db.machine_spares_log.insert_one(log_entry)
    
    return {"message": "Spare added successfully", "spare_id": spare_id}


@app.get("/api/machine-spares/{spare_id}")
async def get_spare_details(
    spare_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get details of a specific spare"""
    spare = await db.machine_spares.find_one({"spare_id": spare_id})
    if not spare:
        raise HTTPException(status_code=404, detail="Spare not found")
    
    spare["_id"] = str(spare["_id"])
    spare["is_low_stock"] = spare.get("quantity", 0) <= spare.get("min_quantity", 5)
    
    return spare


@app.get("/api/machine-spares/{spare_id}/history")
async def get_spare_history(
    spare_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get history log for a specific spare"""
    logs = await db.machine_spares_log.find({"spare_id": spare_id}).sort("timestamp", -1).to_list(length=100)
    
    for log in logs:
        log["_id"] = str(log["_id"])
    
    return logs


@app.post("/api/machine-spares/{spare_id}/take")
async def take_spare(
    spare_id: str,
    request: SpareTakeRequest,
    current_user: User = Depends(get_current_user)
):
    """Take spare from stock (with signature)"""
    spare = await db.machine_spares.find_one({"spare_id": spare_id})
    if not spare:
        raise HTTPException(status_code=404, detail="Spare not found")
    
    current_qty = spare.get("quantity", 0)
    if request.quantity > current_qty:
        raise HTTPException(status_code=400, detail=f"Not enough stock. Available: {current_qty}")
    
    new_qty = current_qty - request.quantity
    
    # Update quantity
    await db.machine_spares.update_one(
        {"spare_id": spare_id},
        {"$set": {"quantity": new_qty, "updated_at": datetime.now(timezone.utc)}}
    )
    
    # Log the take action
    log_entry = {
        "log_id": str(uuid.uuid4()),
        "spare_id": spare_id,
        "machine_id": spare.get("machine_id"),
        "action": "taken",
        "quantity": request.quantity,
        "remaining_qty": new_qty,
        "taken_by_id": current_user.user_id,
        "taken_by_name": current_user.name,
        "timestamp": datetime.now(timezone.utc),
        "timestamp_ist": format_ist_datetime(datetime.now(timezone.utc)),
        "signature": request.signature,
        "remarks": request.remarks
    }
    await db.machine_spares_log.insert_one(log_entry)
    
    return {"message": "Spare taken successfully", "remaining_qty": new_qty}


@app.post("/api/machine-spares/{spare_id}/add-stock")
async def add_spare_stock(
    spare_id: str,
    request: SpareStockUpdate,
    current_user: User = Depends(get_current_user)
):
    """Add stock to a spare"""
    spare = await db.machine_spares.find_one({"spare_id": spare_id})
    if not spare:
        raise HTTPException(status_code=404, detail="Spare not found")
    
    current_qty = spare.get("quantity", 0)
    new_qty = current_qty + request.quantity
    
    # Update quantity
    await db.machine_spares.update_one(
        {"spare_id": spare_id},
        {"$set": {"quantity": new_qty, "updated_at": datetime.now(timezone.utc)}}
    )
    
    # Log the addition
    log_entry = {
        "log_id": str(uuid.uuid4()),
        "spare_id": spare_id,
        "machine_id": spare.get("machine_id"),
        "action": "added",
        "quantity": request.quantity,
        "new_total": new_qty,
        "added_by_id": current_user.user_id,
        "added_by_name": current_user.name,
        "timestamp": datetime.now(timezone.utc),
        "timestamp_ist": format_ist_datetime(datetime.now(timezone.utc)),
        "remarks": request.remarks
    }
    await db.machine_spares_log.insert_one(log_entry)
    
    return {"message": "Stock added successfully", "new_quantity": new_qty}


@app.put("/api/machine-spares/{spare_id}")
async def update_spare(
    spare_id: str,
    spare: MachineSpareCreate,
    current_user: User = Depends(get_current_user)
):
    """Update spare details"""
    existing = await db.machine_spares.find_one({"spare_id": spare_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Spare not found")
    
    update_doc = {
        "category": spare.category,
        "name": spare.name,
        "specs": spare.specs,
        "make": spare.make,
        "location": spare.location,
        "min_quantity": spare.min_quantity,
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.machine_spares.update_one({"spare_id": spare_id}, {"$set": update_doc})
    
    return {"message": "Spare updated successfully"}


@app.delete("/api/machine-spares/{spare_id}")
async def delete_spare(
    spare_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a spare (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await db.machine_spares.delete_one({"spare_id": spare_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Spare not found")
    
    # Also delete history logs
    await db.machine_spares_log.delete_many({"spare_id": spare_id})
    
    return {"message": "Spare deleted successfully"}


# ============================================
# MACHINE STATUS & JOB TRACKING SYSTEM
# ============================================

# Get machine status summary
@app.get("/api/machine-status/summary")
async def get_machine_status_summary(
    current_user: User = Depends(get_current_user)
):
    """Get summary of all machines with their current status"""
    machines = await db.machines.find().to_list(length=100)
    
    running = 0
    idle = 0
    breakdown = 0
    
    machine_details = []
    for machine in machines:
        status = machine.get("status", "idle")
        if status == "running":
            running += 1
        elif status == "breakdown":
            breakdown += 1
        else:
            idle += 1
        
        # Get current job info if running
        current_job = None
        if machine.get("current_job_id"):
            job = await db.jobs.find_one({"job_id": machine["current_job_id"]})
            if job:
                progress = 0
                if job.get("target_qty") and job.get("target_qty") > 0:
                    progress = min(100, int((job.get("produced_qty", 0) / job["target_qty"]) * 100))
                
                # Check if delayed
                is_delayed = False
                if job.get("start_time") and job.get("expected_hours"):
                    start_time = job["start_time"]
                    # Handle timezone-naive datetime from MongoDB
                    if start_time.tzinfo is None:
                        start_time = start_time.replace(tzinfo=timezone.utc)
                    elapsed_hours = (datetime.now(timezone.utc) - start_time).total_seconds() / 3600
                    is_delayed = elapsed_hours > job["expected_hours"]
                
                current_job = {
                    "job_id": job["job_id"],
                    "job_details": job.get("job_details", "")[:50],
                    "progress": progress,
                    "target_qty": job.get("target_qty"),
                    "produced_qty": job.get("produced_qty", 0),
                    "start_time": job.get("start_time").isoformat() if job.get("start_time") else None,
                    "is_delayed": is_delayed
                }
        
        # Get current operator info
        current_operator = None
        if machine.get("current_operator_id"):
            operator = await db.users.find_one({"user_id": machine["current_operator_id"]})
            if operator:
                current_operator = {
                    "user_id": operator["user_id"],
                    "name": operator.get("name", "Unknown"),
                    "picture": operator.get("profile_image") or operator.get("picture")
                }
        
        machine_details.append({
            "machine_id": machine["machine_id"],
            "name": machine["name"],
            "category": machine.get("category", ""),
            "status": status,
            "current_job": current_job,
            "current_operator": current_operator,
            "last_status_update": machine.get("last_status_update").isoformat() if machine.get("last_status_update") else None
        })
    
    return {
        "summary": {
            "total": len(machines),
            "running": running,
            "idle": idle,
            "breakdown": breakdown
        },
        "machines": machine_details
    }


# Public endpoint for TV Mode - No authentication required
@app.get("/api/machine-status/summary-public")
async def get_machine_status_summary_public():
    """Get summary of all machines with their current status - PUBLIC for TV display"""
    machines = await db.machines.find().to_list(length=100)
    
    running = 0
    idle = 0
    breakdown = 0
    
    machine_details = []
    for machine in machines:
        status = machine.get("status", "idle")
        if status == "running":
            running += 1
        elif status == "breakdown":
            breakdown += 1
        else:
            idle += 1
        
        # Get current active work for this machine
        active_work = await db.active_work.find_one({
            "machine_id": machine["machine_id"],
            "status": "in_progress"
        })
        
        machine_info = {
            "machine_id": machine["machine_id"],
            "machine_name": machine["name"],
            "category": machine.get("category", ""),
            "status": status,
        }
        
        if active_work:
            machine_info["operator_name"] = active_work.get("operator_name", "")
            machine_info["job_details"] = active_work.get("job_details", "")[:100]
            if active_work.get("start_time_ist"):
                machine_info["start_time"] = active_work["start_time_ist"]
        
        machine_details.append(machine_info)
    
    return {
        "summary": {
            "total": len(machines),
            "running": running,
            "idle": idle,
            "breakdown": breakdown
        },
        "machines": machine_details
    }


# Start a job
@app.post("/api/jobs/start")
async def start_job(
    request: JobStartRequest,
    current_user: User = Depends(get_current_user)
):
    """Start a job - sets job to in_progress and optionally machine to running. 
    Can either start an existing job or create a new job with manual details.
    Machine is optional for roles like Programmer/Designer."""
    
    now = get_ist_now()  # Use IST timezone
    today = now.strftime("%Y-%m-%d")
    
    machine = None
    machine_name = None
    
    # Verify machine exists only if machine_id is provided
    if request.machine_id:
        machine = await db.machines.find_one({"machine_id": request.machine_id})
        if not machine:
            raise HTTPException(status_code=404, detail="Machine not found")
        
        # Check if machine already has an active job
        if machine.get("current_job_id") and machine.get("status") == "running":
            raise HTTPException(status_code=400, detail="Machine already has an active job. Complete or stop the current job first.")
        
        machine_name = machine.get("name")
    
    job = None
    job_id = request.job_id
    job_details_text = ""
    
    # If job_id is provided, use existing job
    if request.job_id:
        job = await db.jobs.find_one({"job_id": request.job_id})
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")
        
        if job.get("status") == "in_progress":
            raise HTTPException(status_code=400, detail="Job is already in progress")
        
        if job.get("status") == "completed":
            raise HTTPException(status_code=400, detail="Job is already completed")
        
        job_details_text = job.get("job_details", "")
    
    # If manual job_details is provided, create a new job
    elif request.job_details:
        job_id = f"job_{uuid.uuid4().hex[:12]}"
        new_job = {
            "job_id": job_id,
            "job_details": request.job_details,
            "category": "Manual",
            "priority": "medium",
            "assigned_to": current_user.user_id,
            "assigned_to_name": current_user.name,
            "created_by": current_user.user_id,
            "created_by_name": current_user.name,
            "status": "pending",
            "created_date": now,
            "updated_date": now,
            "started_by_name": current_user.name,
            "started_at": now
        }
        await db.jobs.insert_one(new_job)
        job = new_job
        job_details_text = request.job_details
    else:
        raise HTTPException(status_code=400, detail="Either job_id or job_details must be provided")
    
    # Update job to in_progress
    update_data = {
        "status": "in_progress",
        "start_time": now,
        "updated_date": now,
        "operator_id": current_user.user_id,
        "operator_name": current_user.name,
        "started_by_name": current_user.name,
        "started_at": now
    }
    if request.machine_id:
        update_data["machine_id"] = request.machine_id
        update_data["machine_name"] = machine_name
    if request.target_qty:
        update_data["target_qty"] = request.target_qty
    if request.expected_hours:
        update_data["expected_hours"] = request.expected_hours
    if request.estimated_completion_minutes:
        update_data["estimated_completion_minutes"] = request.estimated_completion_minutes
    
    await db.jobs.update_one(
        {"job_id": job_id},
        {"$set": update_data}
    )
    
    # Update machine status only if machine is provided
    if request.machine_id and machine:
        await db.machines.update_one(
            {"machine_id": request.machine_id},
            {"$set": {
                "status": "running",
                "current_job_id": job_id,
                "current_operator_id": current_user.user_id,
                "last_status_update": now
            }}
        )
        
        # Create daily machine status record
        record_id = f"dms_{uuid.uuid4().hex[:12]}"
        daily_record = {
            "record_id": record_id,
            "date": today,
            "machine_id": request.machine_id,
            "machine_name": machine["name"],
            "operator_id": current_user.user_id,
            "operator_name": current_user.name,
            "job_id": job_id,
            "job_details": job_details_text[:100],
            "status": "running",
            "start_time": now,
            "end_time": None,
            "created_at": now
        }
        await db.daily_machine_status.insert_one(daily_record)
    
    return {
        "message": "Job started successfully",
        "job_id": job_id,
        "machine_id": request.machine_id,
        "start_time": now.isoformat()
    }


# Complete a job
@app.post("/api/jobs/complete")
async def complete_job(
    request: JobCompleteRequest,
    current_user: User = Depends(get_current_user)
):
    """Complete a job - sets job to completed and machine to idle"""
    # Verify job exists and is in progress
    job = await db.jobs.find_one({"job_id": request.job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if job.get("status") != "in_progress":
        raise HTTPException(status_code=400, detail="Job is not in progress")
    
    now = datetime.now(timezone.utc)
    
    # Update job
    update_data = {
        "status": "completed",
        "end_time": now,
        "completed_at": now,
        "updated_date": now,
        "completed_by_id": current_user.user_id,
        "completed_by_name": current_user.name
    }
    if request.produced_qty is not None:
        update_data["produced_qty"] = request.produced_qty
    if request.remarks is not None:
        update_data["completion_remarks"] = request.remarks
    
    await db.jobs.update_one(
        {"job_id": request.job_id},
        {"$set": update_data}
    )
    
    # Find the machine with this job and set it to idle
    machine = await db.machines.find_one({"current_job_id": request.job_id})
    if machine:
        await db.machines.update_one(
            {"machine_id": machine["machine_id"]},
            {"$set": {
                "status": "idle",
                "current_job_id": None,
                "current_operator_id": None,
                "last_status_update": now
            }}
        )
        
        # Update daily machine status record
        await db.daily_machine_status.update_one(
            {"job_id": request.job_id, "end_time": None},
            {"$set": {
                "end_time": now,
                "status": "completed"
            }}
        )
    
    return {
        "message": "Job completed successfully",
        "job_id": request.job_id,
        "end_time": now.isoformat()
    }


# Update machine status (manual breakdown, etc.)
@app.put("/api/machines/{machine_id}/status")
async def update_machine_status(
    machine_id: str,
    status_update: MachineStatusUpdate,
    current_user: User = Depends(get_current_user)
):
    """Manually update machine status (e.g., mark as breakdown)"""
    machine = await db.machines.find_one({"machine_id": machine_id})
    if not machine:
        raise HTTPException(status_code=404, detail="Machine not found")
    
    valid_statuses = ["running", "idle", "breakdown"]
    if status_update.status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    now = datetime.now(timezone.utc)
    
    update_data = {
        "status": status_update.status,
        "last_status_update": now
    }
    
    # If marking as breakdown or idle, clear current job
    if status_update.status in ["breakdown", "idle"]:
        if machine.get("current_job_id"):
            # Update the job status if there was an active job
            await db.jobs.update_one(
                {"job_id": machine["current_job_id"], "status": "in_progress"},
                {"$set": {"status": "pending", "updated_date": now}}
            )
        update_data["current_job_id"] = None
        update_data["current_operator_id"] = None
    
    await db.machines.update_one(
        {"machine_id": machine_id},
        {"$set": update_data}
    )
    
    # If breakdown, notify admins
    if status_update.status == "breakdown":
        await notify_admins(
            "🔴 Machine Breakdown",
            f"{machine['name']} has been marked as breakdown by {current_user.name}"
        )
    
    return {"message": f"Machine status updated to {status_update.status}"}


# Get today's machine status records
@app.get("/api/daily-status")
async def get_daily_status(
    date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get daily machine status records"""
    if date is None:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    records = await db.daily_machine_status.find(
        {"date": date}
    ).sort("start_time", -1).to_list(length=200)
    
    result = []
    for record in records:
        result.append({
            "record_id": record["record_id"],
            "date": record["date"],
            "machine_id": record["machine_id"],
            "machine_name": record["machine_name"],
            "operator_id": record["operator_id"],
            "operator_name": record["operator_name"],
            "job_id": record["job_id"],
            "job_details": record["job_details"],
            "status": record["status"],
            "start_time": record["start_time"].isoformat() if record.get("start_time") else None,
            "end_time": record["end_time"].isoformat() if record.get("end_time") else None
        })
    
    return result


# Update job progress (produced qty)
@app.put("/api/jobs/{job_id}/progress")
async def update_job_progress(
    job_id: str,
    target_qty: Optional[int] = None,
    produced_qty: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Update job progress (target and produced quantities)"""
    job = await db.jobs.find_one({"job_id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    update_data = {"updated_date": datetime.now(timezone.utc)}
    
    if target_qty is not None:
        update_data["target_qty"] = target_qty
    
    if produced_qty is not None:
        update_data["produced_qty"] = produced_qty
    
    await db.jobs.update_one(
        {"job_id": job_id},
        {"$set": update_data}
    )
    
    return {"message": "Job progress updated successfully"}


# Get idle machines (available for job assignment)
@app.get("/api/machines/idle")
async def get_idle_machines(
    current_user: User = Depends(get_current_user)
):
    """Get all idle machines available for job assignment"""
    machines = await db.machines.find({"status": {"$ne": "breakdown"}}).to_list(length=100)
    
    result = []
    for machine in machines:
        is_available = machine.get("status") != "running" and not machine.get("current_job_id")
        result.append({
            "machine_id": machine["machine_id"],
            "name": machine["name"],
            "category": machine.get("category", ""),
            "status": machine.get("status", "idle"),
            "is_available": is_available
        })
    
    return result



# ============================================
# TOOLS & INSERTS INVENTORY SYSTEM
# ============================================

# --- VMC TOOLS ENDPOINTS ---

@app.post("/api/tools")
async def create_tool(
    tool: ToolCreate,
    current_user: User = Depends(get_current_user)
):
    """Add a new tool to inventory"""
    tool_id = f"tool_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    tool_doc = {
        "tool_id": tool_id,
        "tool_type": tool.tool_type,
        "subtype": tool.subtype,
        "diameter": tool.diameter,
        "material": tool.material,
        "work_material": tool.work_material,
        "quantity": tool.quantity,
        "min_quantity": tool.min_quantity,
        "location": tool.location,
        "remarks": tool.remarks,
        "created_by": current_user.user_id,
        "created_at": now,
        "updated_at": now
    }
    
    await db.tools.insert_one(tool_doc)
    
    return {"message": "Tool added successfully", "tool_id": tool_id}


@app.get("/api/tools")
async def get_tools(
    low_stock: bool = False,
    current_user: User = Depends(get_current_user)
):
    """Get all tools from inventory"""
    query = {}
    if low_stock:
        # Use $expr to compare two fields
        query = {"$expr": {"$lte": ["$quantity", "$min_quantity"]}}
    
    tools = await db.tools.find(query).sort("tool_type", 1).to_list(length=500)
    
    result = []
    for tool in tools:
        is_low_stock = tool.get("quantity", 0) <= tool.get("min_quantity", 0)
        result.append({
            "tool_id": tool["tool_id"],
            "tool_type": tool.get("tool_type", ""),
            "subtype": tool.get("subtype", ""),
            "diameter": tool.get("diameter", 0),
            "material": tool.get("material", ""),
            "work_material": tool.get("work_material", ""),
            "quantity": tool.get("quantity", 0),
            "min_quantity": tool.get("min_quantity", 0),
            "location": tool.get("location", ""),
            "remarks": tool.get("remarks", ""),
            "is_low_stock": is_low_stock,
            "created_at": tool.get("created_at").isoformat() if tool.get("created_at") else None
        })
    
    return result


@app.put("/api/tools/{tool_id}")
async def update_tool(
    tool_id: str,
    tool_update: ToolUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update tool quantity or details"""
    tool = await db.tools.find_one({"tool_id": tool_id})
    if not tool:
        raise HTTPException(status_code=404, detail="Tool not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc)}
    
    if tool_update.quantity is not None:
        update_data["quantity"] = tool_update.quantity
    if tool_update.min_quantity is not None:
        update_data["min_quantity"] = tool_update.min_quantity
    if tool_update.location is not None:
        update_data["location"] = tool_update.location
    if tool_update.remarks is not None:
        update_data["remarks"] = tool_update.remarks
    
    await db.tools.update_one({"tool_id": tool_id}, {"$set": update_data})
    
    return {"message": "Tool updated successfully"}


@app.delete("/api/tools/{tool_id}")
async def delete_tool(
    tool_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a tool from inventory"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Only Admin can delete tools")
    
    result = await db.tools.delete_one({"tool_id": tool_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tool not found")
    
    return {"message": "Tool deleted successfully"}


# --- CNC LATHE INSERTS ENDPOINTS ---

@app.post("/api/inserts")
async def create_insert(
    insert: InsertCreate,
    current_user: User = Depends(get_current_user)
):
    """Add a new insert to inventory"""
    insert_id = f"insert_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    insert_doc = {
        "insert_id": insert_id,
        "insert_type": insert.insert_type,
        "grade": insert.grade,
        "work_material": insert.work_material,
        "quantity": insert.quantity,
        "min_quantity": insert.min_quantity,
        "location": insert.location,
        "remarks": insert.remarks,
        "created_by": current_user.user_id,
        "created_at": now,
        "updated_at": now
    }
    
    await db.inserts.insert_one(insert_doc)
    
    return {"message": "Insert added successfully", "insert_id": insert_id}


@app.get("/api/inserts")
async def get_inserts(
    low_stock: bool = False,
    current_user: User = Depends(get_current_user)
):
    """Get all inserts from inventory"""
    query = {}
    if low_stock:
        query = {"$expr": {"$lte": ["$quantity", "$min_quantity"]}}
    
    inserts = await db.inserts.find(query).sort("insert_type", 1).to_list(length=500)
    
    result = []
    for insert in inserts:
        is_low_stock = insert.get("quantity", 0) <= insert.get("min_quantity", 0)
        result.append({
            "insert_id": insert["insert_id"],
            "insert_type": insert.get("insert_type", ""),
            "grade": insert.get("grade", ""),
            "work_material": insert.get("work_material", ""),
            "quantity": insert.get("quantity", 0),
            "min_quantity": insert.get("min_quantity", 0),
            "location": insert.get("location", ""),
            "remarks": insert.get("remarks", ""),
            "is_low_stock": is_low_stock,
            "created_at": insert.get("created_at").isoformat() if insert.get("created_at") else None
        })
    
    return result


@app.put("/api/inserts/{insert_id}")
async def update_insert(
    insert_id: str,
    insert_update: InsertUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update insert quantity or details"""
    insert = await db.inserts.find_one({"insert_id": insert_id})
    if not insert:
        raise HTTPException(status_code=404, detail="Insert not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc)}
    
    if insert_update.quantity is not None:
        update_data["quantity"] = insert_update.quantity
    if insert_update.min_quantity is not None:
        update_data["min_quantity"] = insert_update.min_quantity
    if insert_update.location is not None:
        update_data["location"] = insert_update.location
    if insert_update.remarks is not None:
        update_data["remarks"] = insert_update.remarks
    
    await db.inserts.update_one({"insert_id": insert_id}, {"$set": update_data})
    
    return {"message": "Insert updated successfully"}


@app.delete("/api/inserts/{insert_id}")
async def delete_insert(
    insert_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete an insert from inventory"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Only Admin can delete inserts")
    
    result = await db.inserts.delete_one({"insert_id": insert_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Insert not found")
    
    return {"message": "Insert deleted successfully"}


# --- TOOL/INSERT ISSUE ENDPOINTS ---

@app.post("/api/inventory/issue")
async def issue_tool_or_insert(
    request: ToolIssueRequest,
    current_user: User = Depends(get_current_user)
):
    """Issue tool or insert to a user"""
    now = datetime.now(timezone.utc)
    
    if request.item_type == "tool":
        item = await db.tools.find_one({"tool_id": request.item_id})
        if not item:
            raise HTTPException(status_code=404, detail="Tool not found")
        
        if item.get("quantity", 0) < request.quantity:
            raise HTTPException(status_code=400, detail="Insufficient stock")
        
        # Reduce stock
        new_qty = item.get("quantity", 0) - request.quantity
        await db.tools.update_one(
            {"tool_id": request.item_id},
            {"$set": {"quantity": new_qty, "updated_at": now}}
        )
        item_name = f"{item.get('tool_type', '')} {item.get('subtype', '')} Ø{item.get('diameter', '')}mm"
        
    elif request.item_type == "insert":
        item = await db.inserts.find_one({"insert_id": request.item_id})
        if not item:
            raise HTTPException(status_code=404, detail="Insert not found")
        
        if item.get("quantity", 0) < request.quantity:
            raise HTTPException(status_code=400, detail="Insufficient stock")
        
        # Reduce stock
        new_qty = item.get("quantity", 0) - request.quantity
        await db.inserts.update_one(
            {"insert_id": request.item_id},
            {"$set": {"quantity": new_qty, "updated_at": now}}
        )
        item_name = f"{item.get('insert_type', '')} {item.get('grade', '')}"
    else:
        raise HTTPException(status_code=400, detail="Invalid item type")
    
    # Get user info
    issued_user = await db.users.find_one({"user_id": request.issued_to})
    issued_to_name = issued_user.get("name", "Unknown") if issued_user else "Unknown"
    
    # Get machine info if provided
    machine_name = None
    if request.machine_id:
        machine = await db.machines.find_one({"machine_id": request.machine_id})
        machine_name = machine.get("name", "") if machine else None
    
    # Create issue record
    issue_id = f"issue_{uuid.uuid4().hex[:12]}"
    issue_doc = {
        "issue_id": issue_id,
        "item_id": request.item_id,
        "item_type": request.item_type,
        "item_name": item_name,
        "quantity": request.quantity,
        "issued_to": request.issued_to,
        "issued_to_name": issued_to_name,
        "issued_by": current_user.user_id,
        "issued_by_name": current_user.name,
        "machine_id": request.machine_id,
        "machine_name": machine_name,
        "purpose": request.purpose,
        "issued_at": now
    }
    
    await db.tool_issues.insert_one(issue_doc)
    
    # Check if low stock and send alert
    if new_qty <= item.get("min_quantity", 0):
        await notify_admins(
            "⚠️ Low Stock Alert / कम स्टॉक चेतावनी",
            f"{item_name} is low on stock ({new_qty} remaining)"
        )
    
    return {"message": "Item issued successfully", "issue_id": issue_id, "remaining_qty": new_qty}


@app.get("/api/inventory/issues")
async def get_issue_history(
    item_type: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get tool/insert issue history"""
    query = {}
    if item_type:
        query["item_type"] = item_type
    
    issues = await db.tool_issues.find(query).sort("issued_at", -1).to_list(length=200)
    
    result = []
    for issue in issues:
        result.append({
            "issue_id": issue["issue_id"],
            "item_id": issue["item_id"],
            "item_type": issue["item_type"],
            "item_name": issue.get("item_name", ""),
            "quantity": issue["quantity"],
            "issued_to": issue["issued_to"],
            "issued_to_name": issue.get("issued_to_name", ""),
            "issued_by_name": issue.get("issued_by_name", ""),
            "machine_name": issue.get("machine_name", ""),
            "purpose": issue.get("purpose", ""),
            "issued_at": issue.get("issued_at").isoformat() if issue.get("issued_at") else None
        })
    
    return result


@app.get("/api/inventory/low-stock")
async def get_low_stock_items(
    current_user: User = Depends(get_current_user)
):
    """Get all low stock tools and inserts"""
    # Low stock tools
    low_tools = await db.tools.find({
        "$expr": {"$lte": ["$quantity", "$min_quantity"]}
    }).to_list(length=100)
    
    # Low stock inserts
    low_inserts = await db.inserts.find({
        "$expr": {"$lte": ["$quantity", "$min_quantity"]}
    }).to_list(length=100)
    
    tools_result = []
    for tool in low_tools:
        tools_result.append({
            "item_id": tool["tool_id"],
            "name": f"{tool.get('tool_type', '')} {tool.get('subtype', '')} Ø{tool.get('diameter', '')}mm",
            "quantity": tool.get("quantity", 0),
            "min_quantity": tool.get("min_quantity", 0),
            "location": tool.get("location", "")
        })
    
    inserts_result = []
    for insert in low_inserts:
        inserts_result.append({
            "item_id": insert["insert_id"],
            "name": f"{insert.get('insert_type', '')} {insert.get('grade', '')}",
            "quantity": insert.get("quantity", 0),
            "min_quantity": insert.get("min_quantity", 0),
            "location": insert.get("location", "")
        })
    
    return {
        "tools": tools_result,
        "inserts": inserts_result,
        "total_low_stock": len(tools_result) + len(inserts_result)
    }


# ============================================
# UNIFIED TOOLS & INSERTS INVENTORY MODULE
# ============================================

# Pydantic Models for unified Tools & Inserts
class ToolInsertCreate(BaseModel):
    category: str  # 'endmill' or 'insert'
    # Common fields
    quantity: int
    min_quantity: int = 20  # Default minimum qty is 20
    location: Optional[str] = None
    remarks: Optional[str] = None  # Remarks/notes
    tool_image: Optional[str] = None  # base64 image
    # End Mill specific fields
    material: Optional[str] = None  # 'Aluminium' or 'SS'
    endmill_type: Optional[str] = None  # 'Flat', 'Tip Radius', 'Ball Nose'
    diameter: Optional[float] = None  # in mm
    length: Optional[float] = None  # in mm
    grade: Optional[str] = None  # Tool grade
    # Insert specific fields
    insert_type: Optional[str] = None  # 'LOMU', 'R6', 'APKT', etc.
    insert_size: Optional[str] = None
    insert_grade: Optional[str] = None
    tip_radius: Optional[str] = None
    insert_image: Optional[str] = None  # base64
    box_image: Optional[str] = None  # base64

class ToolInsertIssue(BaseModel):
    item_id: str
    quantity: int
    issued_to: Optional[str] = None  # user_id or null for manual name
    issued_to_name: Optional[str] = None  # manual name entry
    signature: Optional[str] = None  # base64 signature image


@app.get("/api/tools-inserts")
async def get_tools_inserts(
    low_stock: bool = False,
    current_user: User = Depends(get_current_user)
):
    """Get all tools and inserts from unified inventory"""
    query = {}
    if low_stock:
        query["is_low_stock"] = True
    
    items = await db.tools_inserts.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=1000)
    
    # Recompute is_low_stock based on current data
    result = []
    for item in items:
        qty = item.get("quantity", 0)
        min_qty = item.get("min_quantity", 20)
        item["is_low_stock"] = qty < min_qty
        
        # Only include if low_stock filter matches
        if low_stock and not item["is_low_stock"]:
            continue
        
        result.append(item)
    
    return result


@app.post("/api/tools-inserts")
async def create_tool_insert(
    item: ToolInsertCreate,
    current_user: User = Depends(get_current_user)
):
    """Add new tool or insert to unified inventory"""
    item_id = f"ti_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    # Compute is_low_stock
    is_low_stock = item.quantity < item.min_quantity
    
    item_doc = {
        "item_id": item_id,
        "category": item.category,
        "quantity": item.quantity,
        "min_quantity": item.min_quantity,
        "location": item.location or "",
        "remarks": item.remarks or "",
        "tool_image": item.tool_image,
        "is_low_stock": is_low_stock,
        "created_by_id": current_user.user_id,
        "created_by_name": current_user.name,
        "created_at": now
    }
    
    if item.category == "endmill":
        item_doc.update({
            "material": item.material,
            "endmill_type": item.endmill_type,
            "diameter": item.diameter,
            "length": item.length,
            "grade": item.grade
        })
    else:  # insert
        item_doc.update({
            "insert_type": item.insert_type,
            "insert_size": item.insert_size,
            "insert_grade": item.insert_grade,
            "tip_radius": item.tip_radius,
            "insert_image": item.insert_image,
            "box_image": item.box_image
        })
    
    await db.tools_inserts.insert_one(item_doc)
    
    # Build response without MongoDB _id
    response = {
        "message": "Item added successfully",
        "item_id": item_id,
        "category": item.category,
        "quantity": item.quantity,
        "min_quantity": item.min_quantity,
        "location": item.location or "",
        "remarks": item.remarks or "",
        "is_low_stock": is_low_stock,
        "created_at": now.isoformat()
    }
    
    if item.category == "endmill":
        response.update({
            "material": item.material,
            "endmill_type": item.endmill_type,
            "diameter": item.diameter,
            "length": item.length,
            "grade": item.grade
        })
    else:
        response.update({
            "insert_type": item.insert_type,
            "insert_size": item.insert_size,
            "insert_grade": item.insert_grade,
            "tip_radius": item.tip_radius
        })
    
    return response


class ToolInsertAddQty(BaseModel):
    quantity: int


@app.post("/api/tools-inserts/{item_id}/add-qty")
async def add_quantity_to_tool_insert(
    item_id: str,
    add_data: ToolInsertAddQty,
    current_user: User = Depends(get_current_user)
):
    """Add quantity to existing tool or insert"""
    item = await db.tools_inserts.find_one({"item_id": item_id})
    
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    current_qty = item.get("quantity", 0)
    new_qty = current_qty + add_data.quantity
    min_qty = item.get("min_quantity", 20)
    is_low_stock = new_qty < min_qty
    
    # Update quantity
    await db.tools_inserts.update_one(
        {"item_id": item_id},
        {"$set": {
            "quantity": new_qty,
            "is_low_stock": is_low_stock,
            "last_added_qty": add_data.quantity,
            "last_added_at": datetime.now(timezone.utc),
            "last_added_by": current_user.user_id
        }}
    )
    
    return {
        "message": "Quantity added successfully",
        "new_quantity": new_qty,
        "is_low_stock": is_low_stock
    }


@app.post("/api/tools-inserts/issue")
async def issue_tool_insert(
    issue: ToolInsertIssue,
    current_user: User = Depends(get_current_user)
):
    """Issue tool or insert from unified inventory"""
    item = await db.tools_inserts.find_one({"item_id": issue.item_id})
    
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # Determine issued_to name
    issued_to_name = issue.issued_to_name
    if issue.issued_to and not issued_to_name:
        # Look up user name
        user = await db.users.find_one({"user_id": issue.issued_to})
        if user:
            issued_to_name = user.get("name", issue.issued_to)
    
    if not issued_to_name:
        raise HTTPException(status_code=400, detail="Recipient name is required")
    
    current_qty = item.get("quantity", 0)
    if current_qty < issue.quantity:
        raise HTTPException(status_code=400, detail=f"Insufficient stock. Available: {current_qty}")
    
    new_qty = current_qty - issue.quantity
    min_qty = item.get("min_quantity", 20)
    is_low_stock = new_qty < min_qty
    
    # Update quantity
    await db.tools_inserts.update_one(
        {"item_id": issue.item_id},
        {"$set": {
            "quantity": new_qty,
            "is_low_stock": is_low_stock,
            "last_issued_at": datetime.now(timezone.utc),
            "last_issued_by": current_user.user_id,
            "last_issued_to": issue.issued_to or issued_to_name,
            "last_issued_to_name": issued_to_name
        }}
    )
    
    # Log the issue
    issue_log = {
        "issue_id": f"issue_{uuid.uuid4().hex[:12]}",
        "item_id": issue.item_id,
        "quantity": issue.quantity,
        "issued_to": issue.issued_to,
        "issued_to_name": issued_to_name,
        "issued_by_id": current_user.user_id,
        "issued_by_name": current_user.name,
        "issued_at": datetime.now(timezone.utc),
        "signature": issue.signature
    }
    await db.tools_inserts_issues.insert_one(issue_log)
    
    return {
        "message": "Item issued successfully",
        "remaining_qty": new_qty,
        "is_low_stock": is_low_stock
    }


class ToolInsertScrap(BaseModel):
    item_id: str
    quantity: int
    returned_by: Optional[str] = None  # user_id or null
    returned_by_name: Optional[str] = None  # manual name entry
    signature: Optional[str] = None  # base64 signature image


@app.post("/api/tools-inserts/scrap")
async def collect_scrap_tool_insert(
    scrap: ToolInsertScrap,
    current_user: User = Depends(get_current_user)
):
    """Collect scrap/broken tools or used inserts from users"""
    item = await db.tools_inserts.find_one({"item_id": scrap.item_id})
    
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    current_scrap = item.get("scrap_qty", 0)
    new_scrap = current_scrap + scrap.quantity
    
    # Update scrap quantity
    await db.tools_inserts.update_one(
        {"item_id": scrap.item_id},
        {"$set": {
            "scrap_qty": new_scrap,
            "last_scrap_at": datetime.now(timezone.utc),
            "last_scrap_by": current_user.user_id,
            "last_scrap_from": scrap.returned_by
        }}
    )
    
    # Get returned_by user name
    returned_by_user = await db.users.find_one({"user_id": scrap.returned_by})
    returned_by_name = returned_by_user.get("name", "Unknown") if returned_by_user else "Unknown"
    
    # Log the scrap collection
    scrap_log = {
        "scrap_id": f"scrap_{uuid.uuid4().hex[:12]}",
        "item_id": scrap.item_id,
        "quantity": scrap.quantity,
        "returned_by": scrap.returned_by,
        "returned_by_name": returned_by_name,
        "collected_by_id": current_user.user_id,
        "collected_by_name": current_user.name,
        "collected_at": datetime.now(timezone.utc),
        "signature": scrap.signature
    }
    await db.tools_inserts_scrap.insert_one(scrap_log)
    
    return {
        "message": "Scrap collected successfully",
        "scrap_qty": new_scrap
    }


@app.get("/api/scrap-store")
async def get_scrap_store(
    current_user: User = Depends(get_current_user)
):
    """Get all scrap collection records with item details"""
    scrap_records = await db.tools_inserts_scrap.find(
        {},
        {"_id": 0}
    ).sort("collected_at", -1).to_list(length=500)
    
    # Enrich with item details and calculate totals
    enriched = []
    total_tools = 0
    total_inserts = 0
    
    for record in scrap_records:
        item_id = record.get("item_id")
        item = await db.tools_inserts.find_one({"item_id": item_id}, {"_id": 0})
        
        if item:
            record["item_details"] = {
                "category": item.get("category"),
                "material": item.get("material"),
                "diameter": item.get("diameter"),
                "insert_type": item.get("insert_type")
            }
            
            if item.get("category") == "endmill":
                total_tools += record.get("quantity", 0)
            else:
                total_inserts += record.get("quantity", 0)
        
        enriched.append(record)
    
    return {
        "records": enriched,
        "totals": {
            "tools": total_tools,
            "inserts": total_inserts
        }
    }


@app.get("/api/tools-daily-summary")
async def get_tools_daily_summary(
    date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get daily summary of tools activity - added, issued, and scrapped. Admin only."""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # If no date provided, use today
    if date:
        target_date = date
    else:
        target_date = get_ist_now().strftime("%Y-%m-%d")
    
    # Parse date range for the day
    start_of_day = datetime.fromisoformat(f"{target_date}T00:00:00+00:00")
    end_of_day = datetime.fromisoformat(f"{target_date}T23:59:59+00:00")
    
    date_query = {"$gte": start_of_day, "$lte": end_of_day}
    
    # Get tools ISSUED today (from tools_inserts_issues collection)
    issued_today = await db.tools_inserts_issues.find(
        {"issued_at": date_query},
        {"_id": 0}
    ).sort("issued_at", -1).to_list(length=100)
    
    # Enrich issued with item details
    for issue in issued_today:
        item = await db.tools_inserts.find_one({"item_id": issue.get("item_id")}, {"_id": 0, "category": 1, "material": 1, "diameter": 1, "insert_type": 1})
        if item:
            issue["item_details"] = item
        # Get issued_to user name
        issued_to_user = await db.users.find_one({"user_id": issue.get("issued_to")})
        issue["issued_to_name"] = issued_to_user.get("name") if issued_to_user else issue.get("issued_to")
    
    # Get SCRAP collected/submitted today (from tools_inserts_scrap collection)
    scrap_today = await db.tools_inserts_scrap.find(
        {"collected_at": date_query},
        {"_id": 0}
    ).sort("collected_at", -1).to_list(length=100)
    
    # Enrich scrap with item details
    for scrap in scrap_today:
        item = await db.tools_inserts.find_one({"item_id": scrap.get("item_id")}, {"_id": 0, "category": 1, "material": 1, "diameter": 1, "insert_type": 1})
        if item:
            scrap["item_details"] = item
    
    # Get tools ADDED today (new items created today)
    new_tools_today = await db.tools_inserts.find(
        {"created_at": date_query},
        {"_id": 0, "item_id": 1, "category": 1, "material": 1, "diameter": 1, 
         "insert_type": 1, "quantity": 1, "created_at": 1, "created_by_name": 1}
    ).sort("created_at", -1).to_list(length=100)
    
    # Get quantity additions today (stock added to existing items)
    qty_added_today = await db.tools_qty_logs.find(
        {"added_at": date_query},
        {"_id": 0}
    ).sort("added_at", -1).to_list(length=100)
    
    # Calculate totals
    total_issued = sum(item.get("quantity", 0) for item in issued_today)
    total_scrapped = sum(item.get("quantity", 0) for item in scrap_today)
    total_new_items = len(new_tools_today)
    total_qty_added = sum(item.get("quantity", 0) for item in qty_added_today)
    
    return {
        "date": target_date,
        "summary": {
            "total_issued": total_issued,
            "total_scrapped": total_scrapped,
            "new_items_added": total_new_items,
            "qty_added_to_existing": total_qty_added,
            "issue_count": len(issued_today),
            "scrap_count": len(scrap_today)
        },
        "issued": issued_today,
        "scrapped": scrap_today,
        "new_items": new_tools_today,
        "qty_additions": qty_added_today
    }


@app.put("/api/tools-inserts/{item_id}")
async def update_tool_insert(
    item_id: str,
    update_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Update tool or insert in unified inventory"""
    item = await db.tools_inserts.find_one({"item_id": item_id})
    
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # Allowed fields to update
    allowed_fields = ["quantity", "min_quantity", "location", "material", "endmill_type", 
                      "diameter", "length", "grade", "insert_type", "insert_size", "insert_grade",
                      "tip_radius", "insert_image", "box_image", "tool_image", "remarks", 
                      "insert_special_details"]
    
    update_dict = {k: v for k, v in update_data.items() if k in allowed_fields}
    
    # Recompute is_low_stock
    new_qty = update_dict.get("quantity", item.get("quantity", 0))
    new_min = update_dict.get("min_quantity", item.get("min_quantity", 20))
    update_dict["is_low_stock"] = new_qty < new_min
    update_dict["updated_at"] = datetime.now(timezone.utc)
    
    await db.tools_inserts.update_one({"item_id": item_id}, {"$set": update_dict})
    
    updated_item = await db.tools_inserts.find_one({"item_id": item_id}, {"_id": 0})
    return updated_item


@app.delete("/api/tools-inserts/{item_id}")
async def delete_tool_insert(
    item_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete tool or insert from unified inventory (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Only Admin can delete items")
    
    result = await db.tools_inserts.delete_one({"item_id": item_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    
    return {"message": "Item deleted successfully"}


@app.get("/api/tools-inserts/low-stock-count")
async def get_low_stock_count(
    current_user: User = Depends(get_current_user)
):
    """Get count of low stock tools and inserts for dashboard"""
    count = await db.tools_inserts.count_documents({"is_low_stock": True})
    return {"low_stock_count": count}


@app.get("/api/tools-inserts/{item_id}/history")
async def get_tool_transaction_history(
    item_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get transaction history for a specific tool/insert item"""
    # Get issues from tools_inserts_issues collection
    issues = await db.tools_inserts_issues.find({"item_id": item_id}).sort("issued_at", -1).to_list(100)
    
    # Get scrap returns from tools_inserts_scrap collection
    scraps = await db.tools_inserts_scrap.find({"item_id": item_id}).sort("returned_at", -1).to_list(100)
    
    # Format and combine transactions
    transactions = []
    
    for issue in issues:
        timestamp = issue.get("issued_at")
        if timestamp:
            try:
                if isinstance(timestamp, str):
                    timestamp_dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                else:
                    timestamp_dt = timestamp
                timestamp_ist = to_ist(timestamp_dt).strftime('%d/%m/%Y %I:%M %p')
            except:
                timestamp_ist = str(timestamp)
        else:
            timestamp_ist = "N/A"
            
        transactions.append({
            "type": "issue",
            "timestamp": str(issue.get("issued_at")),
            "timestamp_ist": timestamp_ist,
            "quantity": issue.get("quantity", 0),
            "issued_to_name": issue.get("issued_to_name"),
            "issued_by_name": issue.get("issued_by_name"),
            "signature": issue.get("signature"),
            "item_details": issue.get("item_details")
        })
    
    for scrap in scraps:
        timestamp = scrap.get("returned_at")
        if timestamp:
            try:
                if isinstance(timestamp, str):
                    timestamp_dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                else:
                    timestamp_dt = timestamp
                timestamp_ist = to_ist(timestamp_dt).strftime('%d/%m/%Y %I:%M %p')
            except:
                timestamp_ist = str(timestamp)
        else:
            timestamp_ist = "N/A"
            
        transactions.append({
            "type": "scrap",
            "timestamp": str(scrap.get("returned_at")),
            "timestamp_ist": timestamp_ist,
            "quantity": scrap.get("quantity", 0),
            "returned_by_name": scrap.get("returned_by_name"),
            "collected_by_name": scrap.get("collected_by_name"),
            "signature": scrap.get("signature"),
            "item_details": scrap.get("item_details")
        })
    
    # Sort by timestamp descending
    transactions.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
    
    return {"transactions": transactions}


# ========================================
# PART LIBRARY MODULE
# ========================================

class PartLocation(BaseModel):
    rack: str
    qty: int

class PartCreate(BaseModel):
    part_id: Optional[str] = None  # If None, auto-generate
    auto_generate_id: bool = True
    part_no: Optional[str] = None  # User-defined part number (PART NO - 01, etc.)
    name: str
    material: Optional[str] = None  # SS 316, Nitrile, etc.
    category: str
    part_type: Optional[str] = None  # aluminium, bearing, bolt, etc.
    size: Optional[str] = None  # Dia/ID/WxB
    length: Optional[str] = None  # Length/Thickness
    variant: Optional[str] = None
    unit: str = "PCS"
    stock: int = 0
    reserved_stock: int = 0
    min_stock: int = 0
    locations: List[PartLocation] = []
    image: Optional[str] = None  # Base64 image
    drawing: Optional[str] = None  # Base64 drawing
    remarks: Optional[str] = None

class PartUpdate(BaseModel):
    part_no: Optional[str] = None
    name: Optional[str] = None
    material: Optional[str] = None
    category: Optional[str] = None
    part_type: Optional[str] = None
    size: Optional[str] = None
    length: Optional[str] = None
    variant: Optional[str] = None
    unit: Optional[str] = None
    stock: Optional[int] = None
    reserved_stock: Optional[int] = None
    min_stock: Optional[int] = None
    locations: Optional[List[PartLocation]] = None
    image: Optional[str] = None
    drawing: Optional[str] = None
    remarks: Optional[str] = None


def generate_part_id(category: str, size: str, variant: str, serial: int) -> str:
    """Generate part ID in format: CATEGORY-SIZE-VARIANT-SERIAL"""
    cat = (category or "GEN")[:3].upper()
    sz = (size or "STD")[:4].upper().replace(" ", "")
    var = (variant or "A1")[:2].upper()
    return f"{cat}-{sz}-{var}-{serial:04d}"


@app.get("/api/parts-library")
async def get_parts_library(
    search: Optional[str] = None,
    category: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all parts with optional search and filter"""
    query = {}
    
    if search:
        query["$or"] = [
            {"part_id": {"$regex": search, "$options": "i"}},
            {"name": {"$regex": search, "$options": "i"}},
            {"category": {"$regex": search, "$options": "i"}}
        ]
    
    if category:
        query["category"] = {"$regex": category, "$options": "i"}
    
    parts = await db.parts_library.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=500)
    
    # Calculate available stock for each part
    for part in parts:
        part["available_stock"] = part.get("stock", 0) - part.get("reserved_stock", 0)
        # Determine status
        if part["available_stock"] <= 0:
            part["status"] = "shortage"
        elif part["available_stock"] <= part.get("min_stock", 0):
            part["status"] = "low"
        else:
            part["status"] = "ok"
    
    return parts


@app.get("/api/parts-library/categories")
async def get_part_categories(current_user: User = Depends(get_current_user)):
    """Get unique categories for filter dropdown"""
    categories = await db.parts_library.distinct("category")
    return [c for c in categories if c]


@app.get("/api/parts-library/types")
async def get_part_types(current_user: User = Depends(get_current_user)):
    """Get unique part types for dropdown with default options"""
    default_types = [
        "aluminium", "bearing", "bolt", "screw", "shaft", 
        "pin", "roller", "plastic", "dowel", "ss", "oring", "seal"
    ]
    
    # Get existing types from database
    db_types = await db.parts_library.distinct("part_type")
    db_types = [t for t in db_types if t and t.strip()]
    
    # Merge and dedupe
    all_types = list(set(default_types + db_types))
    all_types.sort()
    
    return all_types


@app.get("/api/parts-library/next-serial")
async def get_next_serial(
    category: str,
    size: Optional[str] = None,
    variant: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get next serial number for auto-generated part ID"""
    cat = (category or "GEN")[:3].upper()
    sz = (size or "STD")[:4].upper().replace(" ", "")
    var = (variant or "A1")[:2].upper()
    prefix = f"{cat}-{sz}-{var}-"
    
    # Find highest serial with this prefix
    existing = await db.parts_library.find(
        {"part_id": {"$regex": f"^{prefix}"}},
        {"part_id": 1}
    ).sort("part_id", -1).limit(1).to_list(length=1)
    
    if existing:
        try:
            last_serial = int(existing[0]["part_id"].split("-")[-1])
            next_serial = last_serial + 1
        except:
            next_serial = 1
    else:
        next_serial = 1
    
    return {
        "next_serial": next_serial,
        "preview_id": generate_part_id(category, size, variant, next_serial)
    }


@app.post("/api/parts-library")
async def create_part(part: PartCreate, current_user: User = Depends(get_current_user)):
    """Create a new part"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    now = get_ist_now()
    
    # Generate or use provided part_id
    if part.auto_generate_id or not part.part_id:
        # Get next serial
        cat = (part.category or "GEN")[:3].upper()
        sz = (part.size or "STD")[:4].upper().replace(" ", "")
        var = (part.variant or "A1")[:2].upper()
        prefix = f"{cat}-{sz}-{var}-"
        
        existing = await db.parts_library.find(
            {"part_id": {"$regex": f"^{prefix}"}},
            {"part_id": 1}
        ).sort("part_id", -1).limit(1).to_list(length=1)
        
        if existing:
            try:
                last_serial = int(existing[0]["part_id"].split("-")[-1])
                next_serial = last_serial + 1
            except:
                next_serial = 1
        else:
            next_serial = 1
        
        part_id = generate_part_id(part.category, part.size, part.variant, next_serial)
    else:
        part_id = part.part_id.upper().strip()
        # Check if ID already exists
        existing = await db.parts_library.find_one({"part_id": part_id})
        if existing:
            raise HTTPException(status_code=400, detail="Part ID already exists")
    
    # Calculate stock from locations if provided
    location_stock = sum(loc.qty for loc in part.locations) if part.locations else 0
    total_stock = max(part.stock, location_stock)
    
    part_doc = {
        "part_id": part_id,
        "part_no": part.part_no or "",
        "name": part.name,
        "material": part.material or "",
        "category": part.category,
        "part_type": part.part_type or "",
        "size": part.size or "",
        "length": part.length or "",
        "variant": part.variant or "",
        "unit": part.unit,
        "stock": total_stock,
        "reserved_stock": part.reserved_stock,
        "min_stock": part.min_stock,
        "locations": [{"rack": loc.rack, "qty": loc.qty} for loc in part.locations],
        "image": part.image or "",
        "drawing": part.drawing or "",
        "remarks": part.remarks or "",
        "created_at": now,
        "updated_at": now,
        "created_by": current_user.user_id
    }
    
    await db.parts_library.insert_one(part_doc)
    
    return {"message": "Part created successfully", "part_id": part_id}



# Public endpoint for QR code scanning - no auth required
@app.get("/api/public/part/{part_id}")
async def get_part_public(part_id: str):
    """
    Public endpoint for QR code scanning.
    Returns part details without authentication.
    Used when scanning QR codes with external scanner apps.
    """
    part = await db.parts_library.find_one({"part_id": part_id}, {"_id": 0})
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    
    part["available_stock"] = part.get("stock", 0) - part.get("reserved_stock", 0)
    
    if part["available_stock"] <= 0:
        part["status"] = "shortage"
    elif part["available_stock"] <= part.get("min_stock", 0):
        part["status"] = "low"
    else:
        part["status"] = "ok"
    
    # Get assemblies where this part is used
    assembly_usages = await db.assembly_parts.find(
        {"part_id": part_id}, {"_id": 0}
    ).to_list(length=100)
    
    used_in_assemblies = []
    for usage in assembly_usages:
        asm = await db.assemblies.find_one(
            {"assembly_id": usage["assembly_id"]}, 
            {"_id": 0, "assembly_id": 1, "name": 1}
        )
        if asm:
            used_in_assemblies.append({
                "assembly_id": asm["assembly_id"],
                "name": asm.get("name", ""),
                "quantity": usage.get("quantity", 0)
            })
    
    part["used_in_assemblies"] = used_in_assemblies
    
    return part


# Public HTML page for QR code scanning
@app.get("/part/{part_id}", response_class=HTMLResponse)
async def get_part_html_page(part_id: str):
    """
    Public HTML page that displays part details.
    This is what opens when someone scans a QR code with any scanner app.
    """
    part = await db.parts_library.find_one({"part_id": part_id}, {"_id": 0})
    
    if not part:
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>Part Not Found</title>
            <style>
                body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #0F172A; color: white; padding: 20px; margin: 0; }}
                .container {{ max-width: 500px; margin: 0 auto; text-align: center; padding-top: 50px; }}
                h1 {{ color: #EF4444; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Part Not Found</h1>
                <p>Part ID: {part_id}</p>
            </div>
        </body>
        </html>
        """, status_code=404)
    
    # Calculate stock status
    available = part.get("stock", 0) - part.get("reserved_stock", 0)
    if available <= 0:
        status = "SHORTAGE"
        status_color = "#EF4444"
    elif available <= part.get("min_stock", 0):
        status = "LOW STOCK"
        status_color = "#F59E0B"
    else:
        status = "IN STOCK"
        status_color = "#10B981"
    
    # Get assembly usage
    assembly_usages = await db.assembly_parts.find({"part_id": part_id}, {"_id": 0}).to_list(length=50)
    assemblies_html = ""
    for usage in assembly_usages:
        asm = await db.assemblies.find_one({"assembly_id": usage["assembly_id"]}, {"_id": 0, "name": 1})
        if asm:
            assemblies_html += f'<div class="assembly">{asm.get("name", usage["assembly_id"])} (Qty: {usage.get("quantity", 0)})</div>'
    
    if not assemblies_html:
        assemblies_html = '<div class="assembly">Not used in any assembly</div>'
    
    # Build image HTML
    image_html = ""
    if part.get("image"):
        img_src = part["image"] if part["image"].startswith("http") else f"data:image/jpeg;base64,{part['image']}"
        image_html = f'<img src="{img_src}" class="part-image" alt="Part Image">'
    
    # Build drawing HTML
    drawing_html = ""
    if part.get("drawing"):
        drawing_html = f'<a href="{part["drawing"]}" target="_blank" class="drawing-link">📄 View Drawing (PDF)</a>'
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>{part.get('name', part_id)} - Part Details</title>
        <style>
            * {{ box-sizing: border-box; }}
            body {{ 
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; 
                background: #0F172A; 
                color: white; 
                padding: 16px; 
                margin: 0;
                min-height: 100vh;
            }}
            .container {{ max-width: 500px; margin: 0 auto; }}
            .header {{ 
                background: #1E293B; 
                border-radius: 12px; 
                padding: 16px; 
                margin-bottom: 16px;
                text-align: center;
            }}
            .part-id {{ 
                font-size: 14px; 
                color: #64748B; 
                margin-bottom: 8px;
                font-family: monospace;
            }}
            .part-name {{ 
                font-size: 24px; 
                font-weight: bold; 
                margin-bottom: 8px;
            }}
            .part-no {{ 
                font-size: 14px; 
                color: #94A3B8; 
                margin-bottom: 12px;
            }}
            .status {{ 
                display: inline-block;
                background: {status_color}; 
                color: white; 
                padding: 6px 16px; 
                border-radius: 20px; 
                font-size: 12px;
                font-weight: bold;
            }}
            .card {{ 
                background: #1E293B; 
                border-radius: 12px; 
                padding: 16px; 
                margin-bottom: 16px;
            }}
            .card-title {{ 
                font-size: 12px; 
                color: #64748B; 
                text-transform: uppercase;
                margin-bottom: 12px;
                font-weight: 600;
            }}
            .info-row {{ 
                display: flex; 
                justify-content: space-between; 
                padding: 8px 0; 
                border-bottom: 1px solid #334155;
            }}
            .info-row:last-child {{ border-bottom: none; }}
            .info-label {{ color: #94A3B8; font-size: 14px; }}
            .info-value {{ color: white; font-size: 14px; font-weight: 500; }}
            .tags {{ display: flex; flex-wrap: wrap; gap: 8px; margin-top: 8px; }}
            .tag {{ 
                background: #334155; 
                padding: 4px 10px; 
                border-radius: 6px; 
                font-size: 12px;
            }}
            .tag.material {{ background: #10B981; }}
            .tag.category {{ background: #3B82F6; }}
            .assembly {{ 
                background: #0F172A; 
                padding: 10px; 
                border-radius: 8px; 
                margin-bottom: 8px;
                font-size: 14px;
            }}
            .part-image {{ 
                width: 100%; 
                max-height: 200px; 
                object-fit: contain; 
                border-radius: 8px;
                margin-bottom: 12px;
            }}
            .drawing-link {{
                display: block;
                background: #3B82F6;
                color: white;
                text-decoration: none;
                padding: 12px;
                border-radius: 8px;
                text-align: center;
                font-weight: 500;
            }}
            .stock-grid {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 12px; }}
            .stock-item {{ text-align: center; }}
            .stock-value {{ font-size: 24px; font-weight: bold; }}
            .stock-label {{ font-size: 11px; color: #64748B; text-transform: uppercase; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <div class="part-id">{part_id}</div>
                <div class="part-name">{part.get('name', 'Unknown Part')}</div>
                {f'<div class="part-no">Part No: {part.get("part_no")}</div>' if part.get('part_no') else ''}
                <div class="status">{status}</div>
            </div>
            
            {f'<div class="card">{image_html}</div>' if image_html else ''}
            
            <div class="card">
                <div class="card-title">Details</div>
                <div class="tags">
                    {f'<span class="tag material">{part.get("material")}</span>' if part.get('material') else ''}
                    {f'<span class="tag category">{part.get("category")}</span>' if part.get('category') else ''}
                    {f'<span class="tag">{part.get("part_type", "").upper()}</span>' if part.get('part_type') else ''}
                </div>
                <div style="margin-top: 16px;">
                    {f'<div class="info-row"><span class="info-label">Size (Dia/ID)</span><span class="info-value">{part.get("size")}</span></div>' if part.get('size') else ''}
                    {f'<div class="info-row"><span class="info-label">Length</span><span class="info-value">{part.get("length")}</span></div>' if part.get('length') else ''}
                    <div class="info-row"><span class="info-label">Unit</span><span class="info-value">{part.get('unit', 'PCS')}</span></div>
                    {f'<div class="info-row"><span class="info-label">Remarks</span><span class="info-value">{part.get("remarks")}</span></div>' if part.get('remarks') else ''}
                </div>
            </div>
            
            <div class="card">
                <div class="card-title">Stock</div>
                <div class="stock-grid">
                    <div class="stock-item">
                        <div class="stock-value">{part.get('stock', 0)}</div>
                        <div class="stock-label">Total</div>
                    </div>
                    <div class="stock-item">
                        <div class="stock-value">{part.get('reserved_stock', 0)}</div>
                        <div class="stock-label">Reserved</div>
                    </div>
                    <div class="stock-item">
                        <div class="stock-value" style="color: {status_color}">{available}</div>
                        <div class="stock-label">Available</div>
                    </div>
                </div>
            </div>
            
            <div class="card">
                <div class="card-title">Used In Assemblies</div>
                {assemblies_html}
            </div>
            
            {f'<div class="card">{drawing_html}</div>' if drawing_html else ''}
        </div>
    </body>
    </html>
    """
    
    return HTMLResponse(content=html_content)



# ── Literal routes BEFORE the /{part_id} catch-all to avoid shadowing ──

@app.get("/api/parts-library/usage-map")
async def get_parts_usage_map(current_user: User = Depends(get_current_user)):
    """Return a compact map: { part_id: [{assembly_id, name}, ...] } for ALL parts.
    Used by the Parts Library + Assemblies BOM page to show 'linked to N
    assemblies' badges. One DB roundtrip instead of N — required for fast UI.
    """
    rows = await db.assembly_parts.find(
        {}, {"_id": 0, "part_id": 1, "assembly_id": 1}
    ).to_list(length=20000)
    assemblies = await db.assemblies.find(
        {}, {"_id": 0, "assembly_id": 1, "name": 1}
    ).to_list(length=2000)
    asm_lookup = {a["assembly_id"]: a.get("name", a["assembly_id"]) for a in assemblies}

    usage: dict = {}
    for r in rows:
        pid = r.get("part_id")
        aid = r.get("assembly_id")
        if not pid or not aid:
            continue
        usage.setdefault(pid, []).append({
            "assembly_id": aid,
            "name": asm_lookup.get(aid, aid),
        })
    return {"usage": usage}


@app.get("/api/assemblies/{assembly_id}/order-list")
async def get_assembly_order_list(
    assembly_id: str,
    quantity: int = 1,
    cutting_allowance_mm: float = 0,
    current_user: User = Depends(get_current_user),
):
    """Build a purchase/cutting order list for producing N units of an assembly.

    Output sections:
      - parts:        every BOM line × quantity, with current stock & shortage
      - orings:       grouped by (size + material) so purchasing sees totals
      - raw_material: grouped by (dia + material), total length in mm
    """
    if quantity < 1:
        quantity = 1

    assembly = await db.assemblies.find_one({"assembly_id": assembly_id}, {"_id": 0})
    if not assembly:
        raise HTTPException(status_code=404, detail="Assembly not found")

    bom = await db.assembly_parts.find(
        {"assembly_id": assembly_id}, {"_id": 0}
    ).to_list(length=500)

    parts_out: list = []
    oring_groups: dict = {}
    rawmat_groups: dict = {}

    for ap in bom:
        part = await db.parts_library.find_one({"part_id": ap["part_id"]}, {"_id": 0})
        if not part:
            continue
        per_asm = int(ap.get("quantity", 1) or 1)
        total_qty = per_asm * quantity
        stock = int(part.get("stock", 0) or 0)
        short = max(0, total_qty - stock)
        cat = (part.get("category") or "").strip().lower()
        ptype = (part.get("part_type") or "").strip().lower()
        material = part.get("material") or part.get("part_type") or ""
        size = part.get("size") or ""
        length = part.get("length") or ""

        line = {
            "part_id": part.get("part_id"),
            "name": part.get("name", ""),
            "category": part.get("category", ""),
            "material": material,
            "size": size,
            "length": length,
            "per_assembly": per_asm,
            "total_qty": total_qty,
            "in_stock": stock,
            "short": short,
            "unit": part.get("unit", "PCS"),
        }
        parts_out.append(line)

        # O-Ring grouping
        if "o-ring" in cat or "oring" in cat or "oring" in ptype or "o-ring" in ptype:
            key = f"{size} | {material}"
            g = oring_groups.setdefault(key, {
                "size": size, "material": material,
                "total_qty": 0, "in_stock": 0, "short": 0,
                "parts": [],
            })
            g["total_qty"] += total_qty
            g["in_stock"] += stock
            g["short"] = max(0, g["total_qty"] - g["in_stock"])
            g["parts"].append({"part_id": line["part_id"], "name": line["name"], "qty": total_qty})
            continue

        # Raw Material grouping (anything with a numeric length = cut-stock)
        try:
            length_mm = float(str(length).strip())
        except (TypeError, ValueError):
            length_mm = 0
        if length_mm > 0:
            dia = size
            try:
                import re as _re
                m = _re.search(r"[-+]?\d*\.?\d+", str(size))
                if m:
                    dia = m.group(0)
            except Exception:
                pass
            key = f"Ø{dia} | {material}"
            g = rawmat_groups.setdefault(key, {
                "dia": dia, "material": material,
                "unit_length_mm": length_mm,
                "pieces": 0,
                "raw_length_mm": 0,
                "cut_loss_mm": 0,
                "total_length_mm": 0,
                "parts": [],
            })
            pieces = total_qty
            raw = length_mm * pieces
            loss = float(cutting_allowance_mm or 0) * pieces
            g["pieces"] += pieces
            g["raw_length_mm"] += raw
            g["cut_loss_mm"] += loss
            g["total_length_mm"] = g["raw_length_mm"] + g["cut_loss_mm"]
            g["parts"].append({
                "part_id": line["part_id"],
                "name": line["name"],
                "pieces": pieces,
                "length_mm": length_mm,
            })

    return {
        "assembly_id": assembly_id,
        "assembly_name": assembly.get("name", ""),
        "quantity": quantity,
        "cutting_allowance_mm": cutting_allowance_mm,
        "parts": parts_out,
        "orings": list(oring_groups.values()),
        "raw_material": list(rawmat_groups.values()),
        "totals": {
            "total_parts": len(parts_out),
            "total_oring_groups": len(oring_groups),
            "total_rawmat_groups": len(rawmat_groups),
            "total_shortage_items": sum(1 for p in parts_out if p["short"] > 0),
        },
    }




@app.get("/api/parts-library/{part_id}")
async def get_part_detail(part_id: str, current_user: User = Depends(get_current_user)):
    """Get single part details"""
    part = await db.parts_library.find_one({"part_id": part_id}, {"_id": 0})
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    
    part["available_stock"] = part.get("stock", 0) - part.get("reserved_stock", 0)
    
    if part["available_stock"] <= 0:
        part["status"] = "shortage"
    elif part["available_stock"] <= part.get("min_stock", 0):
        part["status"] = "low"
    else:
        part["status"] = "ok"
    
    # Get assemblies where this part is used
    assembly_usages = await db.assembly_parts.find(
        {"part_id": part_id}, {"_id": 0}
    ).to_list(length=100)
    
    used_in_assemblies = []
    for usage in assembly_usages:
        asm = await db.assemblies.find_one(
            {"assembly_id": usage["assembly_id"]}, 
            {"_id": 0, "assembly_id": 1, "name": 1}
        )
        if asm:
            used_in_assemblies.append({
                "assembly_id": asm["assembly_id"],
                "name": asm.get("name", ""),
                "quantity": usage.get("quantity", 0),
                "assembly_part_no": usage.get("assembly_part_no", "")
            })
    
    part["used_in_assemblies"] = used_in_assemblies
    
    # Also get parts with same Common Part ID (cross-assembly linkage)
    common_parts = []
    if part.get("common_part_id"):
        same_common_parts = await db.parts_library.find(
            {"common_part_id": part["common_part_id"], "part_id": {"$ne": part_id}},
            {"_id": 0, "part_id": 1, "name": 1, "source_assembly_name": 1, "material": 1}
        ).to_list(length=50)
        common_parts = same_common_parts
    
    part["linked_common_parts"] = common_parts
    part["total_linked_assemblies"] = len(used_in_assemblies) + len(common_parts)
    
    return part


@app.get("/api/parts-library/where-used/{common_part_id}")
async def get_where_used(common_part_id: str, current_user: User = Depends(get_current_user)):
    """Get all assemblies where a common part is used"""
    # Find all parts with this common_part_id
    parts = await db.parts_library.find(
        {"common_part_id": common_part_id},
        {"_id": 0, "part_id": 1, "name": 1, "material": 1, "source_assembly_name": 1, "source_assembly": 1}
    ).to_list(length=100)
    
    if not parts:
        raise HTTPException(status_code=404, detail="No parts found with this Common Part ID")
    
    # Get assembly details for each part
    where_used = []
    total_qty_needed = 0
    
    for part in parts:
        # Get assembly_parts entries for this part
        usages = await db.assembly_parts.find(
            {"part_id": part["part_id"]},
            {"_id": 0}
        ).to_list(length=50)
        
        for usage in usages:
            asm = await db.assemblies.find_one(
                {"assembly_id": usage["assembly_id"]},
                {"_id": 0, "assembly_id": 1, "name": 1}
            )
            if asm:
                qty = usage.get("quantity", 0)
                total_qty_needed += qty
                where_used.append({
                    "assembly_id": asm["assembly_id"],
                    "assembly_name": asm.get("name", part.get("source_assembly_name", "")),
                    "part_id": part["part_id"],
                    "bom_position": usage.get("assembly_part_no", ""),
                    "quantity": qty
                })
    
    # Get first part for reference info
    ref_part = parts[0] if parts else {}
    
    return {
        "common_part_id": common_part_id,
        "part_name": ref_part.get("name", ""),
        "material": ref_part.get("material", ""),
        "total_assemblies": len(where_used),
        "total_qty_needed_per_set": total_qty_needed,
        "where_used": where_used
    }


@app.put("/api/parts-library/{part_id}")
async def update_part(part_id: str, updates: PartUpdate, current_user: User = Depends(get_current_user)):
    """Update part details (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    part = await db.parts_library.find_one({"part_id": part_id})
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    
    update_data = {"updated_at": get_ist_now()}
    
    if updates.name is not None:
        update_data["name"] = updates.name
    if updates.category is not None:
        update_data["category"] = updates.category
    if updates.part_type is not None:
        update_data["part_type"] = updates.part_type
    if updates.size is not None:
        update_data["size"] = updates.size
    if updates.length is not None:
        update_data["length"] = updates.length
    if updates.variant is not None:
        update_data["variant"] = updates.variant
    if updates.unit is not None:
        update_data["unit"] = updates.unit
    if updates.stock is not None:
        update_data["stock"] = updates.stock
    if updates.reserved_stock is not None:
        update_data["reserved_stock"] = updates.reserved_stock
    if updates.min_stock is not None:
        update_data["min_stock"] = updates.min_stock
    if updates.image is not None:
        update_data["image"] = updates.image
    if updates.drawing is not None:
        update_data["drawing"] = updates.drawing
    if updates.locations is not None:
        update_data["locations"] = [{"rack": loc.rack, "qty": loc.qty} for loc in updates.locations]
        # Recalculate stock from locations
        update_data["stock"] = sum(loc.qty for loc in updates.locations)
    
    await db.parts_library.update_one({"part_id": part_id}, {"$set": update_data})
    
    return {"message": "Part updated successfully"}


@app.delete("/api/parts-library/{part_id}")
async def delete_part(part_id: str, current_user: User = Depends(get_current_user)):
    """Delete a part (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if part is used in any assembly
    assembly_usage = await db.assembly_parts.find_one({"part_id": part_id})
    if assembly_usage:
        raise HTTPException(status_code=400, detail="Cannot delete: Part is used in assemblies")
    
    result = await db.parts_library.delete_one({"part_id": part_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Part not found")
    
    return {"message": "Part deleted successfully"}


@app.post("/api/parts-library/{part_id}/add-stock")
async def add_part_stock(
    part_id: str,
    rack: str,
    qty: int,
    current_user: User = Depends(get_current_user)
):
    """Add stock to a specific location"""
    if qty <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be positive")
    
    part = await db.parts_library.find_one({"part_id": part_id})
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    
    locations = part.get("locations", [])
    
    # Find existing location or add new
    found = False
    for loc in locations:
        if loc["rack"].lower() == rack.lower():
            loc["qty"] += qty
            found = True
            break
    
    if not found:
        locations.append({"rack": rack, "qty": qty})
    
    # Update total stock
    new_stock = part.get("stock", 0) + qty
    
    await db.parts_library.update_one(
        {"part_id": part_id},
        {
            "$set": {
                "stock": new_stock,
                "locations": locations,
                "updated_at": get_ist_now()
            }
        }
    )
    
    return {"message": f"Added {qty} to {rack}", "new_stock": new_stock}


@app.post("/api/parts-library/{part_id}/deduct-stock")
async def deduct_part_stock(
    part_id: str,
    rack: str,
    qty: int,
    current_user: User = Depends(get_current_user)
):
    """Deduct stock from a specific location"""
    if qty <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be positive")
    
    part = await db.parts_library.find_one({"part_id": part_id})
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    
    locations = part.get("locations", [])
    
    # Find location
    found = False
    for loc in locations:
        if loc["rack"].lower() == rack.lower():
            if loc["qty"] < qty:
                raise HTTPException(status_code=400, detail=f"Insufficient stock at {rack}. Available: {loc['qty']}")
            loc["qty"] -= qty
            found = True
            break
    
    if not found:
        raise HTTPException(status_code=404, detail=f"Location {rack} not found")
    
    # Remove location if qty is 0
    locations = [loc for loc in locations if loc["qty"] > 0]
    
    # Update total stock
    new_stock = max(0, part.get("stock", 0) - qty)
    
    await db.parts_library.update_one(
        {"part_id": part_id},
        {
            "$set": {
                "stock": new_stock,
                "locations": locations,
                "updated_at": get_ist_now()
            }
        }
    )
    
    return {"message": f"Deducted {qty} from {rack}", "new_stock": new_stock}


# ========================================
# EXCEL BOM IMPORT
# ========================================

# Type prefix mapping for auto Part ID generation
TYPE_PREFIX_MAP = {
    "bearing": "BEA",
    "bolt": "BOL",
    "screw": "SCR",
    "shaft": "SHF",
    "pin": "PIN",
    "seal": "SEL",
    "oring": "SEL",
    "o-ring": "SEL",
    "roller": "ROL",
    "aluminium": "ALU",
    "aluminum": "ALU",
    "plastic": "PLS",
    "ss": "SST",
    "stainless": "SST",
    "dowel": "DOW",
    "nut": "NUT",
    "washer": "WSH",
    "spring": "SPR",
    "gear": "GER",
    "bush": "BSH",
    "bushing": "BSH",
    "coupling": "CPL",
    "motor": "MTR",
    "pump": "PMP",
    "valve": "VLV",
    "cylinder": "CYL",
    "piston": "PST",
    "rod": "ROD",
    "plate": "PLT",
    "bracket": "BKT",
    "cover": "CVR",
    "gasket": "GSK",
    "ring": "RNG",
    "key": "KEY",
    "pulley": "PUL",
    "belt": "BLT",
    "chain": "CHN",
    "sprocket": "SPK",
}

async def get_next_part_id(prefix: str) -> str:
    """Generate next Part ID with given prefix"""
    # Find highest number for this prefix
    pattern = f"^{prefix}-"
    parts = await db.parts_library.find(
        {"part_id": {"$regex": pattern}},
        {"part_id": 1}
    ).sort("part_id", -1).to_list(length=1)
    
    if parts:
        # Extract number from last part ID
        last_id = parts[0]["part_id"]
        try:
            num = int(last_id.split("-")[-1])
            return f"{prefix}-{str(num + 1).zfill(4)}"
        except:
            pass
    
    return f"{prefix}-0001"

async def get_next_assembly_id() -> str:
    """Generate next Assembly ID"""
    assemblies = await db.assemblies.find(
        {"assembly_id": {"$regex": "^ASM-"}},
        {"assembly_id": 1}
    ).sort("assembly_id", -1).to_list(length=1)
    
    if assemblies:
        try:
            num = int(assemblies[0]["assembly_id"].split("-")[-1])
            return f"ASM-{str(num + 1).zfill(4)}"
        except:
            pass
    
    return "ASM-0001"


def generate_assembly_short_code(assembly_name: str) -> str:
    """Generate short code from assembly name for Part ID"""
    # Remove special characters and take first word or abbreviate
    words = ''.join(c if c.isalnum() or c == ' ' else '' for c in assembly_name).split()
    if not words:
        return "ASM"
    
    if len(words) == 1:
        # Single word - take first 6 chars
        return words[0][:6].upper()
    else:
        # Multiple words - take first letter of each (up to 6)
        short = ''.join(w[0] for w in words[:6]).upper()
        if len(short) < 3:
            short = words[0][:6].upper()
        return short


async def get_next_part_id_with_assembly(type_code: str, assembly_code: str, part_no: str = "") -> str:
    """
    Generate Part ID in format: {ASSEMBLY}-{PART_NO}-F{FACTORY_NO}-{TYPE_CODE}
    Example: WS04-01-F0001-SS
    
    Args:
        type_code: 2-letter type code (SS, OR, GD, BS, BT, etc.)
        assembly_code: Short assembly code (e.g., WS04)
        part_no: Original part number from Excel (optional)
    
    IMPORTANT: Factory number (F0001, F0002...) is GLOBAL across ALL assemblies.
    This ensures unique factory numbers even when parts are in different assemblies.
    """
    # Get next GLOBAL factory number - search ALL parts, not just current assembly
    # This ensures F0001, F0002... are unique across the entire factory
    parts = await db.parts_library.find(
        {"part_id": {"$regex": r"-F\d+"}},  # Match any part with F-number pattern
        {"part_id": 1}
    ).to_list(length=10000)
    
    # Find the highest F number across ALL parts globally
    max_factory_num = 0
    for p in parts:
        try:
            part_id = p.get("part_id", "")
            # Extract F number from part_id like "WS04-01-F0001-SS" or "FACG-1-F0001-SS"
            match = re.search(r'-F(\d+)', part_id)
            if match:
                fnum = int(match.group(1))
                if fnum > max_factory_num:
                    max_factory_num = fnum
        except:
            pass
    
    next_factory_num = max_factory_num + 1
    
    # Clean part_no for use in ID
    if part_no:
        # Remove special chars, keep alphanumeric and hyphens
        clean_part_no = ''.join(c if c.isalnum() or c == '-' else '' for c in str(part_no).strip())
        # Truncate if too long
        clean_part_no = clean_part_no[:20] if len(clean_part_no) > 20 else clean_part_no
        if clean_part_no:
            return f"{assembly_code}-{clean_part_no}-F{str(next_factory_num).zfill(4)}-{type_code}"
    
    # Fallback without part_no
    return f"{assembly_code}-F{str(next_factory_num).zfill(4)}-{type_code}"


# Type code mapping for Part ID suffix
TYPE_CODE_MAP = {
    "seal": "OR",      # O-Ring / Seal
    "bolt": "BT",      # Bolt
    "washer": "WS",    # Washer  
    "ss": "SS",        # Stainless Steel
    "guide": "GD",     # Guide
    "bush": "BS",      # Bush
    "spring": "SP",    # Spring
    "pin": "PN",       # Pin
    "machined": "MC",  # Machined Part
    "casting": "CT",   # Casting
    "": "PT",          # Generic Part
}


class BOMImportRequest(BaseModel):
    assembly_name: str
    assembly_description: Optional[str] = ""
    parts: List[dict]



@app.post("/api/import-bom-preview")
async def import_bom_from_preview(
    request: BOMImportRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Import BOM from previewed/edited data.
    Part IDs include assembly name: TYPE-ASMCODE-0001
    """
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    assembly_name = request.assembly_name.strip()
    assembly_description = request.assembly_description.strip() if request.assembly_description else ""
    
    if not assembly_name:
        raise HTTPException(status_code=400, detail="Assembly name is required")
    
    if not request.parts or len(request.parts) == 0:
        raise HTTPException(status_code=400, detail="At least one part is required")
    
    # Generate assembly short code for Part IDs
    assembly_code = generate_assembly_short_code(assembly_name)
    
    # Create or find assembly (escape regex special chars)
    escaped_asm_name = re.escape(assembly_name)
    existing_assembly = await db.assemblies.find_one({"name": {"$regex": f"^{escaped_asm_name}$", "$options": "i"}})
    
    if existing_assembly:
        assembly_id = existing_assembly["assembly_id"]
        assembly_created = False
    else:
        assembly_id = await get_next_assembly_id()
        await db.assemblies.insert_one({
            "assembly_id": assembly_id,
            "name": assembly_name,
            "description": assembly_description,
            "created_at": get_ist_now(),
            "updated_at": get_ist_now()
        })
        assembly_created = True
    
    parts_created = []
    parts_linked = []
    parts_skipped = []
    
    for part_data in request.parts:
        part_name = str(part_data.get("part_name", "")).strip()
        category = str(part_data.get("category", "")).strip()
        part_type = str(part_data.get("type", "")).strip().lower()
        size = str(part_data.get("size", "")).strip()
        length = str(part_data.get("length", "")).strip()
        qty = int(part_data.get("qty", 1))
        unit = str(part_data.get("unit", "PCS")).strip().upper()
        
        if not part_name:
            continue
        
        # Escape special regex characters in part name
        escaped_name = re.escape(part_name)
        escaped_category = re.escape(category) if category else ""
        
        # Check for duplicate (same name + category + size + length)
        # This allows multiple parts with same name but different dimensions (e.g., O-Ring Seals)
        duplicate_query = {
            "name": {"$regex": f"^{escaped_name}$", "$options": "i"},
        }
        if category:
            duplicate_query["category"] = {"$regex": f"^{escaped_category}$", "$options": "i"}
        if size:
            duplicate_query["size"] = str(size).strip()
        if length:
            duplicate_query["length"] = str(length).strip()
        
        existing_part = await db.parts_library.find_one(duplicate_query)
        
        if existing_part:
            part_id = existing_part["part_id"]
            # Just link to assembly
            existing_link = await db.assembly_parts.find_one({
                "assembly_id": assembly_id,
                "part_id": part_id
            })
            
            if not existing_link:
                await db.assembly_parts.insert_one({
                    "assembly_id": assembly_id,
                    "part_id": part_id,
                    "quantity": qty,
                    "created_at": get_ist_now()
                })
                parts_linked.append({
                    "part_id": part_id,
                    "name": part_name,
                    "quantity": qty,
                    "status": "linked_existing"
                })
            else:
                parts_skipped.append({
                    "part_id": part_id,
                    "name": part_name,
                    "reason": "already_in_bom"
                })
        else:
            # Create new part with new ID format: {ASSEMBLY}-{PART_NO}-F{FACTORY_NO}-{TYPE_CODE}
            # Get type code based on part_type, material, category or name
            type_code = TYPE_CODE_MAP.get(part_type, "")
            if not type_code:
                # Try to infer from category or name
                cat_lower = category.lower()
                name_lower = part_name.lower()
                if 'o-ring' in cat_lower or 'seal' in cat_lower or 'o-ring' in name_lower:
                    type_code = "OR"
                elif 'bolt' in cat_lower or 'screw' in cat_lower:
                    type_code = "BT"
                elif 'washer' in cat_lower:
                    type_code = "WS"
                elif 'guide' in cat_lower or 'guide' in name_lower:
                    type_code = "GD"
                elif 'bush' in cat_lower or 'bush' in name_lower:
                    type_code = "BS"
                elif 'spring' in cat_lower or 'spring' in name_lower:
                    type_code = "SP"
                elif 'pin' in cat_lower or 'pin' in name_lower:
                    type_code = "PN"
                else:
                    type_code = "PT"  # Generic Part
            
            # Get part_no from data if available
            part_no = str(part_data.get("part_no", "")).strip()
            part_id = await get_next_part_id_with_assembly(type_code, assembly_code, part_no)
            
            await db.parts_library.insert_one({
                "part_id": part_id,
                "name": part_name,
                "category": category,
                "part_type": part_type,
                "size": size,
                "length": length,
                "variant": "",
                "unit": unit,
                "stock": 0,
                "reserved_stock": 0,
                "min_stock": 0,
                "locations": [],
                "image": "",
                "drawing": "",
                "created_at": get_ist_now(),
                "updated_at": get_ist_now()
            })
            
            # Link to assembly
            await db.assembly_parts.insert_one({
                "assembly_id": assembly_id,
                "part_id": part_id,
                "quantity": qty,
                "created_at": get_ist_now()
            })
            
            parts_created.append({
                "part_id": part_id,
                "name": part_name,
                "category": category,
                "type": part_type,
                "size": size,
                "length": length,
                "quantity": qty,
                "unit": unit,
                "status": "created",
                "image": "",
                "drawing": ""
            })
    
    return {
        "success": True,
        "assembly": {
            "assembly_id": assembly_id,
            "name": assembly_name,
            "description": assembly_description,
            "created": assembly_created
        },
        "summary": {
            "parts_created": len(parts_created),
            "parts_linked": len(parts_linked),
            "parts_skipped": len(parts_skipped),
            "total_processed": len(parts_created) + len(parts_linked) + len(parts_skipped)
        },
        "parts_created": parts_created,
        "parts_linked": parts_linked,
        "parts_skipped": parts_skipped
    }


@app.post("/api/import-bom-excel")
async def import_bom_from_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Import BOM from Excel file.
    Excel format:
    - Row 1: Assembly Name: | <name>
    - Row 2: Description: | <description>
    - Row 3: (empty)
    - Row 4: Part Name | Category | Type | Size | Length | Qty | Unit (headers)
    - Row 5+: Part data
    """
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        # Read Excel file
        contents = await file.read()
        logger.info(f"Excel file received: {file.filename}, size: {len(contents)} bytes")
        
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        
        # Parse Assembly info from Row 1 and 2
        assembly_name = None
        assembly_description = None
        
        # Row 1: Assembly Name - Try multiple formats
        row1 = [cell.value for cell in ws[1]]
        logger.info(f"Row 1 content: {row1}")
        
        if row1:
            # Format 1: "Assembly Name:" in A1, value in B1
            if len(row1) >= 2 and row1[0] and "assembly" in str(row1[0]).lower():
                assembly_name = str(row1[1]).strip() if row1[1] else None
            # Format 2: Just the name in A1 (no label)
            elif row1[0] and not any(x in str(row1[0]).lower() for x in ['part', 'category', 'type', 'size']):
                # Check if A1 looks like a name (not a header)
                first_cell = str(row1[0]).strip()
                if first_cell and len(first_cell) > 2:
                    assembly_name = first_cell
        
        # Row 2: Description
        row2 = [cell.value for cell in ws[2]]
        logger.info(f"Row 2 content: {row2}")
        
        if row2 and len(row2) >= 2:
            if row2[0] and "description" in str(row2[0]).lower():
                assembly_description = str(row2[1]).strip() if row2[1] else None
        
        if not assembly_name:
            # Try to get from filename if not in cells
            if file.filename:
                name_from_file = file.filename.replace('.xlsx', '').replace('.xls', '').replace('_', ' ')
                if name_from_file and len(name_from_file) > 2:
                    assembly_name = name_from_file
                    logger.info(f"Using filename as assembly name: {assembly_name}")
        
        if not assembly_name:
            raise HTTPException(status_code=400, detail="Assembly Name not found. Put 'Assembly Name:' in A1 and the name in B1, OR just put the assembly name in A1")
        
        # Create or find assembly (escape regex special chars)
        escaped_asm_name = re.escape(assembly_name)
        existing_assembly = await db.assemblies.find_one({"name": {"$regex": f"^{escaped_asm_name}$", "$options": "i"}})
        
        if existing_assembly:
            assembly_id = existing_assembly["assembly_id"]
            assembly_created = False
        else:
            assembly_id = await get_next_assembly_id()
            await db.assemblies.insert_one({
                "assembly_id": assembly_id,
                "name": assembly_name,
                "description": assembly_description or "",
                "created_at": get_ist_now(),
                "updated_at": get_ist_now()
            })
            assembly_created = True
        
        # Generate assembly short code for Part IDs (includes assembly name)
        assembly_code = generate_assembly_short_code(assembly_name)
        
        # Find the header row (contains "Part Name" or similar)
        header_row = None
        data_start_row = 5  # Default
        
        for row_num in range(1, 10):  # Check first 10 rows
            row = [cell.value for cell in ws[row_num]]
            if row:
                row_lower = [str(c).lower() if c else '' for c in row]
                # Check if this looks like a header row
                if any('part' in c and 'name' in c for c in row_lower) or \
                   any(c == 'part name' for c in row_lower) or \
                   (row_lower[0] and 'part' in row_lower[0]):
                    header_row = row_num
                    data_start_row = row_num + 1
                    logger.info(f"Found header row at {row_num}: {row}")
                    break
        
        logger.info(f"Data starts at row {data_start_row}")
        
        # Parse parts starting from data row
        # Template column order: Part No | Part Name | Material | Category | Dia/ID/WxB | Length/Thickness | Qty | Unit | Remarks
        parts_created = []
        parts_linked = []
        parts_skipped = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            # Skip empty rows
            if not row or (not row[0] and (len(row) < 2 or not row[1])):
                continue
            
            # Correct column mapping based on unified template:
            # Col 0: Part No, Col 1: Part Name, Col 2: Material, Col 3: Category
            # Col 4: Dia/ID/WxB (size), Col 5: Length/Thickness, Col 6: Qty, Col 7: Unit, Col 8: Remarks
            part_no_raw = str(row[0]).strip() if row[0] else ""
            part_name = str(row[1]).strip() if len(row) > 1 and row[1] else None
            material = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            category = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            size = str(row[4]).strip() if len(row) > 4 and row[4] else ""  # Keep as string - may have Ø, fractions
            length = str(row[5]).strip() if len(row) > 5 and row[5] else ""  # Keep as string
            
            # Safe qty conversion - handle decimals and fractions
            qty = 1
            if len(row) > 6 and row[6]:
                try:
                    if isinstance(row[6], (int, float)):
                        qty = int(row[6])
                    else:
                        qty_str = str(row[6]).strip()
                        if '/' not in qty_str:
                            qty = int(float(qty_str))
                except:
                    qty = 1
            
            unit = str(row[7]).strip().upper() if len(row) > 7 and row[7] else "PCS"
            remarks = str(row[8]).strip() if len(row) > 8 and row[8] else ""
            
            # Fallback: If part_name is empty but part_no has content, maybe old format
            if not part_name and part_no_raw:
                # Check if this might be old format where col 0 was part_name
                # Look at row[0] - if it doesn't look like a part number, treat it as name
                if not any(c.isdigit() for c in part_no_raw) or len(part_no_raw) > 30:
                    part_name = part_no_raw
                    part_no_raw = ""
            
            if not part_name:
                continue
            
            # Determine part_type from material/category for Part ID prefix
            mat_lower = material.lower()
            cat_lower = category.lower()
            name_lower = part_name.lower()
            
            part_type = ""
            if 'o-ring' in cat_lower or 'oring' in cat_lower or 'o-ring' in name_lower or 'seal' in cat_lower:
                part_type = "seal"
            elif 'nitrile' in mat_lower or 'viton' in mat_lower or 'epdm' in mat_lower or 'silicon' in mat_lower:
                part_type = "seal"
            elif 'bolt' in cat_lower or 'bolt' in name_lower or 'screw' in cat_lower:
                part_type = "bolt"
            elif 'washer' in cat_lower or 'washer' in name_lower:
                part_type = "washer"
            elif 'ss' in mat_lower or 'stainless' in mat_lower:
                part_type = "ss"
            
            # Escape special regex characters in part name
            escaped_name = re.escape(part_name)
            escaped_category = re.escape(category) if category else ""
            
            # Check for duplicate (same name + category + size + length)
            # This allows multiple parts with same name but different dimensions (e.g., O-Ring Seals)
            duplicate_query = {
                "name": {"$regex": f"^{escaped_name}$", "$options": "i"},
            }
            if category:
                duplicate_query["category"] = {"$regex": f"^{escaped_category}$", "$options": "i"}
            if size:
                duplicate_query["size"] = str(size).strip()
            if length:
                duplicate_query["length"] = str(length).strip()
            
            existing_part = await db.parts_library.find_one(duplicate_query)
            
            if existing_part:
                part_id = existing_part["part_id"]
                # Just link to assembly
                existing_link = await db.assembly_parts.find_one({
                    "assembly_id": assembly_id,
                    "part_id": part_id
                })
                
                if not existing_link:
                    await db.assembly_parts.insert_one({
                        "assembly_id": assembly_id,
                        "part_id": part_id,
                        "quantity": qty,
                        "created_at": get_ist_now()
                    })
                    parts_linked.append({
                        "part_id": part_id,
                        "name": part_name,
                        "quantity": qty,
                        "status": "linked_existing"
                    })
                else:
                    parts_skipped.append({
                        "part_id": part_id,
                        "name": part_name,
                        "reason": "already_in_bom"
                    })
            else:
                # Create new part with new ID format: {ASSEMBLY}-{PART_NO}-F{FACTORY_NO}-{TYPE_CODE}
                # Determine type code from part_type, material, or category
                type_code = TYPE_CODE_MAP.get(part_type, "")
                if not type_code:
                    cat_lower = category.lower()
                    name_lower = part_name.lower()
                    mat_lower = material.lower()
                    if 'o-ring' in cat_lower or 'seal' in cat_lower or 'o-ring' in name_lower:
                        type_code = "OR"
                    elif 'bolt' in cat_lower or 'screw' in cat_lower:
                        type_code = "BT"
                    elif 'washer' in cat_lower:
                        type_code = "WS"
                    elif 'guide' in cat_lower or 'guide' in name_lower:
                        type_code = "GD"
                    elif 'bush' in cat_lower or 'bush' in name_lower:
                        type_code = "BS"
                    elif 'spring' in cat_lower or 'spring' in name_lower:
                        type_code = "SP"
                    elif 'pin' in cat_lower or 'pin' in name_lower:
                        type_code = "PN"
                    elif 'ss' in mat_lower or 'stainless' in mat_lower:
                        type_code = "SS"
                    elif 'nitrile' in mat_lower or 'viton' in mat_lower or 'rubber' in mat_lower:
                        type_code = "OR"
                    else:
                        type_code = "PT"  # Generic Part
                
                part_id = await get_next_part_id_with_assembly(type_code, assembly_code, part_no_raw)
                
                await db.parts_library.insert_one({
                    "part_id": part_id,
                    "part_no": part_no_raw,  # Original part number from Excel
                    "name": part_name,
                    "material": material,
                    "category": category,
                    "part_type": part_type,
                    "size": size,  # Dia/ID/WxB - kept as string
                    "length": length,  # Length/Thickness - kept as string
                    "variant": "",
                    "unit": unit,
                    "remarks": remarks,
                    "stock": 0,
                    "reserved_stock": 0,
                    "min_stock": 0,
                    "locations": [],
                    "image": "",
                    "drawing": "",
                    "source_assembly": assembly_code,
                    "created_at": get_ist_now(),
                    "updated_at": get_ist_now()
                })
                
                # Link to assembly
                await db.assembly_parts.insert_one({
                    "assembly_id": assembly_id,
                    "part_id": part_id,
                    "quantity": qty,
                    "created_at": get_ist_now()
                })
                
                parts_created.append({
                    "part_id": part_id,
                    "part_no": part_no_raw,
                    "name": part_name,
                    "material": material,
                    "category": category,
                    "type": part_type,
                    "size": size,
                    "length": length,
                    "quantity": qty,
                    "unit": unit,
                    "remarks": remarks,
                    "status": "created"
                })
        
        wb.close()
        
        return {
            "success": True,
            "assembly": {
                "assembly_id": assembly_id,
                "name": assembly_name,
                "description": assembly_description,
                "created": assembly_created
            },
            "summary": {
                "parts_created": len(parts_created),
                "parts_linked": len(parts_linked),
                "parts_skipped": len(parts_skipped),
                "total_processed": len(parts_created) + len(parts_linked) + len(parts_skipped)
            },
            "parts_created": parts_created,
            "parts_linked": parts_linked,
            "parts_skipped": parts_skipped
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel import error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to parse Excel file: {str(e)}")


# ========================================
# PUMP & SPRAY GUN EXCEL IMPORT WITH AUTO-DETECT
# ========================================

async def get_next_factory_part_no(prefix: str = None) -> int:
    """Get next GLOBAL factory-wide part number (shared across ALL assemblies)"""
    # Find highest factory number across ALL parts (not filtered by prefix)
    # This ensures factory numbers are unique globally: F0001, F0002... across all assemblies
    
    # Search for ALL parts that have factory number pattern
    parts = await db.parts_library.find(
        {"part_id": {"$regex": r"-F\d+"}},
        {"part_id": 1}
    ).sort("created_at", -1).to_list(length=5000)
    
    max_factory_no = 0
    for part in parts:
        part_id = part.get("part_id", "")
        # Extract factory number (F0001, F0002, etc.) - can be anywhere in the ID
        match = re.search(r'-F(\d+)', part_id)
        if match:
            factory_no = int(match.group(1))
            if factory_no > max_factory_no:
                max_factory_no = factory_no
    
    return max_factory_no + 1


def generate_pump_spray_part_id(
    prefix: str,  # FTP or FTSG
    assembly_code: str,
    assembly_part_no: int,
    size_details: str,
    factory_no: int
) -> str:
    """
    Generate Part ID in format: PREFIX-ASSEMBLY-PARTNO-SIZE-FACTORYNO
    Example: FTP-DOSPA-001-25MM-F0001
    """
    parts = [prefix, assembly_code, str(assembly_part_no).zfill(3)]
    
    # Add size if provided
    if size_details and size_details.strip():
        # Clean size - remove spaces, make uppercase
        clean_size = size_details.strip().upper().replace(' ', '').replace('.', '')
        parts.append(clean_size)
    
    # Add factory number
    parts.append(f"F{str(factory_no).zfill(4)}")
    
    return "-".join(parts)


def format_oring_size(od: str, inner_id: str, thickness: str) -> str:
    """Format O-Ring size as OD12ID10TH1"""
    size_parts = []
    if od:
        size_parts.append(f"OD{od.replace('mm', '').replace('MM', '').strip()}")
    if inner_id:
        size_parts.append(f"ID{inner_id.replace('mm', '').replace('MM', '').strip()}")
    if thickness:
        size_parts.append(f"TH{thickness.replace('mm', '').replace('MM', '').strip()}")
    return "".join(size_parts)


def detect_template_type(headers: list) -> str:
    """
    Auto-detect if Excel is Pump or Spray Gun template based on headers.
    Returns 'pump', 'spray_gun', or 'generic'
    """
    headers_lower = [str(h).lower() if h else '' for h in headers]
    headers_str = ' '.join(headers_lower)
    
    # Check for pump-specific columns
    pump_keywords = ['dosing', 'pump', 'ftp', 'metering']
    spray_keywords = ['spray', 'gun', 'ftsg', 'nozzle', 'trigger']
    
    # Check headers for keywords
    is_pump = any(kw in headers_str for kw in pump_keywords)
    is_spray = any(kw in headers_str for kw in spray_keywords)
    
    # Also check for O-Ring specific columns which both might have
    has_oring_cols = any('o-ring' in h or 'oring' in h or 'od' in h.split() for h in headers_lower)
    
    if is_pump and not is_spray:
        return 'pump'
    elif is_spray and not is_pump:
        return 'spray_gun'
    else:
        return 'generic'


def generate_common_part_id(part_name: str, material: str) -> str:
    """
    Generate Common Part ID for linking same parts across assemblies.
    Format: NORMALIZED_PART_NAME-MATERIAL_CODE
    Example: M4-BOLT-10MM-SS304, PLAIN-WASHER-SS304
    """
    if not part_name:
        return ""
    
    # Normalize part name: uppercase, remove special chars, replace spaces with dash
    import re
    clean_name = re.sub(r'[^A-Z0-9\s-]', '', part_name.upper())
    clean_name = re.sub(r'\s+', '-', clean_name.strip())
    clean_name = re.sub(r'-+', '-', clean_name)  # Remove multiple dashes
    
    # Get material code
    material_upper = (material or "").upper().strip()
    material_codes = {
        'ALUMINUM': 'ALU',
        'ALUMINIUM': 'ALU',
        'SS304': 'SS304',
        'SS316': 'SS316',
        'STAINLESS STEEL': 'SS',
        'NYLON': 'NYL',
        'PP': 'PP',
        'PPS': 'PPS',
        'PPF': 'PPF',
        'PPF/NYLON': 'PPN',
        'SILICON': 'SIL',
        'SILICONE': 'SIL',
        'TEFLON': 'TEF',
        'PTFE': 'TEF',
        'DELLDRIN': 'DEL',
        'DELRIN': 'DEL',
        'VITON': 'VIT',
        'RUBBER': 'RUB',
        'BRASS': 'BRS',
        'BRONZE': 'BRZ',
        'COPPER': 'COP',
        'STEEL': 'STL',
        'CARBON STEEL': 'CS',
        'CAST IRON': 'CI',
        'HDPE': 'HDPE',
        'PVC': 'PVC',
        'ACETAL': 'ACL',
        'PEEK': 'PEEK',
    }
    
    mat_code = material_codes.get(material_upper, material_upper[:4] if material_upper else 'GEN')
    
    if mat_code:
        return f"{clean_name}-{mat_code}"
    return clean_name


def get_material_code(material: str) -> str:
    """Get short material code from material name"""
    material_upper = (material or "").upper().strip()
    material_codes = {
        'ALUMINUM': 'ALU', 'ALUMINIUM': 'ALU',
        'SS304': 'SS304', 'SS316': 'SS316', 'STAINLESS STEEL': 'SS',
        'NYLON': 'NYL', 'PP': 'PP', 'PPS': 'PPS', 'PPF': 'PPF',
        'PPF/NYLON': 'PPN', 'SILICON': 'SIL', 'SILICONE': 'SIL',
        'TEFLON': 'TEF', 'PTFE': 'TEF', 'DELLDRIN': 'DEL', 'DELRIN': 'DEL',
        'VITON': 'VIT', 'RUBBER': 'RUB', 'BRASS': 'BRS', 'BRONZE': 'BRZ',
        'COPPER': 'COP', 'STEEL': 'STL', 'CARBON STEEL': 'CS', 'CAST IRON': 'CI',
        'HDPE': 'HDPE', 'PVC': 'PVC', 'ACETAL': 'ACL', 'PEEK': 'PEEK',
    }
    return material_codes.get(material_upper, material_upper[:4] if material_upper else '')


@app.post("/api/parse-bom-excel")
async def parse_bom_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Parse unified BOM Excel template.
    
    Format:
    Row 1: Assembly Name: | [Name]
    Row 2: Description: | [Description]
    Row 3: Headers (Part No | Part Name | Material | Category | Dia/ID/WxB | Length/Thickness | Qty | Unit | Remarks)
    Row 4+: Data
    
    Part ID Format: {ASSEMBLY}-{PART_NO}-F{FACTORY_NO}-{TYPE_CODE}
    Example: WS04-01-F0001-SS
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx)")
    
    def get_type_code(material: str, category: str, part_name: str) -> str:
        """Get type code based on material/category"""
        mat_lower = material.lower() if material else ""
        cat_lower = category.lower() if category else ""
        name_lower = part_name.lower() if part_name else ""
        
        # Check for O-Ring/Seal
        if 'o-ring' in cat_lower or 'oring' in cat_lower or 'o-ring' in name_lower or 'seal' in cat_lower:
            return "SEAL"
        if 'nitrile' in mat_lower or 'viton' in mat_lower or 'epdm' in mat_lower or 'silicon' in mat_lower:
            return "SEAL"
        
        # Check for fasteners
        if 'bolt' in cat_lower or 'bolt' in name_lower or 'screw' in cat_lower or 'screw' in name_lower:
            return "BOLT"
        if 'washer' in cat_lower or 'washer' in name_lower:
            return "WASHER"
        if 'fastener' in cat_lower:
            if 'washer' in name_lower:
                return "WASHER"
            return "BOLT"
        
        # Check for SS materials
        if 'ss' in mat_lower or 'stainless' in mat_lower or 'ss304' in mat_lower or 'ss316' in mat_lower or 'ss 304' in mat_lower or 'ss 316' in mat_lower:
            return "SS"
        
        # Default
        return "PART"
    
    def safe_str(val) -> str:
        """Safely convert any value to string without int conversion"""
        if val is None:
            return ""
        return str(val).strip()
    
    def safe_int(val, default=1) -> int:
        """Safely convert to int, handling fractions and special chars"""
        if val is None:
            return default
        if isinstance(val, (int, float)):
            return int(val)
        s = str(val).strip()
        if not s:
            return default
        # If contains fraction or special chars, return default
        if '/' in s or 'x' in s.lower() or 'Ø' in s or '⌀' in s or 'M' in s:
            return default
        try:
            return int(float(s))
        except:
            return default
    
    def extract_part_number(part_no_str: str) -> str:
        """Extract just the number from PART NO - 01 format"""
        if not part_no_str:
            return "00"
        # Try to extract number after dash or hyphen
        import re
        match = re.search(r'(\d+)\s*$', part_no_str)
        if match:
            return match.group(1).zfill(2)
        return "00"
    
    try:
        contents = await file.read()
        logger.info(f"Parsing BOM Excel: {file.filename}, size: {len(contents)} bytes")
        
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        
        # Row 1: Assembly Name
        row1 = [cell.value for cell in ws[1]]
        assembly_name = None
        if row1 and len(row1) >= 2:
            if row1[0] and 'assembly' in str(row1[0]).lower():
                assembly_name = safe_str(row1[1])
        
        # Row 2: Description
        row2 = [cell.value for cell in ws[2]]
        assembly_description = ""
        if row2 and len(row2) >= 2:
            if row2[0] and 'description' in str(row2[0]).lower():
                assembly_description = safe_str(row2[1])
            elif row2[0] and 'assembly' in str(row2[0]).lower() and not assembly_name:
                assembly_name = safe_str(row2[1])
        
        if not assembly_name:
            assembly_name = file.filename.replace('.xlsx', '').replace('.xls', '').replace('_', ' ')
        
        # Generate assembly code (e.g., WS04 SPRAY -> WS04)
        assembly_code = generate_assembly_short_code(assembly_name)
        
        # Find header row (row 3)
        header_row = [cell.value for cell in ws[3]]
        header_row_num = 3
        
        if not header_row or not any('part' in str(c).lower() for c in header_row if c):
            # Try row 4
            header_row = [cell.value for cell in ws[4]]
            header_row_num = 4
        
        if not header_row:
            raise HTTPException(status_code=400, detail="Could not find header row")
        
        logger.info(f"Headers: {header_row}")
        
        # Map columns by index (0-based) - use separate loops to avoid elif issues
        col_map = {}
        logger.info(f"Scanning headers: {header_row}")
        
        for idx, header in enumerate(header_row):
            if not header:
                continue
            h = str(header).lower().strip()
            
            # Part Name - check first with multiple variations
            if 'part_name' not in col_map:
                if h == 'part name' or h == 'partname' or h == 'name' or 'part name' in h:
                    col_map['part_name'] = idx
                    logger.info(f"Found part_name at index {idx}: '{header}'")
            
            # Part No
            if 'part_no' not in col_map:
                if h == 'part no' or h == 'partno' or 'part no' in h or h == 'part number':
                    col_map['part_no'] = idx
            
            # Material
            if 'material' not in col_map and h == 'material':
                col_map['material'] = idx
            
            # Category
            if 'category' not in col_map and h == 'category':
                col_map['category'] = idx
            
            # Type
            if 'type' not in col_map and h == 'type':
                col_map['type'] = idx
            
            # Size fields - Dia/ID/WxB
            if 'size1' not in col_map:
                if 'dia' in h or 'wxb' in h or h == 'size' or h == 'diameter':
                    col_map['size1'] = idx
            
            # Length/Thickness
            if 'size2' not in col_map:
                if 'length' in h or 'thick' in h:
                    col_map['size2'] = idx
            
            # Qty
            if 'qty' not in col_map:
                if h == 'qty' or h == 'quantity' or h == 'qty.':
                    col_map['qty'] = idx
            
            # Unit
            if 'unit' not in col_map and h == 'unit':
                col_map['unit'] = idx
            
            # Remarks
            if 'remarks' not in col_map and 'remark' in h:
                col_map['remarks'] = idx
        
        logger.info(f"Final column mapping: {col_map}")
        
        if 'part_name' not in col_map:
            # Last resort: if there's a "Name" column that wasn't matched due to "Part Name" not existing
            for idx, header in enumerate(header_row):
                if header and str(header).lower().strip() == 'name':
                    col_map['part_name'] = idx
                    logger.info(f"Found 'name' column as part_name at index {idx}")
                    break
        
        if 'part_name' not in col_map:
            raise HTTPException(status_code=400, detail=f"'Part Name' column not found. Headers found: {header_row}")
        
        # Get next factory number
        factory_no_start = await get_next_factory_part_no('BOM')
        
        # Parse data rows
        parts = []
        current_factory_no = factory_no_start
        data_start_row = header_row_num + 1
        
        for row_num, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            if not row:
                continue
            
            # Get part name (required)
            part_name_idx = col_map.get('part_name', 1)
            if len(row) <= part_name_idx or not row[part_name_idx]:
                continue
            
            part_name = safe_str(row[part_name_idx])
            if not part_name:
                continue
            
            # Get all fields as strings
            part_no_raw = safe_str(row[col_map['part_no']]) if 'part_no' in col_map and len(row) > col_map['part_no'] else ""
            material = safe_str(row[col_map['material']]) if 'material' in col_map and len(row) > col_map['material'] else ""
            category = safe_str(row[col_map['category']]) if 'category' in col_map and len(row) > col_map['category'] else ""
            size1 = safe_str(row[col_map['size1']]) if 'size1' in col_map and len(row) > col_map['size1'] else ""
            size2 = safe_str(row[col_map['size2']]) if 'size2' in col_map and len(row) > col_map['size2'] else ""
            unit = safe_str(row[col_map['unit']]) if 'unit' in col_map and len(row) > col_map['unit'] else "PCS"
            remarks = safe_str(row[col_map['remarks']]) if 'remarks' in col_map and len(row) > col_map['remarks'] else ""
            
            # Qty - safe int conversion
            qty = 1
            if 'qty' in col_map and len(row) > col_map['qty'] and row[col_map['qty']]:
                qty = safe_int(row[col_map['qty']], 1)
            
            # Extract part number for ID
            part_num = extract_part_number(part_no_raw)
            
            # Get type code
            type_code = get_type_code(material, category, part_name)
            
            # Generate Part ID: ASSEMBLY-PARTNO-FFACTORYNO-TYPECODE
            # Example: WS04-01-F0001-SS
            generated_part_id = f"{assembly_code}-{part_num}-F{current_factory_no:04d}-{type_code}"
            
            # Check for O-Ring
            is_oring = 'o-ring' in category.lower() or 'oring' in category.lower() or 'o-ring' in part_name.lower()
            
            # Format size display
            if is_oring and size1 and size2:
                size_display = f"ID{size1} x T{size2}"
            elif size1 and size2:
                size_display = f"{size1} x {size2}"
            elif size1:
                size_display = size1
            else:
                size_display = ""
            
            # Check if part exists
            common_part_id = generate_common_part_id(part_name, material)
            existing_part = None
            if common_part_id:
                existing_part = await db.parts_library.find_one({"common_part_id": common_part_id})
            
            parts.append({
                "row": row_num,
                "part_no": part_no_raw,
                "part_num": part_num,
                "part_name": part_name,
                "material": material,
                "category": category,
                "size1": size1,
                "size2": size2,
                "size": size_display,
                "is_oring": is_oring,
                "qty": qty,
                "unit": unit.upper() if unit else "PCS",
                "remarks": remarks,
                "type_code": type_code,
                "factory_no": current_factory_no,
                "status": "exists" if existing_part else "new",
                "existing_id": existing_part["part_id"] if existing_part else None,
                "generated_id": generated_part_id,
                "common_part_id": common_part_id
            })
            
            current_factory_no += 1
        
        wb.close()
        
        if not parts:
            raise HTTPException(status_code=400, detail="No parts found in Excel file")
        
        # Detect material conflicts (same Name + Size but different Material)
        material_conflicts = []
        parts_by_name_size = {}
        
        for part in parts:
            # Normalize for comparison
            name_lower = part["part_name"].lower().strip()
            size_key = part["size1"].strip() if part.get("size1") else ""
            key = f"{name_lower}|{size_key}"
            
            if key not in parts_by_name_size:
                parts_by_name_size[key] = []
            parts_by_name_size[key].append(part)
        
        # Check for conflicts (same name+size, different material)
        for key, matching_parts in parts_by_name_size.items():
            if len(matching_parts) > 1:
                # Check if materials are different
                materials = set()
                for p in matching_parts:
                    mat = p.get("material", "").lower().strip()
                    if mat:
                        materials.add(mat)
                
                if len(materials) > 1:
                    # Material conflict detected!
                    material_conflicts.append({
                        "key": f"{matching_parts[0]['part_name']}|{matching_parts[0].get('size1', '')}",  # For frontend to use directly
                        "name": matching_parts[0]["part_name"],
                        "size": matching_parts[0].get("size1", ""),
                        "parts": [{"row": p["row"], "material": p.get("material", ""), "qty": p.get("qty", 1)} for p in matching_parts],
                        "message": f"Parts with same name '{matching_parts[0]['part_name']}' and size '{matching_parts[0].get('size1', '')}' have different materials"
                    })
                elif len(materials) <= 1:
                    # Same material - these will be auto-merged (mark them)
                    total_qty = sum(p.get("qty", 1) for p in matching_parts)
                    for p in matching_parts[1:]:
                        p["will_merge"] = True
                        p["merge_into_row"] = matching_parts[0]["row"]
                    matching_parts[0]["merged_qty"] = total_qty
                    matching_parts[0]["merge_note"] = f"Will merge {len(matching_parts)} rows into qty={total_qty}"
        
        logger.info(f"Parsed {len(parts)} parts from {file.filename}, found {len(material_conflicts)} material conflicts")
        
        return {
            "success": True,
            "assembly_name": assembly_name,
            "assembly_code": assembly_code,
            "assembly_description": assembly_description,
            "parts": parts,
            "total_parts": len(parts),
            "new_parts": len([p for p in parts if p["status"] == "new"]),
            "existing_parts": len([p for p in parts if p["status"] == "exists"]),
            "next_factory_no": factory_no_start,
            "material_conflicts": material_conflicts,
            "has_conflicts": len(material_conflicts) > 0
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"BOM Excel parse error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to parse Excel: {str(e)}")


# Keep old endpoint for backward compatibility
@app.post("/api/parse-pump-spray-excel")
async def parse_pump_spray_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Parse Pump or Spray Gun Excel template with auto-detection.
    
    Expected columns:
    - Part Name (required)
    - Assembly Part No (001, 002, etc.)
    - Category
    - Type (oring, bearing, etc.)
    - Size (for regular parts)
    - O-Ring OD (for O-rings)
    - O-Ring ID (for O-rings)
    - O-Ring Thickness (for O-rings)
    - Qty
    - Unit
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx)")
    
    try:
        contents = await file.read()
        logger.info(f"Parsing Pump/Spray Excel: {file.filename}, size: {len(contents)} bytes")
        
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        
        # Row 1: Template Type (PUMP or SPRAY GUN)
        row1 = [cell.value for cell in ws[1]]
        logger.info(f"Row 1: {row1}")
        
        template_type = 'generic'
        if row1 and row1[0]:
            first_cell = str(row1[0]).strip().upper()
            if 'PUMP' in first_cell or 'FTP' in first_cell:
                template_type = 'pump'
            elif 'SPRAY' in first_cell or 'FTSG' in first_cell:
                template_type = 'spray_gun'
        
        # Row 2: Assembly Name
        row2 = [cell.value for cell in ws[2]]
        logger.info(f"Row 2: {row2}")
        
        assembly_name = None
        if row2:
            if len(row2) >= 2 and row2[0] and 'assembly' in str(row2[0]).lower():
                assembly_name = str(row2[1]).strip() if row2[1] else None
            elif row2[0]:
                assembly_name = str(row2[0]).strip()
        
        if not assembly_name:
            assembly_name = file.filename.replace('.xlsx', '').replace('.xls', '').replace('_', ' ')
        
        # Row 3: Description (optional)
        row3 = [cell.value for cell in ws[3]]
        assembly_description = ""
        if row3 and len(row3) >= 2 and row3[0] and 'description' in str(row3[0]).lower():
            assembly_description = str(row3[1]).strip() if row3[1] else ""
        
        # Row 4: Headers
        header_row = [cell.value for cell in ws[4]]
        logger.info(f"Headers: {header_row}")
        
        # Auto-detect template type from headers if not determined from row 1
        if template_type == 'generic':
            template_type = detect_template_type(header_row)
        
        # Set prefix based on template type
        prefix = 'FTP' if template_type == 'pump' else ('FTSG' if template_type == 'spray_gun' else 'PRT')
        
        # Generate assembly short code
        assembly_code = generate_assembly_short_code(assembly_name)
        
        # Map column indices from headers
        col_map = {}
        for idx, header in enumerate(header_row):
            if header:
                header_lower = str(header).lower().strip()
                if 'part' in header_lower and 'name' in header_lower:
                    col_map['part_name'] = idx
                elif header_lower == 'assembly part no' or 'asm part' in header_lower or header_lower == 'part no':
                    col_map['assembly_part_no'] = idx
                elif header_lower == 'category':
                    col_map['category'] = idx
                elif header_lower == 'type':
                    col_map['type'] = idx
                elif header_lower == 'material':
                    col_map['material'] = idx
                elif header_lower == 'size' and 'o-ring' not in header_lower and 'oring' not in header_lower:
                    col_map['size'] = idx
                elif 'o-ring' in header_lower or 'oring' in header_lower:
                    if 'od' in header_lower:
                        col_map['oring_od'] = idx
                    elif 'id' in header_lower:
                        col_map['oring_id'] = idx
                    elif 'thick' in header_lower or 'th' in header_lower:
                        col_map['oring_th'] = idx
                elif header_lower == 'od':
                    col_map['oring_od'] = idx
                elif header_lower == 'id' and 'part' not in header_lower:
                    col_map['oring_id'] = idx
                elif header_lower in ['thickness', 'th']:
                    col_map['oring_th'] = idx
                elif header_lower in ['qty', 'quantity']:
                    col_map['qty'] = idx
                elif header_lower == 'unit':
                    col_map['unit'] = idx
        
        logger.info(f"Column mapping: {col_map}")
        
        if 'part_name' not in col_map:
            raise HTTPException(status_code=400, detail="'Part Name' column not found in headers")
        
        # Get next factory number
        factory_no_start = await get_next_factory_part_no(prefix)
        
        # Parse parts from Row 5 onwards
        parts = []
        current_factory_no = factory_no_start
        assembly_part_counter = 1
        
        for row_num, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
            if not row or not row[col_map.get('part_name', 0)]:
                continue
            
            part_name = str(row[col_map['part_name']]).strip()
            if not part_name:
                continue
            
            # Get assembly part no from Excel or auto-generate
            assembly_part_no = assembly_part_counter
            if 'assembly_part_no' in col_map and row[col_map['assembly_part_no']]:
                try:
                    assembly_part_no = int(row[col_map['assembly_part_no']])
                except:
                    assembly_part_no = assembly_part_counter
            
            category = str(row[col_map.get('category', 1)]).strip() if col_map.get('category') is not None and len(row) > col_map.get('category', 0) and row[col_map.get('category')] else ""
            part_type = str(row[col_map.get('type', 2)]).strip().lower() if col_map.get('type') is not None and len(row) > col_map.get('type', 0) and row[col_map.get('type')] else ""
            
            # Get Material
            material = ""
            if 'material' in col_map and len(row) > col_map['material'] and row[col_map['material']]:
                material = str(row[col_map['material']]).strip()
            
            # Determine size based on part type
            size_details = ""
            if 'oring' in part_type or 'o-ring' in part_type or part_name.lower().startswith('o-ring') or part_name.lower().startswith('oring'):
                # O-Ring - use OD, ID, Thickness
                od = str(row[col_map.get('oring_od', -1)]).strip() if col_map.get('oring_od') is not None and len(row) > col_map.get('oring_od', 0) and row[col_map.get('oring_od')] else ""
                inner_id = str(row[col_map.get('oring_id', -1)]).strip() if col_map.get('oring_id') is not None and len(row) > col_map.get('oring_id', 0) and row[col_map.get('oring_id')] else ""
                th = str(row[col_map.get('oring_th', -1)]).strip() if col_map.get('oring_th') is not None and len(row) > col_map.get('oring_th', 0) and row[col_map.get('oring_th')] else ""
                size_details = format_oring_size(od, inner_id, th)
            else:
                # Regular part - use Size column
                if 'size' in col_map and len(row) > col_map['size'] and row[col_map['size']]:
                    size_details = str(row[col_map['size']]).strip()
            
            qty = 1
            if 'qty' in col_map and len(row) > col_map['qty'] and row[col_map['qty']]:
                try:
                    qty_val = row[col_map['qty']]
                    if isinstance(qty_val, (int, float)):
                        qty = int(qty_val)
                    elif isinstance(qty_val, str):
                        qty_str = qty_val.strip()
                        if '/' in qty_str:
                            qty = 1
                        else:
                            qty = int(float(qty_str))
                except:
                    qty = 1
            
            unit = "PCS"
            if 'unit' in col_map and len(row) > col_map['unit'] and row[col_map['unit']]:
                unit = str(row[col_map['unit']]).strip().upper()
            
            # Generate Common Part ID for linking same parts across assemblies
            common_part_id = generate_common_part_id(part_name, material)
            
            # Check if part with same Common Part ID already exists
            existing_part = None
            if common_part_id:
                existing_part = await db.parts_library.find_one({"common_part_id": common_part_id})
            
            # If not found by common_part_id, try by name
            if not existing_part:
                escaped_name = re.escape(part_name)
                existing_part = await db.parts_library.find_one({
                    "name": {"$regex": f"^{escaped_name}$", "$options": "i"}
                })
            
            # Generate Part ID
            generated_part_id = generate_pump_spray_part_id(
                prefix=prefix,
                assembly_code=assembly_code,
                assembly_part_no=assembly_part_no,
                size_details=size_details,
                factory_no=current_factory_no
            )
            
            parts.append({
                "row": row_num,
                "part_name": part_name,
                "assembly_part_no": assembly_part_no,
                "category": category,
                "part_type": part_type,
                "material": material,
                "material_code": get_material_code(material),
                "common_part_id": common_part_id,
                "size": size_details,
                "oring_od": row[col_map.get('oring_od')] if col_map.get('oring_od') is not None and len(row) > col_map.get('oring_od', 0) else None,
                "oring_id": row[col_map.get('oring_id')] if col_map.get('oring_id') is not None and len(row) > col_map.get('oring_id', 0) else None,
                "oring_th": row[col_map.get('oring_th')] if col_map.get('oring_th') is not None and len(row) > col_map.get('oring_th', 0) else None,
                "qty": qty,
                "unit": unit,
                "factory_no": current_factory_no,
                "status": "exists" if existing_part else "new",
                "existing_id": existing_part["part_id"] if existing_part else None,
                "generated_id": generated_part_id
            })
            
            current_factory_no += 1
            assembly_part_counter += 1
        
        wb.close()
        
        if not parts:
            raise HTTPException(status_code=400, detail="No parts found in Excel file")
        
        return {
            "success": True,
            "template_type": template_type,
            "prefix": prefix,
            "assembly_name": assembly_name,
            "assembly_code": assembly_code,
            "assembly_description": assembly_description,
            "parts": parts,
            "total_parts": len(parts),
            "new_parts": len([p for p in parts if p["status"] == "new"]),
            "existing_parts": len([p for p in parts if p["status"] == "exists"]),
            "next_factory_no": factory_no_start
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Pump/Spray Excel parse error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to parse Excel: {str(e)}")


@app.post("/api/import-pump-spray-excel")
async def import_pump_spray_excel(
    file: UploadFile = File(...),
    conflict_resolutions: str = Form(default="{}"),
    current_user: User = Depends(get_current_user)
):
    """
    Import Pump or Spray Gun parts from Excel with the new Part ID format.
    
    Args:
        file: Excel file to import
        conflict_resolutions: JSON string with user choices for material conflicts
            Format: {"part_name|size": "keep_separate" | "merge", ...}
    """
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Parse conflict resolutions
    try:
        user_conflict_choices = json.loads(conflict_resolutions) if conflict_resolutions else {}
    except:
        user_conflict_choices = {}
    
    # First parse the file
    await file.seek(0)
    contents = await file.read()
    
    # Reparse for import
    try:
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        
        # Row 1: Template Type
        row1 = [cell.value for cell in ws[1]]
        template_type = 'generic'
        if row1 and row1[0]:
            first_cell = str(row1[0]).strip().upper()
            if 'PUMP' in first_cell or 'FTP' in first_cell:
                template_type = 'pump'
            elif 'SPRAY' in first_cell or 'FTSG' in first_cell:
                template_type = 'spray_gun'
        
        # Row 2: Assembly Name - could be in different positions
        row2 = [cell.value for cell in ws[2]]
        assembly_name = None
        if row2:
            if len(row2) >= 2 and row2[0] and 'assembly' in str(row2[0]).lower():
                assembly_name = str(row2[1]).strip() if row2[1] else None
            elif len(row2) >= 2 and row2[1]:
                # Check if row2 cell A looks like a label
                cell_a = str(row2[0]).lower() if row2[0] else ""
                if 'name' in cell_a or 'assembly' in cell_a or cell_a.endswith(':'):
                    assembly_name = str(row2[1]).strip()
                else:
                    assembly_name = str(row2[0]).strip()
            elif row2[0]:
                assembly_name = str(row2[0]).strip()
        
        if not assembly_name:
            # Try Row 1 column B
            if row1 and len(row1) >= 2 and row1[1]:
                assembly_name = str(row1[1]).strip()
            elif row1 and row1[0] and 'assembly' not in str(row1[0]).lower():
                assembly_name = str(row1[0]).strip()
        
        if not assembly_name:
            assembly_name = file.filename.replace('.xlsx', '').replace('.xls', '').replace('_', ' ')
        
        # Row 3: Description (or could be Row 2 if no separate description row)
        row3 = [cell.value for cell in ws[3]]
        assembly_description = ""
        if row3 and len(row3) >= 2 and row3[0] and 'description' in str(row3[0]).lower():
            assembly_description = str(row3[1]).strip() if row3[1] else ""
        
        # Find header row - it could be Row 3 or Row 4
        # Check Row 3 first for headers
        header_row = None
        header_row_num = 3
        row3_check = [cell.value for cell in ws[3]]
        if row3_check and any(h and 'part' in str(h).lower() and 'name' in str(h).lower() for h in row3_check):
            header_row = row3_check
            header_row_num = 3
        else:
            # Try Row 4
            row4_check = [cell.value for cell in ws[4]]
            if row4_check and any(h and 'part' in str(h).lower() and 'name' in str(h).lower() for h in row4_check):
                header_row = row4_check
                header_row_num = 4
        
        if not header_row:
            raise HTTPException(status_code=400, detail="Could not find header row with 'Part Name' column")
        
        data_start_row = header_row_num + 1
        
        logger.info(f"import-pump-spray-excel: Found headers at row {header_row_num}: {header_row}")
        
        if template_type == 'generic':
            template_type = detect_template_type(header_row)
        
        prefix = 'FTP' if template_type == 'pump' else ('FTSG' if template_type == 'spray_gun' else 'PRT')
        assembly_code = generate_assembly_short_code(assembly_name)
        
        # Create or find assembly
        escaped_asm_name = re.escape(assembly_name)
        existing_assembly = await db.assemblies.find_one({"name": {"$regex": f"^{escaped_asm_name}$", "$options": "i"}})
        
        if existing_assembly:
            assembly_id = existing_assembly["assembly_id"]
            assembly_created = False
        else:
            assembly_id = await get_next_assembly_id()
            await db.assemblies.insert_one({
                "assembly_id": assembly_id,
                "name": assembly_name,
                "description": assembly_description,
                "template_type": template_type,
                "prefix": prefix,
                "created_at": get_ist_now(),
                "created_by": current_user.user_id,
                "updated_at": get_ist_now()
            })
            assembly_created = True
        
        # Map column indices - Enhanced for new unified template
        col_map = {}
        for idx, header in enumerate(header_row):
            if header:
                header_lower = str(header).lower().strip()
                # Part Name - required
                if ('part' in header_lower and 'name' in header_lower) or header_lower == 'name':
                    col_map['part_name'] = idx
                # Part No
                elif header_lower == 'part no' or header_lower == 'partno' or 'part no' in header_lower:
                    col_map['part_no'] = idx
                elif header_lower == 'assembly part no' or 'asm part' in header_lower:
                    col_map['assembly_part_no'] = idx
                # Material - new field
                elif header_lower == 'material':
                    col_map['material'] = idx
                # Category
                elif header_lower == 'category':
                    col_map['category'] = idx
                # Type
                elif header_lower == 'type':
                    col_map['type'] = idx
                # Size / Dia/ID/WxB
                elif 'dia' in header_lower or 'wxb' in header_lower or header_lower == 'size' or header_lower == 'diameter':
                    col_map['size'] = idx
                # Length / Thickness
                elif 'length' in header_lower or 'thick' in header_lower:
                    col_map['length'] = idx
                # O-Ring specific
                elif 'o-ring' in header_lower or 'oring' in header_lower:
                    if 'od' in header_lower:
                        col_map['oring_od'] = idx
                    elif 'id' in header_lower:
                        col_map['oring_id'] = idx
                    elif 'thick' in header_lower:
                        col_map['oring_th'] = idx
                elif header_lower == 'od':
                    col_map['oring_od'] = idx
                elif header_lower == 'id':
                    col_map['oring_id'] = idx
                elif header_lower in ['thickness', 'th']:
                    col_map['oring_th'] = idx
                # Qty
                elif header_lower in ['qty', 'quantity', 'qty.']:
                    col_map['qty'] = idx
                # Unit
                elif header_lower == 'unit':
                    col_map['unit'] = idx
                # Remarks
                elif 'remark' in header_lower:
                    col_map['remarks'] = idx
        
        logger.info(f"import-pump-spray-excel column mapping: {col_map}")
        
        if 'part_name' not in col_map:
            raise HTTPException(status_code=400, detail="'Part Name' column not found")
        
        factory_no_start = await get_next_factory_part_no(prefix)
        current_factory_no = factory_no_start
        assembly_part_counter = 1
        
        parts_created = []
        parts_linked = []
        parts_merged = []  # Track parts merged due to same Name+Size+Material
        parts_skipped = []
        processed_parts_in_import = {}  # Track unique_key_full -> part_id for exact duplicate detection
        processed_partial_keys = {}  # Track unique_key_partial (name+size) -> part_id for conflict resolution
        
        for row_num, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            if not row or not row[col_map.get('part_name', 0)]:
                continue
            
            part_name = str(row[col_map['part_name']]).strip()
            if not part_name:
                continue
            
            assembly_part_no = assembly_part_counter
            if 'assembly_part_no' in col_map and row[col_map['assembly_part_no']]:
                try:
                    assembly_part_no = int(row[col_map['assembly_part_no']])
                except:
                    assembly_part_no = assembly_part_counter
            
            category = str(row[col_map.get('category', 1)]).strip() if col_map.get('category') is not None and len(row) > col_map.get('category', 0) and row[col_map.get('category')] else ""
            part_type = str(row[col_map.get('type', 2)]).strip().lower() if col_map.get('type') is not None and len(row) > col_map.get('type', 0) and row[col_map.get('type')] else ""
            
            # Determine size
            size_details = ""
            oring_od = None
            oring_id = None
            oring_th = None
            
            if 'oring' in part_type or 'o-ring' in part_type or part_name.lower().startswith('o-ring'):
                od = str(row[col_map.get('oring_od')]).strip() if col_map.get('oring_od') is not None and len(row) > col_map.get('oring_od', 0) and row[col_map.get('oring_od')] else ""
                inner_id = str(row[col_map.get('oring_id')]).strip() if col_map.get('oring_id') is not None and len(row) > col_map.get('oring_id', 0) and row[col_map.get('oring_id')] else ""
                th = str(row[col_map.get('oring_th')]).strip() if col_map.get('oring_th') is not None and len(row) > col_map.get('oring_th', 0) and row[col_map.get('oring_th')] else ""
                size_details = format_oring_size(od, inner_id, th)
                oring_od = od if od and od != 'None' else None
                oring_id = inner_id if inner_id and inner_id != 'None' else None
                oring_th = th if th and th != 'None' else None
            else:
                if 'size' in col_map and len(row) > col_map['size'] and row[col_map['size']]:
                    size_details = str(row[col_map['size']]).strip()
            
            qty = 1
            if 'qty' in col_map and len(row) > col_map['qty'] and row[col_map['qty']]:
                try:
                    qty_val = row[col_map['qty']]
                    if isinstance(qty_val, (int, float)):
                        qty = int(qty_val)
                    elif isinstance(qty_val, str):
                        qty_str = qty_val.strip()
                        if '/' in qty_str:
                            qty = 1
                        else:
                            qty = int(float(qty_str))
                except:
                    qty = 1
            
            unit = "PCS"
            if 'unit' in col_map and len(row) > col_map['unit'] and row[col_map['unit']]:
                unit = str(row[col_map['unit']]).strip().upper()
            
            # Get Material
            material = ""
            if 'material' in col_map and len(row) > col_map['material'] and row[col_map['material']]:
                material = str(row[col_map['material']]).strip()
            
            # Get Length/Thickness
            length = ""
            if 'length' in col_map and len(row) > col_map['length'] and row[col_map['length']]:
                length = str(row[col_map['length']]).strip()
            
            # Get Part No
            part_no = ""
            if 'part_no' in col_map and len(row) > col_map['part_no'] and row[col_map['part_no']]:
                part_no = str(row[col_map['part_no']]).strip()
            
            # Get Remarks
            remarks = ""
            if 'remarks' in col_map and len(row) > col_map['remarks'] and row[col_map['remarks']]:
                remarks = str(row[col_map['remarks']]).strip()
            
            # Generate Common Part ID for linking same parts across assemblies
            common_part_id = generate_common_part_id(part_name, material)
            
            # Normalize values for duplicate detection
            normalized_name = part_name.lower().strip()
            normalized_size = str(size_details).strip() if size_details else ""
            normalized_material = material.lower().strip() if material else ""
            
            # Create unique key: Name + Size + Material (full key for exact match)
            unique_key_full = f"{normalized_name}|{normalized_size}|{normalized_material}"
            # Partial key for conflict detection (Name + Size only)
            unique_key_partial = f"{normalized_name}|{normalized_size}"
            
            # Check user's conflict resolution choice for this Name+Size combo
            # Key format in conflict_resolutions can vary:
            # - "O RING SEAL|15" (original from parse endpoint)
            # - "o ring seal|15" (lowercase)
            # - "o_ring_seal__15" (web app normalized format with underscores)
            # We need to check all possible formats
            
            def normalize_conflict_key(name, size):
                """Generate all possible key formats for matching"""
                keys = []
                name_lower = name.lower().strip()
                size_str = str(size).strip() if size else ""
                # Format 1: lowercase with pipe
                keys.append(f"{name_lower}|{size_str}")
                # Format 2: original case with pipe
                keys.append(f"{name}|{size_str}")
                # Format 3: underscores instead of spaces, double underscore as separator
                name_underscore = name_lower.replace(" ", "_")
                size_underscore = size_str.replace(".", "_").replace(" ", "_").replace("x", "x")
                keys.append(f"{name_underscore}__{size_underscore}")
                # Format 4: with 'x' in size normalized
                keys.append(f"{name_underscore}__{size_str}")
                return keys
            
            possible_keys = normalize_conflict_key(part_name, size_details)
            user_choice = None
            for pk in possible_keys:
                if pk in user_conflict_choices:
                    user_choice = user_conflict_choices[pk]
                    break
            # Also try lowercase keys from user choices
            if not user_choice:
                for ck, cv in user_conflict_choices.items():
                    ck_normalized = ck.lower().replace("_", " ").replace("__", "|")
                    for pk in possible_keys:
                        if ck_normalized == pk.lower() or ck.lower() == pk.lower():
                            user_choice = cv
                            break
                    if user_choice:
                        break
            
            # Check if we've already processed this exact part (same Name + Size + Material)
            if unique_key_full in processed_parts_in_import:
                # Exact duplicate - merge by adding quantity
                existing_part_id = processed_parts_in_import[unique_key_full]
                await db.assembly_parts.update_one(
                    {"assembly_id": assembly_id, "part_id": existing_part_id},
                    {"$inc": {"quantity": qty}},
                    upsert=False
                )
                parts_merged.append({
                    "part_id": existing_part_id,
                    "name": part_name,
                    "merged_qty": qty,
                    "reason": "same_name_size_material"
                })
                continue
            
            # Check if user wants to MERGE parts with same Name+Size but different Material
            if user_choice == "merge" and unique_key_partial in processed_partial_keys:
                # User chose to merge despite different materials
                existing_part_id = processed_partial_keys[unique_key_partial]
                await db.assembly_parts.update_one(
                    {"assembly_id": assembly_id, "part_id": existing_part_id},
                    {"$inc": {"quantity": qty}},
                    upsert=False
                )
                parts_merged.append({
                    "part_id": existing_part_id,
                    "name": part_name,
                    "merged_qty": qty,
                    "reason": "user_chose_merge_different_materials"
                })
                continue
            
            # Check if part exists in global parts library (for cross-assembly reuse)
            # Match by Name + Size + Material (exact)
            existing_part = await db.parts_library.find_one({
                "name": {"$regex": f"^{re.escape(part_name)}$", "$options": "i"},
                "size": normalized_size,
                "material": {"$regex": f"^{re.escape(material)}$", "$options": "i"} if material else {"$in": ["", None]}
            })
            
            if existing_part:
                # Link existing part to assembly (Where Used tracking)
                await db.assembly_parts.update_one(
                    {"assembly_id": assembly_id, "part_id": existing_part["part_id"]},
                    {"$set": {
                        "quantity": qty,
                        "assembly_part_no": assembly_part_no,
                        "updated_at": get_ist_now()
                    }},
                    upsert=True
                )
                
                # Update part's common_part_id if not set
                if not existing_part.get("common_part_id") and common_part_id:
                    await db.parts_library.update_one(
                        {"part_id": existing_part["part_id"]},
                        {"$set": {"common_part_id": common_part_id, "material": material}}
                    )
                
                # Track this part as processed for duplicate detection
                processed_parts_in_import[unique_key_full] = existing_part["part_id"]
                processed_partial_keys[unique_key_partial] = existing_part["part_id"]  # Track for conflict resolution
                
                parts_linked.append({
                    "part_id": existing_part["part_id"],
                    "name": part_name,
                    "assembly_part_no": assembly_part_no,
                    "common_part_id": common_part_id,
                    "linked_to_assembly": assembly_name,
                    "reused_from": existing_part.get("source_assembly_name", "Parts Library"),
                    "status": "linked_existing"
                })
            else:
                # Create new part with new Part ID format
                part_id = generate_pump_spray_part_id(
                    prefix=prefix,
                    assembly_code=assembly_code,
                    assembly_part_no=assembly_part_no,
                    size_details=size_details,
                    factory_no=current_factory_no
                )
                
                await db.parts_library.insert_one({
                    "part_id": part_id,
                    "part_no": part_no,
                    "name": part_name,
                    "category": category,
                    "type": part_type,
                    "material": material,
                    "material_code": get_material_code(material),
                    "common_part_id": common_part_id,
                    "size": size_details,
                    "length": length,
                    "oring_od": oring_od,
                    "oring_id": oring_id,
                    "oring_thickness": oring_th,
                    "unit": unit,
                    "remarks": remarks,
                    "stock": 0,
                    "min_stock": 0,
                    "reserved_stock": 0,
                    "assembly_part_no": assembly_part_no,
                    "factory_no": current_factory_no,
                    "template_type": template_type,
                    "source_assembly": assembly_code,
                    "source_assembly_name": assembly_name,
                    "locations": [],
                    "image": "",
                    "drawing": "",
                    "created_at": get_ist_now(),
                    "created_by": current_user.user_id,
                    "updated_at": get_ist_now()
                })
                
                # Link to assembly
                await db.assembly_parts.insert_one({
                    "assembly_id": assembly_id,
                    "part_id": part_id,
                    "quantity": qty,
                    "assembly_part_no": assembly_part_no,
                    "created_at": get_ist_now()
                })
                
                parts_created.append({
                    "part_id": part_id,
                    "name": part_name,
                    "assembly_part_no": assembly_part_no,
                    "size": size_details,
                    "material": material,
                    "factory_no": current_factory_no,
                    "status": "created_new"
                })
                
                # Track this part as processed for duplicate detection
                processed_parts_in_import[unique_key_full] = part_id
                processed_partial_keys[unique_key_partial] = part_id  # Track for conflict resolution
                
                current_factory_no += 1
            
            assembly_part_counter += 1
        
        wb.close()
        
        return {
            "success": True,
            "template_type": template_type,
            "prefix": prefix,
            "assembly": {
                "assembly_id": assembly_id,
                "name": assembly_name,
                "description": assembly_description,
                "created": assembly_created
            },
            "summary": {
                "parts_created": len(parts_created),
                "parts_linked": len(parts_linked),
                "parts_merged": len(parts_merged),
                "total_unique_parts": len(parts_created) + len(parts_linked),
                "total_rows_processed": len(parts_created) + len(parts_linked) + len(parts_merged)
            },
            "parts_created": parts_created,
            "parts_linked": parts_linked,
            "parts_merged": parts_merged,
            "message": f"Successfully imported {len(parts_created)} new parts, linked {len(parts_linked)} existing parts" + (f", merged {len(parts_merged)} duplicate rows" if parts_merged else "")
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Pump/Spray import error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")


@app.get("/api/templates/bom-excel")
async def download_bom_template():
    """Download unified BOM Excel template - 9 columns matching user format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Parts"
    
    # Styling
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Row 1: Assembly Name
    ws['A1'] = "Assembly Name:"
    ws['A1'].font = Font(bold=True)
    ws['B1'] = "WS04 SPRAY"
    ws.merge_cells('B1:I1')
    
    # Row 2: Description
    ws['A2'] = "Description:"
    ws['A2'].font = Font(bold=True)
    ws['B2'] = "Spray Gun Assembly"
    ws.merge_cells('B2:I2')
    
    # Row 3: Headers (9 columns - matching screenshot)
    headers = [
        "Part No", "Part Name", "Material", "Category", 
        "Dia/ID/WxB", "Length/Thickness", "Qty", "Unit", "Remarks"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Sample data - matching user's screenshot exactly
    sample_data = [
        ["PART NO - 01", "Main Body", "SS 316", "Machined Part", "", "", 1, "PCS", ""],
        ["PART NO - 02", "Nozzle Insert", "SS 316", "Machined Part", "8", "25", 1, "PCS", ""],
        ["PART NO - 03", "Front Cap", "SS 316", "Cap / Cover", "", "", 1, "PCS", ""],
        ["PART NO - 04", "Outer Sleeve", "SS 316", "Housing", "12", "40", 1, "PCS", ""],
        ["PART NO - 05", "Retainer Ring", "SS 316", "Ring", "15", "", 1, "PCS", ""],
        ["PART NO - 06", "Needle Rod", "SS 316", "Shaft / Rod", "6", "50", 1, "PCS", ""],
        ["PART NO - 07", "Spring Guide", "SS 316", "Guide", "", "", 1, "PCS", ""],
        ["PART NO - 08", "Compression Spring", "SS 316", "Spring", "8", "15", 1, "PCS", ""],
        ["PART NO - 09", "Back Cap", "SS 316", "Cap / Cover", "", "", 1, "PCS", ""],
        ["PART NO - 10", "Washer", "SS 316", "Fastener", "M4", "", 2, "PCS", ""],
        ["PART NO - 11", "Hex Screw", "SS 316", "Fastener", "M4", "10", 4, "PCS", ""],
        ["PART NO - 15", "Connector Nipple", "SS 316", "Fitting", "1/4", "", 1, "PCS", ""],
        ["PART NO - 17", "Nozzle Holder", "SS 316", "Machined Part", "", "", 1, "PCS", ""],
        ["PART NO - 18", "Ball End Nozzle", "SS 316", "Nozzle", "2", "", 1, "PCS", ""],
        ["PART NO - 21", "O-Ring Small", "Nitrile", "O-Ring", "10", "1.5", 2, "PCS", ""],
        ["PART NO - 22", "O-Ring Medium", "Nitrile", "O-Ring", "12", "2", 1, "PCS", ""],
        ["PART NO - 23", "O-Ring Type A", "Nitrile", "O-Ring", "15", "1.5", 1, "PCS", ""],
        ["PART NO - 24", "O-Ring Type B", "Nitrile", "O-Ring", "18", "2", 1, "PCS", ""],
        ["PART NO - 25", "O-Ring Large", "Nitrile", "O-Ring", "20", "2.5", 1, "PCS", ""],
        ["PART NO - 26", "Back O-Ring", "Nitrile", "O-Ring", "25", "3", 1, "PCS", ""],
        ["PART NO - 28", "Air Connector", "SS 316", "Fitting", "1/8", "", 1, "PCS", ""],
        ["PART NO - 30", "Square Block", "SS 316", "Machined Part", "20x15", "30", 1, "PCS", ""],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Adjust column widths
    column_widths = [14, 22, 10, 14, 14, 16, 5, 5, 12]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    
    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        "BOM PARTS TEMPLATE INSTRUCTIONS",
        "",
        "Row 1: Enter your Assembly Name in column B",
        "Row 2: Enter description in column B (optional)",
        "Row 3: Headers (do not modify)",
        "Row 4+: Enter your parts data",
        "",
        "COLUMNS (9 total):",
        "- Part No: PART NO - 01, PART NO - 02, etc.",
        "- Part Name: Name of the part",
        "- Material: SS 316, SS 304, Nitrile, etc.",
        "- Category: Machined Part, Fastener, O-Ring, Fitting, etc.",
        "- Dia/ID/WxB: Diameter, O-Ring ID, or Width x Breadth",
        "- Length/Thickness: Length for parts, Thickness for O-Rings",
        "- Qty: Quantity",
        "- Unit: PCS, SET, MTR, etc.",
        "- Remarks: Additional notes",
        "",
        "SIZE EXAMPLES:",
        "  O-Ring: Dia/ID/WxB = 10, Length/Thickness = 1.5 (ID x Thick)",
        "  Rod: Dia/ID/WxB = 8, Length/Thickness = 50",
        "  Fitting: Dia/ID/WxB = 1/4 or 1/8",
        "  Square: Dia/ID/WxB = 20x15, Length/Thickness = 30",
        "  Bolt: Dia/ID/WxB = M4, Length/Thickness = 10",
        "",
        "GENERATED PART ID FORMAT:",
        "  {Assembly}-{PartNo}-F{FactoryNo}-{TypeCode}",
        "  Example: WS04-01-F0001-SS",
        "",
        "TYPE CODES:",
        "  SS = Stainless Steel parts",
        "  SEAL = O-Rings, Seals (Nitrile, Viton, etc.)",
        "  BOLT = Bolts, Screws",
        "  WASHER = Washers",
        "  PART = Other parts"
    ]
    for idx, line in enumerate(instructions, 1):
        ws2.cell(row=idx, column=1, value=line)
    ws2.column_dimensions['A'].width = 55
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bom_parts_template.xlsx"}
    )


# Keep old endpoints for backward compatibility (redirect to new template)
@app.get("/api/templates/pump-excel")
async def download_pump_template():
    """Redirect to unified BOM template"""
    return await download_bom_template()


@app.get("/api/templates/spray-gun-excel")
async def download_spray_gun_template():
    """Redirect to unified BOM template"""
    return await download_bom_template()


# ========================================
# ASSEMBLY MODULE
# ========================================

class AssemblyCreate(BaseModel):
    name: str
    description: Optional[str] = None

class AssemblyPartAdd(BaseModel):
    part_id: str
    quantity: int


@app.get("/api/assemblies")
async def get_assemblies(
    search: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all assemblies"""
    query = {}
    if search:
        query["$or"] = [
            {"assembly_id": {"$regex": search, "$options": "i"}},
            {"name": {"$regex": search, "$options": "i"}}
        ]
    
    assemblies = await db.assemblies.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=200)
    
    # Get part count for each assembly
    for asm in assemblies:
        parts_count = await db.assembly_parts.count_documents({"assembly_id": asm["assembly_id"]})
        asm["parts_count"] = parts_count
    
    return assemblies


@app.post("/api/assemblies")
async def create_assembly(assembly: AssemblyCreate, current_user: User = Depends(get_current_user)):
    """Create a new assembly"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    now = get_ist_now()
    
    # Generate assembly ID
    count = await db.assemblies.count_documents({})
    assembly_id = f"ASM-{count + 1:04d}"
    
    assembly_doc = {
        "assembly_id": assembly_id,
        "name": assembly.name,
        "description": assembly.description or "",
        "created_at": now,
        "updated_at": now,
        "created_by": current_user.user_id
    }
    
    await db.assemblies.insert_one(assembly_doc)
    
    return {"message": "Assembly created successfully", "assembly_id": assembly_id}


@app.get("/api/assemblies/{assembly_id}")
async def get_assembly_detail(assembly_id: str, current_user: User = Depends(get_current_user)):
    """Get assembly with BOM (parts list)"""
    assembly = await db.assemblies.find_one({"assembly_id": assembly_id}, {"_id": 0})
    if not assembly:
        raise HTTPException(status_code=404, detail="Assembly not found")
    
    # Get parts in this assembly
    assembly_parts = await db.assembly_parts.find(
        {"assembly_id": assembly_id}, {"_id": 0}
    ).to_list(length=100)
    
    # Enrich with part details including image and drawing
    for ap in assembly_parts:
        part = await db.parts_library.find_one({"part_id": ap["part_id"]}, {"_id": 0})
        if part:
            ap["part_name"] = part.get("name", "")
            ap["category"] = part.get("category", "")
            ap["part_type"] = part.get("part_type", "")
            ap["size"] = part.get("size", "")
            ap["length"] = part.get("length", "")
            ap["unit"] = part.get("unit", "PCS")
            ap["stock"] = part.get("stock", 0)
            ap["available_stock"] = part.get("stock", 0) - part.get("reserved_stock", 0)
            ap["locations"] = part.get("locations", [])
            ap["image"] = part.get("image", "")
            ap["drawing"] = part.get("drawing", "")
    
    assembly["parts"] = assembly_parts
    
    return assembly


@app.put("/api/assemblies/{assembly_id}")
async def update_assembly(
    assembly_id: str,
    name: Optional[str] = None,
    description: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Update assembly details"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    assembly = await db.assemblies.find_one({"assembly_id": assembly_id})
    if not assembly:
        raise HTTPException(status_code=404, detail="Assembly not found")
    
    update_data = {"updated_at": get_ist_now()}
    if name:
        update_data["name"] = name
    if description is not None:
        update_data["description"] = description
    
    await db.assemblies.update_one({"assembly_id": assembly_id}, {"$set": update_data})
    
    return {"message": "Assembly updated successfully"}


@app.delete("/api/assemblies/{assembly_id}")
async def delete_assembly(
    assembly_id: str, 
    delete_parts: bool = True,
    current_user: User = Depends(get_current_user)
):
    """Delete an assembly and optionally all its parts from Part Library"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get assembly first
    assembly = await db.assemblies.find_one({"assembly_id": assembly_id})
    if not assembly:
        raise HTTPException(status_code=404, detail="Assembly not found")
    
    # Get all parts linked to this assembly
    assembly_parts = await db.assembly_parts.find({"assembly_id": assembly_id}).to_list(length=1000)
    part_ids = [ap["part_id"] for ap in assembly_parts]
    
    parts_deleted = 0
    if delete_parts and part_ids:
        # Delete all parts that belong ONLY to this assembly
        # (Check if part is used in other assemblies before deleting)
        for part_id in part_ids:
            # Check if part is used in other assemblies
            other_usage = await db.assembly_parts.find_one({
                "part_id": part_id,
                "assembly_id": {"$ne": assembly_id}
            })
            
            if not other_usage:
                # Part is only used in this assembly, safe to delete
                await db.parts_library.delete_one({"part_id": part_id})
                parts_deleted += 1
    
    # Delete assembly parts links
    await db.assembly_parts.delete_many({"assembly_id": assembly_id})
    
    # Delete assembly
    await db.assemblies.delete_one({"assembly_id": assembly_id})
    
    return {
        "message": "Assembly deleted successfully",
        "assembly_id": assembly_id,
        "parts_deleted": parts_deleted,
        "parts_kept": len(part_ids) - parts_deleted
    }


@app.post("/api/assemblies/{assembly_id}/parts")
async def add_part_to_assembly(
    assembly_id: str,
    part_data: AssemblyPartAdd,
    current_user: User = Depends(get_current_user)
):
    """Add a part to assembly BOM"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Verify assembly exists
    assembly = await db.assemblies.find_one({"assembly_id": assembly_id})
    if not assembly:
        raise HTTPException(status_code=404, detail="Assembly not found")
    
    # Verify part exists
    part = await db.parts_library.find_one({"part_id": part_data.part_id})
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    
    # Check if part already in assembly
    existing = await db.assembly_parts.find_one({
        "assembly_id": assembly_id,
        "part_id": part_data.part_id
    })
    
    if existing:
        # Update quantity
        await db.assembly_parts.update_one(
            {"assembly_id": assembly_id, "part_id": part_data.part_id},
            {"$set": {"quantity": part_data.quantity}}
        )
        return {"message": "Part quantity updated"}
    else:
        # Add new
        await db.assembly_parts.insert_one({
            "assembly_id": assembly_id,
            "part_id": part_data.part_id,
            "quantity": part_data.quantity,
            "added_at": get_ist_now()
        })
        return {"message": "Part added to assembly"}


@app.delete("/api/assemblies/{assembly_id}/parts/{part_id}")
async def remove_part_from_assembly(
    assembly_id: str,
    part_id: str,
    current_user: User = Depends(get_current_user)
):
    """Remove a part from assembly BOM"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await db.assembly_parts.delete_one({
        "assembly_id": assembly_id,
        "part_id": part_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Part not found in assembly")
    
    return {"message": "Part removed from assembly"}


@app.post("/api/assemblies/{assembly_id}/parts/search-similar")
async def search_similar_parts(
    assembly_id: str,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Search for similar existing parts before adding new one"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    data = await request.json()
    name = data.get("name", "").strip().upper()
    size = data.get("size", "").strip()
    length = data.get("length", "").strip()
    category = data.get("category", "").strip()
    
    if not name:
        return {"similar_parts": []}
    
    # Build search query - look for parts with similar name
    query = {
        "$or": [
            {"name": {"$regex": name, "$options": "i"}},
            {"name": name}
        ]
    }
    
    # If size provided, narrow down further
    if size:
        query["size"] = size
    
    # If length provided (for pipes), include in search
    if length:
        query["length"] = length
    
    similar_parts = await db.parts_library.find(query, {"_id": 0}).to_list(length=20)
    
    # Also search with exact match on name + size + length for perfect duplicates
    exact_query = {"name": {"$regex": f"^{name}$", "$options": "i"}}
    if size:
        exact_query["size"] = size
    if length:
        exact_query["length"] = length
    
    exact_matches = await db.parts_library.find(exact_query, {"_id": 0}).to_list(length=10)
    
    # Mark exact matches
    exact_part_ids = {p["part_id"] for p in exact_matches}
    for part in similar_parts:
        part["is_exact_match"] = part["part_id"] in exact_part_ids
        # Get which assemblies use this part
        usages = await db.assembly_parts.find({"part_id": part["part_id"]}).to_list(length=10)
        assembly_names = []
        for u in usages:
            asm = await db.assemblies.find_one({"assembly_id": u["assembly_id"]}, {"name": 1})
            if asm:
                assembly_names.append(asm.get("name", u["assembly_id"]))
        part["used_in_assemblies"] = assembly_names
    
    return {"similar_parts": similar_parts}


@app.post("/api/assemblies/{assembly_id}/parts/create-and-add")
async def create_part_and_add_to_assembly(
    assembly_id: str,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Create a new part with auto-generated Part ID and add to assembly BOM"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Verify assembly exists
    assembly = await db.assemblies.find_one({"assembly_id": assembly_id})
    if not assembly:
        raise HTTPException(status_code=404, detail="Assembly not found")
    
    data = await request.json()
    name = data.get("name", "").strip()
    category = data.get("category", "")
    part_type = data.get("part_type", "")  # Material type (ss, seal, bolt, etc.)
    size = data.get("size", "")  # Dia/OD x Thickness
    length = data.get("length", "")  # Length (especially for pipes)
    quantity = data.get("quantity", 1)
    unit = data.get("unit", "PCS")
    remarks = data.get("remarks", "")
    
    if not name:
        raise HTTPException(status_code=400, detail="Part name is required")
    
    # Generate globally unique Part ID (F0001, F0002, ...)
    # Find the highest F-number across all parts
    last_part = await db.parts_library.find(
        {"part_id": {"$regex": "^F\\d+"}}
    ).sort("part_id", -1).limit(1).to_list(length=1)
    
    if last_part:
        # Extract number from last part_id (e.g., "F0045-SS" -> 45)
        import re
        match = re.search(r'F(\d+)', last_part[0]["part_id"])
        if match:
            next_num = int(match.group(1)) + 1
        else:
            next_num = 1
    else:
        next_num = 1
    
    # Determine suffix based on part_type
    suffix_map = {
        "ss": "SS",
        "stainless": "SS",
        "seal": "OR",
        "o-ring": "OR",
        "oring": "OR",
        "bolt": "BT",
        "nut": "NT",
        "spring": "SP",
        "pipe": "PP",
        "bush": "BH",
        "pin": "PN",
    }
    
    part_type_lower = part_type.lower() if part_type else ""
    suffix = suffix_map.get(part_type_lower, "SS")
    
    # Also check category for suffix
    category_lower = category.lower() if category else ""
    if "o-ring" in category_lower or "oring" in category_lower:
        suffix = "OR"
    elif "spring" in category_lower:
        suffix = "SP"
    elif "pipe" in category_lower:
        suffix = "PP"
    elif "bush" in category_lower:
        suffix = "BH"
    elif "pin" in category_lower:
        suffix = "PN"
    
    new_part_id = f"F{str(next_num).zfill(4)}-{suffix}"
    
    # Create the part in parts_library
    new_part = {
        "part_id": new_part_id,
        "name": name.upper(),
        "category": category,
        "part_type": part_type,
        "size": size,
        "length": length,
        "unit": unit,
        "stock": 0,
        "reserved_stock": 0,
        "remarks": remarks,
        "created_at": get_ist_now(),
        "created_by": current_user.user_id,
        "manually_added": True
    }
    
    await db.parts_library.insert_one(new_part)
    
    # Add to assembly
    await db.assembly_parts.insert_one({
        "assembly_id": assembly_id,
        "part_id": new_part_id,
        "quantity": quantity,
        "added_at": get_ist_now(),
        "added_by": current_user.user_id
    })
    
    logger.info(f"Created new part {new_part_id} and added to assembly {assembly_id}")
    
    return {
        "message": "Part created and added to assembly",
        "part_id": new_part_id,
        "part": new_part
    }


@app.put("/api/assemblies/{assembly_id}/parts/{part_id}")
async def update_part_in_assembly(
    assembly_id: str,
    part_id: str,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Update part details and quantity in assembly BOM"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    data = await request.json()
    
    # Update part details in parts_library
    part_updates = {}
    if "name" in data:
        part_updates["name"] = data["name"].strip().upper()
    if "category" in data:
        part_updates["category"] = data["category"]
    if "part_type" in data:
        part_updates["part_type"] = data["part_type"]
    if "size" in data:
        part_updates["size"] = data["size"]
    if "length" in data:
        part_updates["length"] = data["length"]
    if "unit" in data:
        part_updates["unit"] = data["unit"]
    if "remarks" in data:
        part_updates["remarks"] = data["remarks"]
    
    if part_updates:
        part_updates["updated_at"] = get_ist_now()
        part_updates["updated_by"] = current_user.user_id
        await db.parts_library.update_one(
            {"part_id": part_id},
            {"$set": part_updates}
        )
    
    # Update quantity in assembly_parts
    if "quantity" in data:
        await db.assembly_parts.update_one(
            {"assembly_id": assembly_id, "part_id": part_id},
            {"$set": {"quantity": data["quantity"], "updated_at": get_ist_now()}}
        )
    
    return {"message": "Part updated successfully"}


@app.post("/api/assemblies/{assembly_id}/parts/link-existing")
async def link_existing_part_to_assembly(
    assembly_id: str,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Link an existing part from parts library to this assembly"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Verify assembly exists
    assembly = await db.assemblies.find_one({"assembly_id": assembly_id})
    if not assembly:
        raise HTTPException(status_code=404, detail="Assembly not found")
    
    data = await request.json()
    part_id = data.get("part_id")
    quantity = data.get("quantity", 1)
    
    if not part_id:
        raise HTTPException(status_code=400, detail="part_id is required")
    
    # Verify part exists
    part = await db.parts_library.find_one({"part_id": part_id})
    if not part:
        raise HTTPException(status_code=404, detail="Part not found in library")
    
    # Check if already linked
    existing = await db.assembly_parts.find_one({
        "assembly_id": assembly_id,
        "part_id": part_id
    })
    
    if existing:
        # Update quantity
        await db.assembly_parts.update_one(
            {"assembly_id": assembly_id, "part_id": part_id},
            {"$set": {"quantity": quantity, "updated_at": get_ist_now()}}
        )
        return {"message": "Part quantity updated", "action": "updated"}
    else:
        # Add link
        await db.assembly_parts.insert_one({
            "assembly_id": assembly_id,
            "part_id": part_id,
            "quantity": quantity,
            "added_at": get_ist_now(),
            "added_by": current_user.user_id
        })
        return {"message": "Part linked to assembly", "action": "linked"}


# ========================================
# MANAGE ASSEMBLIES - ADMIN SECTION
# ========================================

@app.get("/api/manage/assemblies")
async def get_all_assemblies_with_stats(current_user: User = Depends(get_current_user)):
    """Get all assemblies with part counts and linked parts info"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    assemblies = await db.assemblies.find({}).to_list(length=500)
    
    result = []
    for asm in assemblies:
        assembly_id = asm.get("assembly_id")
        
        # Get parts for this assembly
        parts = await db.assembly_parts.find({"assembly_id": assembly_id}).to_list(length=500)
        part_ids = [p["part_id"] for p in parts]
        
        # Count how many parts are linked to other assemblies
        linked_count = 0
        for part_id in part_ids:
            other_usage = await db.assembly_parts.count_documents({
                "part_id": part_id,
                "assembly_id": {"$ne": assembly_id}
            })
            if other_usage > 0:
                linked_count += 1
        
        result.append({
            "assembly_id": assembly_id,
            "name": asm.get("name", ""),
            "description": asm.get("description", ""),
            "total_parts": len(part_ids),
            "linked_parts_count": linked_count,
            "created_at": asm.get("created_at"),
        })
    
    # Sort by name
    result.sort(key=lambda x: (x.get("name") or "").lower())
    
    return result


@app.get("/api/manage/assemblies/{assembly_id}/parts")
async def get_assembly_parts_with_links(
    assembly_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get all parts of an assembly with linked info"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get assembly
    assembly = await db.assemblies.find_one({"assembly_id": assembly_id})
    if not assembly:
        raise HTTPException(status_code=404, detail="Assembly not found")
    
    # Get parts
    assembly_parts = await db.assembly_parts.find({"assembly_id": assembly_id}).to_list(length=500)
    
    parts_with_links = []
    for ap in assembly_parts:
        part_id = ap["part_id"]
        
        # Get part details
        part = await db.parts_library.find_one({"part_id": part_id})
        if not part:
            continue
        
        # Check which other assemblies use this part
        other_usages = await db.assembly_parts.find({
            "part_id": part_id,
            "assembly_id": {"$ne": assembly_id}
        }).to_list(length=50)
        
        linked_assemblies = []
        for usage in other_usages:
            other_asm = await db.assemblies.find_one(
                {"assembly_id": usage["assembly_id"]},
                {"name": 1, "assembly_id": 1}
            )
            if other_asm:
                linked_assemblies.append({
                    "assembly_id": other_asm["assembly_id"],
                    "name": other_asm.get("name", other_asm["assembly_id"])
                })
        
        parts_with_links.append({
            "part_id": part_id,
            "name": part.get("name", ""),
            "category": part.get("category", ""),
            "part_type": part.get("part_type", ""),
            "size": part.get("size", ""),
            "length": part.get("length", ""),
            "quantity": ap.get("quantity", 1),
            "unit": part.get("unit", "PCS"),
            "stock": part.get("stock", 0),
            "is_linked": len(linked_assemblies) > 0,
            "linked_to": linked_assemblies,
            "remarks": part.get("remarks", "")
        })
    
    return {
        "assembly": {
            "assembly_id": assembly_id,
            "name": assembly.get("name", ""),
            "description": assembly.get("description", "")
        },
        "parts": parts_with_links,
        "total_parts": len(parts_with_links),
        "linked_parts_count": sum(1 for p in parts_with_links if p["is_linked"])
    }


@app.post("/api/manage/assemblies")
async def create_assembly(request: Request, current_user: User = Depends(get_current_user)):
    """Create a new assembly"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    data = await request.json()
    assembly_id = data.get("assembly_id", "").strip()
    name = data.get("name", "").strip()
    description = data.get("description", "").strip()
    
    if not assembly_id or not name:
        raise HTTPException(status_code=400, detail="assembly_id and name are required")
    
    # Check if exists
    existing = await db.assemblies.find_one({"assembly_id": assembly_id})
    if existing:
        raise HTTPException(status_code=400, detail="Assembly ID already exists")
    
    assembly = {
        "assembly_id": assembly_id,
        "name": name,
        "description": description,
        "created_at": get_ist_now(),
        "created_by": current_user.user_id
    }
    
    await db.assemblies.insert_one(assembly)
    
    return {"message": "Assembly created", "assembly_id": assembly_id}


@app.put("/api/manage/assemblies/{assembly_id}")
async def update_assembly(
    assembly_id: str,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Update assembly details"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    assembly = await db.assemblies.find_one({"assembly_id": assembly_id})
    if not assembly:
        raise HTTPException(status_code=404, detail="Assembly not found")
    
    data = await request.json()
    updates = {"updated_at": get_ist_now(), "updated_by": current_user.user_id}
    
    if "name" in data:
        updates["name"] = data["name"].strip()
    if "description" in data:
        updates["description"] = data["description"].strip()
    
    await db.assemblies.update_one(
        {"assembly_id": assembly_id},
        {"$set": updates}
    )
    
    return {"message": "Assembly updated"}


@app.get("/api/manage/assemblies/{assembly_id}/delete-preview")
async def preview_assembly_deletion(
    assembly_id: str,
    current_user: User = Depends(get_current_user)
):
    """Preview what will happen when deleting an assembly"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    assembly = await db.assemblies.find_one({"assembly_id": assembly_id})
    if not assembly:
        raise HTTPException(status_code=404, detail="Assembly not found")
    
    # Get all parts
    assembly_parts = await db.assembly_parts.find({"assembly_id": assembly_id}).to_list(length=500)
    
    unique_parts = []  # Only used in this assembly - will be deleted
    shared_parts = []  # Used in other assemblies - user chooses
    
    for ap in assembly_parts:
        part_id = ap["part_id"]
        part = await db.parts_library.find_one({"part_id": part_id})
        if not part:
            continue
        
        # Check other usages
        other_usages = await db.assembly_parts.find({
            "part_id": part_id,
            "assembly_id": {"$ne": assembly_id}
        }).to_list(length=50)
        
        part_info = {
            "part_id": part_id,
            "name": part.get("name", ""),
            "size": part.get("size", ""),
            "category": part.get("category", "")
        }
        
        if len(other_usages) == 0:
            unique_parts.append(part_info)
        else:
            # Get assembly names
            also_in = []
            for usage in other_usages:
                other_asm = await db.assemblies.find_one(
                    {"assembly_id": usage["assembly_id"]},
                    {"name": 1}
                )
                if other_asm:
                    also_in.append(other_asm.get("name", usage["assembly_id"]))
            
            part_info["also_used_in"] = also_in
            shared_parts.append(part_info)
    
    return {
        "assembly_id": assembly_id,
        "name": assembly.get("name", ""),
        "total_parts": len(assembly_parts),
        "unique_parts": unique_parts,
        "unique_parts_count": len(unique_parts),
        "shared_parts": shared_parts,
        "shared_parts_count": len(shared_parts)
    }


@app.delete("/api/manage/assemblies/{assembly_id}")
async def delete_assembly_managed(
    assembly_id: str,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Delete assembly with control over shared parts"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    assembly = await db.assemblies.find_one({"assembly_id": assembly_id})
    if not assembly:
        raise HTTPException(status_code=404, detail="Assembly not found")
    
    data = await request.json() if request.headers.get("content-type") == "application/json" else {}
    delete_shared_parts = data.get("delete_shared_parts", [])  # List of part_ids to delete
    
    # Get all parts
    assembly_parts = await db.assembly_parts.find({"assembly_id": assembly_id}).to_list(length=500)
    
    parts_deleted = 0
    parts_kept = 0
    
    for ap in assembly_parts:
        part_id = ap["part_id"]
        
        # Check if part is used elsewhere
        other_usage = await db.assembly_parts.find_one({
            "part_id": part_id,
            "assembly_id": {"$ne": assembly_id}
        })
        
        if not other_usage:
            # Unique to this assembly - delete it
            await db.parts_library.delete_one({"part_id": part_id})
            parts_deleted += 1
        elif part_id in delete_shared_parts:
            # User chose to delete this shared part
            # Remove from all assemblies first
            await db.assembly_parts.delete_many({"part_id": part_id})
            await db.parts_library.delete_one({"part_id": part_id})
            parts_deleted += 1
        else:
            # Keep the part (just unlink from this assembly)
            parts_kept += 1
    
    # Delete assembly parts links
    await db.assembly_parts.delete_many({"assembly_id": assembly_id})
    
    # Delete assembly
    await db.assemblies.delete_one({"assembly_id": assembly_id})
    
    return {
        "message": "Assembly deleted",
        "assembly_id": assembly_id,
        "parts_deleted": parts_deleted,
        "parts_kept": parts_kept
    }


@app.delete("/api/manage/assemblies/{assembly_id}/parts/{part_id}")
async def remove_part_from_assembly_managed(
    assembly_id: str,
    part_id: str,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Remove part from assembly with option to delete from library"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    data = await request.json() if request.headers.get("content-type") == "application/json" else {}
    delete_from_library = data.get("delete_from_library", False)
    
    # Remove from assembly
    result = await db.assembly_parts.delete_one({
        "assembly_id": assembly_id,
        "part_id": part_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Part not found in assembly")
    
    deleted_from_library = False
    still_used_in = []
    
    if delete_from_library:
        # Check if used elsewhere
        other_usage = await db.assembly_parts.find({
            "part_id": part_id
        }).to_list(length=50)
        
        if len(other_usage) == 0:
            # Safe to delete
            await db.parts_library.delete_one({"part_id": part_id})
            deleted_from_library = True
        else:
            # Part is still used elsewhere
            for usage in other_usage:
                asm = await db.assemblies.find_one(
                    {"assembly_id": usage["assembly_id"]},
                    {"name": 1}
                )
                if asm:
                    still_used_in.append(asm.get("name", usage["assembly_id"]))
    
    return {
        "message": "Part removed from assembly",
        "deleted_from_library": deleted_from_library,
        "still_used_in": still_used_in
    }


@app.put("/api/manage/parts/{part_id}/correct-id")
async def correct_part_id(
    part_id: str,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Correct/change part ID across all references"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    data = await request.json()
    new_part_id = data.get("new_part_id", "").strip()
    
    if not new_part_id:
        raise HTTPException(status_code=400, detail="new_part_id is required")
    
    if new_part_id == part_id:
        raise HTTPException(status_code=400, detail="New Part ID is same as current")
    
    # Check if new ID already exists
    existing = await db.parts_library.find_one({"part_id": new_part_id})
    if existing:
        raise HTTPException(status_code=400, detail=f"Part ID {new_part_id} already exists")
    
    # Get the part
    part = await db.parts_library.find_one({"part_id": part_id})
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    
    # Update in parts_library
    await db.parts_library.update_one(
        {"part_id": part_id},
        {"$set": {"part_id": new_part_id, "old_part_id": part_id, "updated_at": get_ist_now()}}
    )
    
    # Update in assembly_parts
    asm_result = await db.assembly_parts.update_many(
        {"part_id": part_id},
        {"$set": {"part_id": new_part_id}}
    )
    
    # Update in storage locations
    storage_result = await db.storage_locations.update_many(
        {"part_id": part_id},
        {"$set": {"part_id": new_part_id}}
    )
    
    # Update in gauge_storage if exists
    gauge_result = await db.gauge_storage.update_many(
        {"part_id": part_id},
        {"$set": {"part_id": new_part_id}}
    )
    
    # Update in production records
    prod_result = await db.production_entries.update_many(
        {"part_id": part_id},
        {"$set": {"part_id": new_part_id}}
    )
    
    # Log the change
    await db.part_id_changes.insert_one({
        "old_part_id": part_id,
        "new_part_id": new_part_id,
        "changed_by": current_user.user_id,
        "changed_at": get_ist_now(),
        "assemblies_updated": asm_result.modified_count,
        "storage_updated": storage_result.modified_count,
        "production_updated": prod_result.modified_count
    })
    
    return {
        "message": "Part ID corrected successfully",
        "old_part_id": part_id,
        "new_part_id": new_part_id,
        "assemblies_updated": asm_result.modified_count,
        "storage_locations_updated": storage_result.modified_count,
        "production_records_updated": prod_result.modified_count
    }


@app.get("/api/manage/parts/search")
async def search_parts_for_linking(
    q: str = "",
    current_user: User = Depends(get_current_user)
):
    """Search parts library for linking to assembly"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if q:
        query["$or"] = [
            {"part_id": {"$regex": q, "$options": "i"}},
            {"name": {"$regex": q, "$options": "i"}}
        ]
    
    parts = await db.parts_library.find(query, {"_id": 0}).limit(50).to_list(length=50)
    
    return parts


# ========================================
# END MANAGE ASSEMBLIES
# ========================================


# ========================================
# PRODUCTION PLANNING
# ========================================

@app.post("/api/production-planning/calculate")
async def calculate_production_requirements(
    assembly_id: str,
    quantity: int,
    current_user: User = Depends(get_current_user)
):
    """Calculate required parts for planned production"""
    if quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be positive")
    
    # Get assembly
    assembly = await db.assemblies.find_one({"assembly_id": assembly_id})
    if not assembly:
        raise HTTPException(status_code=404, detail="Assembly not found")
    
    # Get assembly parts
    assembly_parts = await db.assembly_parts.find(
        {"assembly_id": assembly_id}
    ).to_list(length=100)
    
    if not assembly_parts:
        raise HTTPException(status_code=400, detail="Assembly has no parts defined")
    
    available_parts = []
    shortage_parts = []
    
    for ap in assembly_parts:
        part = await db.parts_library.find_one({"part_id": ap["part_id"]}, {"_id": 0})
        if not part:
            continue
        
        required_qty = ap["quantity"] * quantity
        available_stock = part.get("stock", 0) - part.get("reserved_stock", 0)
        
        part_info = {
            "part_id": part["part_id"],
            "name": part.get("name", ""),
            "category": part.get("category", ""),
            "unit": part.get("unit", "PCS"),
            "required": required_qty,
            "per_assembly": ap["quantity"],
            "stock": part.get("stock", 0),
            "reserved": part.get("reserved_stock", 0),
            "available": available_stock,
            "locations": part.get("locations", [])
        }
        
        if available_stock >= required_qty:
            part_info["status"] = "ok"
            part_info["shortage"] = 0
            available_parts.append(part_info)
        else:
            part_info["status"] = "shortage"
            part_info["shortage"] = required_qty - available_stock
            shortage_parts.append(part_info)
    
    return {
        "assembly_id": assembly_id,
        "assembly_name": assembly.get("name", ""),
        "planned_quantity": quantity,
        "total_parts": len(assembly_parts),
        "available_count": len(available_parts),
        "shortage_count": len(shortage_parts),
        "can_produce": len(shortage_parts) == 0,
        "available_parts": available_parts,
        "shortage_parts": shortage_parts
    }


@app.post("/api/production-planning/reserve")
async def reserve_stock_for_production(
    assembly_id: str,
    quantity: int,
    current_user: User = Depends(get_current_user)
):
    """Reserve stock for planned production"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # First calculate to verify stock availability
    calc_result = await calculate_production_requirements(assembly_id, quantity, current_user)
    
    if not calc_result["can_produce"]:
        raise HTTPException(status_code=400, detail="Cannot reserve: Stock shortage exists")
    
    # Reserve stock for each part
    for part_info in calc_result["available_parts"]:
        await db.parts_library.update_one(
            {"part_id": part_info["part_id"]},
            {"$inc": {"reserved_stock": part_info["required"]}}
        )
    
    # Create planning record
    now = get_ist_now()
    plan_id = f"PLN-{now.strftime('%Y%m%d%H%M%S')}"
    
    await db.production_plans.insert_one({
        "plan_id": plan_id,
        "assembly_id": assembly_id,
        "assembly_name": calc_result["assembly_name"],
        "planned_quantity": quantity,
        "status": "reserved",
        "parts": calc_result["available_parts"],
        "created_at": now,
        "created_by": current_user.user_id
    })
    
    return {
        "message": "Stock reserved successfully",
        "plan_id": plan_id,
        "parts_reserved": len(calc_result["available_parts"])
    }


@app.post("/api/production-planning/release/{plan_id}")
async def release_reserved_stock(plan_id: str, current_user: User = Depends(get_current_user)):
    """Release reserved stock (cancel plan)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    plan = await db.production_plans.find_one({"plan_id": plan_id})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    if plan.get("status") != "reserved":
        raise HTTPException(status_code=400, detail="Plan is not in reserved status")
    
    # Release reserved stock for each part
    for part_info in plan.get("parts", []):
        await db.parts_library.update_one(
            {"part_id": part_info["part_id"]},
            {"$inc": {"reserved_stock": -part_info["required"]}}
        )
    
    # Update plan status
    await db.production_plans.update_one(
        {"plan_id": plan_id},
        {"$set": {"status": "cancelled", "cancelled_at": get_ist_now()}}
    )
    
    return {"message": "Reserved stock released successfully"}


# ========================================
# STATIC FILE SERVING FOR EXPO WEB BUILD
# ========================================
# This must be at the END of all routes to act as a catch-all

DIST_DIR = "/app/frontend/dist"
PUBLIC_DIR = "/app/frontend/public"

# Mount static assets directories if they exist
if os.path.exists(f"{DIST_DIR}/assets"):
    app.mount("/assets", StaticFiles(directory=f"{DIST_DIR}/assets"), name="assets")
if os.path.exists(f"{DIST_DIR}/_expo"):
    app.mount("/_expo", StaticFiles(directory=f"{DIST_DIR}/_expo"), name="expo_internal")

# PWA Routes - serve manifest.json and service worker
@app.get("/manifest.json")
async def serve_manifest():
    """Serve PWA manifest file."""
    manifest_path = f"{PUBLIC_DIR}/manifest.json"
    if os.path.isfile(manifest_path):
        return FileResponse(manifest_path, media_type="application/manifest+json")
    raise HTTPException(status_code=404, detail="Manifest not found")

@app.get("/sw.js")
async def serve_service_worker():
    """Serve PWA service worker file."""
    sw_path = f"{PUBLIC_DIR}/sw.js"
    if os.path.isfile(sw_path):
        return FileResponse(sw_path, media_type="application/javascript")
    raise HTTPException(status_code=404, detail="Service worker not found")

# HTML template for the SPA app (used when dist/index.html doesn't exist)
def get_spa_html(title: str = "VMC Job Shop"):
    """Generate fallback HTML - should only be used if Expo build failed"""
    html = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>''' + title + '''</title>
    <style>
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: #0F172A;
            color: white;
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
            margin: 0;
        }
        .container { text-align: center; max-width: 400px; padding: 20px; }
        .error { color: #EF4444; }
    </style>
</head>
<body>
    <div class="container">
        <h1>Build Error</h1>
        <p class="error">The Expo web build is missing. Please redeploy the application.</p>
        <p>Contact support if this issue persists.</p>
    </div>
</body>
</html>'''
    return html

# ========================================
# AI AUDIT MODULE
# ========================================

# Initialize AI Audit Engine after db is available
ai_audit_engine = None

@app.on_event("startup")
async def init_ai_audit():
    global ai_audit_engine
    ai_audit_engine = AIAuditEngine(db)
    logger.info("AI Audit Engine initialized")


class VoiceCommandRequest(BaseModel):
    command: str


@app.get("/api/ai/audit-report")
async def get_ai_audit_report(
    date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get AI-generated audit report for a specific date"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if not ai_audit_engine:
        raise HTTPException(status_code=500, detail="AI Audit Engine not initialized")
    
    try:
        report = await ai_audit_engine.generate_daily_report(date)
        return report
    except Exception as e:
        logger.error(f"AI audit report error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ai/audit-report/latest")
async def get_latest_audit_report(
    current_user: User = Depends(get_current_user)
):
    """Get the latest cached audit report"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get today's report from cache
    today = datetime.now(IST).strftime('%Y-%m-%d')
    report = await db.audit_reports.find_one({"date": today})
    
    if report:
        report["_id"] = str(report["_id"])
        return report
    
    # If no report exists, generate one
    if ai_audit_engine:
        return await ai_audit_engine.generate_daily_report(today)
    
    return {"message": "No report available", "date": today}


@app.post("/api/ai/voice-command")
async def process_voice_command(
    request: VoiceCommandRequest,
    current_user: User = Depends(get_current_user)
):
    """Process voice command and return action"""
    if not ai_audit_engine:
        raise HTTPException(status_code=500, detail="AI not initialized")
    
    try:
        result = await ai_audit_engine.process_voice_command(
            request.command, 
            current_user.user_id
        )
        return result
    except Exception as e:
        logger.error(f"Voice command error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ai/quick-insights")
async def get_quick_insights(
    current_user: User = Depends(get_current_user)
):
    """Get quick AI insights for dashboard (lighter version)"""
    today = datetime.now(IST).strftime('%Y-%m-%d')
    
    # Get counts for quick insights
    entry_issues = []
    tool_alerts = []
    
    # Check for low stock tools
    low_stock_tools = await db.tools_inserts.count_documents({"is_low_stock": True})
    
    # Check for missing signatures today
    storage_deductions = await db.storage_deductions.find({
        "taken_date": {"$regex": f"^{today}"}
    }).to_list(length=100)
    missing_sigs = sum(1 for d in storage_deductions if not d.get("signature"))
    
    # Production stats
    production = await db.production_entries.find({"date": today}).to_list(length=500)
    system_ended = sum(1 for p in production if p.get("system_ended"))
    missing_remarks = sum(1 for p in production if not p.get("end_remarks"))
    
    # Build issues list
    if missing_sigs > 0:
        entry_issues.append({
            "type": "missing_signature",
            "count": missing_sigs,
            "message": f"{missing_sigs} parts taken without signature"
        })
    
    if system_ended > 0:
        entry_issues.append({
            "type": "system_ended",
            "count": system_ended,
            "message": f"{system_ended} jobs ended by system"
        })
    
    if missing_remarks > 0:
        entry_issues.append({
            "type": "missing_remarks",
            "count": missing_remarks,
            "message": f"{missing_remarks} jobs missing end remarks"
        })
    
    if low_stock_tools > 0:
        tool_alerts.append({
            "type": "low_stock",
            "count": low_stock_tools,
            "message": f"{low_stock_tools} tools low in stock"
        })
    
    return {
        "date": today,
        "total_issues": len(entry_issues) + len(tool_alerts),
        "entry_issues": entry_issues,
        "tool_alerts": tool_alerts,
        "production_count": len(production),
        "has_full_report": await db.audit_reports.find_one({"date": today}) is not None
    }


# ============== MOULD MANAGEMENT APIs ==============
# MOVED TO: routes/moulds.py
# Router is included at the top of this file via app.include_router(moulds_router)


# ============== PARTS LIBRARY EXCEL UPLOAD ==============

class PartCreate(BaseModel):
    category: str  # "pump" or "spray_gun"
    part_type: str  # shaft, plastic_parts, o_ring, etc.
    part_name: str
    part_number: Optional[str] = None
    # Common fields
    quantity: int = 0
    min_quantity: int = 5
    location: Optional[str] = None
    # O-Ring specific
    od: Optional[float] = None  # Outer diameter
    id: Optional[float] = None  # Inner diameter
    thickness: Optional[float] = None
    # Bearing specific
    bearing_size: Optional[str] = None
    bearing_type: Optional[str] = None
    # Nozzle specific
    nozzle_size: Optional[str] = None
    nozzle_type: Optional[str] = None
    # General
    material: Optional[str] = None
    specifications: Optional[str] = None
    image: Optional[str] = None
    remarks: Optional[str] = None
    # Association
    associated_product: Optional[str] = None  # Which pump/spray gun this part belongs to

# Pump part types
PUMP_PART_TYPES = ["shaft", "plastic_parts", "aluminium_parts", "bearings", "bush", "bolts", "screw", "oil_seal"]
# Spray gun part types
SPRAY_GUN_PART_TYPES = ["pipe", "manifold", "o_ring", "gasket", "nozzles", "ss_parts", "circlips", "nuts"]

@app.post("/api/parts-library/upload-excel")
async def upload_parts_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload Excel file to add parts - auto-detects Pump vs Spray Gun"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        import pandas as pd
        from io import BytesIO
        
        content = await file.read()
        df = pd.read_excel(BytesIO(content))
        
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        # Auto-detect category based on part_type column
        detected_category = None
        parts_added = 0
        parts_updated = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                part_type = str(row.get('part_type', '')).strip().lower().replace(' ', '_')
                
                # Detect category
                if part_type in PUMP_PART_TYPES:
                    category = "pump"
                elif part_type in SPRAY_GUN_PART_TYPES:
                    category = "spray_gun"
                else:
                    # Try to infer from other columns or default
                    category = str(row.get('category', 'pump')).strip().lower()
                
                if detected_category is None:
                    detected_category = category
                
                part_name = str(row.get('part_name', '')).strip()
                if not part_name:
                    errors.append(f"Row {idx + 2}: Missing part_name")
                    continue
                
                # Build part document
                part_doc = {
                    "part_id": f"part_{uuid.uuid4().hex[:12]}",
                    "category": category,
                    "part_type": part_type,
                    "part_name": part_name,
                    "part_number": str(row.get('part_number', '')).strip() or None,
                    "quantity": int(row.get('quantity', 0)) if pd.notna(row.get('quantity')) else 0,
                    "min_quantity": int(row.get('min_quantity', 5)) if pd.notna(row.get('min_quantity')) else 5,
                    "location": str(row.get('location', '')).strip() or None,
                    "material": str(row.get('material', '')).strip() or None,
                    "specifications": str(row.get('specifications', '')).strip() or None,
                    "associated_product": str(row.get('associated_product', '')).strip() or None,
                    "remarks": str(row.get('remarks', '')).strip() or None,
                    "is_low_stock": False,
                    "created_at": datetime.now(timezone.utc),
                    "created_by": current_user.user_id,
                    "updated_at": datetime.now(timezone.utc)
                }
                
                # O-Ring specific fields
                if part_type == 'o_ring':
                    part_doc["od"] = float(row.get('od', 0)) if pd.notna(row.get('od')) else None
                    part_doc["id"] = float(row.get('id', 0)) if pd.notna(row.get('id')) else None
                    part_doc["thickness"] = float(row.get('thickness', 0)) if pd.notna(row.get('thickness')) else None
                
                # Bearing specific fields
                if part_type == 'bearings':
                    part_doc["bearing_size"] = str(row.get('bearing_size', '')).strip() or None
                    part_doc["bearing_type"] = str(row.get('bearing_type', '')).strip() or None
                
                # Nozzle specific fields
                if part_type == 'nozzles':
                    part_doc["nozzle_size"] = str(row.get('nozzle_size', '')).strip() or None
                    part_doc["nozzle_type"] = str(row.get('nozzle_type', '')).strip() or None
                
                # Check low stock
                part_doc["is_low_stock"] = part_doc["quantity"] < part_doc["min_quantity"]
                
                # Check if part already exists (by part_number or part_name + category)
                existing = None
                if part_doc.get("part_number"):
                    existing = await db.parts_library.find_one({"part_number": part_doc["part_number"]})
                
                if not existing:
                    existing = await db.parts_library.find_one({
                        "part_name": part_name,
                        "category": category,
                        "part_type": part_type
                    })
                
                if existing:
                    # Update existing
                    update_fields = {k: v for k, v in part_doc.items() if k not in ["part_id", "created_at", "created_by"]}
                    await db.parts_library.update_one({"part_id": existing["part_id"]}, {"$set": update_fields})
                    parts_updated += 1
                else:
                    # Insert new
                    await db.parts_library.insert_one(part_doc)
                    parts_added += 1
                    
            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")
        
        return {
            "message": "Excel processed successfully",
            "detected_category": detected_category,
            "parts_added": parts_added,
            "parts_updated": parts_updated,
            "total_processed": parts_added + parts_updated,
            "errors": errors[:10] if errors else []  # Return first 10 errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing Excel: {str(e)}")

@app.get("/api/parts-library/order-list")
async def get_parts_order_list(
    category: Optional[str] = None,
    part_type: Optional[str] = None,
    associated_product: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get order list for low stock parts - optionally filtered"""
    query = {"is_low_stock": True}
    
    if category:
        query["category"] = category
    if part_type:
        query["part_type"] = part_type
    if associated_product:
        query["associated_product"] = {"$regex": associated_product, "$options": "i"}
    
    parts = await db.parts_library.find(query, {"_id": 0}).sort("part_type", 1).to_list(length=500)
    
    # Group by part_type for easier ordering
    order_list = {}
    for part in parts:
        pt = part.get("part_type", "other")
        if pt not in order_list:
            order_list[pt] = []
        
        order_item = {
            "part_name": part.get("part_name"),
            "part_number": part.get("part_number"),
            "current_qty": part.get("quantity", 0),
            "min_qty": part.get("min_quantity", 5),
            "need_qty": max(0, part.get("min_quantity", 5) - part.get("quantity", 0)),
            "associated_product": part.get("associated_product"),
            "specifications": part.get("specifications")
        }
        
        # Add O-Ring specifics
        if pt == "o_ring":
            order_item["od"] = part.get("od")
            order_item["id"] = part.get("id")
            order_item["thickness"] = part.get("thickness")
        
        order_list[pt].append(order_item)
    
    return {
        "total_items": len(parts),
        "order_list": order_list
    }


@app.get("/api/parts-library/excel-template/{category}")
async def get_excel_template(category: str):
    """Get Excel template for Pump or Spray Gun parts"""
    import pandas as pd
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    if category == "pump":
        columns = ["part_type", "part_name", "part_number", "quantity", "min_quantity", 
                   "location", "material", "bearing_size", "bearing_type", "specifications", 
                   "associated_product", "remarks"]
        example_data = [
            {"part_type": "shaft", "part_name": "Main Shaft 25mm", "part_number": "SH-001", "quantity": 10, "min_quantity": 5, "location": "Rack A1", "material": "Steel", "associated_product": "Pump Model X"},
            {"part_type": "bearings", "part_name": "Ball Bearing 6205", "part_number": "BR-001", "quantity": 20, "min_quantity": 10, "bearing_size": "6205", "bearing_type": "Ball", "associated_product": "Pump Model X"},
            {"part_type": "oil_seal", "part_name": "Oil Seal 25x35x7", "part_number": "OS-001", "quantity": 15, "min_quantity": 10, "specifications": "25x35x7mm", "associated_product": "Pump Model X"},
        ]
    elif category == "spray_gun":
        columns = ["part_type", "part_name", "part_number", "quantity", "min_quantity",
                   "location", "material", "od", "id", "thickness", "nozzle_size", "nozzle_type",
                   "specifications", "associated_product", "remarks"]
        example_data = [
            {"part_type": "o_ring", "part_name": "O-Ring 10x2", "part_number": "OR-001", "quantity": 50, "min_quantity": 20, "od": 10, "id": 8, "thickness": 2, "associated_product": "Spray Gun SG-100"},
            {"part_type": "nozzles", "part_name": "Spray Nozzle 1.5mm", "part_number": "NZ-001", "quantity": 10, "min_quantity": 5, "nozzle_size": "1.5mm", "nozzle_type": "Fan", "associated_product": "Spray Gun SG-100"},
            {"part_type": "gasket", "part_name": "Gasket Set", "part_number": "GS-001", "quantity": 25, "min_quantity": 10, "material": "Rubber", "associated_product": "Spray Gun SG-100"},
        ]
    else:
        raise HTTPException(status_code=400, detail="Category must be 'pump' or 'spray_gun'")
    
    df = pd.DataFrame(example_data, columns=columns)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Parts')
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={category}_parts_template.xlsx"}
    )



# ============================================
# BIOMETRIC ATTENDANCE MODULE (eSSL Integration)
# ============================================

# eSSL Device Configuration
# NOTE: timeout reduced from 10s to 2s to prevent backend blocking
# Cloud server cannot reach local LAN device (192.168.x.x), so fail fast
ESSL_CONFIG = {
    "ip": "192.168.0.201",
    "port": 4370,
    "timeout": 2  # Reduced from 10 to prevent 10s blocking on unreachable device
}

# Shift Configuration
SHIFT_CONFIG = {
    "day": {
        "name": "Day Shift",
        "start": "08:00",
        "end": "20:00",
        "late_grace_minutes": 15
    },
    "night": {
        "name": "Night Shift", 
        "start": "20:00",
        "end": "06:00",
        "late_grace_minutes": 15
    }
}

# Indian Public Holidays / Festivals (2025-2026)
# Format: "YYYY-MM-DD": "Holiday Name"
HOLIDAYS = {
    # 2025 Holidays
    "2025-01-14": "Makar Sankranti",
    "2025-01-26": "Republic Day",
    "2025-03-14": "Holi",
    "2025-03-31": "Id-ul-Fitr",
    "2025-04-06": "Ram Navami",
    "2025-04-10": "Mahavir Jayanti",
    "2025-04-14": "Ambedkar Jayanti",
    "2025-04-18": "Good Friday",
    "2025-05-01": "May Day",
    "2025-05-12": "Buddha Purnima",
    "2025-06-07": "Eid ul-Adha",
    "2025-07-06": "Muharram",
    "2025-08-15": "Independence Day",
    "2025-08-16": "Janmashtami",
    "2025-09-05": "Milad un-Nabi",
    "2025-10-02": "Gandhi Jayanti",
    "2025-10-20": "Dussehra",
    "2025-10-21": "Dussehra Holiday",
    "2025-11-01": "Diwali",
    "2025-11-02": "Diwali Holiday",
    "2025-11-03": "Diwali Holiday",
    "2025-11-05": "Bhai Dooj",
    "2025-11-15": "Guru Nanak Jayanti",
    "2025-12-25": "Christmas",
    # 2026 Holidays
    "2026-01-14": "Makar Sankranti",
    "2026-01-26": "Republic Day",
    "2026-03-03": "Holi",
    "2026-03-20": "Id-ul-Fitr",
    "2026-03-26": "Ram Navami",
    "2026-03-31": "Mahavir Jayanti",
    "2026-04-03": "Good Friday",
    "2026-04-14": "Ambedkar Jayanti",
    "2026-05-01": "May Day",
    "2026-05-31": "Buddha Purnima",
    "2026-05-27": "Eid ul-Adha",
    "2026-06-26": "Muharram",
    "2026-08-15": "Independence Day",
    "2026-08-25": "Janmashtami",
    "2026-10-02": "Gandhi Jayanti",
    "2026-10-09": "Dussehra",
    "2026-10-20": "Diwali",
    "2026-10-21": "Diwali Holiday",
    "2026-10-22": "Diwali Holiday",
    "2026-11-04": "Guru Nanak Jayanti",
    "2026-12-25": "Christmas",
}

def get_holiday_name(date_str: str) -> str:
    """Check if a date is a holiday and return the holiday name"""
    return HOLIDAYS.get(date_str, None)


def get_shift_for_time(punch_time: datetime) -> dict:
    """
    Determine which shift a punch time belongs to.
    
    Day Shift: 8 AM - 8 PM (workers arrive 5 AM - 2 PM, considered day shift)
    Night Shift: 8 PM - 6 AM (workers arrive 5 PM - midnight, considered night shift)
    
    Early morning punches (midnight - 5 AM) are ambiguous - could be:
    - OUT punch for previous night shift
    - Very early IN for day shift (rare)
    
    This function is used when inserting new records. For determining shift
    context in attendance reports, use the logic in get_monthly_attendance.
    """
    hour = punch_time.hour
    
    # 5 AM to 5 PM (17:00) -> Day Shift arrival window
    if 5 <= hour < 17:
        return SHIFT_CONFIG["day"]
    # 5 PM (17:00) to midnight -> Night Shift arrival window  
    elif 17 <= hour <= 23:
        return SHIFT_CONFIG["night"]
    # Midnight to 5 AM -> Likely Night Shift OUT, but default to Night for record keeping
    else:
        return SHIFT_CONFIG["night"]


def determine_status(punch_time: datetime, shift: dict) -> str:
    """Determine if punch is on-time, late, or early"""
    shift_start_str = shift["start"]
    shift_hour, shift_min = map(int, shift_start_str.split(":"))
    
    # Create shift start datetime for the same day
    shift_start = punch_time.replace(hour=shift_hour, minute=shift_min, second=0, microsecond=0)
    
    # Add grace period
    grace_deadline = shift_start + timedelta(minutes=shift["late_grace_minutes"])
    
    if punch_time <= shift_start:
        return "on_time"
    elif punch_time <= grace_deadline:
        return "on_time"  # Within grace period
    else:
        return "late"


@app.get("/api/attendance/device-status")
async def check_essl_device_status():
    """
    Check if eSSL device is reachable and return last sync time.
    """
    import socket
    
    ip = ESSL_CONFIG["ip"]
    port = ESSL_CONFIG["port"]
    timeout = ESSL_CONFIG.get("timeout", 5)
    
    # Get last sync time from database
    last_sync = None
    try:
        # Try synced_at first, fall back to created_at, then punch_time_ist
        latest_record = await db.attendance_records.find_one(
            sort=[("synced_at", -1)]
        )
        sync_time = None
        
        if latest_record:
            sync_time = latest_record.get("synced_at") or latest_record.get("created_at") or latest_record.get("punch_time_ist")
        
        if sync_time:
            # Ensure it has timezone info (IST)
            if isinstance(sync_time, datetime):
                if sync_time.tzinfo is None:
                    sync_time = sync_time.replace(tzinfo=IST)
                last_sync = sync_time.isoformat()
            elif isinstance(sync_time, str):
                # Already a string, add IST if no timezone
                if '+' not in sync_time and 'Z' not in sync_time:
                    last_sync = sync_time + '+05:30'
                else:
                    last_sync = sync_time
    except Exception as e:
        logger.error(f"Error getting last sync time: {e}")
        pass
    
    # Step 1: TCP Socket Check (no ping dependency)
    tcp_ok = False
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        result = sock.connect_ex((ip, port))
        sock.close()
        tcp_ok = (result == 0)
    except Exception as e:
        logger.error(f"TCP socket check failed: {e}")
        tcp_ok = False
    
    if not tcp_ok:
        return {
            "status": "unreachable",
            "status_display": "Device Unreachable",
            "ip": ip,
            "port": port,
            "message": f"Cannot establish TCP connection to {ip}:{port}",
            "last_sync": last_sync
        }
    
    # Step 2: pyzk Connection Check (actual device authentication)
    try:
        from zk import ZK
        
        zk = ZK(ip, port=port, timeout=timeout)
        conn = zk.connect()
        
        if conn:
            # Successfully connected and authenticated
            conn.disconnect()
            return {
                "status": "connected",
                "status_display": "TCP Connected",
                "ip": ip,
                "port": port,
                "message": f"Device connected and authenticated at {ip}:{port}",
                "last_sync": last_sync
            }
        else:
            return {
                "status": "auth_failed",
                "status_display": "Authentication Failed",
                "ip": ip,
                "port": port,
                "message": "TCP connected but device authentication failed",
                "last_sync": last_sync
            }
            
    except ImportError:
        return {
            "status": "error",
            "status_display": "Library Missing",
            "ip": ip,
            "port": port,
            "message": "pyzk library not installed",
            "last_sync": last_sync
        }
    except Exception as e:
        error_msg = str(e).lower()
        # Check for auth-related errors
        if "password" in error_msg or "auth" in error_msg or "permission" in error_msg:
            return {
                "status": "auth_failed",
                "status_display": "Authentication Failed",
                "ip": ip,
                "port": port,
                "message": f"TCP connected but authentication failed: {str(e)}",
                "last_sync": last_sync
            }
        else:
            return {
                "status": "error",
                "status_display": "Connection Error",
                "ip": ip,
                "port": port,
                "message": f"pyzk connection error: {str(e)}",
                "last_sync": last_sync
            }


@app.post("/api/attendance/sync")
async def sync_attendance_from_essl(current_user: User = Depends(get_current_user)):
    """
    Sync attendance data from eSSL biometric device.
    Uses TCP socket check + pyzk connect() - NO ping dependency.
    Returns specific error messages:
    - "Device Unreachable" - Cannot establish TCP connection
    - "Authentication Failed" - TCP works but pyzk auth failed
    """
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    import socket
    
    ip = ESSL_CONFIG["ip"]
    port = ESSL_CONFIG["port"]
    timeout = ESSL_CONFIG.get("timeout", 5)
    
    # Step 1: TCP Socket Check (no ping dependency)
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        result = sock.connect_ex((ip, port))
        sock.close()
        
        if result != 0:
            raise HTTPException(
                status_code=503, 
                detail="Device Unreachable - Cannot establish TCP connection to eSSL device"
            )
    except socket.timeout:
        raise HTTPException(
            status_code=503, 
            detail="Device Unreachable - Connection timed out"
        )
    except socket.error as e:
        raise HTTPException(
            status_code=503, 
            detail=f"Device Unreachable - {str(e)}"
        )
    
    # Step 2: pyzk Connection
    try:
        from zk import ZK
        
        zk = ZK(ip, port=port, timeout=timeout)
        conn = zk.connect()
        
        if not conn:
            raise HTTPException(
                status_code=401, 
                detail="Authentication Failed - Device rejected connection"
            )
        
        # Get all users from device
        users = conn.get_users()
        user_map = {user.user_id: user.name for user in users}
        
        # Sync users to database
        for user in users:
            await db.biometric_employees.update_one(
                {"biometric_id": str(user.user_id)},
                {"$set": {
                    "biometric_id": str(user.user_id),
                    "name": user.name or f"Employee {user.user_id}",
                    "card_no": getattr(user, 'card', None),
                    "privilege": getattr(user, 'privilege', 0),
                    "updated_at": get_ist_now()
                },
                "$setOnInsert": {
                    "created_at": get_ist_now()
                }},
                upsert=True
            )
        
        # Get attendance records
        attendance = conn.get_attendance()
        
        records_synced = 0
        records_new = 0
        
        for record in attendance:
            user_id = str(record.user_id)
            punch_time = record.timestamp
            
            # Make timezone aware (IST)
            if punch_time.tzinfo is None:
                punch_time = punch_time.replace(tzinfo=IST)
            
            # Create unique record ID
            record_id = f"{user_id}_{punch_time.strftime('%Y%m%d%H%M%S')}"
            
            # Check if record exists
            existing = await db.attendance_records.find_one({"record_id": record_id})
            
            if not existing:
                # Determine shift and status
                shift = get_shift_for_time(punch_time)
                
                await db.attendance_records.insert_one({
                    "record_id": record_id,
                    "biometric_id": user_id,
                    "employee_name": user_map.get(int(user_id), f"Employee {user_id}"),
                    "punch_time": punch_time,
                    "punch_time_ist": punch_time.strftime('%d/%m/%Y %I:%M %p'),
                    "date": punch_time.strftime('%Y-%m-%d'),
                    "time": punch_time.strftime('%H:%M:%S'),
                    "shift": shift["name"],
                    "punch_type": record.punch,  # 0=Check-In, 1=Check-Out
                    "synced_at": get_ist_now()
                })
                records_new += 1
            
            records_synced += 1
        
        conn.disconnect()
        
        return {
            "success": True,
            "status": "TCP Connected",
            "message": f"Synced {records_synced} records ({records_new} new)",
            "total_employees": len(users),
            "total_records": records_synced,
            "new_records": records_new,
            "synced_at": get_ist_now().isoformat()
        }
        
    except ImportError:
        raise HTTPException(status_code=500, detail="pyzk library not installed")
    except HTTPException:
        # Re-raise HTTP exceptions as-is
        raise
    except Exception as e:
        error_msg = str(e).lower()
        # Check for authentication-related errors
        if "password" in error_msg or "auth" in error_msg or "permission" in error_msg or "access" in error_msg:
            raise HTTPException(
                status_code=401, 
                detail=f"Authentication Failed - {str(e)}"
            )
        else:
            raise HTTPException(
                status_code=500, 
                detail=f"Sync failed: {str(e)}"
            )


class PushSyncData(BaseModel):
    """Data structure for local sync script to push attendance"""
    sync_key: Optional[str] = None  # Simple API key for authentication
    employees: Optional[List[dict]] = []  # List of employees from device
    attendance_records: Optional[List[dict]] = []  # List of attendance records
    # Alternative field names some scripts use
    records: Optional[List[dict]] = []
    users: Optional[List[dict]] = []
    synced_at: Optional[str] = None

@app.post("/api/attendance/push-sync")
async def push_sync_attendance(request: Request):
    """
    Receive attendance data pushed from local sync script.
    This endpoint is used when the biometric device is on a local network
    and cannot be reached directly from the cloud server.
    Accepts flexible data formats from various sync scripts.
    """
    try:
        # Parse raw JSON to handle various formats
        raw_data = await request.json()
    except:
        raise HTTPException(status_code=400, detail="Invalid JSON data")
    
    # Simple API key validation (optional - skip if not provided)
    SYNC_API_KEY = os.getenv("ATTENDANCE_SYNC_KEY", "shopfloor2024sync")
    sync_key = raw_data.get("sync_key")
    if sync_key and sync_key != SYNC_API_KEY:
        raise HTTPException(status_code=401, detail="Invalid sync key")
    
    try:
        # Get employees from various possible field names
        employees = raw_data.get("employees") or raw_data.get("users") or []
        
        # Get attendance records from various possible field names
        attendance_records = raw_data.get("attendance_records") or raw_data.get("records") or raw_data.get("attendance") or []
        
        # Sync employees
        employees_synced = 0
        for emp in employees:
            emp_id = str(emp.get("user_id") or emp.get("biometric_id") or emp.get("id", ""))
            if not emp_id:
                continue
            await db.biometric_employees.update_one(
                {"biometric_id": emp_id},
                {"$set": {
                    "biometric_id": emp_id,
                    "name": emp.get("name") or f"Employee {emp_id}",
                    "card_no": emp.get("card") or emp.get("card_no"),
                    "privilege": emp.get("privilege", 0),
                    "updated_at": get_ist_now()
                },
                "$setOnInsert": {
                    "created_at": get_ist_now()
                }},
                upsert=True
            )
            employees_synced += 1
        
        # Build user map for names
        user_map = {}
        for emp in employees:
            emp_id = str(emp.get("user_id") or emp.get("biometric_id") or emp.get("id", ""))
            if emp_id:
                user_map[emp_id] = emp.get("name") or f"Employee {emp_id}"
        
        # Sync attendance records
        records_synced = 0
        records_new = 0
        
        for record in attendance_records:
            user_id = str(record.get("user_id") or record.get("biometric_id") or record.get("id", ""))
            if not user_id:
                continue
            
            # Parse timestamp from various formats
            punch_time_str = record.get("timestamp") or record.get("punch_time") or record.get("time")
            if isinstance(punch_time_str, str):
                try:
                    # Try various datetime formats
                    punch_time_str = punch_time_str.replace('Z', '+00:00')
                    if 'T' in punch_time_str:
                        punch_time = datetime.fromisoformat(punch_time_str)
                    else:
                        # Try common formats
                        for fmt in ['%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%Y-%m-%d %H:%M']:
                            try:
                                punch_time = datetime.strptime(punch_time_str, fmt)
                                break
                            except:
                                continue
                        else:
                            punch_time = get_ist_now()
                    
                    if punch_time.tzinfo is None:
                        punch_time = punch_time.replace(tzinfo=IST)
                except:
                    punch_time = get_ist_now()
            else:
                punch_time = get_ist_now()
            
            # Create unique record ID
            record_id = f"{user_id}_{punch_time.strftime('%Y%m%d%H%M%S')}"
            
            # Check if record exists
            existing = await db.attendance_records.find_one({"record_id": record_id})
            
            if not existing:
                # Determine shift
                shift = get_shift_for_time(punch_time)
                
                # Get punch type (0=Check-In, 1=Check-Out)
                punch_type = record.get("punch") or record.get("punch_type") or record.get("status") or 0
                if isinstance(punch_type, str):
                    punch_type = int(punch_type) if punch_type.isdigit() else 0
                
                await db.attendance_records.insert_one({
                    "record_id": record_id,
                    "biometric_id": user_id,
                    "employee_name": user_map.get(user_id, f"Employee {user_id}"),
                    "punch_time": punch_time,
                    "punch_time_ist": punch_time.strftime('%d/%m/%Y %I:%M %p'),
                    "date": punch_time.strftime('%Y-%m-%d'),
                    "time": punch_time.strftime('%H:%M:%S'),
                    "shift": shift["name"],
                    "punch_type": punch_type,
                    "synced_at": get_ist_now(),
                    "sync_source": "local_script"
                })
                records_new += 1
            
            records_synced += 1
        
        logger.info(f"Push sync completed: {employees_synced} employees, {records_new} new records out of {records_synced}")
        
        return {
            "success": True,
            "message": f"Synced {records_synced} records ({records_new} new)",
            "total_employees": employees_synced,
            "total_records": records_synced,
            "new_records": records_new,
            "synced_at": get_ist_now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Push sync failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")


@app.get("/api/attendance/employees")
async def get_biometric_employees(current_user: User = Depends(get_optional_user)):
    """Get all employees registered in biometric system"""
    employees = await db.biometric_employees.find(
        {}, {"_id": 0}
    ).sort("name", 1).to_list(length=500)
    
    return employees


@app.delete("/api/attendance/employees/{biometric_id}")
async def delete_biometric_employee(biometric_id: str, current_user: User = Depends(get_current_user)):
    """Delete a single employee from biometric records"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Delete employee
    result = await db.biometric_employees.delete_one({"biometric_id": biometric_id})
    
    # Optionally delete their attendance records too
    # await db.attendance_records.delete_many({"biometric_id": biometric_id})
    
    return {"success": True, "deleted": result.deleted_count}


class EmployeeNameUpdate(BaseModel):
    biometric_id: str
    name: str


class EmployeeShiftUpdate(BaseModel):
    biometric_id: str
    shifts: List[str]  # List of shift IDs: ["day_8_8", "night_8_6", etc.]


class UserBiometricLink(BaseModel):
    user_id: str
    biometric_id: str


# Shift configuration
EMPLOYEE_SHIFT_CONFIG = {
    "day_8_8": {"name": "Day (8AM-8PM)", "start": "08:00", "end": "20:00", "late_threshold": 15, "type": "day"},
    "day_9_7": {"name": "Day (9AM-7PM)", "start": "09:00", "end": "19:00", "late_threshold": 15, "type": "day"},
    "night_8_6": {"name": "Night (8PM-6AM)", "start": "20:00", "end": "06:00", "late_threshold": 15, "type": "night"},
    "rotating": {"name": "Rotating (Day/Night)", "start": None, "end": None, "late_threshold": 15, "type": "auto"},
    "double": {"name": "Double Shift (Day+Night)", "start": "08:00", "end": "06:00", "late_threshold": 15, "type": "auto"},
}


def get_employee_shift_config(employee_shifts: List[str], punch_hour: int) -> dict:
    """
    Determine which shift config to use for an employee based on their assigned shifts and punch time.
    
    For fixed shifts: Use the exact start time from SHIFT_CONFIG
    For rotating/double: Auto-detect based on punch-in hour (before noon = day, after noon = night)
    
    Returns: {"start_hour": int, "start_min": int, "late_threshold": int, "shift_type": str}
    """
    if not employee_shifts:
        # Default: auto-detect like rotating
        employee_shifts = ["rotating"]
    
    # Check if employee has a fixed shift
    fixed_day_shifts = [s for s in employee_shifts if s in ["day_8_8", "day_9_7"]]
    fixed_night_shifts = [s for s in employee_shifts if s == "night_8_6"]
    auto_shifts = [s for s in employee_shifts if s in ["rotating", "double"]]
    
    # If employee has rotating or double shift, auto-detect based on punch time
    if auto_shifts:
        # Auto-detect: punch before 12 noon = day shift, after = night shift
        if 5 <= punch_hour < 14:  # 5 AM to 2 PM = Day shift (using 2 PM as cutoff)
            # Default day shift times for auto-detect
            return {"start_hour": 8, "start_min": 0, "late_threshold": 15, "shift_type": "Day", "shift_id": "auto_day"}
        else:
            # Night shift
            return {"start_hour": 20, "start_min": 0, "late_threshold": 15, "shift_type": "Night", "shift_id": "auto_night"}
    
    # For fixed shifts, determine which one applies based on punch time
    if fixed_day_shifts and fixed_night_shifts:
        # Employee has both day and night fixed shifts - use punch time to determine
        if 5 <= punch_hour < 14:
            shift_id = fixed_day_shifts[0]
        else:
            shift_id = fixed_night_shifts[0]
    elif fixed_day_shifts:
        shift_id = fixed_day_shifts[0]
    elif fixed_night_shifts:
        shift_id = fixed_night_shifts[0]
    else:
        # Fallback
        return {"start_hour": 8, "start_min": 0, "late_threshold": 15, "shift_type": "Day", "shift_id": "default"}
    
    config = EMPLOYEE_SHIFT_CONFIG[shift_id]
    start_parts = config["start"].split(":")
    return {
        "start_hour": int(start_parts[0]),
        "start_min": int(start_parts[1]),
        "late_threshold": config["late_threshold"],
        "shift_type": config["type"].capitalize() if config["type"] != "auto" else ("Day" if int(start_parts[0]) < 12 else "Night"),
        "shift_id": shift_id
    }


def check_if_late(in_time: str, shift_config: dict) -> bool:
    """Check if employee is late based on their assigned shift config"""
    try:
        in_parts = in_time.split(":")
        in_hour = int(in_parts[0])
        in_min = int(in_parts[1]) if len(in_parts) > 1 else 0
        
        expected_hour = shift_config["start_hour"]
        expected_min = shift_config["start_min"]
        late_threshold = shift_config["late_threshold"]
        
        # Calculate expected time + threshold
        threshold_min = expected_min + late_threshold
        threshold_hour = expected_hour
        if threshold_min >= 60:
            threshold_min -= 60
            threshold_hour += 1
        
        # For night shift (expected hour >= 18), handle the punch time comparison
        if expected_hour >= 18:
            # Night shift: late if punch is after expected + threshold
            if in_hour > threshold_hour or (in_hour == threshold_hour and in_min > threshold_min):
                return True
        else:
            # Day shift: late if punch is after expected + threshold
            if in_hour > threshold_hour or (in_hour == threshold_hour and in_min > threshold_min):
                return True
        
        return False
    except:
        return False


@app.put("/api/attendance/employees/update-shift")
async def update_employee_shift(data: EmployeeShiftUpdate, current_user: User = Depends(get_current_user)):
    """Update employee shift types (can have multiple)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Validate all shift types
    for shift_id in data.shifts:
        if shift_id not in SHIFT_CONFIG:
            raise HTTPException(status_code=400, detail=f"Invalid shift type: {shift_id}. Valid: {list(SHIFT_CONFIG.keys())}")
    
    # Get shift names for display
    shift_names = [EMPLOYEE_SHIFT_CONFIG[s]["name"] for s in data.shifts]
    
    result = await db.biometric_employees.update_one(
        {"biometric_id": data.biometric_id},
        {"$set": {
            "shifts": data.shifts,
            "shift_names": shift_names,
            "updated_at": get_ist_now()
        }}
    )
    
    if result.modified_count == 0 and result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    return {"success": True, "message": f"Shifts updated: {', '.join(shift_names)}"}


@app.get("/api/attendance/shift-types")
async def get_shift_types():
    """Get available shift types"""
    return {
        "shifts": [
            {"id": k, "name": v["name"], "start": v["start"], "end": v["end"]}
            for k, v in SHIFT_CONFIG.items()
        ]
    }


@app.put("/api/attendance/employees/update-name")
async def update_employee_name(data: EmployeeNameUpdate, current_user: User = Depends(get_current_user)):
    """Update employee name"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Update in biometric_employees
    result = await db.biometric_employees.update_one(
        {"biometric_id": data.biometric_id},
        {"$set": {"name": data.name, "updated_at": get_ist_now()}}
    )
    
    # Also update in attendance_records
    await db.attendance_records.update_many(
        {"biometric_id": data.biometric_id},
        {"$set": {"employee_name": data.name}}
    )
    

# =============================================================================
# USER-BIOMETRIC LINKING & PERSONAL ATTENDANCE
# =============================================================================

@app.post("/api/attendance/link-user-biometric")
async def link_user_to_biometric(
    data: UserBiometricLink,
    current_user: User = Depends(get_current_user)
):
    """
    Link a user account to a biometric employee ID.
    Admin only - allows operators to view their own attendance.
    """
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Only Admin can link users to biometric IDs")
    
    # Verify user exists
    user = await db.users.find_one({"user_id": data.user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify biometric employee exists
    emp = await db.biometric_employees.find_one({"biometric_id": data.biometric_id})
    if not emp:
        raise HTTPException(status_code=404, detail="Biometric employee not found")
    
    # Update user with biometric_id
    await db.users.update_one(
        {"user_id": data.user_id},
        {"$set": {
            "biometric_id": data.biometric_id,
            "updated_at": get_ist_now()
        }}
    )
    
    # Also update biometric_employees with user_id link
    await db.biometric_employees.update_one(
        {"biometric_id": data.biometric_id},
        {"$set": {
            "linked_user_id": data.user_id,
            "updated_at": get_ist_now()
        }}
    )
    
    return {
        "status": "linked",
        "user_id": data.user_id,
        "user_name": user.get("name"),
        "biometric_id": data.biometric_id,
        "employee_name": emp.get("name")
    }


@app.delete("/api/attendance/unlink-user-biometric/{user_id}")
async def unlink_user_biometric(
    user_id: str,
    current_user: User = Depends(get_current_user)
):
    """Remove biometric link from a user"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Only Admin can unlink users")
    
    user = await db.users.find_one({"user_id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    old_biometric_id = user.get("biometric_id")
    
    # Remove biometric_id from user
    await db.users.update_one(
        {"user_id": user_id},
        {"$unset": {"biometric_id": ""}}
    )
    
    # Remove linked_user_id from biometric_employees
    if old_biometric_id:
        await db.biometric_employees.update_one(
            {"biometric_id": old_biometric_id},
            {"$unset": {"linked_user_id": ""}}
        )
    
    return {"status": "unlinked", "user_id": user_id}


@app.get("/api/attendance/users-biometric-mapping")
async def get_users_biometric_mapping(current_user: User = Depends(get_current_user)):
    """
    Get all users with their biometric ID mappings.
    Admin only - for managing user-biometric links.
    """
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Only Admin can view all mappings")
    
    # Get all users
    users = await db.users.find({}, {"_id": 0, "user_id": 1, "name": 1, "email": 1, "role": 1, "biometric_id": 1}).to_list(500)
    
    # Get all biometric employees
    employees = await db.biometric_employees.find({}, {"_id": 0, "biometric_id": 1, "name": 1, "linked_user_id": 1}).to_list(500)
    
    return {
        "users": users,
        "biometric_employees": employees
    }


@app.get("/api/attendance/my-attendance")
async def get_my_attendance(
    month: Optional[str] = None,  # Format: YYYY-MM
    year: Optional[str] = None,   # Format: YYYY (for yearly summary)
    current_user: User = Depends(get_current_user)
):
    """
    Get attendance for the currently logged-in user.
    User must have a linked biometric_id.
    
    If month is provided: Returns daily attendance for that month
    If year is provided: Returns yearly summary
    If neither: Returns current month
    """
    # Get user's biometric_id
    user = await db.users.find_one({"user_id": current_user.user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    biometric_id = user.get("biometric_id")
    if not biometric_id:
        raise HTTPException(status_code=400, detail="Your account is not linked to a biometric ID. Contact admin.")
    
    # Get employee name
    emp = await db.biometric_employees.find_one({"biometric_id": biometric_id})
    emp_name = emp.get("name", current_user.name) if emp else current_user.name
    emp_shifts = emp.get("shifts", []) if emp else []
    
    now = get_ist_now()
    
    if year and not month:
        # Yearly summary
        year_int = int(year)
        
        # Get all records for the year
        start_date = f"{year_int}-01-01"
        end_date = f"{year_int}-12-31"
        
        records = await db.attendance_records.find({
            "biometric_id": biometric_id,
            "date": {"$gte": start_date, "$lte": end_date}
        }).to_list(length=400)
        
        # Calculate yearly stats
        monthly_stats = {}
        for i in range(1, 13):
            month_key = f"{year_int}-{str(i).zfill(2)}"
            monthly_stats[month_key] = {
                "present": 0,
                "late": 0,
                "absent": 0,
                "total_hours": 0
            }
        
        # Group records by date to count days
        days_with_records = {}
        for r in records:
            date = r.get("date", "")
            if date and len(date) >= 7:
                month_key = date[:7]
                if date not in days_with_records:
                    days_with_records[date] = {"month": month_key, "times": []}
                days_with_records[date]["times"].append(r.get("time", ""))
        
        # Process each day
        for date, data in days_with_records.items():
            month_key = data["month"]
            if month_key in monthly_stats:
                times = sorted(data["times"])
                in_time = times[0] if times else None
                out_time = times[-1] if len(times) > 1 else None
                
                # Check if late
                is_late = False
                if in_time:
                    try:
                        in_h = int(in_time.split(":")[0])
                        shift_config = get_employee_shift_config(emp_shifts, in_h)
                        is_late = check_if_late(in_time, shift_config)
                    except:
                        pass
                
                if is_late:
                    monthly_stats[month_key]["late"] += 1
                else:
                    monthly_stats[month_key]["present"] += 1
                
                # Calculate hours
                if in_time and out_time and in_time != out_time:
                    try:
                        in_parts = in_time.split(":")
                        out_parts = out_time.split(":")
                        in_mins = int(in_parts[0]) * 60 + int(in_parts[1])
                        out_mins = int(out_parts[0]) * 60 + int(out_parts[1])
                        if out_mins < in_mins:
                            out_mins += 24 * 60  # Next day
                        hours = (out_mins - in_mins) / 60
                        monthly_stats[month_key]["total_hours"] += round(hours, 1)
                    except:
                        pass
        
        return {
            "type": "yearly",
            "year": year_int,
            "employee": {
                "biometric_id": biometric_id,
                "name": emp_name
            },
            "monthly_stats": monthly_stats,
            "totals": {
                "present": sum(m["present"] for m in monthly_stats.values()),
                "late": sum(m["late"] for m in monthly_stats.values()),
                "total_hours": round(sum(m["total_hours"] for m in monthly_stats.values()), 1)
            }
        }
    
    else:
        # Monthly detailed view
        if month:
            target_month = month
        else:
            target_month = now.strftime("%Y-%m")
        
        year_int, month_int = map(int, target_month.split("-"))
        
        # Get all records for the month
        records = await db.attendance_records.find({
            "biometric_id": biometric_id,
            "date": {"$regex": f"^{target_month}"}
        }).sort("date", 1).to_list(length=100)
        
        # Group by date
        days = {}
        for r in records:
            date = r.get("date", "")
            time = r.get("time", "")
            if date and time:
                if date not in days:
                    days[date] = []
                days[date].append(time)
        
        # Build daily attendance
        import calendar
        num_days = calendar.monthrange(year_int, month_int)[1]
        
        daily_attendance = []
        total_present = 0
        total_late = 0
        total_absent = 0
        total_hours = 0
        
        for day in range(1, num_days + 1):
            date_str = f"{target_month}-{str(day).zfill(2)}"
            day_date = datetime(year_int, month_int, day)
            
            # Skip future dates
            if day_date.date() > now.date():
                continue
            
            # Skip weekends (optional - uncomment if needed)
            # if day_date.weekday() >= 5:
            #     continue
            
            day_data = {
                "date": date_str,
                "day": day,
                "weekday": day_date.strftime("%a"),
                "status": "absent",
                "in_time": None,
                "out_time": None,
                "hours": 0,
                "is_late": False,
                "shift": None
            }
            
            if date_str in days:
                times = sorted(days[date_str])
                in_time = times[0]
                out_time = times[-1] if len(times) > 1 else None
                
                day_data["in_time"] = in_time
                day_data["out_time"] = out_time if out_time != in_time else None
                
                # Determine shift and late status
                try:
                    in_h = int(in_time.split(":")[0])
                    shift_config = get_employee_shift_config(emp_shifts, in_h)
                    day_data["shift"] = shift_config.get("shift_type", "Day")
                    day_data["is_late"] = check_if_late(in_time, shift_config)
                except:
                    pass
                
                # Calculate hours
                if in_time and out_time and out_time != in_time:
                    try:
                        in_parts = in_time.split(":")
                        out_parts = out_time.split(":")
                        in_mins = int(in_parts[0]) * 60 + int(in_parts[1])
                        out_mins = int(out_parts[0]) * 60 + int(out_parts[1])
                        if out_mins < in_mins:
                            out_mins += 24 * 60
                        hours = (out_mins - in_mins) / 60
                        day_data["hours"] = round(hours, 1)
                        total_hours += hours
                    except:
                        pass
                
                if day_data["is_late"]:
                    day_data["status"] = "late"
                    total_late += 1
                else:
                    day_data["status"] = "present"
                    total_present += 1
            else:
                total_absent += 1
            
            daily_attendance.append(day_data)
        
        return {
            "type": "monthly",
            "month": target_month,
            "employee": {
                "biometric_id": biometric_id,
                "name": emp_name,
                "shifts": emp_shifts
            },
            "summary": {
                "present": total_present,
                "late": total_late,
                "absent": total_absent,
                "total_hours": round(total_hours, 1),
                "working_days": total_present + total_late + total_absent
            },
            "daily": daily_attendance
        }



@app.post("/api/attendance/import-data")
async def import_attendance_data(current_user: User = Depends(get_optional_user)):
    """Import attendance data from JSON files for migration"""
    import httpx
    
    base_url = "https://shop-floor-track.preview.emergentagent.com"
    
    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            # Import employees
            print("Fetching employees...")
            emp_response = await client.get(f"{base_url}/attendance_employees.json")
            employees = emp_response.json()
            
            if employees:
                await db.biometric_employees.delete_many({})
                await db.biometric_employees.insert_many(employees)
                emp_count = len(employees)
            else:
                emp_count = 0
            
            # Import attendance records
            print("Fetching attendance records...")
            att_response = await client.get(f"{base_url}/attendance_records.json")
            records = att_response.json()
            
            if records:
                await db.attendance_records.delete_many({})
                # Insert in batches
                batch_size = 2000
                for i in range(0, len(records), batch_size):
                    batch = records[i:i+batch_size]
                    await db.attendance_records.insert_many(batch)
                att_count = len(records)
            else:
                att_count = 0
            
            return {
                "success": True,
                "employees_imported": emp_count,
                "attendance_records_imported": att_count
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


    return {"success": True, "updated": result.modified_count}


@app.delete("/api/attendance/employees-clear-all")
async def clear_all_biometric_employees(current_user: User = Depends(get_current_user)):
    """Clear ALL employees from biometric records (use with caution)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await db.biometric_employees.delete_many({})
    
    return {"success": True, "deleted": result.deleted_count, "message": "All employees cleared"}


@app.delete("/api/attendance/records-clear-all") 
async def clear_all_attendance_records(current_user: User = Depends(get_current_user)):
    """Clear ALL attendance records (use with caution)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await db.attendance_records.delete_many({})
    
    return {"success": True, "deleted": result.deleted_count, "message": "All attendance records cleared"}


# Temporary migration endpoint - delete after use
@app.post("/api/attendance/migration-clear-all")
async def migration_clear_all_attendance():
    """One-time migration endpoint to clear all attendance data"""
    emp_result = await db.biometric_employees.delete_many({})
    att_result = await db.attendance_records.delete_many({})
    
    return {
        "success": True,
        "employees_deleted": emp_result.deleted_count,
        "records_deleted": att_result.deleted_count,
        "message": "All attendance data cleared for migration"
    }


@app.post("/api/attendance/sync-clean")
async def sync_and_clean_employees(current_user: User = Depends(get_current_user)):
    """Re-sync employees - removes old ones not in latest sync"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # This will be called after a fresh sync to clean up
    # For now, just return current count
    count = await db.biometric_employees.count_documents({})
    return {"success": True, "current_employees": count}


class PunchTimeEdit(BaseModel):
    biometric_id: str
    date: str  # YYYY-MM-DD
    in_time: Optional[str] = None  # HH:MM:SS or HH:MM
    out_time: Optional[str] = None  # HH:MM:SS or HH:MM


@app.post("/api/attendance/edit-punch")
async def edit_punch_time(data: PunchTimeEdit, current_user: User = Depends(get_current_user)):
    """Edit or add punch in/out time for an employee"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        # Parse date
        date_obj = datetime.strptime(data.date, '%Y-%m-%d')
        date_obj = date_obj.replace(tzinfo=IST)
        
        # Get employee name
        employee = await db.biometric_employees.find_one({"biometric_id": data.biometric_id})
        emp_name = employee.get("name", f"Employee {data.biometric_id}") if employee else f"Employee {data.biometric_id}"
        
        # Process IN time
        if data.in_time:
            in_time_parts = data.in_time.split(":")
            in_hour = int(in_time_parts[0])
            in_min = int(in_time_parts[1])
            in_sec = int(in_time_parts[2]) if len(in_time_parts) > 2 else 0
            
            in_datetime = date_obj.replace(hour=in_hour, minute=in_min, second=in_sec)
            in_record_id = f"{data.biometric_id}_{in_datetime.strftime('%Y%m%d%H%M%S')}"
            
            # Check if record exists for this date (any time)
            existing_in = await db.attendance_records.find_one({
                "biometric_id": data.biometric_id,
                "date": data.date,
                "punch_type": 0
            })
            
            if existing_in:
                # Update existing
                await db.attendance_records.update_one(
                    {"_id": existing_in["_id"]},
                    {"$set": {
                        "time": f"{in_hour:02d}:{in_min:02d}:{in_sec:02d}",
                        "punch_time": in_datetime,
                        "punch_time_ist": in_datetime.strftime('%d/%m/%Y %I:%M %p'),
                        "updated_at": get_ist_now(),
                        "updated_by": current_user.email,
                        "manual_edit": True
                    }}
                )
            else:
                # Create new
                shift = get_shift_for_time(in_datetime)
                await db.attendance_records.insert_one({
                    "record_id": in_record_id,
                    "biometric_id": data.biometric_id,
                    "employee_name": emp_name,
                    "punch_time": in_datetime,
                    "punch_time_ist": in_datetime.strftime('%d/%m/%Y %I:%M %p'),
                    "date": data.date,
                    "time": f"{in_hour:02d}:{in_min:02d}:{in_sec:02d}",
                    "shift": shift["name"],
                    "punch_type": 0,  # Check-In
                    "synced_at": get_ist_now(),
                    "sync_source": "manual",
                    "created_by": current_user.email,
                    "manual_edit": True
                })
        
        # Process OUT time
        if data.out_time:
            out_time_parts = data.out_time.split(":")
            out_hour = int(out_time_parts[0])
            out_min = int(out_time_parts[1])
            out_sec = int(out_time_parts[2]) if len(out_time_parts) > 2 else 0
            
            out_datetime = date_obj.replace(hour=out_hour, minute=out_min, second=out_sec)
            out_record_id = f"{data.biometric_id}_{out_datetime.strftime('%Y%m%d%H%M%S')}"
            
            # Check if out record exists
            existing_out = await db.attendance_records.find_one({
                "biometric_id": data.biometric_id,
                "date": data.date,
                "punch_type": 1
            })
            
            if existing_out:
                # Update existing
                await db.attendance_records.update_one(
                    {"_id": existing_out["_id"]},
                    {"$set": {
                        "time": f"{out_hour:02d}:{out_min:02d}:{out_sec:02d}",
                        "punch_time": out_datetime,
                        "punch_time_ist": out_datetime.strftime('%d/%m/%Y %I:%M %p'),
                        "updated_at": get_ist_now(),
                        "updated_by": current_user.email,
                        "manual_edit": True
                    }}
                )
            else:
                # Create new
                shift = get_shift_for_time(out_datetime)
                await db.attendance_records.insert_one({
                    "record_id": out_record_id,
                    "biometric_id": data.biometric_id,
                    "employee_name": emp_name,
                    "punch_time": out_datetime,
                    "punch_time_ist": out_datetime.strftime('%d/%m/%Y %I:%M %p'),
                    "date": data.date,
                    "time": f"{out_hour:02d}:{out_min:02d}:{out_sec:02d}",
                    "shift": shift["name"],
                    "punch_type": 1,  # Check-Out
                    "synced_at": get_ist_now(),
                    "sync_source": "manual",
                    "created_by": current_user.email,
                    "manual_edit": True
                })
        
        return {
            "success": True,
            "message": "Punch time updated successfully",
            "biometric_id": data.biometric_id,
            "date": data.date
        }
        
    except Exception as e:
        logger.error(f"Error editing punch time: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update: {str(e)}")


@app.get("/api/attendance/export-pdf")
async def export_attendance_pdf(
    biometric_id: Optional[str] = None,
    period: str = "monthly",  # monthly or yearly
    month: Optional[str] = None,  # YYYY-MM for monthly
    year: Optional[str] = None,  # YYYY for yearly
    current_user: User = Depends(get_optional_user)
):
    """Export attendance statement as PDF for a single employee or all employees"""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from calendar import monthrange
    
    now = get_ist_now()
    
    # Determine date range
    if period == "yearly":
        # Yearly = Last 12 months from current month
        end_date = datetime(now.year, now.month, 1, tzinfo=IST) + timedelta(days=32)
        end_date = end_date.replace(day=1)  # First of next month
        
        # Go back 12 months
        start_year = now.year - 1 if now.month == 1 else now.year
        start_month = now.month  # Same month last year
        if now.month == 12:
            start_date = datetime(now.year - 1, 12, 1, tzinfo=IST)
        else:
            start_date = datetime(now.year - 1, now.month + 1, 1, tzinfo=IST)
        
        # Actually simpler: go back 11 months from current month
        start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        for _ in range(11):
            start_date = (start_date - timedelta(days=1)).replace(day=1)
        
        period_display = f"{start_date.strftime('%b %Y')} - {now.strftime('%b %Y')}"
    else:  # monthly
        if month:
            try:
                y, m = map(int, month.split('-'))
                start_date = datetime(y, m, 1, tzinfo=IST)
            except:
                start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        else:
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year + 1, month=1)
        else:
            end_date = start_date.replace(month=start_date.month + 1)
        period_display = start_date.strftime('%B %Y')
    
    # Get employee(s)
    if biometric_id:
        employees = await db.biometric_employees.find({"biometric_id": biometric_id}).to_list(1)
    else:
        employees = await db.biometric_employees.find().sort("name", 1).to_list(500)
    
    if not employees:
        raise HTTPException(status_code=404, detail="No employees found")
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        topMargin=0.4*cm, 
        bottomMargin=0.4*cm,
        leftMargin=0.5*cm,
        rightMargin=0.5*cm
    )
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=12, alignment=1, spaceAfter=2)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=9, alignment=1, textColor=colors.grey, spaceAfter=4)
    month_header_style = ParagraphStyle('MonthHeader', parent=styles['Heading3'], fontSize=9, spaceAfter=2, spaceBefore=4)
    
    # Extended date range for night shift handling
    extended_start = start_date - timedelta(days=1)
    extended_end = end_date + timedelta(days=1)
    
    for emp in employees:
        emp_id = emp.get("biometric_id")
        emp_name = emp.get("name", f"Employee {emp_id}")
        
        # Title
        title = f"Attendance Statement - {emp_name} (ID: {emp_id})"
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f"Period: {period_display}", subtitle_style))
        
        # Get attendance records
        records = await db.attendance_records.find({
            "biometric_id": emp_id,
            "date": {
                "$gte": extended_start.strftime('%Y-%m-%d'),
                "$lt": extended_end.strftime('%Y-%m-%d')
            }
        }).sort([("date", 1), ("time", 1)]).to_list(2000)
        
        # Group records by date
        punches_by_date = {}
        for rec in records:
            date = rec.get("date")
            time = rec.get("time", "")[:5]
            if date not in punches_by_date:
                punches_by_date[date] = []
            punches_by_date[date].append(time)
        
        # Process attendance with night shift handling
        attendance_by_date = {}
        for date_str, times in sorted(punches_by_date.items()):
            if times:
                times.sort()
                first_punch = times[0]
                last_punch = times[-1] if len(times) > 1 else None
                
                first_hour = int(first_punch.split(":")[0])
                
                if first_hour < 8:
                    prev_date = (datetime.strptime(date_str, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
                    if prev_date in attendance_by_date:
                        prev_att = attendance_by_date[prev_date]
                        if prev_att.get("in") and not prev_att.get("out"):
                            prev_in_hour = int(prev_att["in"].split(":")[0])
                            if prev_in_hour >= 17:
                                prev_att["out"] = first_punch
                                prev_att["out_next_day"] = True
                                if len(times) > 1:
                                    first_punch = times[1]
                                    last_punch = times[-1] if len(times) > 2 else None
                                else:
                                    continue
                
                attendance_by_date[date_str] = {
                    "in": first_punch,
                    "out": last_punch,
                    "out_next_day": False
                }
        
        # Check for night shifts needing OUT from next day
        for date_str in list(attendance_by_date.keys()):
            att = attendance_by_date[date_str]
            if att.get("in") and not att.get("out"):
                in_hour = int(att["in"].split(":")[0])
                if in_hour >= 17:
                    next_date = (datetime.strptime(date_str, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')
                    if next_date in punches_by_date:
                        next_times = sorted(punches_by_date[next_date])
                        if next_times:
                            next_first_hour = int(next_times[0].split(":")[0])
                            if next_first_hour < 8:
                                att["out"] = next_times[0]
                                att["out_next_day"] = True
        
        # YEARLY REPORT - Last 12 months, each month on one page
        if period == "yearly":
            grand_total_present = 0
            grand_total_late = 0
            grand_total_absent = 0
            grand_total_working_days = 0
            grand_total_hours = 0
            monthly_summaries = []
            
            # Generate list of 12 months from start_date
            months_to_process = []
            current_month = start_date
            for _ in range(12):
                if current_month <= now:
                    months_to_process.append(current_month)
                # Move to next month
                if current_month.month == 12:
                    current_month = current_month.replace(year=current_month.year + 1, month=1)
                else:
                    current_month = current_month.replace(month=current_month.month + 1)
            
            for month_idx, month_start in enumerate(months_to_process):
                month_year = month_start.year
                month_num = month_start.month
                days_in_month = monthrange(month_year, month_num)[1]
                month_end = month_start + timedelta(days=days_in_month)
                month_name = month_start.strftime('%B %Y')
                
                # Page break before each month (except first)
                if month_idx > 0:
                    elements.append(PageBreak())
                
                # Month title
                elements.append(Paragraph(f"<b>{month_name}</b>", ParagraphStyle('MonthTitle', fontSize=11, alignment=1, spaceAfter=6, spaceBefore=2)))
                
                # Working days for this month (exclude Sundays and Holidays)
                working_days = 0
                holidays_in_month = 0
                for d in range(days_in_month):
                    check_date = month_start + timedelta(days=d)
                    check_date_str = check_date.strftime('%Y-%m-%d')
                    if check_date.weekday() != 6:  # Not Sunday
                        if get_holiday_name(check_date_str):
                            holidays_in_month += 1
                        else:
                            working_days += 1
                
                table_data = [["Date", "Day", "IN", "OUT", "Status", "Hours"]]
                month_present = 0
                month_late = 0
                month_hours = 0
                month_holidays = 0
                
                current = month_start
                while current < month_end and current <= now:
                    date_str = current.strftime('%Y-%m-%d')
                    day_name = current.strftime('%a')
                    holiday_name = get_holiday_name(date_str)
                    
                    if current.weekday() == 6:  # Sunday
                        table_data.append([current.strftime('%d'), day_name, "-", "-", "Off", "-"])
                    elif holiday_name:  # Festival/Holiday
                        month_holidays += 1
                        # Truncate holiday name if too long
                        short_name = holiday_name[:8] if len(holiday_name) > 8 else holiday_name
                        table_data.append([current.strftime('%d'), day_name, "-", "-", short_name, "-"])
                    elif date_str in attendance_by_date:
                        att = attendance_by_date[date_str]
                        in_time = att.get("in") or "-"
                        out_time = att.get("out") or "-"
                        out_next = att.get("out_next_day", False)
                        
                        if out_next and out_time != "-":
                            out_time = out_time + "*"
                        
                        hours_str = "-"
                        if att.get("in") and att.get("out"):
                            try:
                                in_parts = att["in"].split(":")
                                out_parts = att["out"].split(":")
                                in_mins = int(in_parts[0]) * 60 + int(in_parts[1])
                                out_mins = int(out_parts[0]) * 60 + int(out_parts[1])
                                if out_next or out_mins < in_mins:
                                    out_mins += 24 * 60
                                diff_hours = (out_mins - in_mins) / 60
                                hours_str = f"{diff_hours:.1f}"
                                month_hours += diff_hours
                            except:
                                pass
                        
                        status = "P"
                        if att.get("in"):
                            try:
                                in_hour = int(att["in"].split(":")[0])
                                in_min = int(att["in"].split(":")[1])
                                if in_hour < 17:
                                    if in_hour > 8 or (in_hour == 8 and in_min > 15):
                                        status = "L"
                                        month_late += 1
                                else:
                                    if in_hour > 20 or (in_hour == 20 and in_min > 15):
                                        status = "L"
                                        month_late += 1
                                month_present += 1
                            except:
                                month_present += 1
                        
                        table_data.append([current.strftime('%d'), day_name, in_time, out_time, status, hours_str])
                    else:
                        table_data.append([current.strftime('%d'), day_name, "-", "-", "A", "-"])
                    
                    current += timedelta(days=1)
                
                month_absent = working_days - month_present
                
                # Month summary row
                att_pct = round(month_present / working_days * 100, 1) if working_days > 0 else 0
                table_data.append(["", "", "", "", "", ""])
                table_data.append([f"P:{month_present}", f"L:{month_late}", f"A:{month_absent}", f"H:{month_holidays}", f"{att_pct}%", f"{month_hours:.0f}h"])
                
                col_widths = [0.9*cm, 0.9*cm, 1.3*cm, 1.3*cm, 1*cm, 1*cm]
                table = Table(table_data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('GRID', (0, 0), (-1, -3), 0.4, colors.grey),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E2E8F0')),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ]))
                
                elements.append(table)
                
                grand_total_present += month_present
                grand_total_late += month_late
                grand_total_absent += month_absent
                grand_total_working_days += working_days
                grand_total_hours += month_hours
                
                monthly_summaries.append({
                    "month": month_start.strftime('%b %Y'),
                    "present": month_present,
                    "late": month_late,
                    "absent": month_absent,
                    "holidays": month_holidays,
                    "working_days": working_days,
                    "hours": month_hours,
                    "att_pct": att_pct
                })
            
            # DETAILED YEARLY SUMMARY - on new page
            elements.append(PageBreak())
            elements.append(Paragraph("<b>YEARLY SUMMARY</b>", ParagraphStyle('SummaryTitle', fontSize=14, alignment=1, spaceAfter=6)))
            elements.append(Paragraph(f"{emp_name} (ID: {emp_id})", ParagraphStyle('EmpSubtitle', fontSize=10, alignment=1, textColor=colors.grey, spaceAfter=10)))
            elements.append(Paragraph(f"Period: {period_display}", ParagraphStyle('Period', fontSize=9, alignment=1, textColor=colors.grey, spaceAfter=15)))
            
            # Calculate total holidays
            grand_total_holidays = sum(ms.get("holidays", 0) for ms in monthly_summaries)
            
            # Detailed Monthly Summary Table
            summary_data = [["Month", "Work\nDays", "Present", "Late", "Absent", "Holidays", "Att%", "Hours"]]
            
            for ms in monthly_summaries:
                summary_data.append([
                    ms["month"],
                    str(ms["working_days"]),
                    str(ms["present"]),
                    str(ms["late"]),
                    str(ms["absent"]),
                    str(ms.get("holidays", 0)),
                    f"{ms['att_pct']}%",
                    f"{ms['hours']:.0f}h"
                ])
            
            # Grand total row
            overall_att_pct = round(grand_total_present / grand_total_working_days * 100, 1) if grand_total_working_days > 0 else 0
            overall_avg_hrs = grand_total_hours / grand_total_present if grand_total_present > 0 else 0
            
            summary_data.append(["", "", "", "", "", "", "", ""])
            summary_data.append([
                "TOTAL",
                str(grand_total_working_days),
                str(grand_total_present),
                str(grand_total_late),
                str(grand_total_absent),
                str(grand_total_holidays),
                f"{overall_att_pct}%",
                f"{grand_total_hours:.0f}h"
            ])
            
            summary_col_widths = [1.6*cm, 1.1*cm, 1.1*cm, 0.9*cm, 1.1*cm, 1.1*cm, 1*cm, 1.1*cm]
            summary_table = Table(summary_data, colWidths=summary_col_widths)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -3), 0.5, colors.grey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF3C7')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            
            elements.append(summary_table)
            elements.append(Spacer(1, 0.5*cm))
            
            # Key Statistics Box
            stats_text = f"""
            <b>KEY STATISTICS</b><br/><br/>
            <b>Total Working Days:</b> {grand_total_working_days}<br/>
            <b>Days Present:</b> {grand_total_present}<br/>
            <b>Days Absent:</b> {grand_total_absent}<br/>
            <b>Late Arrivals:</b> {grand_total_late}<br/>
            <b>Paid Holidays:</b> {grand_total_holidays}<br/><br/>
            <b>Overall Attendance:</b> {overall_att_pct}%<br/>
            <b>Total Hours Worked:</b> {grand_total_hours:.0f} hours<br/>
            <b>Average Hours/Day:</b> {overall_avg_hrs:.1f} hours<br/>
            """
            
            stats_style = ParagraphStyle('Stats', fontSize=9, leading=14, leftIndent=20)
            elements.append(Paragraph(stats_text, stats_style))
            
            # Night shift footnote
            if any(att.get("out_next_day") for att in attendance_by_date.values()):
                elements.append(Spacer(1, 0.3*cm))
                footnote = Paragraph("<i>* OUT time is from next day (night shift)</i>", 
                                   ParagraphStyle('Footnote', fontSize=7, textColor=colors.grey))
                elements.append(footnote)
        
        else:
            # MONTHLY REPORT
            # Calculate working days (exclude Sundays and Holidays)
            working_days = 0
            total_holidays = 0
            for d in range((end_date - start_date).days):
                check_date = start_date + timedelta(days=d)
                check_date_str = check_date.strftime('%Y-%m-%d')
                if check_date.weekday() != 6:  # Not Sunday
                    if get_holiday_name(check_date_str):
                        total_holidays += 1
                    else:
                        working_days += 1
            
            table_data = [["Date", "Day", "IN", "OUT", "Status", "Hrs"]]
            total_present = 0
            total_late = 0
            total_hours = 0
            
            current = start_date
            while current < end_date:
                date_str = current.strftime('%Y-%m-%d')
                day_name = current.strftime('%a')
                holiday_name = get_holiday_name(date_str)
                
                if current.weekday() == 6:  # Sunday
                    table_data.append([current.strftime('%d'), day_name, "-", "-", "Off", "-"])
                elif holiday_name:  # Festival/Holiday
                    short_name = holiday_name[:8] if len(holiday_name) > 8 else holiday_name
                    table_data.append([current.strftime('%d'), day_name, "-", "-", short_name, "-"])
                elif date_str in attendance_by_date:
                    att = attendance_by_date[date_str]
                    in_time = att.get("in") or "-"
                    out_time = att.get("out") or "-"
                    out_next = att.get("out_next_day", False)
                    
                    if out_next and out_time != "-":
                        out_time = out_time + "*"
                    
                    hours_str = "-"
                    if att.get("in") and att.get("out"):
                        try:
                            in_parts = att["in"].split(":")
                            out_parts = att["out"].split(":")
                            in_mins = int(in_parts[0]) * 60 + int(in_parts[1])
                            out_mins = int(out_parts[0]) * 60 + int(out_parts[1])
                            if out_next or out_mins < in_mins:
                                out_mins += 24 * 60
                            diff_hours = (out_mins - in_mins) / 60
                            hours_str = f"{diff_hours:.1f}"
                            total_hours += diff_hours
                        except:
                            pass
                    
                    status = "P"
                    if att.get("in"):
                        try:
                            in_hour = int(att["in"].split(":")[0])
                            in_min = int(att["in"].split(":")[1])
                            if in_hour < 17:
                                if in_hour > 8 or (in_hour == 8 and in_min > 15):
                                    status = "L"
                                    total_late += 1
                            else:
                                if in_hour > 20 or (in_hour == 20 and in_min > 15):
                                    status = "L"
                                    total_late += 1
                            total_present += 1
                        except:
                            total_present += 1
                    
                    table_data.append([current.strftime('%d'), day_name, in_time, out_time, status, hours_str])
                else:
                    table_data.append([current.strftime('%d'), day_name, "-", "-", "A", "-"])
                
                current += timedelta(days=1)
            
            att_pct = round(total_present / working_days * 100, 1) if working_days > 0 else 0
            table_data.append(["", "", "", "", "", ""])
            table_data.append([f"P:{total_present}", f"L:{total_late}", f"A:{working_days - total_present}", f"H:{total_holidays}", f"{att_pct}%", f"{total_hours:.0f}h"])
            
            col_widths = [0.9*cm, 0.9*cm, 1.3*cm, 1.3*cm, 1*cm, 1*cm]
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
                ('GRID', (0, 0), (-1, -3), 0.5, colors.grey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            
            elements.append(table)
            
            if any(att.get("out_next_day") for att in attendance_by_date.values()):
                elements.append(Spacer(1, 0.1*cm))
                footnote = Paragraph("<i>* OUT from next day (night shift)</i>", 
                                   ParagraphStyle('Footnote', fontSize=6, textColor=colors.grey))
                elements.append(footnote)
        
        if not biometric_id and emp != employees[-1]:
            elements.append(PageBreak())
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"attendance_{biometric_id or 'all'}_{period}_{period_display.replace(' ', '_')}.pdf"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



@app.get("/api/attendance/debug-user")
async def debug_user_attendance(
    email: str,
    month: Optional[str] = None,  # YYYY-MM
    current_user: User = Depends(get_current_user)
):
    """Debug endpoint to check attendance data for a specific user (Admin only)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    now = get_ist_now()
    
    # Parse month or use current
    if month:
        try:
            year, mon = map(int, month.split('-'))
            month_start = datetime(year, mon, 1, tzinfo=IST)
        except:
            month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Calculate month end
    if month_start.month == 12:
        month_end = month_start.replace(year=month_start.year + 1, month=1)
    else:
        month_end = month_start.replace(month=month_start.month + 1)
    
    # Find user by email
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user:
        return {"error": f"User with email {email} not found in users collection"}
    
    # Find biometric link
    biometric_link = await db.biometric_user_links.find_one({"user_id": user.get("user_id")}, {"_id": 0})
    biometric_id = biometric_link.get("biometric_id") if biometric_link else None
    
    # Find biometric employee directly by name match
    biometric_emp_by_name = await db.biometric_employees.find_one(
        {"name": {"$regex": user.get("name", "").split()[0], "$options": "i"}},
        {"_id": 0}
    )
    
    # Count punch records
    punch_count = 0
    punch_records = []
    if biometric_id:
        punch_count = await db.attendance_records.count_documents({
            "biometric_id": str(biometric_id),
            "date": {"$gte": month_start.strftime('%Y-%m-%d'), "$lt": month_end.strftime('%Y-%m-%d')}
        })
        punch_records = await db.attendance_records.find({
            "biometric_id": str(biometric_id),
            "date": {"$gte": month_start.strftime('%Y-%m-%d'), "$lt": month_end.strftime('%Y-%m-%d')}
        }, {"_id": 0}).limit(10).to_list(length=10)
    
    # Also check attendance by name
    punch_by_name_count = await db.attendance_records.count_documents({
        "name": {"$regex": user.get("name", "").split()[0], "$options": "i"},
        "date": {"$gte": month_start.strftime('%Y-%m-%d'), "$lt": month_end.strftime('%Y-%m-%d')}
    })
    
    return {
        "user": user,
        "biometric_link": biometric_link,
        "biometric_id_linked": biometric_id,
        "biometric_employee_by_name": biometric_emp_by_name,
        "month_range": f"{month_start.strftime('%Y-%m-%d')} to {month_end.strftime('%Y-%m-%d')}",
        "punch_records_count": punch_count,
        "punch_records_sample": punch_records,
        "punch_by_name_count": punch_by_name_count,
        "diagnosis": {
            "has_biometric_link": biometric_link is not None,
            "has_punch_records": punch_count > 0,
            "possible_issue": (
                "No biometric link found" if not biometric_link else
                "No punch records for this biometric_id - employee may be absent OR biometric_id is wrong" if punch_count == 0 else
                f"Has {punch_count} punch records for the month"
            )
        }
    }


@app.get("/api/attendance/monthly")
async def get_monthly_attendance(
    month: Optional[str] = None,
    employee_id: Optional[str] = None,
    current_user: User = Depends(get_optional_user)
):
    """Get monthly attendance summary"""
    now = get_ist_now()
    
    # Parse month or use current
    if month:
        try:
            year, mon = map(int, month.split('-'))
            month_start = datetime(year, mon, 1, tzinfo=IST)
        except:
            month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Calculate month end
    if month_start.month == 12:
        month_end = month_start.replace(year=month_start.year + 1, month=1)
    else:
        month_end = month_start.replace(month=month_start.month + 1)
    
    # Build query
    query = {
        "date": {
            "$gte": month_start.strftime('%Y-%m-%d'),
            "$lt": month_end.strftime('%Y-%m-%d')
        }
    }
    
    if employee_id:
        query["biometric_id"] = employee_id
    
    # Get all records for the month (include 1 day before and after for night shift handling)
    extended_start = month_start - timedelta(days=1)
    extended_end = month_end + timedelta(days=1)
    
    extended_query = {"date": {"$gte": extended_start.strftime('%Y-%m-%d'), "$lt": extended_end.strftime('%Y-%m-%d')}}
    if employee_id:
        extended_query["biometric_id"] = employee_id
    
    records = await db.attendance_records.find(
        extended_query, {"_id": 0}
    ).sort([("biometric_id", 1), ("punch_time", 1)]).to_list(length=15000)
    
    # Get all employee names and shifts from biometric_employees table
    emp_name_map = {}
    emp_shift_map = {}
    all_employees = await db.biometric_employees.find({}, {"biometric_id": 1, "name": 1, "shifts": 1}).to_list(500)
    for emp in all_employees:
        emp_name_map[emp.get("biometric_id")] = emp.get("name", f"Employee {emp.get('biometric_id')}")
        emp_shift_map[emp.get("biometric_id")] = emp.get("shifts", [])
    
    # Group records by employee first
    records_by_emp = {}
    for record in records:
        emp_id = record.get("biometric_id")
        if emp_id not in records_by_emp:
            records_by_emp[emp_id] = []
        records_by_emp[emp_id].append(record)
    
    # Process records by employee
    attendance_by_employee = {}
    
    for emp_id, emp_records in records_by_emp.items():
        emp_name = emp_name_map.get(emp_id, f"Employee {emp_id}")
        
        attendance_by_employee[emp_id] = {
            "biometric_id": emp_id,
            "name": emp_name,
            "days": {},
            "total_present": 0,
            "total_late": 0,
            "total_absent": 0
        }
        
        # Group punches by work session (a session can span midnight for night shift)
        # Logic: 
        # - Punch between 6am-8pm = Day shift IN
        # - Punch between 8pm-midnight = Night shift IN
        # - Punch between midnight-6am = Likely OUT for previous night shift
        
        i = 0
        while i < len(emp_records):
            record = emp_records[i]
            current_time = record.get("time", "")
            current_date = record.get("date", "")
            punch_time = record.get("punch_time")
            
            try:
                hour = int(current_time.split(":")[0])
            except:
                hour = 12
            
            # Determine if this is an IN or potential OUT for previous day
            # Early morning punch (midnight to 8am) - likely OUT for night shift
            if 0 <= hour < 8:
                # Check if previous day has an open night shift (IN after 6pm, no OUT)
                prev_date = (datetime.strptime(current_date, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
                
                if prev_date in attendance_by_employee[emp_id]["days"]:
                    prev_day = attendance_by_employee[emp_id]["days"][prev_date]
                    # If previous day has IN after 6pm and no OUT, this is the OUT
                    if prev_day["in_time"] and not prev_day["out_time"]:
                        try:
                            prev_in_hour = int(prev_day["in_time"].split(":")[0])
                            if prev_in_hour >= 18:  # Night shift started after 6pm
                                prev_day["out_time"] = current_time
                                prev_day["out_time_display"] = record.get("punch_time_ist")
                                prev_day["out_next_day"] = True
                                i += 1
                                continue
                        except:
                            pass
            
            # This is an IN punch - create/update day record
            work_date = current_date
            
            # Skip if outside the requested month
            if work_date < month_start.strftime('%Y-%m-%d') or work_date >= month_end.strftime('%Y-%m-%d'):
                i += 1
                continue
            
            if work_date not in attendance_by_employee[emp_id]["days"]:
                attendance_by_employee[emp_id]["days"][work_date] = {
                    "date": work_date,
                    "in_time": None,
                    "out_time": None,
                    "status": "absent",
                    "shift": "Day",
                    "is_late": False
                }
            
            day_data = attendance_by_employee[emp_id]["days"][work_date]
            
            # First punch of the day = IN time
            if day_data["in_time"] is None:
                day_data["in_time"] = current_time
                day_data["in_time_display"] = record.get("punch_time_ist")
                day_data["status"] = "present"
                
                # Get employee's assigned shifts and use the shift config
                emp_shifts = emp_shift_map.get(emp_id, [])
                shift_config = get_employee_shift_config(emp_shifts, hour)
                
                day_data["shift"] = shift_config["shift_type"]
                day_data["shift_id"] = shift_config.get("shift_id")
                day_data["expected_start"] = f"{shift_config['start_hour']:02d}:{shift_config['start_min']:02d}"
                
                # Check if late using the employee-specific shift config
                is_late = check_if_late(current_time, shift_config)
                if is_late:
                    day_data["is_late"] = True
                    day_data["status"] = "late"
            else:
                # Subsequent punch - could be OUT
                if current_time > day_data["in_time"]:
                    day_data["out_time"] = current_time
                    day_data["out_time_display"] = record.get("punch_time_ist")
            
            i += 1
    
    # Calculate totals (excluding Sundays as weekly off)
    num_days = (month_end - month_start).days
    working_days = 0
    sundays = 0
    
    for i in range(num_days):
        day = month_start + timedelta(days=i)
        if day.weekday() == 6:  # Sunday
            sundays += 1
        else:
            working_days += 1
    
    for emp_id, emp_data in attendance_by_employee.items():
        present_days = [d for d in emp_data["days"].values() if d["status"] in ["present", "late"]]
        late_days = [d for d in emp_data["days"].values() if d["status"] == "late"]
        
        emp_data["total_present"] = len(present_days)
        emp_data["total_late"] = len(late_days)
        emp_data["total_absent"] = max(0, working_days - len(present_days))  # Exclude Sundays
        emp_data["working_days"] = working_days
        emp_data["sundays"] = sundays
        emp_data["attendance_percentage"] = round(len(present_days) / working_days * 100, 1) if working_days > 0 else 0
    
    return {
        "month": month_start.strftime('%Y-%m'),
        "month_display": month_start.strftime('%B %Y'),
        "total_days": num_days,
        "working_days": working_days,
        "sundays": sundays,
        "employees": list(attendance_by_employee.values())
    }


@app.get("/api/attendance/daily")
async def get_daily_attendance(
    date: Optional[str] = None,
    current_user: User = Depends(get_optional_user)
):
    """Get daily attendance with proper night shift handling and employee-specific shift times"""
    now = get_ist_now()
    target_date = date or now.strftime('%Y-%m-%d')
    
    # Calculate previous and next dates for night shift handling
    target_dt = datetime.strptime(target_date, '%Y-%m-%d')
    prev_date = (target_dt - timedelta(days=1)).strftime('%Y-%m-%d')
    next_date = (target_dt + timedelta(days=1)).strftime('%Y-%m-%d')
    
    # Get all employees with their assigned shifts
    employees = await db.biometric_employees.find({}, {"_id": 0}).to_list(length=500)
    emp_name_map = {e.get("biometric_id"): e.get("name", f"Employee {e.get('biometric_id')}") for e in employees}
    emp_shift_map = {e.get("biometric_id"): e.get("shifts", []) for e in employees}
    
    # Get records for previous, current, and next date
    records = await db.attendance_records.find(
        {"date": {"$in": [prev_date, target_date, next_date]}},
        {"_id": 0}
    ).sort("punch_time", 1).to_list(length=10000)
    
    # Group by employee
    records_by_emp = {}
    for record in records:
        emp_id = record.get("biometric_id")
        if emp_id not in records_by_emp:
            records_by_emp[emp_id] = []
        records_by_emp[emp_id].append(record)
    
    attendance = {}
    
    # Initialize all employees as absent
    for emp_id, emp_name in emp_name_map.items():
        attendance[emp_id] = {
            "biometric_id": emp_id,
            "name": emp_name,
            "status": "absent",
            "in_time": None,
            "out_time": None,
            "is_late": False,
            "shift": None
        }
    
    for emp_id, emp_records in records_by_emp.items():
        emp_name = emp_name_map.get(emp_id, f"Employee {emp_id}")
        
        # Separate punches by date
        prev_day_punches = [r for r in emp_records if r.get("date") == prev_date]
        target_day_punches = [r for r in emp_records if r.get("date") == target_date]
        next_day_punches = [r for r in emp_records if r.get("date") == next_date]
        
        in_time = None
        out_time = None
        shift_type = None
        out_next_day = False
        
        # Check for night shift from previous day ending on target date morning
        prev_night_in = None
        prev_day_in = None
        prev_day_has_out = False
        
        for p in prev_day_punches:
            t = p.get("time", "")
            try:
                h = int(t.split(":")[0])
                if h >= 18:  # Night shift IN (6pm onwards)
                    prev_night_in = p
                elif 5 <= h < 18:  # Day shift IN
                    if prev_day_in is None:
                        prev_day_in = p
                # Check if there's any OUT punch after 17:00
                if h >= 17 and prev_day_in:
                    prev_day_has_out = True
            except:
                pass
        
        # Check if previous day's shift (day or night) extended to target day morning
        prev_shift_extends_to_today = False
        if prev_night_in:
            prev_shift_extends_to_today = True
        elif prev_day_in and not prev_day_has_out:
            # Day shift started but no OUT - check if OUT is in target day morning (overtime)
            for p in target_day_punches:
                t = p.get("time", "")
                try:
                    h = int(t.split(":")[0])
                    if 0 <= h < 8:  # Early morning punch
                        prev_shift_extends_to_today = True
                        break
                except:
                    pass
        
        # If previous day had night shift IN, check if target day morning has OUT
        if prev_night_in:
            for p in target_day_punches:
                t = p.get("time", "")
                try:
                    h = int(t.split(":")[0])
                    if 0 <= h < 8:  # Early morning = OUT for previous night
                        # Skip this punch as IN for target day
                        pass
                except:
                    pass
        
        # Process target day punches
        for p in target_day_punches:
            t = p.get("time", "")
            try:
                h = int(t.split(":")[0])
            except:
                h = 12
            
            # Very early morning punch (0-6am) - likely OUT for previous shift
            if 0 <= h < 6:
                # Check if this is OUT for previous shift (night or overtime)
                if prev_shift_extends_to_today:
                    continue  # Skip, this belongs to previous day
                else:
                    # Unusual early IN - still count it
                    if in_time is None:
                        in_time = t
                        shift_type = "Day"
            
            # Early day shift arrival (6am - 8am) - this is someone coming early
            elif 6 <= h < 8:
                # If previous shift extends AND this is before 6:30 AM, it might be OUT for previous
                if prev_shift_extends_to_today and h < 7:
                    continue  # Skip, likely belongs to previous night shift
                else:
                    # This is an early IN for day shift (e.g., 7:45 AM)
                    if in_time is None:
                        in_time = t
                        shift_type = "Day"
            
            # Day time punch (8am - 6pm)
            elif 8 <= h < 18:
                if in_time is None:
                    in_time = t
                    shift_type = "Day"
                else:
                    out_time = t
            
            # Evening punch (6pm onwards)
            elif h >= 18:
                if in_time is None:
                    in_time = t
                    shift_type = "Night"
                elif shift_type == "Day":
                    out_time = t
        
        # If night shift, check next day for OUT
        if shift_type == "Night" and not out_time:
            for p in next_day_punches:
                t = p.get("time", "")
                try:
                    h = int(t.split(":")[0])
                    if 0 <= h < 8:
                        out_time = t
                        out_next_day = True
                        break
                except:
                    pass
        
        # DOUBLE SHIFT / OVERTIME HANDLING:
        # If Day shift started but NO OUT on same day, check next day morning for OUT
        if shift_type == "Day" and in_time and not out_time:
            for p in next_day_punches:
                t = p.get("time", "")
                try:
                    h = int(t.split(":")[0])
                    if 0 <= h < 8:  # Next day early morning = OUT for previous day overtime
                        out_time = t
                        out_next_day = True
                        shift_type = "Double"  # Mark as double/overtime shift
                        break
                except:
                    pass
        
        if in_time:
            # Get employee's assigned shifts and determine which shift config to use
            emp_shifts = emp_shift_map.get(emp_id, [])
            try:
                in_h, in_m = map(int, in_time.split(":")[:2])
            except:
                in_h, in_m = 12, 0
            
            shift_config = get_employee_shift_config(emp_shifts, in_h)
            is_late = check_if_late(in_time, shift_config)
            
            # Determine shift type based on config
            shift_type = shift_config.get("shift_type", "Day")
            
            attendance[emp_id] = {
                "biometric_id": emp_id,
                "name": emp_name,
                "status": "late" if is_late else "present",
                "in_time": in_time,
                "out_time": out_time,
                "is_late": is_late,
                "shift": shift_type,
                "shift_id": shift_config.get("shift_id"),
                "expected_start": f"{shift_config['start_hour']:02d}:{shift_config['start_min']:02d}",
                "out_next_day": out_next_day
            }
    
    # Sort: present first, then late, then absent
    attendance_list = sorted(
        attendance.values(), 
        key=lambda x: (x["status"] == "absent", x["status"] == "late", x["name"])
    )
    
    return {
        "date": target_date,
        "date_display": target_dt.strftime('%d %B %Y'),
        "summary": {
            "total": len(attendance),
            "present": len([e for e in attendance_list if e["status"] == "present"]),
            "late": len([e for e in attendance_list if e["status"] == "late"]),
            "absent": len([e for e in attendance_list if e["status"] == "absent"])
        },
        "attendance": attendance_list
    }


@app.get("/api/attendance/export")
async def export_attendance(
    month: str,
    format: str = "xlsx",
    current_user: User = Depends(get_current_user)
):
    """Export monthly attendance to Excel"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get monthly data
    data = await get_monthly_attendance(month=month, current_user=current_user)
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    
    # Styling
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    present_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    late_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    absent_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    
    # Title
    ws['A1'] = f"Attendance Report - {data['month_display']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # Headers
    headers = ["Employee ID", "Employee Name", "Present Days", "Late Days", "Absent Days", "Attendance %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    row = 4
    for emp in data["employees"]:
        ws.cell(row=row, column=1, value=emp["biometric_id"])
        ws.cell(row=row, column=2, value=emp["name"])
        ws.cell(row=row, column=3, value=emp["total_present"])
        ws.cell(row=row, column=4, value=emp["total_late"])
        ws.cell(row=row, column=5, value=emp["total_absent"])
        ws.cell(row=row, column=6, value=f"{emp['attendance_percentage']}%")
        row += 1
    
    # Detailed sheet
    ws2 = wb.create_sheet("Daily Details")
    ws2['A1'] = f"Daily Attendance - {data['month_display']}"
    ws2['A1'].font = Font(bold=True, size=14)
    
    detail_headers = ["Date", "Employee ID", "Employee Name", "In Time", "Out Time", "Status", "Shift"]
    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row = 4
    for emp in data["employees"]:
        for date, day_data in sorted(emp["days"].items()):
            ws2.cell(row=row, column=1, value=date)
            ws2.cell(row=row, column=2, value=emp["biometric_id"])
            ws2.cell(row=row, column=3, value=emp["name"])
            ws2.cell(row=row, column=4, value=day_data.get("in_time_display") or day_data.get("in_time") or "-")
            ws2.cell(row=row, column=5, value=day_data.get("out_time_display") or day_data.get("out_time") or "-")
            
            status_cell = ws2.cell(row=row, column=6, value=day_data["status"].upper())
            if day_data["status"] == "present":
                status_cell.fill = present_fill
            elif day_data["status"] == "late":
                status_cell.fill = late_fill
            else:
                status_cell.fill = absent_fill
            
            ws2.cell(row=row, column=7, value=day_data.get("shift", "Day Shift"))
            row += 1
    
    # Adjust column widths
    for ws_sheet in [ws, ws2]:
        for col in ws_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_sheet.column_dimensions[column].width = max_length + 2
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"attendance_{month}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.post("/api/attendance/manual-entry")
async def add_manual_attendance(
    biometric_id: str = Body(...),
    date: str = Body(...),
    in_time: str = Body(None),
    out_time: str = Body(None),
    status: str = Body("present"),
    remarks: str = Body(None),
    current_user: User = Depends(get_current_user)
):
    """Add manual attendance entry (for corrections)"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get employee name
    employee = await db.biometric_employees.find_one({"biometric_id": biometric_id})
    emp_name = employee.get("name", f"Employee {biometric_id}") if employee else f"Employee {biometric_id}"
    
    record_id = f"manual_{biometric_id}_{date}_{get_ist_now().strftime('%H%M%S')}"
    
    await db.attendance_records.insert_one({
        "record_id": record_id,
        "biometric_id": biometric_id,
        "employee_name": emp_name,
        "date": date,
        "time": in_time,
        "punch_time_ist": f"{date} {in_time}" if in_time else None,
        "shift": "Day Shift",
        "punch_type": 0,
        "is_manual": True,
        "manual_in_time": in_time,
        "manual_out_time": out_time,
        "manual_status": status,
        "remarks": remarks,
        "created_by": current_user.user_id,
        "created_at": get_ist_now()
    })
    
    return {"success": True, "message": "Manual entry added"}


@app.get("/api/attendance/config")
async def get_attendance_config():
    """Get attendance configuration"""
    return {
        "device": ESSL_CONFIG,
        "shifts": EMPLOYEE_SHIFT_CONFIG
    }




# ============================================================================
# CNC MACHINE MONITORING API - MOVED TO: routes/cnc.py
# ============================================================================



# ==================== SALARY MANAGEMENT MODULE ====================
# MOVED TO: routes/salary.py
# ==================== END SALARY MODULE ====================


# ==================== LEAVE MANAGEMENT MODULE ====================
# MOVED TO: routes/leave.py
# ==================== END LEAVE MANAGEMENT MODULE ====================


# ==================== END LEAVE MANAGEMENT MODULE ====================


# Serve Expo web build for non-API routes (SPA fallback)
# NOTE: This MUST be the last route defined to avoid catching API routes

# ==================== HANDOFF FILE ENDPOINTS ====================
@app.get("/api/handoff/backend-zip")
async def download_backend_zip():
    """Download backend handoff zip file"""
    zip_path = "/app/backend/static_backend_handoff.zip"
    if os.path.exists(zip_path):
        return FileResponse(zip_path, filename="backend_handoff.zip", media_type="application/zip")
    raise HTTPException(status_code=404, detail="Backend zip not found")

@app.get("/api/handoff/mongo-backup-zip")
async def download_mongo_backup_zip():
    """Download MongoDB backup zip file"""
    zip_path = "/app/backend/static_mongo_backup.zip"
    if os.path.exists(zip_path):
        return FileResponse(zip_path, filename="mongo_backup.zip", media_type="application/zip")
    raise HTTPException(status_code=404, detail="Mongo backup zip not found")
# ==================== END HANDOFF ====================

@app.get("/{full_path:path}")
async def serve_expo_app(full_path: str, request: Request):
    """Serve Expo static build files. Falls back to index.html for SPA routing."""
    # Skip API routes - they should already be handled above
    if full_path.startswith("api/"):
        raise HTTPException(status_code=404, detail="Not Found")
    
    # Try to serve exact file match first
    if full_path:
        # Check for .html file
        html_file = DIST_DIR + "/" + full_path + ".html"
        if os.path.isfile(html_file):
            return FileResponse(html_file, media_type="text/html")
        
        # Check for exact file
        exact_file = DIST_DIR + "/" + full_path
        if os.path.isfile(exact_file):
            return FileResponse(exact_file)
    
    # Check if dist/index.html exists (Expo web build)
    index_file = DIST_DIR + "/index.html"
    if os.path.isfile(index_file):
        return FileResponse(index_file, media_type="text/html")
    
    # Fallback: Return SPA HTML that handles auth check and redirects appropriately
    # This ensures the app works even if Expo build is missing
    return HTMLResponse(content=get_spa_html(), status_code=200)



# Production entry point - use PORT environment variable
if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8001))
    logger.info(f"Starting server on port {port}")
    uvicorn.run(app, host="0.0.0.0", port=port, log_level="info")


