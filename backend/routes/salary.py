"""Salary Management API Routes - PIN Protected"""
from fastapi import APIRouter, HTTPException, Request, Depends
from fastapi.responses import StreamingResponse, HTMLResponse
from datetime import datetime, timedelta
from typing import Optional
import logging
import secrets
import calendar

from utils.database import get_db, get_ist_now
from utils.dependencies import User, get_current_user, get_optional_user

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/salary", tags=["salary"])

# Salary Configuration
SALARY_PIN = "110513"  # 6-digit PIN for salary access
SALARY_SESSION_EXPIRY = 3600  # 1 hour session expiry

# In-memory session storage (for production, use Redis or DB)
salary_sessions = {}


def verify_salary_session(session_token: str) -> bool:
    """Verify if salary session is valid"""
    if not session_token or session_token not in salary_sessions:
        return False
    session = salary_sessions[session_token]
    if datetime.now() > session["expires_at"]:
        del salary_sessions[session_token]
        return False
    return True


@router.post("/verify-pin")
async def verify_salary_pin(request: Request, current_user: User = Depends(get_current_user)):
    """Verify PIN for salary module access"""
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    data = await request.json()
    pin = data.get("pin", "")
    
    if pin == SALARY_PIN:
        session_token = secrets.token_hex(32)
        salary_sessions[session_token] = {
            "user_id": current_user.user_id,
            "created_at": datetime.now(),
            "expires_at": datetime.now() + timedelta(seconds=SALARY_SESSION_EXPIRY)
        }
        return {"valid": True, "session_token": session_token}
    
    return {"valid": False}


@router.get("/staff-list")
async def get_salary_staff_list(current_user: User = Depends(get_current_user)):
    """Get list of staff with their monthly salaries"""
    db = get_db()
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    staff_map = {}
    
    # Source 1: biometric_employees collection
    biometric_employees = await db.biometric_employees.find({}).to_list(length=500)
    for emp in biometric_employees:
        bio_id = str(emp.get("biometric_id") or emp.get("_id", ""))
        if bio_id and bio_id not in staff_map:
            staff_map[bio_id] = {
                "biometric_id": bio_id,
                "name": emp.get("name") or f"Staff {bio_id}",
            }
    
    # Source 2: attendance_records
    biometric_staff = await db.attendance_records.aggregate([
        {"$match": {"biometric_id": {"$exists": True, "$ne": None, "$ne": ""}}},
        {"$group": {"_id": "$biometric_id", "name": {"$first": "$name"}}},
    ]).to_list(length=500)
    
    for bio in biometric_staff:
        bio_id = str(bio.get("_id", ""))
        if bio_id and bio_id not in staff_map:
            staff_map[bio_id] = {
                "biometric_id": bio_id,
                "name": bio.get("name") or f"Staff {bio_id}",
            }
    
    # Source 3: biometric_user_links
    biometric_links = await db.biometric_user_links.find({}).to_list(length=500)
    biometric_to_user = {}
    for link in biometric_links:
        bio_id = str(link.get("biometric_id", ""))
        user_id = link.get("user_id")
        if bio_id:
            biometric_to_user[bio_id] = user_id
            if bio_id not in staff_map:
                staff_map[bio_id] = {
                    "biometric_id": bio_id,
                    "name": link.get("name") or f"Staff {bio_id}",
                }
    
    if not staff_map:
        return {"staff": []}
    
    # Get salary settings
    salary_settings = await db.salary_settings.find({}).to_list(length=500)
    salary_map = {}
    for s in salary_settings:
        bio_id = s.get("biometric_id")
        user_id = s.get("user_id")
        if bio_id:
            salary_map[str(bio_id)] = s
        if user_id:
            salary_map[str(user_id)] = s
    
    # Build final staff list
    staff_list = []
    for bio_id, info in staff_map.items():
        user_id = biometric_to_user.get(bio_id) or f"bio_{bio_id}"
        salary_info = salary_map.get(bio_id) or salary_map.get(user_id) or {}
        
        staff_list.append({
            "user_id": user_id,
            "biometric_id": bio_id,
            "name": info.get("name") or f"Staff {bio_id}",
            "email": "",
            "role": "Staff",
            "monthly_salary": salary_info.get("monthly_salary"),
        })
    
    staff_list.sort(key=lambda x: (x.get("name") or "").lower())
    return {"staff": staff_list}


@router.post("/set-staff-salary")
async def set_staff_salary(request: Request, current_user: User = Depends(get_current_user)):
    """Set monthly salary for a staff member"""
    db = get_db()
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    data = await request.json()
    user_id = data.get("user_id")
    monthly_salary = data.get("monthly_salary")
    
    if not user_id or monthly_salary is None:
        raise HTTPException(status_code=400, detail="user_id and monthly_salary required")
    
    await db.salary_settings.update_one(
        {"user_id": user_id},
        {"$set": {
            "user_id": user_id,
            "monthly_salary": float(monthly_salary),
            "updated_at": get_ist_now(),
            "updated_by": current_user.user_id
        }},
        upsert=True
    )
    
    return {"success": True}


@router.get("/staff-attendance")
async def get_staff_attendance_for_salary(
    user_id: str,
    month: int,
    year: int,
    current_user: User = Depends(get_current_user)
):
    """Get attendance data for a staff member for salary calculation"""
    db = get_db()
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get biometric ID
    biometric_id = None
    if user_id.startswith("bio_"):
        biometric_id = user_id.replace("bio_", "")
    else:
        biometric_link = await db.biometric_user_links.find_one({"user_id": user_id})
        if biometric_link:
            biometric_id = str(biometric_link.get("biometric_id", ""))
        else:
            biometric_id = user_id
    
    # Calculate date range
    actual_month = month + 1
    _, last_day = calendar.monthrange(year, actual_month)
    start_date = f"{year}-{str(actual_month).zfill(2)}-01"
    end_date = f"{year}-{str(actual_month).zfill(2)}-{str(last_day).zfill(2)}"
    
    # Get attendance records
    punches_by_date = {}
    if biometric_id:
        bio_ids_to_try = [biometric_id, str(biometric_id)]
        
        records = await db.attendance_records.find({
            "biometric_id": {"$in": bio_ids_to_try},
            "date": {"$gte": start_date, "$lte": end_date}
        }).sort([("date", 1), ("time", 1)]).to_list(length=2000)
        
        for r in records:
            date = r.get("date")
            if not date:
                continue
            
            time_str = r.get("time") or r.get("punch_time") or r.get("timestamp") or ""
            if not time_str:
                continue
            
            time_str = str(time_str).strip()
            if len(time_str) == 5:
                time_str = time_str + ":00"
            
            if date not in punches_by_date:
                punches_by_date[date] = []
            punches_by_date[date].append(time_str)
    
    # Aggregate: first punch = IN, last punch = OUT
    attendance_records = {}
    for date, times in punches_by_date.items():
        if not times:
            continue
        
        sorted_times = sorted(times)
        first_punch = sorted_times[0]
        last_punch = sorted_times[-1] if len(sorted_times) >= 2 else None
        
        if last_punch and last_punch == first_punch:
            last_punch = None
        
        attendance_records[date] = {
            "in": first_punch,
            "out": last_punch,
            "punch_count": len(sorted_times),
            "single_punch": len(sorted_times) == 1
        }
    
    # Check for edited records
    edited_records = await db.salary_attendance_edits.find({
        "user_id": user_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }).to_list(length=100)
    
    edit_map = {e["date"]: e for e in edited_records}
    
    # Build full month attendance
    day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    full_attendance = []
    
    for day in range(1, last_day + 1):
        date_str = f"{year}-{str(month + 1).zfill(2)}-{str(day).zfill(2)}"
        
        if date_str in edit_map:
            edit = edit_map[date_str]
            punch_in = edit.get("punch_in")
            punch_out = edit.get("punch_out")
            is_edited = True
        else:
            record = attendance_records.get(date_str, {})
            punch_in = record.get("in")
            punch_out = record.get("out")
            is_edited = False
        
        # Calculate hours
        total_hours = 0
        if punch_in and punch_out:
            try:
                def parse_time(t):
                    t = t.strip().upper()
                    for fmt in ["%H:%M", "%I:%M %p", "%I:%M%p", "%H:%M:%S"]:
                        try:
                            return datetime.strptime(t, fmt)
                        except:
                            continue
                    return None
                
                in_time = parse_time(punch_in)
                out_time = parse_time(punch_out)
                
                if in_time and out_time:
                    delta = out_time - in_time
                    total_hours = delta.total_seconds() / 3600
                    if total_hours < 0:
                        total_hours += 24
            except:
                pass
        
        from datetime import date as dt_date
        d = dt_date(year, month + 1, day)
        day_name = day_names[d.weekday()]
        
        full_attendance.append({
            "date": date_str,
            "day_name": day_name,
            "punch_in": punch_in,
            "punch_out": punch_out,
            "total_hours": round(total_hours, 2),
            "is_present": punch_in is not None or punch_out is not None,
            "is_edited": is_edited
        })
    
    # Check for saved record
    saved_record = await db.salary_records.find_one({
        "user_id": user_id,
        "month": month,
        "year": year
    }, {"_id": 0})
    
    return {
        "attendance": full_attendance,
        "saved_record": saved_record
    }


@router.post("/edit-attendance")
async def edit_salary_attendance(request: Request, current_user: User = Depends(get_current_user)):
    """Edit attendance record for salary calculation"""
    db = get_db()
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    data = await request.json()
    user_id = data.get("user_id")
    date = data.get("date")
    punch_in = data.get("punch_in")
    punch_out = data.get("punch_out")
    
    if not user_id or not date:
        raise HTTPException(status_code=400, detail="user_id and date required")
    
    await db.salary_attendance_edits.update_one(
        {"user_id": user_id, "date": date},
        {"$set": {
            "user_id": user_id,
            "date": date,
            "punch_in": punch_in,
            "punch_out": punch_out,
            "edited_at": get_ist_now(),
            "edited_by": current_user.user_id
        }},
        upsert=True
    )
    
    return {"success": True}


@router.post("/save-record")
async def save_salary_record(request: Request, current_user: User = Depends(get_current_user)):
    """Save salary calculation record for a staff member"""
    db = get_db()
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    data = await request.json()
    user_id = data.get("user_id")
    name = data.get("name")
    month = data.get("month")
    year = data.get("year")
    biometric_id = data.get("biometric_id")
    attendance_data = data.get("attendance_data", [])
    calculation = data.get("calculation", {})
    food_expense = data.get("food_expense", 0)
    loan_deduction = data.get("loan_deduction", 0)
    standard_hours_per_day = data.get("standard_hours_per_day", 8)
    
    if not biometric_id and user_id and user_id.startswith("bio_"):
        biometric_id = user_id.replace("bio_", "")
    
    await db.salary_records.update_one(
        {"user_id": user_id, "month": month, "year": year},
        {"$set": {
            "user_id": user_id,
            "biometric_id": biometric_id,
            "name": name,
            "month": month,
            "year": year,
            "attendance_data": attendance_data,
            "present_days": calculation.get("present_days", 0),
            "total_hours": calculation.get("total_hours_worked", 0),
            "overtime_hours": calculation.get("overtime_hours", 0),
            "base_pay": calculation.get("base_pay", 0),
            "overtime_pay": calculation.get("overtime_pay", 0),
            "food_expense": food_expense,
            "loan_deduction": loan_deduction,
            "final_salary": calculation.get("final_salary", 0),
            "standard_hours_per_day": standard_hours_per_day,
            "monthly_salary": calculation.get("monthly_salary", 0),
            "saved_at": get_ist_now(),
            "saved_by": current_user.user_id
        }},
        upsert=True
    )
    
    return {"success": True}


@router.get("/month-records")
async def get_month_salary_records(
    month: int,
    year: int,
    current_user: User = Depends(get_current_user)
):
    """Get all saved salary records for a month"""
    db = get_db()
    if (current_user.role or "").lower() not in {"admin", "super_admin", "manager", "hr"}:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    records = await db.salary_records.find(
        {"month": month, "year": year},
        {"_id": 0, "attendance_data": 0}
    ).to_list(length=500)
    
    return {"records": records}


@router.get("/download-slip")
async def download_salary_slip(
    user_id: str,
    month: int,
    year: int,
    format: str = "pdf",
    current_user: User = Depends(get_optional_user)
):
    """Download salary slip for a staff member"""
    db = get_db()
    
    record = await db.salary_records.find_one(
        {"user_id": user_id, "month": month, "year": year},
        {"_id": 0}
    )
    
    if not record:
        raise HTTPException(status_code=404, detail="Salary record not found. Please save the record first.")
    
    months = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']
    month_name = months[month]
    
    if format == "excel":
        from io import BytesIO
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel library not available")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Salary Slip"
        
        ws['A1'] = "SALARY SLIP"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"{month_name} {year}"
        ws.merge_cells('A2:F2')
        
        ws['A4'] = "Employee:"
        ws['B4'] = record.get("name", "")
        
        row = 6
        ws[f'A{row}'] = "Date"
        ws[f'B{row}'] = "Day"
        ws[f'C{row}'] = "Punch In"
        ws[f'D{row}'] = "Punch Out"
        ws[f'E{row}'] = "Hours"
        
        for cell in ws[row]:
            cell.font = Font(bold=True)
        
        for att in record.get("attendance_data", []):
            row += 1
            ws[f'A{row}'] = att.get("date", "")
            ws[f'B{row}'] = att.get("day_name", "")
            ws[f'C{row}'] = att.get("punch_in", "-")
            ws[f'D{row}'] = att.get("punch_out", "-")
            ws[f'E{row}'] = att.get("total_hours", 0)
        
        row += 2
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Present Days:"
        ws[f'B{row}'] = record.get("present_days", 0)
        
        row += 1
        ws[f'A{row}'] = "Total Hours:"
        ws[f'B{row}'] = record.get("total_hours", 0)
        
        row += 1
        ws[f'A{row}'] = "Overtime Hours:"
        ws[f'B{row}'] = record.get("overtime_hours", 0)
        
        row += 1
        ws[f'A{row}'] = "Base Pay:"
        ws[f'B{row}'] = f"₹{record.get('base_pay', 0):,.2f}"
        
        row += 1
        ws[f'A{row}'] = "Overtime Pay:"
        ws[f'B{row}'] = f"₹{record.get('overtime_pay', 0):,.2f}"
        
        row += 1
        ws[f'A{row}'] = "Food Expense:"
        ws[f'B{row}'] = f"₹{record.get('food_expense', 0):,.2f}"
        
        row += 1
        ws[f'A{row}'] = "Loan Deduction:"
        ws[f'B{row}'] = f"-₹{record.get('loan_deduction', 0):,.2f}"
        
        row += 1
        ws[f'A{row}'] = "NET SALARY:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"₹{record.get('final_salary', 0):,.2f}"
        ws[f'B{row}'].font = Font(bold=True, color="008000")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"salary_slip_{record.get('name', 'staff').replace(' ', '_')}_{month_name}_{year}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    else:
        # Generate PDF (HTML)
        attendance_rows = ""
        for att in record.get("attendance_data", []):
            status = "Present" if att.get("is_present") else "Absent"
            status_color = "#10B981" if att.get("is_present") else "#EF4444"
            attendance_rows += f"""
            <tr>
                <td>{att.get('date', '')}</td>
                <td>{att.get('day_name', '')}</td>
                <td>{att.get('punch_in', '-') or '-'}</td>
                <td>{att.get('punch_out', '-') or '-'}</td>
                <td>{att.get('total_hours', 0):.1f}</td>
                <td style="color: {status_color}; font-weight: bold;">{status}</td>
            </tr>
            """
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Salary Slip - {record.get('name')} - {month_name} {year}</title>
            <style>
                body {{ font-family: Arial, sans-serif; padding: 20px; max-width: 800px; margin: 0 auto; }}
                h1 {{ text-align: center; color: #1E293B; }}
                h2 {{ text-align: center; color: #3B82F6; font-weight: normal; }}
                .info {{ background: #f8fafc; padding: 15px; border-radius: 8px; margin-bottom: 20px; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; font-size: 12px; }}
                th {{ background: #1E293B; color: white; }}
                tr:nth-child(even) {{ background: #f9f9f9; }}
                .summary {{ background: #f8fafc; padding: 20px; border-radius: 8px; }}
                .summary-row {{ display: flex; justify-content: space-between; padding: 8px 0; border-bottom: 1px solid #e2e8f0; }}
                .final {{ font-size: 20px; font-weight: bold; color: #10B981; }}
                .deduction {{ color: #EF4444; }}
                .addition {{ color: #10B981; }}
                @media print {{ button {{ display: none; }} }}
            </style>
        </head>
        <body>
            <h1>SALARY SLIP</h1>
            <h2>{month_name} {year}</h2>
            <div class="info">
                <p><strong>Employee Name:</strong> {record.get('name', '')}</p>
                <p><strong>Monthly Salary:</strong> ₹{record.get('monthly_salary', 0):,.2f}</p>
            </div>
            <h3>Attendance Records</h3>
            <table>
                <thead><tr><th>Date</th><th>Day</th><th>Punch In</th><th>Punch Out</th><th>Hours</th><th>Status</th></tr></thead>
                <tbody>{attendance_rows}</tbody>
            </table>
            <div class="summary">
                <h3>Salary Summary</h3>
                <div class="summary-row"><span>Present Days</span><span><strong>{record.get('present_days', 0)}</strong></span></div>
                <div class="summary-row"><span>Total Working Hours</span><span>{record.get('total_hours', 0):.1f} hrs</span></div>
                <div class="summary-row"><span>Overtime Hours</span><span>{record.get('overtime_hours', 0):.1f} hrs</span></div>
                <hr>
                <div class="summary-row"><span>Base Pay</span><span>₹{record.get('base_pay', 0):,.2f}</span></div>
                <div class="summary-row"><span>Overtime Pay</span><span class="addition">+₹{record.get('overtime_pay', 0):,.2f}</span></div>
                <div class="summary-row"><span>Food Expense</span><span class="addition">+₹{record.get('food_expense', 0):,.2f}</span></div>
                <div class="summary-row"><span>Loan Deduction</span><span class="deduction">-₹{record.get('loan_deduction', 0):,.2f}</span></div>
                <hr>
                <div class="summary-row"><span style="font-size: 18px;"><strong>NET SALARY</strong></span><span class="final">₹{record.get('final_salary', 0):,.2f}</span></div>
            </div>
            <p style="text-align: center; margin-top: 30px; color: #64748B; font-size: 12px;">Generated on: {get_ist_now().strftime('%d/%m/%Y %I:%M %p')}</p>
            <button onclick="window.print()" style="display: block; margin: 20px auto; padding: 10px 30px; background: #3B82F6; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px;">Print / Save as PDF</button>
        </body>
        </html>
        """
        
        return HTMLResponse(content=html)


@router.get("/download-all")
async def download_all_salary_slips(
    month: int,
    year: int,
    format: str = "excel",
    current_user: User = Depends(get_optional_user)
):
    """Download all salary slips for a month"""
    db = get_db()
    
    records = await db.salary_records.find(
        {"month": month, "year": year},
        {"_id": 0}
    ).to_list(length=500)
    
    if not records:
        raise HTTPException(status_code=404, detail="No salary records found for this month")
    
    months = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']
    month_name = months[month]
    
    if format == "excel":
        from io import BytesIO
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel library not available")
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Salary Summary {month_name} {year}"
        
        headers = ["S.No", "Employee Name", "Present Days", "Total Hours", "OT Hours",
                   "Base Pay", "OT Pay", "Food Exp", "Loan Ded", "Net Salary"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        total_salary = 0
        for idx, record in enumerate(records, 1):
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=record.get("name", ""))
            ws.cell(row=idx+1, column=3, value=record.get("present_days", 0))
            ws.cell(row=idx+1, column=4, value=round(record.get("total_hours", 0), 1))
            ws.cell(row=idx+1, column=5, value=round(record.get("overtime_hours", 0), 1))
            ws.cell(row=idx+1, column=6, value=round(record.get("base_pay", 0), 2))
            ws.cell(row=idx+1, column=7, value=round(record.get("overtime_pay", 0), 2))
            ws.cell(row=idx+1, column=8, value=round(record.get("food_expense", 0), 2))
            ws.cell(row=idx+1, column=9, value=round(record.get("loan_deduction", 0), 2))
            ws.cell(row=idx+1, column=10, value=round(record.get("final_salary", 0), 2))
            total_salary += record.get("final_salary", 0)
        
        row = len(records) + 2
        ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=10, value=round(total_salary, 2)).font = Font(bold=True)
        
        ws.column_dimensions['B'].width = 25
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws.column_dimensions[col].width = 12
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"salary_summary_{month_name}_{year}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    else:
        # Generate HTML
        rows_html = ""
        total_salary = 0
        for idx, record in enumerate(records, 1):
            total_salary += record.get("final_salary", 0)
            rows_html += f"""
            <tr>
                <td>{idx}</td>
                <td>{record.get('name', '')}</td>
                <td>{record.get('present_days', 0)}</td>
                <td>{record.get('total_hours', 0):.1f}</td>
                <td>{record.get('overtime_hours', 0):.1f}</td>
                <td>₹{record.get('base_pay', 0):,.2f}</td>
                <td>₹{record.get('overtime_pay', 0):,.2f}</td>
                <td>₹{record.get('food_expense', 0):,.2f}</td>
                <td>₹{record.get('loan_deduction', 0):,.2f}</td>
                <td style="font-weight: bold; color: #10B981;">₹{record.get('final_salary', 0):,.2f}</td>
            </tr>
            """
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Salary Summary - {month_name} {year}</title>
            <style>
                body {{ font-family: Arial, sans-serif; padding: 20px; }}
                h1 {{ text-align: center; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; font-size: 12px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: right; }}
                th {{ background: #1E293B; color: white; }}
                td:nth-child(2) {{ text-align: left; }}
                tr:nth-child(even) {{ background: #f9f9f9; }}
                .total {{ font-weight: bold; background: #f0f9ff !important; }}
                @media print {{ button {{ display: none; }} }}
            </style>
        </head>
        <body>
            <h1>Salary Summary - {month_name} {year}</h1>
            <p>Total Staff: {len(records)} | Total Salary: ₹{total_salary:,.2f}</p>
            <table>
                <thead><tr><th>#</th><th style="text-align: left;">Employee</th><th>Days</th><th>Hours</th><th>OT Hrs</th><th>Base Pay</th><th>OT Pay</th><th>Food</th><th>Loan</th><th>Net Salary</th></tr></thead>
                <tbody>
                    {rows_html}
                    <tr class="total"><td colspan="9" style="text-align: right;">TOTAL</td><td style="color: #10B981;">₹{total_salary:,.2f}</td></tr>
                </tbody>
            </table>
            <button onclick="window.print()" style="display: block; margin: 20px auto; padding: 10px 30px; background: #3B82F6; color: white; border: none; border-radius: 8px; cursor: pointer;">Print / Save as PDF</button>
        </body>
        </html>
        """
        
        return HTMLResponse(content=html)
